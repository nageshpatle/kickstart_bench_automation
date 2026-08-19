from __future__ import annotations

import os
from pathlib import Path
import threading
import time

import pytest
from openpyxl import load_workbook

from sid_bench_gui import (
    MEAS_HEADERS, RUN_HEADERS, SweepWorker, WorkbookStore, calculate_measurement,
    capture_root_for_source, efficiency_axis_bounds, parse_points,
)
from sid_instruments import InstrumentHub, InstrumentSnapshot, SupplyChannel, VisaInstrument


@pytest.fixture(autouse=True)
def clean_qt_state(tmp_path, monkeypatch):
    # No GUI test may ever touch the operator's configured campaign workbook.
    monkeypatch.setenv("KICKSTART_WORKBOOK_PATH", str(tmp_path / "pytest_campaign.xlsx"))
    config_file = Path(__file__).resolve().parent / "bench_config.json"
    orig_config = config_file.read_text(encoding="utf-8") if config_file.exists() else None
    yield
    if orig_config is not None and config_file.exists():
        try:
            config_file.write_text(orig_config, encoding="utf-8")
        except Exception:
            pass
    try:
        from PyQt6 import QtCore, QtWidgets
        app = QtWidgets.QApplication.instance()
        if app is not None:
            pool = QtCore.QThreadPool.globalInstance()
            if pool is not None:
                pool.waitForDone(1000)
            app.processEvents()
    except Exception:
        pass




def run_record(run_id: str, source: str = "Simulation") -> dict:
    record = {name: "" for name in RUN_HEADERS}
    record.update({"RunID": run_id, "Created": "2026-08-16T00:00:00Z", "Status": "Valid", "DataSource": source, "Mode": "Continuous", "VinTarget_V": 48.0, "ModulationLabel": "test-custom-profile", "Frequency_Hz": 100000.0})
    return record


def test_simulation_and_hardware_use_separate_workbooks_and_combined_history(tmp_path: Path):
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtCore, QtWidgets
    from sid_bench_gui import MainWindow, WorkbookStore

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.hardware_store = WorkbookStore(tmp_path / "hw.xlsx")
    window.simulation_store = WorkbookStore(tmp_path / "sim.xlsx")
    window.store = window.hardware_store
    assert window.hardware_store.path != window.simulation_store.path
    window.hardware_store.create_run({**run_record("HW-RUN", "Hardware"), "CampaignName": "Measured"})
    window.simulation_store.create_run({**run_record("SIM-RUN", "Simulation"), "CampaignName": "Demo"})
    measurement = {name: "" for name in MEAS_HEADERS}
    window.simulation_store.append_measurement({
        **measurement, "PointID": "SIM-P1", "RunID": "SIM-RUN", "Status": "Valid",
        "DataSource": "Simulation", "Iout_A": 2.0,
        "EfficiencyConverter_pct": 97.8, "EfficiencySystem_pct": 96.9,
        "LossConverter_W": 2.1, "LossSystem_W": 2.7,
        "PinConverter_W": 100.0, "Pout_W": 97.9, "Paux_W": 0.6,
    })
    window.hardware_store.append_measurement({
        **measurement, "PointID": "HW-P1", "RunID": "HW-RUN", "Status": "Valid",
        "DataSource": "Hardware", "Iout_A": 2.0,
        "EfficiencyConverter_pct": 97.5, "EfficiencySystem_pct": 96.5,
        "LossConverter_W": 2.2, "LossSystem_W": 2.9,
        "PinConverter_W": 101.0, "Pout_W": 98.8, "Paux_W": 0.7,
    })
    window._load_history()
    sources = {window.history_table.item(row, 4).text() for row in range(window.history_table.rowCount())}
    assert sources == {"Hardware", "Simulation"}
    sim_row = next(row for row in range(window.history_table.rowCount()) if window.history_table.item(row, 4).text() == "Simulation")
    hw_row = next(row for row in range(window.history_table.rowCount()) if window.history_table.item(row, 4).text() == "Hardware")
    assert window._store_for_history_row(sim_row) is window.simulation_store
    assert window._store_for_history_row(hw_row) is window.hardware_store
    window.history_table.selectRow(sim_row)
    window._history_selection_changed()
    assert len(window.comp_plot_widget.listDataItems()) == 2
    assert window.comp_metric_combo.itemText(1) == "Loss (W)"
    assert window.comp_metric_combo.itemText(2) == "Power (W)"
    assert window.comp_legend.offset == (-12, -12)
    assert vars(window.comp_legend)["_GraphicsWidgetAnchor__parentAnchor"] == (1, 1)
    for metric_index, expected_count in ((0, 2), (1, 2), (2, 3)):
        window.comp_metric_combo.setCurrentIndex(metric_index)
        window._history_selection_changed()
        items = window.comp_plot_widget.listDataItems()
        assert len(items) == expected_count
        expected_colors = ["#002676", "#d97706", "#0f766e"][:expected_count]
        assert [item.opts["pen"].color().name() for item in items] == expected_colors
        assert all(item.opts["pen"].style() == QtCore.Qt.PenStyle.SolidLine for item in items)

    # Multi-run overlays retain metric color families and use marker identity for runs.
    window.history_table.selectAll()
    window.comp_metric_combo.setCurrentIndex(2)
    window._history_selection_changed()
    items = window.comp_plot_widget.listDataItems()
    assert len(items) == 6
    assert [item.opts["symbol"] for item in items[:3]] == [items[0].opts["symbol"]] * 3
    assert [item.opts["symbol"] for item in items[3:]] == [items[3].opts["symbol"]] * 3
    assert items[0].opts["symbol"] != items[3].opts["symbol"]
    assert items[0].opts["pen"].color().hue() == items[3].opts["pen"].color().hue()
    assert items[1].opts["pen"].color().hue() == items[4].opts["pen"].color().hue()
    assert items[2].opts["pen"].color().hue() == items[5].opts["pen"].color().hue()
    window.close()


def test_busy_live_workbook_uses_one_pending_state_without_fallback_files(tmp_path: Path, monkeypatch):
    target = tmp_path / "campaign.xlsx"
    store = WorkbookStore(target)
    store.create_run(run_record("BASE", "Hardware"))
    real_replace = os.replace

    def reject_live_replace(source, destination):
        if Path(destination) == target:
            raise PermissionError("simulated Excel lock")
        return real_replace(source, destination)

    monkeypatch.setattr(os, "replace", reject_live_replace)
    assert store.create_run(run_record("QUEUED", "Hardware")) is True
    assert store.pending_path.exists()
    assert {run["RunID"] for run in store.list_runs()} == {"BASE", "QUEUED"}
    assert list(tmp_path.glob("fallback_*.xlsx")) == []


def test_duplicate_points_are_scoped_to_run_mode(tmp_path: Path):
    store = WorkbookStore(tmp_path / "mode_identity.xlsx")
    store.create_run(run_record("STEP-RUN"))
    record = {name: "" for name in MEAS_HEADERS}
    record.update({
        "PointID": "STEP-P001", "RunID": "STEP-RUN", "Status": "Valid",
        "DataSource": "Simulation", "Mode": "Step Current", "VinTarget_V": 48.0,
        "ModulationLabel": "same-name", "Frequency_Hz": 100000.0, "RequestedIout_A": 2.0,
    })
    store.append_measurement(record)
    base = {"DataSource": "Simulation", "VinTarget_V": 48.0, "ModulationLabel": "same-name", "Frequency_Hz": 100000.0}
    assert store.find_duplicates({**base, "Mode": "Step Current"}, [2.0])
    assert store.find_duplicates({**base, "Mode": "Pulse"}, [2.0]) == []


def test_full_run_names_include_mode_and_minimal_parameters():
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.test_name.setText("test1")
    window.btn_mode_cont.click()
    window.cont_stop.setValue(min(20.0, window.cap_val))
    continuous = window._collect_settings()
    assert continuous["run_record"]["CampaignName"].startswith("test1_Continuous_")
    window.btn_mode_pulse.click()
    window.pulse_stop.setValue(min(20.0, window.cap_val))
    pulse = window._collect_settings()
    assert "test1_Pulse_" in pulse["run_record"]["CampaignName"]
    assert pulse["mode"] == "Pulse"
    window.close()


def test_switching_frequency_is_shown_in_khz_and_stored_in_hz():
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    assert window.frequency.suffix() == " kHz"
    assert window.frequency.value() == pytest.approx(float(window.config["frequency_hz"]) / 1000.0)
    window.frequency.setValue(200.0)
    assert window.frequency.text() == "200 kHz"
    settings = window._collect_settings(manual=True)
    assert settings["frequency"] == 200000.0
    assert settings["run_record"]["Frequency_Hz"] == 200000.0
    window.close()


def test_manual_action_error_releases_busy_controls():
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.simulation.setChecked(True)
    window.btn_mode_direct.click()
    window.point_action_busy = True
    window._manual_active_task = True
    window.update_enabled_states()
    assert not window.btn_direct_set.isEnabled()
    assert window.btn_direct_zero.isEnabled()
    window._manual_action_failed("simulated save failure", window._manual_point_token)
    assert window.btn_direct_set.isEnabled()
    assert window.btn_direct_zero.isEnabled()
    window.close()



def test_parse_points_supports_ranges_and_values():
    assert parse_points("0:6:2, 9.5") == [0.0, 2.0, 4.0, 6.0, 9.5]


def test_calculation_preserves_signed_validity_and_auxiliary_power():
    pa = InstrumentSnapshot("pa", {"vin": 48.0, "iin": 10.0, "vout": 12.0})
    load = InstrumentSnapshot("load", {"current": 38.0, "voltage": 12.0})
    psu = InstrumentSnapshot("psu", {
        "ch1_voltage": 5.0, "ch1_current": 0.1,
        "ch2_voltage": 12.0, "ch2_current": 0.05,
        "ch3_voltage": 6.0, "ch3_current": 0.2,
    })
    # CH1 not contributing to loss, CH2 not contributing, CH3 contributing to loss
    channels = [
        SupplyChannel(1, "Vdrv_A", True, True, False),
        SupplyChannel(2, "Vdrv_B", True, True, False),
        SupplyChannel(3, "Vdrv_C", True, True, True),
    ]
    result, warnings = calculate_measurement(pa, load, psu, channels, (24.0, 16.0, 3.4))
    assert warnings == []
    assert result["Pout_W"] == 456.0
    assert result["PinConverter_W"] == 480.0

    # Raw individual channel measurements recorded regardless of contributes_loss
    assert result["Vdrv_A_V"] == 5.0
    assert result["Idrv_A_A"] == 0.1
    assert result["Pdrv_A_W"] == pytest.approx(0.5)

    assert result["Vdrv_B_V"] == 12.0
    assert result["Idrv_B_A"] == 0.05
    assert result["Pdrv_B_W"] == pytest.approx(0.6)

    assert result["Vdrv_C_V"] == 6.0
    assert result["Idrv_C_A"] == 0.2
    assert result["Pdrv_C_W"] == pytest.approx(1.2)

    # Paux_W only includes CH3 (1.2 W)
    assert result["Paux_W"] == pytest.approx(1.2)

    # Loss equations
    assert result["LossConverter_W"] == pytest.approx(480.0 - 456.0) # 24.0 W
    assert result["LossSystem_W"] == pytest.approx(24.0 + 1.2)       # 25.2 W
    assert result["EfficiencyConverter_pct"] == 95.0


def test_system_efficiency_is_missing_without_aux_supply_data():
    pa = InstrumentSnapshot("pa", {"vin": 48.0, "iin": 2.0, "vout": 12.0})
    load = InstrumentSnapshot("load", {"current": 7.5, "voltage": 12.0})
    result, warnings = calculate_measurement(pa, load, None, [])
    assert any("auxiliary loss is incomplete" in warning for warning in warnings)
    assert result["EfficiencyConverter_pct"] == pytest.approx(93.75)
    assert result["EfficiencySystem_pct"] is None



def test_invalid_sign_does_not_become_plausible_data():
    result, warnings = calculate_measurement(InstrumentSnapshot("pa", {"vin": 48.0, "iin": -1.0, "vout": 12.0}), InstrumentSnapshot("load", {"current": 10.0}), None, [], (24.0, 16.0, 3.4))
    assert result == {}
    assert "sign" in warnings[0]


def test_simulated_stale_measurement_is_invalid():
    hub = InstrumentHub(True, {"simulation_scenario": "Stale measurement"})
    pa = hub.instruments["pa"].read_snapshot()
    load = hub.instruments["load"].read_snapshot()
    result, warnings = calculate_measurement(pa, load, None, [], (24.0, 16.0, 3.4))
    assert result == {}
    assert "stale" in warnings[0].lower()


def test_release_sends_local_and_closes_session():
    class FakeSession:
        def __init__(self):
            self.closed = False
            self.writes = []
            self.timeout = 0
        def query(self, command):
            return "VENDOR,MODEL,SERIAL,1.0"
        def write(self, command):
            self.writes.append(command)
        def close(self):
            self.closed = True
    class FakeResourceManager:
        def __init__(self, session): self.fake = session
        def open_resource(self, address): return self.fake
    class FakeManager:
        def __init__(self, session): self.rm = FakeResourceManager(session)
        def open(self): return self.rm
    session = FakeSession()
    instrument = VisaInstrument(FakeManager(session), "USB::TEST", ("MODEL",))
    instrument.connect(persistent=True)
    instrument.release()
    assert session.closed
    assert session.writes[0] in {"SYSTem:LOCal", "SYST:LOC"}
    assert not instrument.connected


def test_workbook_append_and_supersede(tmp_path: Path):
    store = WorkbookStore(tmp_path / "campaign.xlsx")
    store.create_run(run_record("R1"))
    base = {name: "" for name in MEAS_HEADERS}
    base.update({"PointID": "P1", "RunID": "R1", "Status": "Valid", "DataSource": "Simulation", "VinTarget_V": 48.0, "ModulationLabel": "test-custom-profile", "Frequency_Hz": 100000.0, "RequestedIout_A": 20.0})
    store.append_measurement(base)
    store.create_run(run_record("R2"))
    store.append_measurement({**base, "PointID": "P2", "RunID": "R2"}, "supersede")
    sheet = load_workbook(store.path, data_only=True)["Measurements"]
    assert sheet.cell(2, MEAS_HEADERS.index("Status") + 1).value == "Superseded"
    assert sheet.cell(3, MEAS_HEADERS.index("SupersedesPointID") + 1).value == "P1"
    assert store.path.with_suffix(".xlsx.bak").exists()


def test_scope_failure_is_nonfatal_and_measurement_is_saved(tmp_path: Path):
    hub = InstrumentHub(True, {"simulation_scenario": "Scope capture failure"})
    store = WorkbookStore(tmp_path / "campaign.xlsx")
    run_id = "SIM-RUN"
    settings = {"run_id": run_id, "run_record": run_record(run_id), "points": [10.0], "capture_points": {10.0}, "mode": "Continuous", "settle": 0.0, "dwell": 0.1, "sample_window": 0.0, "sample_count": 1, "cooldown": 0.0, "working_cap": 70.0, "vin_target": 48.0, "modulation": "test-custom-profile", "frequency": 100000.0, "dimensions": (24.0, 16.0, 3.4), "supply_channels": [SupplyChannel(3, "gate", True, True, True)], "psu_required": False, "data_source": "Simulation", "duplicate_action": "keep"}
    SweepWorker(hub, store, settings).run()
    sheet = load_workbook(store.path, data_only=True)["Measurements"]
    row = {name: sheet.cell(2, index + 1).value for index, name in enumerate(MEAS_HEADERS)}
    assert row["Status"] == "Valid"
    assert row["ScopeCaptureStatus"] == "Failed"
    assert "scope capture failure" in row["ScopeCaptureError"].lower()


def test_pause_is_dropped_and_two_run_controls_used():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()

    # Verify Pause is dropped from the GUI
    assert not hasattr(window, "strip_pause")

    # Verify the 2 run controls: Start and Stop & Return to Zero
    assert hasattr(window, "run_sweep_btn")
    assert hasattr(window, "stop_sweep_btn")
    assert hasattr(window, "strip_stop_btn")
    assert "STOP & RETURN TO ZERO" in window.stop_sweep_btn.text()
    assert "STOP & RETURN TO ZERO" in window.strip_stop_btn.text()
    window.close()



def test_required_device_failure_aborts_and_turns_load_off(tmp_path: Path):
    hub = InstrumentHub(True, {"simulation_scenario": "Required device failure"})
    store = WorkbookStore(tmp_path / "abort.xlsx")
    run_id = "ABORT-RUN"
    settings = {"run_id": run_id, "run_record": run_record(run_id), "points": [10.0, 20.0], "capture_points": set(), "mode": "Continuous", "settle": 0.0, "dwell": 0.0, "sample_window": 0.0, "sample_count": 1, "cooldown": 0.0, "working_cap": 70.0, "vin_target": 48.0, "modulation": "test-custom-profile", "frequency": 100000.0, "dimensions": (24.0, 16.0, 3.4), "supply_channels": [], "psu_required": False, "data_source": "Simulation", "duplicate_action": "keep"}
    SweepWorker(hub, store, settings).run()
    run_sheet = load_workbook(store.path, data_only=True)["Runs"]
    assert run_sheet.cell(2, RUN_HEADERS.index("Status") + 1).value == "Aborted"
    assert hub.environment.load_enabled is False


def test_simulated_scope_creates_valid_artifacts(tmp_path: Path):
    hub = InstrumentHub(True)
    png, csv = tmp_path / "capture.png", tmp_path / "capture.csv"
    channels, rows = hub.instruments["scope"].capture(png, csv)
    assert png.read_bytes().startswith(b"\x89PNG")
    assert csv.read_text(encoding="utf-8").startswith("CH1_Time_s")
    assert channels == [1, 2, 3, 4]
    assert rows == 1000


def test_e36312a_scpi_queries_use_channel_list_syntax():
    class FakePsuSession:
        def __init__(self):
            self.queries = []
            self.closed = False
        def query(self, command: str) -> str:
            self.queries.append(command)
            if "*IDN?" in command:
                return "Keysight Technologies,E36312A,TESTPSU001,1.0.0"
            if "VOLT" in command:
                return "6.002"
            if "CURR" in command:
                return "0.150"
            if "OUTP" in command:
                return "1"
            return "0"
        def close(self):
            self.closed = True
        def write(self, command: str) -> None:
            pass

    class FakeManager:
        def __init__(self, session):
            self.sess = session
        def open(self):
            class RM:
                def __init__(self, s): self.s = s
                def open_resource(self, _): return self.s
            return RM(self.sess)

    sess = FakePsuSession()
    from sid_instruments import E36312A
    psu = E36312A(FakeManager(sess), "USB::PSU", ("E36312",))
    psu.connect(persistent=True)
    snap = psu.read_snapshot(channels=[1, 3])
    psu.release()

    assert any("MEASure:VOLTage? (@1)" in q or "MEAS:VOLT? (@1)" in q for q in sess.queries)
    assert any("MEASure:VOLTage? (@3)" in q or "MEAS:VOLT? (@3)" in q for q in sess.queries)
    assert any("OUTPut? (@1)" in q or "OUTP? (@1)" in q for q in sess.queries)
    assert snap.values["ch1_voltage"] == 6.002
    assert snap.values["ch3_enabled"] is True


def test_sim_psu_reports_channel_enabled_states():
    hub = InstrumentHub(True)
    snap = hub.instruments["psu"].read_snapshot(channels=[1, 2, 3])
    assert "ch1_enabled" in snap.values
    assert "ch2_enabled" in snap.values
    assert "ch3_enabled" in snap.values
    assert snap.values["ch3_enabled"] is True
    assert snap.values["ch1_enabled"] is False


def test_instrument_card_and_bench_check():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, InstrumentCard
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    hub = InstrumentHub(True)
    card = InstrumentCard("load", "Load", [("Iout", "A")], lambda: hub.instruments["load"])
    card.read_once()
    QtCore.QThreadPool.globalInstance().waitForDone(2000)
    QtWidgets.QApplication.processEvents()
    assert card.last_snapshot is not None
    assert card.status_badge.text() in {"Ready", "Updated", "Connected"}




    
    window = MainWindow()
    window.simulation.setChecked(True)
    window._check_entire_bench()
    QtCore.QThreadPool.globalInstance().waitForDone(2000)
    QtWidgets.QApplication.processEvents()
    assert "● READY" in window.readiness_status.text() or "● NOT READY" in window.readiness_status.text()
    window.close()


def test_mode_dependent_execution_box_visibility():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.show()
    
    # Mode 0 = SET CURRENT -> sweep_exec_box hidden
    window._mode_selected(0)
    assert window.sweep_exec_box.isHidden()
    
    # Mode 1 = STEP CURRENT -> sweep_exec_box hidden
    window._mode_selected(1)
    assert window.sweep_exec_box.isHidden()

    # Mode 2 = CONTINUOUS -> sweep_exec_box shown
    window._mode_selected(2)
    assert not window.sweep_exec_box.isHidden()
    
    # Mode 3 = PULSE -> sweep_exec_box shown
    window._mode_selected(3)
    assert not window.sweep_exec_box.isHidden()
    window.close()



def test_supply_card_embedded_controls():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, SupplyCard
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.simulation.setChecked(True)
    card = window.supply_card
    
    assert len(card.channel_controls) == 3
    card.channel_controls[0]["desired_out"].setChecked(True)
    card.channel_controls[0]["voltage"].setValue(5.25)
    card._apply()
    QtCore.QThreadPool.globalInstance().waitForDone(2000)
    QtWidgets.QApplication.processEvents()

    assert window.hub.environment.channels[1]["enabled"] is True
    assert "ON" in card.channel_controls[0]["live_lbl"].text()
    assert "5.25" in card.channel_controls[0]["live_lbl"].text() or window.hub.environment.channels[1]["voltage"] == 5.25
    window.close()



def test_blank_naming_auto_generates_defaults():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.test_name.setText("")

    settings = window._collect_settings(manual=True)
    assert settings["run_record"]["CampaignName"].startswith("Test_")
    assert window.test_name.text().startswith("Test_")
    window.close()



def test_history_is_read_only_and_idle_run_strip_is_hidden():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    assert window.history_table.editTriggers() == QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers
    assert not window.run_strip.isVisible()
    window.close()


def test_workbook_uses_readable_groups_and_frozen_identifiers(tmp_path: Path):
    from openpyxl import load_workbook
    from sid_bench_gui import WorkbookStore

    path = tmp_path / "campaign.xlsx"
    store = WorkbookStore(path)
    store.create_run(run_record("STYLE"))
    store.append_measurement({
        "PointID": "STYLE-P001", "RunID": "STYLE", "Timestamp": "2026-08-17T00:00:00Z",
        "Status": "Valid", "DataSource": "Hardware", "Iout_A": 10.0,
        "EfficiencyConverter_pct": 97.5,
    })
    store.finish_run("STYLE", "Valid")

    wb = load_workbook(path)
    # Column A is frozen horizontally, row 1 header frozen vertically
    assert wb["Runs"].freeze_panes == "B2"
    assert wb["Measurements"].freeze_panes == "B2"
    assert wb["Events"].freeze_panes == "A2"
    assert wb["Runs"].column_dimensions["A"].width >= 20
    assert wb["Measurements"].column_dimensions["A"].width >= 20
    assert wb["Plots"].max_row == 2
    assert len(wb["Plots"]._charts) == 1


def test_aux_supply_logging_and_workbook_freeze_and_upgrade(tmp_path: Path):
    from openpyxl import Workbook, load_workbook
    from sid_bench_gui import MEAS_HEADERS, RUN_HEADERS, WorkbookStore

    wb_path = tmp_path / "legacy_campaign.xlsx"
    
    # 1. Create a legacy workbook with old columns (missing Vdrv_A_V, etc.)
    legacy_wb = Workbook()
    legacy_wb.remove(legacy_wb.active)
    
    old_meas_headers = [
        "PointID", "RunID", "Timestamp", "Status", "DataSource", "Mode", "VinTarget_V",
        "ModulationLabel", "Frequency_Hz", "RequestedIout_A", "Iout_A", "Vin_V", "Iin_A",
        "Vout_V", "PinConverter_W", "Pout_W", "Paux_W", "LossConverter_W", "LossSystem_W",
        "EfficiencyConverter_pct", "EfficiencySystem_pct", "PowerDensity_W_per_in3",
        "SupplyMeasurements", "Quality", "Warning", "ScopeCaptureStatus", "ScopeCaptureError",
        "ScopePNG", "ScopeCSV", "SupersedesPointID",
    ]
    old_run_headers = [
        "RunID", "CampaignName", "Created", "Completed", "Status", "DataSource", "Mode", "VinTarget_V",
        "ModulationLabel", "Frequency_Hz", "ModulationMetadata", "SupplyConfiguration",
        "Length_mm", "Width_mm", "Height_mm", "WorkingCap_A", "Notes", "InstrumentIdentities",
        "FPGASnapshotStatus", "FPGASnapshot", "Warnings", "SupersedesRunID",
    ]
    
    m_sheet = legacy_wb.create_sheet("Measurements")
    m_sheet.append(old_meas_headers)
    m_sheet.freeze_panes = "F2" # Old freeze pane with multiple frozen columns
    m_sheet.append([
        "LEG-P001", "LEG-RUN", "2026-08-16T12:00:00Z", "Valid", "Hardware", "Continuous", 48.0,
        "Profile-A", 100000.0, 10.0, 10.0, 48.0, 2.5,
        12.0, 120.0, 115.0, 1.5, 5.0, 6.5,
        95.83, 94.65, 50.0,
        "{}", "Valid", "", "Skipped", "",
        "", "", "",
    ])
    
    r_sheet = legacy_wb.create_sheet("Runs")
    r_sheet.append(old_run_headers)
    r_sheet.freeze_panes = "F2"
    r_sheet.append([
        "LEG-RUN", "Legacy Test", "2026-08-16T12:00:00Z", "2026-08-16T12:05:00Z", "Valid", "Hardware", "Continuous", 48.0,
        "Profile-A", 100000.0, "", "[]",
        24.0, 16.0, 3.4, 70.0, "Historical test", "{}",
        "Captured", "{}", "", "",
    ])
    legacy_wb.save(wb_path)
    
    # 2. Open with WorkbookStore - should automatically upgrade headers, fix freeze panes to B2, and preserve data
    store = WorkbookStore(wb_path)
    meas_rows = store.get_run_measurements("LEG-RUN")
    assert len(meas_rows) == 1
    assert meas_rows[0]["PointID"] == "LEG-P001"
    assert meas_rows[0]["Vin_V"] == 48.0
    assert meas_rows[0]["Paux_W"] == 1.5
    # Newly added fields exist in the dict and are None / blank (not fabricated)
    assert meas_rows[0]["Vdrv_A_V"] in (None, "")
    assert meas_rows[0]["Pdrv_C_W"] in (None, "")
    
    runs = store.list_runs()
    assert len(runs) == 1
    assert runs[0]["RunID"] == "LEG-RUN"
    assert runs[0]["AuxA_Included"] in (None, "")
    
    # 3. Append a new measurement with full Vdrv data and verify Excel structure
    new_record = {
        "PointID": "LEG-P002", "RunID": "LEG-RUN", "Timestamp": "2026-08-17T05:00:00Z",
        "Status": "Valid", "DataSource": "Hardware", "Mode": "Continuous", "VinTarget_V": 48.0,
        "ModulationLabel": "Profile-A", "Frequency_Hz": 100000.0, "RequestedIout_A": 20.0,
        "Iout_A": 20.0, "Vin_V": 48.0, "Iin_A": 5.0, "Vout_V": 12.0, "PinConverter_W": 240.0, "Pout_W": 230.0,
        "Vdrv_A_V": 5.01, "Idrv_A_A": 0.08, "Pdrv_A_W": 0.4008,
        "Vdrv_B_V": 0.0, "Idrv_B_A": 0.0, "Pdrv_B_W": 0.0,
        "Vdrv_C_V": 6.02, "Idrv_C_A": 0.15, "Pdrv_C_W": 0.903,
        "Paux_W": 0.903, "LossConverter_W": 10.0, "LossSystem_W": 10.903,
        "EfficiencyConverter_pct": 95.83, "EfficiencySystem_pct": 95.47, "PowerDensity_W_per_in3": 100.0,
        "SupplyMeasurements": {}, "Quality": "Valid", "Warning": "",
        "ScopeCaptureStatus": "Skipped", "ScopeCaptureError": "", "ScopePNG": "", "ScopeCSV": "",
    }
    store.append_measurement(new_record)
    
    reloaded_wb = load_workbook(wb_path, data_only=True)
    assert reloaded_wb["Measurements"].freeze_panes == "B2"
    assert reloaded_wb["Runs"].freeze_panes == "B2"
    
    # Verify header row matches MEAS_HEADERS
    loaded_headers = [reloaded_wb["Measurements"].cell(1, c).value for c in range(1, len(MEAS_HEADERS) + 1)]
    assert loaded_headers == MEAS_HEADERS
    
    # Check row 2 (legacy) has empty Vdrv
    vdrv_a_idx = MEAS_HEADERS.index("Vdrv_A_V") + 1
    assert reloaded_wb["Measurements"].cell(2, vdrv_a_idx).value in (None, "")
    
    # Check row 3 (new) has exact measured Vdrv_A_V, Idrv_A_A, Pdrv_A_W
    assert reloaded_wb["Measurements"].cell(3, vdrv_a_idx).value == pytest.approx(5.01)
    pdrv_c_idx = MEAS_HEADERS.index("Pdrv_C_W") + 1
    assert reloaded_wb["Measurements"].cell(3, pdrv_c_idx).value == pytest.approx(0.903)



def test_generate_points_inclusive_and_cap_validation():
    from sid_bench_gui import generate_points
    points, summary = generate_points(0.0, 40.0, 5.0, 70.0)
    assert points == [0.0, 5.0, 10.0, 15.0, 20.0, 25.0, 30.0, 35.0, 40.0]
    assert "0 A → 40 A in 5 A steps · 9 points" in summary

    # 2-point pulse test (step >= span)
    pts_2pt, summary_2pt = generate_points(0.0, 60.0, 80.0, 70.0)
    assert pts_2pt == [0.0, 60.0]
    assert summary_2pt == "0 A → 60 A · 2 points"

    # Single point
    pts_single, summary_single = generate_points(10.0, 10.0, 2.0, 70.0)
    assert pts_single == [10.0]
    assert summary_single == "10 A · 1 point"

    # Rejection above current cap
    with pytest.raises(ValueError, match="exceeds maximum allowed load current"):
        generate_points(0.0, 80.0, 5.0, 70.0)

    # Rejection invalid step
    with pytest.raises(ValueError, match="strictly positive"):
        generate_points(0.0, 10.0, 0.0, 70.0)



def test_mode_indicator_states():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import ModeIndicator
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    indicator = ModeIndicator()
    
    indicator.set_mode("Continuous")
    assert indicator.mode == "Continuous"
    
    indicator.set_mode("Pulse")
    assert indicator.mode == "Pulse"
    
    indicator.set_mode("Manual")
    assert indicator.mode == "Manual"


def test_single_emergency_stop_button_only():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    
    # Assert only one emergency stop button exists in the main window
    buttons = window.findChildren(QtWidgets.QPushButton, "emergency_stop_btn")
    assert len(buttons) == 1
    assert buttons[0].text() == "LOAD OFF"
    assert "Esc" in buttons[0].shortcut().toString()
    window.close()


def test_smart_enabled_states_during_run():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    
    # In hardware mode before verifying load, run sweep is enabled and clickable with informative tooltip
    window.simulation.setChecked(False)
    window.chk_load.setChecked(False)
    window.update_enabled_states()
    assert window.run_sweep_btn.isEnabled()
    assert "Hardware write locked" in window.run_sweep_btn.toolTip()
    
    # Once verified, run sweep is enabled
    window.chk_load.setChecked(True)
    window.update_enabled_states()
    assert window.run_sweep_btn.isEnabled()
    
    # In simulation mode, run sweep is enabled without hardware gate
    window.simulation.setChecked(True)
    window.chk_load.setChecked(False)
    window.update_enabled_states()
    assert window.run_sweep_btn.isEnabled()
    window.close()


def test_low_current_verification_gating_across_all_run_modes(monkeypatch):
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()
    window.show()

    warnings: list[tuple[str, str]] = []
    monkeypatch.setattr(QtWidgets.QMessageBox, "warning", lambda parent, title, text, *args: warnings.append((title, text)))

    load_commands: list[tuple[str, Any]] = []
    real_set_current = window.hub.instruments["load"].set_current
    real_set_input = window.hub.instruments["load"].set_input
    monkeypatch.setattr(window.hub.instruments["load"], "set_current", lambda v: (load_commands.append(("set_current", v)), real_set_current(v)))
    monkeypatch.setattr(window.hub.instruments["load"], "set_input", lambda s: (load_commands.append(("set_input", s)), real_set_input(s)))

    # 1. Hardware Mode + Verification UNCHECKED
    window.simulation.setChecked(False)
    window.chk_load.setChecked(False)
    window.update_enabled_states()

    # A. SET CURRENT (Mode 0)
    window.btn_mode_direct.click()
    QtWidgets.QApplication.processEvents()
    window.manual_target_spin.setValue(10.0)
    warnings.clear()
    load_commands.clear()
    window.btn_direct_set.click()
    QtWidgets.QApplication.processEvents()

    assert len(warnings) == 1
    assert warnings[0][0] == "Hardware write locked"
    assert "Check 'I verified low-current load control on this bench' in Bench Setup first." in warnings[0][1]
    assert len([c for c in load_commands if c[0] == "set_input" and c[1] is True]) == 0
    assert window._manual_target_current == 0.0

    # B. STEP CURRENT + (Mode 1)
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()
    window.manual_step_inc.setValue(2.0)
    warnings.clear()
    load_commands.clear()
    window.btn_plus_step.click()
    QtWidgets.QApplication.processEvents()

    assert len(warnings) == 1
    assert warnings[0][0] == "Hardware write locked"
    assert "Check 'I verified low-current load control on this bench' in Bench Setup first." in warnings[0][1]
    assert len([c for c in load_commands if c[0] == "set_input" and c[1] is True]) == 0
    assert window._manual_target_current == 0.0

    # C. CONTINUOUS SWEEP START (Mode 2)
    window.btn_mode_cont.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_sweep_btn.isEnabled() is True
    warnings.clear()
    load_commands.clear()
    window.run_sweep_btn.click()
    QtWidgets.QApplication.processEvents()

    assert len(warnings) == 1
    assert warnings[0][0] == "Hardware write locked"
    assert "Check 'I verified low-current load control on this bench' in Bench Setup first." in warnings[0][1]
    assert window.worker is None or not window.worker.isRunning()
    assert window.run_sweep_btn.isEnabled() is True
    assert "START SWEEP" in window.run_sweep_btn.text()
    assert len([c for c in load_commands if c[0] == "set_input" and c[1] is True]) == 0

    # D. PULSE SWEEP START (Mode 3)
    window.btn_mode_pulse.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_sweep_btn.isEnabled() is True
    warnings.clear()
    load_commands.clear()
    window.run_sweep_btn.click()
    QtWidgets.QApplication.processEvents()

    assert len(warnings) == 1
    assert warnings[0][0] == "Hardware write locked"
    assert "Check 'I verified low-current load control on this bench' in Bench Setup first." in warnings[0][1]
    assert window.worker is None or not window.worker.isRunning()
    assert window.run_sweep_btn.isEnabled() is True
    assert "START PULSE" in window.run_sweep_btn.text()
    assert len([c for c in load_commands if c[0] == "set_input" and c[1] is True]) == 0

    # 2. Safety Actions ALWAYS Permitted without warning when unchecked
    warnings.clear()
    window.btn_direct_zero.click()
    QtWidgets.QApplication.processEvents()
    assert len(warnings) == 0

    warnings.clear()
    window.btn_step_zero.click()
    QtWidgets.QApplication.processEvents()
    assert len(warnings) == 0

    warnings.clear()
    window.emergency_stop_action()
    QtWidgets.QApplication.processEvents()
    assert len(warnings) == 0

    # 3. With Verification CHECKED: All operate without warning
    window.chk_load.setChecked(True)
    window.update_enabled_states()
    assert window.require_load_control_verified() is True

    # 4. Demo Mode ON: Operates without requiring verification checkbox
    window.simulation.setChecked(True)
    window.chk_load.setChecked(False)
    window.update_enabled_states()
    assert window.require_load_control_verified() is True

    window.close()


def test_supply_card_config_dictionary_persistence():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.simulation.setChecked(True)
    
    card = window.supply_card
    # Check ratings: CH1 is 6V, CH2 and CH3 are 25V
    assert card.channel_controls[0]["voltage"].maximum() == 6.0
    assert card.channel_controls[1]["voltage"].maximum() == 25.0
    assert card.channel_controls[2]["voltage"].maximum() == 25.0

    # Edit role, voltage, and loss contribution
    card.channel_controls[0]["role_edit"].setText("Vdrv_Primary")
    card.channel_controls[0]["voltage"].setValue(5.5)
    card.channel_controls[0]["loss_chk"].setChecked(False)
    card._apply()
    QtCore.QThreadPool.globalInstance().waitForDone(2000)
    QtWidgets.QApplication.processEvents()

    assert window.config["supply_channels"][0]["role"] == "Vdrv_Primary"
    assert window.config["supply_channels"][0]["voltage_set"] == 5.5
    assert window.config["supply_channels"][0]["contributes_loss"] is False
    window.close()




def test_live_plot_ranges_and_negative_current_prevention():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    
    # Mouse interaction on axes is controlled/disabled
    assert not window.live_plot_widget.plotItem.vb.state["mouseEnabled"][0]
    assert not window.live_plot_widget.plotItem.vb.state["mouseEnabled"][1]
    
    # Simulate measurements with high efficiency > 100%
    window.plot_rows = [
        {"Status": "Valid", "Iout_A": 0.0, "EfficiencyConverter_pct": 0.0, "EfficiencySystem_pct": None},
        {"Status": "Valid", "Iout_A": 10.0, "EfficiencyConverter_pct": 102.5,
         "EfficiencySystem_pct": 99.0, "LossConverter_W": -1.5, "LossSystem_W": -1.0,
         "PinConverter_W": 125.0, "Pout_W": 120.0, "Paux_W": 0.7},
    ]
    window._switch_live_plot(0) # Efficiency
    assert window.live_metric_combo.itemText(0) == "Efficiency (%)"
    assert list(window.live_curve.getData()[1]) == [0.0, 102.5]
    assert list(window.live_system_curve.getData()[1]) == [99.0]
    # X min is 0.0, never negative
    x_range = window.live_plot_widget.getPlotItem().getViewBox().viewRange()[0]
    assert x_range[0] <= 0.01
    
    # Y range accommodates > 100% with margin
    y_range = window.live_plot_widget.getPlotItem().getViewBox().viewRange()[1]
    assert y_range[1] >= 102.5
    assert window.live_metric_combo.itemText(1) == "Loss (W)"
    assert window.live_metric_combo.itemText(2) == "Power (W)"
    assert window.live_legend.offset == (-12, -12)
    assert vars(window.live_legend)["_GraphicsWidgetAnchor__parentAnchor"] == (1, 1)
    window._switch_live_plot(1)
    assert list(window.live_curve.getData()[1]) == [-1.5]
    assert list(window.live_system_curve.getData()[1]) == [-1.0]
    window._switch_live_plot(2)
    assert list(window.live_curve.getData()[1]) == [125.0]
    assert list(window.live_system_curve.getData()[1]) == [120.0]
    assert list(window.live_aux_curve.getData()[1]) == [0.7]
    window.close()


def test_efficiency_axis_rounding_and_capture_source_folders(tmp_path: Path):
    assert efficiency_axis_bounds([93.6, 96.1, 97.0]) == (90.0, 100.0)
    assert efficiency_axis_bounds([95.0, 100.0]) == (90.0, 105.0)
    workbook = tmp_path / "results" / "hardware_campaign.xlsx"
    assert capture_root_for_source(workbook, "Hardware") == tmp_path / "results" / "captures" / "hardware"
    assert capture_root_for_source(workbook, "Simulation") == tmp_path / "results" / "captures" / "simulation"


def test_plot_compact_live_status_and_sweep_progress_tracker():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow, SweepProgressTracker
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()
    
    # 1. Idle state verification
    assert window.live_stat_tag.text() == "IDLE ○"
    assert window.live_point_lbl.text() == "Point — / —"
    assert window.live_cmd_lbl.text() == "Command: —"
    assert window.live_act_lbl.text() == "Actual: —"
    assert isinstance(window.plot_progress_marker, SweepProgressTracker)
    assert not window.plot_progress_marker.active

    # 2. Immediate Start/Stop range synchronization on field edit (e.g. 0 A to 20 A)
    window.cont_start.setValue(0.0)
    window.cont_stop.setValue(20.0)
    window.cont_step.setValue(2.0)
    window.cont_settle.setValue(3.0)
    QtWidgets.QApplication.processEvents()

    assert window.plot_progress_marker.start_val == 0.0
    assert window.plot_progress_marker.stop_val == 20.0
    assert "0 → 20 A" in window.cont_summary_lbl.text()
    assert "11 points" in window.cont_summary_lbl.text()
    assert "~33 s" in window.cont_summary_lbl.text()

    # 3. Simulate sweep progress update: Point 5 of 11, command 8 A
    window._run_progress(5, 11, 8.0, 10.0, "Testing")
    assert window.live_stat_tag.text() == "RUNNING ●"
    assert window.live_point_lbl.text() == "Point 5 / 11"
    assert window.live_cmd_lbl.text() == "Command: 8 A"
    assert window.plot_progress_marker.active
    assert window.plot_progress_marker.current_val == 8.0
    assert window.plot_progress_marker.start_val == 0.0
    assert window.plot_progress_marker.stop_val == 20.0

    # 4. Simulate measurement received with actual current 7.98 A
    window._measurement_received({"Iout_A": 7.98, "EfficiencyConverter_pct": 98.2})
    assert window.live_act_lbl.text() == "Actual: 7.98 A"

    # 5. Simulate run completion
    window._run_completed("Valid", "")
    assert window.live_stat_tag.text() == "COMPLETED ●"
    assert not window.plot_progress_marker.active

    window.close()




def test_default_sweep_ranges_and_buttonless_cap():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.load_card.cap_spin.setValue(70.0)
    window.load_card.apply_cap_btn.click()
    window.cont_stop.setValue(60.0)
    window.pulse_stop.setValue(60.0)
    window._update_sweep_summary()

    # 4-Peer Segmented Mode Selector: Compact buttons with tooltips
    assert window.btn_mode_direct.text() == "SET CURRENT"
    assert window.btn_mode_step.text() == "STEP CURRENT"
    assert window.btn_mode_cont.text() == "CONTINUOUS"
    assert window.btn_mode_pulse.text() == "PULSE"
    assert len(window.btn_mode_direct.toolTip()) > 5
    assert len(window.btn_mode_step.toolTip()) > 5
    assert len(window.btn_mode_cont.toolTip()) > 5
    assert len(window.btn_mode_pulse.toolTip()) > 5

    # Continuous default view: 4 primary fields + summary strip + Advanced (collapsed by default)
    assert window.btn_mode_cont.isChecked()
    assert window.cont_start.value() == 0.0
    assert window.cont_stop.value() == 60.0
    assert window.cont_step.value() == 2.0
    assert window.cont_settle.value() == 5.0
    assert window.cont_sample_window.value() == 3.0
    assert window.direct_auto_delay.value() == 4.0
    assert window.step_auto_delay.value() == 4.0
    assert "0 → 60 A" in window.cont_summary_lbl.text()
    assert "31 points" in window.cont_summary_lbl.text()
    assert window.cont_adv.isCheckable()
    assert not window.cont_adv_box.isVisible() # Collapsed by default!
    assert "START SWEEP" in window.run_sweep_btn.text()

    # Pulse view: 5 primary fields + summary strip + Advanced (collapsed by default)
    window.btn_mode_pulse.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_mode_pulse.isChecked()
    assert window.pulse_start.value() == 0.0
    assert window.pulse_stop.value() == 60.0
    assert window.pulse_step.value() == 2.0
    assert window.pulse_dwell.value() == 5.0
    assert window.pulse_cooldown.value() == 5.0
    assert window.pulse_sample_window.value() == 3.0
    assert "0 → 60 A" in window.pulse_summary_lbl.text()
    assert "31 pulses" in window.pulse_summary_lbl.text()
    assert window.pulse_adv.isCheckable()
    assert not window.pulse_adv_box.isVisible() # Collapsed by default!
    assert "START PULSE TEST" in window.run_sweep_btn.text()

    # Chroma safety limit cap spinbox has NoButtons
    assert window.load_card.cap_spin.buttonSymbols() == QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons

    # 4 equal-dimension column cards on Bench tab
    assert len(window.cards) == 3
    assert window.scope_card is not None

    # Tabs are exactly 4: Bench Setup, Run, History, Help
    assert window.tabs.count() == 4
    assert window.tabs.tabText(0) == "Bench Setup"
    assert window.tabs.tabText(1) == "Run"
    assert window.tabs.tabText(2) == "History"
    # Restore default 60 A limit
    window.load_card.cap_spin.setValue(60.0)
    window.load_card.apply_cap_btn.click()
    window.close()



def test_manual_mode_direct_set_and_step_control():
    import os, time
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()

    def wait_for_ui(condition, timeout=5.0):
        start = time.time()
        while not condition() and (time.time() - start) < timeout:
            QtWidgets.QApplication.processEvents()
            QtCore.QThread.msleep(15)
        QtWidgets.QApplication.processEvents()

    # Switch to Simulation mode and SET CURRENT mode
    window.simulation.setChecked(True)
    window.direct_auto_save.setChecked(False)
    window.direct_auto_capture.setChecked(False)
    window.step_auto_save.setChecked(False)
    window.step_auto_capture.setChecked(False)
    window.btn_mode_direct.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_stack.currentIndex() == 0

    # Verify local Load ON / Load OFF buttons are removed from Manual Mode
    assert not hasattr(window, "btn_manual_load_on")
    assert not hasattr(window, "btn_manual_load_off")
    assert not hasattr(window, "btn_step_load_on")
    assert not hasattr(window, "btn_step_load_off")

    # Direct Set > 0 A: sets current and turns load ON automatically
    window.manual_target_spin.setValue(12.5)
    window.btn_direct_set.click()
    wait_for_ui(lambda: window.hub.environment.load_enabled and window.hub.environment.current_set == 12.5)
    assert window._manual_target_current == 12.5
    assert "12.50 A · ON" in window.step_present_lbl.text()
    assert window.hub.environment.current_set == 12.5
    assert window.hub.environment.load_enabled is True

    # Direct Set 0 A: sets current to 0 A and turns load OFF automatically
    window.manual_target_spin.setValue(0.0)
    window.btn_direct_set.click()
    wait_for_ui(lambda: not window.hub.environment.load_enabled and window.hub.environment.current_set == 0.0)
    assert window._manual_target_current == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert window.hub.environment.current_set == 0.0
    assert window.hub.environment.load_enabled is False

    # Direct Set > 0 A then click ZERO / OFF button: sets 0 A and turns load OFF
    window.manual_target_spin.setValue(10.0)
    window.btn_direct_set.click()
    wait_for_ui(lambda: window.hub.environment.load_enabled and window.hub.environment.current_set == 10.0)
    assert window.hub.environment.load_enabled is True

    window.btn_direct_zero.click()
    wait_for_ui(lambda: not window.hub.environment.load_enabled and window.hub.environment.current_set == 0.0)
    assert window._manual_target_current == 0.0
    assert window.manual_target_spin.value() == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert window.hub.environment.load_enabled is False

    # 2. Switch to STEP CURRENT peer mode
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_mode_step.isChecked()
    assert window.run_stack.currentIndex() == 1

    # Verify default asymmetric step sizes and button labels
    assert window.manual_step_inc.value() == 2.0
    assert window.manual_step_dec.value() == 5.0
    assert window.btn_minus_step.text() == "−5 A"
    assert window.btn_step_zero.text() == "ZERO / OFF"
    assert window.btn_plus_step.text() == "+2 A"

    # Sequence from user specification:
    # 0 A / OFF → +2 A → 2 A / ON
    window.btn_plus_step.click()
    wait_for_ui(lambda: window.hub.environment.load_enabled and window.hub.environment.current_set == 2.0)
    assert window._manual_target_current == 2.0
    assert window.manual_target_spin.value() == 2.0
    assert "2.00 A · ON" in window.step_present_lbl.text()
    assert window.hub.environment.current_set == 2.0
    assert window.hub.environment.load_enabled is True

    # Direct set to 20 A in SET CURRENT mode to test user example: 20 A → −5 A → 15 A
    window.btn_mode_direct.click()
    QtWidgets.QApplication.processEvents()
    window.manual_target_spin.setValue(20.0)
    window.btn_direct_set.click()
    wait_for_ui(lambda: window.hub.environment.current_set == 20.0)
    assert window._manual_target_current == 20.0
    assert "20.00 A · ON" in window.step_present_lbl.text()

    # Switch to STEP CURRENT to step: 20 A → −5 A → 15 A
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()
    window.btn_minus_step.click()
    wait_for_ui(lambda: window.hub.environment.current_set == 15.0)
    assert window._manual_target_current == 15.0
    assert "15.00 A · ON" in window.step_present_lbl.text()
    assert window.hub.environment.current_set == 15.0
    assert window.hub.environment.load_enabled is True

    # 15 A → +2 A → 17 A
    window.btn_plus_step.click()
    wait_for_ui(lambda: window.hub.environment.current_set == 17.0)
    assert window._manual_target_current == 17.0
    assert "17.00 A · ON" in window.step_present_lbl.text()
    assert window.hub.environment.current_set == 17.0
    assert window.hub.environment.load_enabled is True

    # Direct set to 3 A in SET CURRENT to test: 3 A → −5 A → 0 A / OFF
    window.btn_mode_direct.click()
    QtWidgets.QApplication.processEvents()
    window.manual_target_spin.setValue(3.0)
    window.btn_direct_set.click()
    wait_for_ui(lambda: window.hub.environment.current_set == 3.0)
    assert window._manual_target_current == 3.0

    # Switch to STEP CURRENT and step: 3 A → −5 A → 0 A / OFF
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()
    window.btn_minus_step.click()
    wait_for_ui(lambda: not window.hub.environment.load_enabled and window.hub.environment.current_set == 0.0)
    assert window._manual_target_current == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert window.hub.environment.current_set == 0.0
    assert window.hub.environment.load_enabled is False

    # Dynamic button label updating when asymmetric step sizes change
    window.manual_step_inc.setValue(1.0)
    window.manual_step_dec.setValue(10.0)
    QtWidgets.QApplication.processEvents()
    assert window.btn_plus_step.text() == "+1 A"
    assert window.btn_minus_step.text() == "−10 A"

    # Step Control: +1 A step from 0 A -> 1.0 A · ON
    window.btn_plus_step.click()
    wait_for_ui(lambda: window.hub.environment.load_enabled and window.hub.environment.current_set == 1.0)
    assert window._manual_target_current == 1.0
    assert "1.00 A · ON" in window.step_present_lbl.text()
    assert window.hub.environment.load_enabled is True

    # Step Control: −10 A step from 1 A -> 0 A / OFF
    window.btn_minus_step.click()
    wait_for_ui(lambda: not window.hub.environment.load_enabled and window.hub.environment.current_set == 0.0)
    assert window._manual_target_current == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert window.hub.environment.load_enabled is False

    # Emergency stop top-right button while load is ON
    window.btn_plus_step.click()
    wait_for_ui(lambda: window.hub.environment.load_enabled and window.hub.environment.current_set == 1.0)
    assert window.hub.environment.load_enabled is True

    window.emergency_stop_btn.click()
    wait_for_ui(lambda: not window.hub.environment.load_enabled and window.hub.environment.current_set == 0.0)
    assert window._manual_target_current == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert window.hub.environment.load_enabled is False

    # Step size and secondary actions are tucked inside collapsed Advanced box
    assert window.btn_adv_step.isCheckable()
    assert not window.btn_adv_step.isChecked()
    assert window.btn_adv_direct.isCheckable()
    assert not window.btn_adv_direct.isChecked()

    # Cluttered legacy actions removed from primary surface
    assert not hasattr(window, "manual_measure_btn")
    assert not hasattr(window, "spot_pulse_btn")

    window.close()





def test_no_critical_clipping_at_supported_resolutions():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    
    resolutions = [(1150, 720), (1366, 768), (1420, 880), (1920, 1080)]
    for width, height in resolutions:
        window.resize(width, height)
        window.show()
        QtWidgets.QApplication.processEvents()
        
        # Verify critical header & controls have non-zero geometry and are visible
        assert window.emergency_stop_btn.width() > 50
        assert window.emergency_stop_btn.height() > 20
        assert window.tabs.count() == 4
        
        # Test tab switching across all 4 tabs without crash or invalid geometry
        for tab_idx in range(4):
            window.tabs.setCurrentIndex(tab_idx)
            QtWidgets.QApplication.processEvents()
            current_widget = window.tabs.currentWidget()
            assert current_widget.width() > 0
            assert current_widget.height() > 0
    window.close()



def test_msox4024a_capture_stops_and_restores_run(tmp_path: Path):
    from sid_instruments import MSOX4024A
    class FakeScopeSession:
        def __init__(self):
            self.writes = []
            self.queries = []
            self.closed = False
            self.timeout = 3000
        def query(self, command: str) -> str:
            self.queries.append(command)
            if "*IDN?" in command:
                return "Keysight Technologies,MSOX4024A,TESTSCOPE001,1.0.0"
            if ":CHANnel1:DISPlay?" in command:
                return "1"
            if ":CHANnel" in command and ":DISPlay?" in command:
                return "0"
            if ":WAVeform:PREamble?" in command:
                return "0,0,1000,1,1e-6,-5e-4,0,1e-3,0,0"
            return "0"
        def write(self, command: str) -> None:
            self.writes.append(command)
        def read_raw(self) -> bytes:
            return b"#800000008PNGDATA"
        def query_binary_values(self, *args, **kwargs) -> list:
            return [100, 200, 300]
        def close(self):
            self.closed = True

    class FakeManager:
        def __init__(self, session): self.s = session
        def open(self):
            class RM:
                def __init__(self, s): self.s = s
                def open_resource(self, _): return self.s
            return RM(self.s)

    sess = FakeScopeSession()
    scope = MSOX4024A(FakeManager(sess), "USB::SCOPE", ("MSOX4024",))
    scope.connect(persistent=True)
    png_path = tmp_path / "scope.png"
    csv_path = tmp_path / "scope.csv"
    channels, count = scope.capture(png_path, csv_path)
    scope.release()

    assert ":STOP" in sess.writes
    assert ":SINGle" not in sess.writes
    assert "*OPC?" not in sess.queries
    assert ":RUN" in sess.writes
    stop_idx = sess.writes.index(":STOP")
    run_idx = sess.writes.index(":RUN")
    assert run_idx > stop_idx
    assert png_path.exists()
    assert csv_path.exists()



def test_msox4024a_capture_reports_stage_and_runs_after_failure(tmp_path: Path):
    from sid_instruments import InstrumentError, MSOX4024A

    class FailingSession:
        def __init__(self): self.writes = []; self.queries = []; self.timeout = 3000
        def query(self, command):
            self.queries.append(command)
            if command == "*IDN?": return "Keysight,MSOX4024A,MY1,1.0"
            if command == ":CHANnel1:DISPlay?": return "1"
            if ":DISPlay?" in command: return "0"
            return "0"
        def write(self, command): self.writes.append(command)
        def read_raw(self): raise TimeoutError("binary transfer timeout")
        def close(self): pass

    class Manager:
        def __init__(self, session): self.session = session
        def open(self):
            session = self.session
            class RM:
                def open_resource(self, _): return session
            return RM()

    session = FailingSession()
    scope = MSOX4024A(Manager(session), "USB::SCOPE", ("MSOX4024",))
    scope.connect(persistent=True)
    with pytest.raises(InstrumentError, match="saving screen PNG"):
        scope.capture(tmp_path / "failed.png", tmp_path / "failed.csv")
    assert ":STOP" in session.writes
    assert ":SINGle" not in session.writes
    assert session.writes.index(":RUN") > session.writes.index(":STOP")



def test_all_spinboxes_have_no_buttons_and_wheel_is_ignored():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore, QtGui
    from sid_bench_gui import MainWindow, NoWheelFilter

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.show()

    # 1. Assert ALL spinboxes across the entire window have NoButtons
    spinboxes = window.findChildren(QtWidgets.QAbstractSpinBox)
    assert len(spinboxes) >= 10, f"Expected >= 10 spinboxes, found {len(spinboxes)}"
    for sp in spinboxes:
        assert sp.buttonSymbols() == QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons, f"Expected NoButtons for {sp}"

    # 2. Test wheel event filtering on a representative spinbox
    test_spin = window.vin_target
    test_spin.setValue(48.0)
    wheel_event = QtGui.QWheelEvent(
        QtCore.QPointF(10, 10), QtCore.QPointF(10, 10),
        QtCore.QPoint(0, 120), QtCore.QPoint(0, 120),
        QtCore.Qt.MouseButton.NoButton, QtCore.Qt.KeyboardModifier.NoModifier,
        QtCore.Qt.ScrollPhase.NoScrollPhase, False
    )
    # Event filter should intercept and return True (ignored)
    filtered = window.wheel_filter.eventFilter(test_spin, wheel_event)
    assert filtered is True
    assert test_spin.value() == 48.0

    # 3. Test direct typing/setting works accurately
    test_spin.setValue(36.5)
    window.close()


def test_pilawa_outer_shell_and_chrome_formatting():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtGui
    from sid_bench_gui import MainWindow, apply_forced_light_theme, BERKELEY_BLUE, CALIFORNIA_GOLD
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    apply_forced_light_theme(app)

    window = MainWindow()
    window.show()

    # Assert gold accent bar exists below header
    assert hasattr(window, "toolbar_accent")
    assert window.toolbar_accent.height() == 2
    assert "background: #FDB515" in window.toolbar_accent.styleSheet()

    # Assert DWM native chrome methods and color conversion
    assert MainWindow._windows_colorref("#002676") == (0x00 | (0x26 << 8) | (0x76 << 16))
    window._apply_native_windows_chrome()

    # Assert Light Palette is enforced
    palette = app.palette()
    assert palette.color(QtGui.QPalette.ColorRole.Window).name().upper() == "#F7F8FA"
    assert palette.color(QtGui.QPalette.ColorRole.Base).name().upper() == "#FFFFFF"

    window.close()


def test_kpi_header_breathing_layout_and_prominent_values():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow, InstrumentSnapshot
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()
    
    # 1. Direct and derived KPIs share the single-row header.
    assert set(window.kpi_labels.keys()) == {"Vin", "Iin", "Vout", "Iout", "Pin", "Pout", "Eff"}
    assert window.kpi_labels["Pin"].text() == "—"
    assert window.kpi_labels["Pout"].text() == "—"
    assert window.kpi_labels["Eff"].text() == "—"

    # 2. Check prominent styling on value label
    vin_lbl = window.kpi_labels["Vin"]
    assert "20px" in vin_lbl.styleSheet()
    assert "font-weight: 800" in vin_lbl.styleSheet()

    iin_lbl = window.kpi_labels["Iin"]
    assert "20px" in iin_lbl.styleSheet()
    assert "font-weight: 800" in iin_lbl.styleSheet()

    # 3. Simulate PA snapshot
    pa_snap = InstrumentSnapshot(
        instrument="pa",
        values={"vin": 48.02, "vout": 12.01, "iin": 2.50, "pin": 120.0},
        timestamp="2026-08-17T00:00:00Z",
        valid=True,
    )
    window._snapshot_received("pa", pa_snap)
    assert window.kpi_labels["Vin"].text() == "48.02 V"
    assert window.kpi_labels["Iin"].text() == "2.500 A"
    assert window.kpi_labels["Vout"].text() == "12.01 V"
    assert window.kpi_labels["Pin"].text() == "—"

    # Derived values update together only from one valid calculated point.
    window._measurement_received({
        "Status": "Valid", "Iout_A": 4.99, "PinConverter_W": 29.69,
        "Pout_W": 28.99, "EfficiencyConverter_pct": 97.68,
    })
    assert window.kpi_labels["Pin"].text() == "29.69 W"
    assert window.kpi_labels["Pout"].text() == "28.99 W"
    assert window.kpi_labels["Eff"].text() == "97.68%"
    assert "#B45309" in window.kpi_labels["Pin"].styleSheet()

    # A failed/invalid later point preserves the last successful trio.
    window._measurement_received({
        "Status": "Invalid", "PinConverter_W": 1.0, "Pout_W": 1.0,
        "EfficiencyConverter_pct": 1.0,
    })
    assert window.kpi_labels["Pin"].text() == "29.69 W"
    assert window.kpi_labels["Pout"].text() == "28.99 W"
    assert window.kpi_labels["Eff"].text() == "97.68%"

    window.close()


def test_demo_mode_on_run_tab_and_warning_states():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()
    window.show()
    window.tabs.setCurrentIndex(1)  # Switch to Run tab
    QtWidgets.QApplication.processEvents()

    # 1. Verify Demo Mode toggle is NOT on Run tab setup strip (setup strip contains only Test Name, Vin, Freq)
    setup_strip = window.findChild(QtWidgets.QFrame, "run_setup_strip")
    assert window.simulation not in setup_strip.findChildren(QtWidgets.QCheckBox)
    assert not hasattr(window, "demo_btn")
    assert not hasattr(window, "scenario")

    # 2. In Hardware Mode (Demo Mode OFF by default):
    window.simulation.setChecked(False)
    QtWidgets.QApplication.processEvents()
    assert window.simulation.text() == "Demo Mode"
    assert not window.sim_warning_banner.isVisible()
    assert window.plot_watermark_lbl.isHidden()
    assert "START SWEEP" in window.run_sweep_btn.text()

    # 3. Enable Demo Mode:
    top_widgets_before = [w for w in QtWidgets.QApplication.topLevelWidgets() if w.isVisible()]
    window.simulation.setChecked(True)
    QtWidgets.QApplication.processEvents()
    top_widgets_after = [w for w in QtWidgets.QApplication.topLevelWidgets() if w.isVisible()]
    # Verify no orphan top-level widgets or floating windows were created
    assert len(top_widgets_before) == len(top_widgets_after)

    assert window.simulation.text() == "Demo Mode"
    assert "#D97706" in window.simulation.styleSheet()
    assert window.sim_warning_banner.isVisible()
    banner_labels = window.sim_warning_banner.findChildren(QtWidgets.QLabel)
    assert any("DEMO MODE · SYNTHETIC DATA" in lbl.text() for lbl in banner_labels)
    assert any("NO HARDWARE COMMANDS" in lbl.text() for lbl in banner_labels)

    # Plot watermark & action button
    assert window.plot_watermark_lbl.isVisible()
    assert window.plot_watermark_lbl.text() == "DEMO DATA"
    assert window.run_sweep_btn.text() == "▶  START DEMO"

    # Verify settings DataSource = Simulation
    settings = window._collect_settings()
    assert settings["data_source"] == "Simulation"

    # 4. Disable Demo Mode: returns to clean passive hardware mode
    window.simulation.setChecked(False)
    QtWidgets.QApplication.processEvents()
    assert not window.sim_warning_banner.isVisible()
    assert not window.plot_watermark_lbl.isVisible()
    assert "START SWEEP" in window.run_sweep_btn.text()

    window.close()


def test_bench_readiness_status_indicator_and_release_all_devices():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, SUCCESS_GREEN, WARNING_AMBER
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()
    window.show()

    # 1. Initial State: Compact status indicator shows "● Not checked"
    assert hasattr(window, "readiness_status")
    assert window.readiness_status.text() == "● Not checked"

    # 2. Check / Refresh Entire Bench is the primary button action
    assert hasattr(window, "check_bench_btn")
    assert window.check_bench_btn.text() == "Check / Refresh Entire Bench"

    # 3. Simulate successful readiness check
    window.simulation.setChecked(True)
    window.vin_target.setValue(48.0)
    window.chk_load.setChecked(True)
    window._check_entire_bench()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()

    assert "● READY" in window.readiness_status.text()

    # 4. Test incomplete check (e.g. load not verified)
    window.chk_load.setChecked(False)
    window._check_entire_bench()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()

    assert "● NOT READY (Load not verified)" in window.readiness_status.text()

    # 5. Verify [ Discover VISA Devices ] and [ Release All Devices ] buttons at bottom
    assert hasattr(window, "disc_btn")
    assert hasattr(window, "release_btn")
    assert window.release_btn.text() == "Release All Devices"
    assert "Close bench VISA sessions" in window.release_btn.toolTip()

    # 6. Click Release All Devices
    window.release_btn.click()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()

    assert window.readiness_status.text() == "● Not checked"
    assert "released" in window.statusBar().currentMessage().lower()

    # 7. Verify "Proceed to Run" button is completely removed
    buttons = window.findChildren(QtWidgets.QPushButton)
    assert not any("Proceed to Run" in b.text() for b in buttons)

    window.close()




def test_measurement_sampling_and_scope_capture_currents():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow, parse_capture_points
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()

    # 1. Verify default Readings per point is 1 for both Continuous and Pulse
    assert window.cont_sample_count.value() == 1
    assert window.pulse_sample_count.value() == 1

    # 2. Measure Last defaults to 3 s inside the 5 s dwell/ON interval.
    assert window.cont_sample_window.value() == 3.0
    assert window.pulse_sample_window.value() == 3.0
    assert window.cont_sample_window.isEnabled()

    # Switch to Pulse mode to verify pulse_sample_window is enabled
    window.btn_mode_pulse.setChecked(True)
    window._mode_selected(3)
    assert window.pulse_sample_window.isEnabled()

    # Switch back to Continuous
    window.btn_mode_cont.setChecked(True)
    window._mode_selected(2)
    assert window.cont_sample_window.isEnabled()

    # 3. Readings count changes preserve enabled Measure Last control
    window.cont_sample_count.setValue(4)
    assert window.cont_sample_window.isEnabled()

    window.cont_sample_count.setValue(1)
    assert window.cont_sample_window.isEnabled()

    # 4. Verify scope capture default text is '0, 10, 20, 30'
    assert window.cont_capture_points.text() == "0, 10, 20, 30"
    assert window.pulse_capture_points.text() == "0, 10, 20, 30"
    # In default 0..60 2A step, 0, 10, 20, 30 are all valid sweep points
    assert "4 scope captures" in window.cont_scope_summary_lbl.text()

    # If sweep stops at 20 A, then 30 A is beyond stop and flagged
    window.cont_stop.setValue(20.0)
    QtWidgets.QApplication.processEvents()
    assert "Unaligned" in window.cont_scope_summary_lbl.text()
    assert "30A" in window.cont_scope_summary_lbl.text()

    # 5. Live validation against aligned sweep points (e.g. 0 to 30 in step of 5 A)
    window.cont_start.setValue(0.0)
    window.cont_stop.setValue(30.0)
    window.cont_step.setValue(5.0)
    window.cont_capture_points.setText("0, 5, 10, 15, 20, 25, 30")
    QtWidgets.QApplication.processEvents()
    assert "7 scope captures" in window.cont_scope_summary_lbl.text()

    # If capture current is not in sweep (e.g. 7 A with step of 5 A), warning is shown
    window.cont_capture_points.setText("0, 5, 7, 10")
    QtWidgets.QApplication.processEvents()
    assert "Unaligned" in window.cont_scope_summary_lbl.text()
    assert "7A" in window.cont_scope_summary_lbl.text()


    # 6. Strict parse validation: invalid tokens raise ValueError and are not silently ignored
    with pytest.raises(ValueError, match="Invalid scope capture current"):
        parse_capture_points("0, invalid_val, 10")

    window.close()


def test_simplified_test_setup_strip_and_blank_auto_name():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()
    window.show()

    # 1. Verify compact single horizontal strip exists with the 3 main inputs
    setup_strip = window.findChild(QtWidgets.QFrame, "run_setup_strip")
    assert setup_strip is not None

    assert hasattr(window, "test_name")
    assert hasattr(window, "vin_target")
    assert hasattr(window, "frequency")
    assert not hasattr(window, "modulation")  # Modulation removed from GUI

    # 2. Test auto-generation of blank Test Name: Test_YYYYMMDD_HHMM
    window.test_name.setText("")
    settings = window._collect_settings()
    assert settings["run_record"]["CampaignName"].startswith("Test_")
    assert window.test_name.text().startswith("Test_")
    assert window.test_name.text() in settings["run_record"]["CampaignName"]

    window.close()


def test_chroma_safety_limit_card_and_bench_configuration():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    window = MainWindow()

    # 1. Verify Tab 0 is named "Bench Setup"
    assert window.tabs.tabText(0) == "Bench Setup"

    # 2. Verify LoadCard exists with safety limit section and low-current load verification checkbox
    assert hasattr(window, "load_card")
    assert hasattr(window.load_card, "cap_spin")
    assert hasattr(window.load_card, "apply_cap_btn")
    assert hasattr(window.load_card, "chk_load")
    assert window.chk_load is window.load_card.chk_load
    assert "low-current load control" in window.chk_load.text()

    # 3. Initial safety limit from config
    assert window.cap_val == float(window.config.get("working_current_cap_a", 60.0))

    # 4. Change safety limit in LoadCard and apply
    window.load_card.cap_spin.setValue(45.0)
    window.load_card.apply_cap_btn.click()

    # 5. Verify limits updated everywhere
    assert window.cap_val == 45.0
    assert window.config["working_current_cap_a"] == 45.0
    assert "45" in window.chk_cap.text()
    assert window.manual_target_spin.maximum() == 45.0
    assert window.cont_stop.maximum() == 45.0
    assert window.pulse_stop.maximum() == 45.0

    # Restore default 60 A limit
    window.load_card.cap_spin.setValue(60.0)
    window.load_card.apply_cap_btn.click()
    assert window.cap_val == 60.0

    window.close()



def test_graceful_stop_and_return_to_zero_sweep_worker(tmp_path: Path):
    from sid_bench_gui import SweepWorker, InstrumentHub, WorkbookStore
    from sid_instruments import InstrumentSnapshot
    import threading

    class FakeLoad:
        def __init__(self):
            self.currents_commanded = []
            self.input_states = []
            self.safe_off_called = False
            self.identity = "Chroma,63206A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def set_current(self, amps: float): self.currents_commanded.append(amps)
        def set_input(self, state: bool): self.input_states.append(state)
        def safe_off(self):
            self.safe_off_called = True
            self.input_states.append(False)
        def read_snapshot(self, **_: Any):
            return InstrumentSnapshot("load", {"current": self.currents_commanded[-1] if self.currents_commanded else 0.0, "enabled": True})

    class FakePA:
        def __init__(self): self.identity = "Keysight,PA2201A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def read_snapshot(self):
            return InstrumentSnapshot("pa", {"vin": 48.0, "vout": 12.0, "iin": 2.5, "pin": 120.0})

    class FakeScope:
        def __init__(self): self.identity = "Keysight,MSOX4024A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def capture(self, png, csv): pass

    hub = InstrumentHub(True, {})
    fake_load = FakeLoad()
    fake_pa = FakePA()
    hub.instruments["load"] = fake_load
    hub.instruments["pa"] = fake_pa
    hub.instruments["scope"] = FakeScope()

    store = WorkbookStore(tmp_path / "test.xlsx")

    # Configure Continuous sweep at 47 A with 5 A return step
    settings = {
        "run_id": "TEST-RAMP-RUN",
        "run_record": {
            "RunID": "TEST-RAMP-RUN", "CampaignName": "RampTest", "Created": "2026-08-17T00:00:00Z",
            "Status": "Aborted", "DataSource": "Simulation", "Mode": "Continuous", "VinTarget_V": 48.0,
            "ModulationLabel": "RampTest", "Frequency_Hz": 100000.0, "ModulationMetadata": "",
            "SupplyConfiguration": [], "Length_mm": 24.0, "Width_mm": 16.0, "Height_mm": 3.4,
            "WorkingCap_A": 60.0, "Notes": "", "InstrumentIdentities": {},
            "FPGASnapshotStatus": "Skipped", "FPGASnapshot": {}, "Warnings": "",
        },
        "points": [47.0, 50.0, 55.0],
        "capture_points": set(),
        "mode": "Continuous",
        "settle": 0.01,
        "dwell": 0.01,
        "sample_window": 0.01,
        "sample_count": 1,
        "cooldown": 0.01,
        "working_cap": 60.0,
        "vin_target": 48.0,
        "modulation": "RampTest",
        "frequency": 100000.0,
        "dimensions": (24.0, 16.0, 3.4),
        "supply_channels": [],
        "psu_required": False,
        "data_source": "Simulation",
        "duplicate_action": "keep",
        "return_to_zero_step": 5.0,
    }

    worker = SweepWorker(hub, store, settings)
    states_received = []
    ramp_amps = []
    worker.state_changed.connect(lambda s, d: states_received.append(s))
    worker.ramp_progress.connect(lambda a: ramp_amps.append(a))

    completed_status = []
    worker.completed.connect(lambda s, w: completed_status.append(s))

    # Trigger stop after first point is applied
    def trigger_stop(curr, total, amps, next_a, text):
        if amps == 47.0:
            worker.stop_and_return_to_zero()

    worker.progress.connect(trigger_stop)
    worker.run()

    # Verify graceful ramp sequence from 47 A down to 0 A with 5 A step:
    # 47 -> 42 -> 37 -> 32 -> 27 -> 22 -> 17 -> 12 -> 7 -> 2 -> 0 A
    assert "RETURNING TO ZERO" in states_received
    assert 42.0 in ramp_amps
    assert 37.0 in ramp_amps
    assert 0.0 in ramp_amps
    assert ramp_amps[-1] == 0.0
    assert fake_load.input_states[-1] is False  # Load turned OFF at 0 A
    assert completed_status == ["Stopped"]

    # Verify only the 1 measured point (47 A) was saved in the store, ramp points were NOT saved
    runs = store.list_runs()
    assert len(runs) == 1
    assert runs[0]["Status"] == "Stopped"


def test_continuous_normal_completion_auto_ramp(tmp_path: Path):
    from sid_bench_gui import SweepWorker, InstrumentHub, WorkbookStore
    from sid_instruments import InstrumentSnapshot

    class FakeLoad:
        def __init__(self):
            self.currents_commanded = []
            self.input_states = []
            self.identity = "Chroma,63206A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def set_current(self, amps: float): self.currents_commanded.append(amps)
        def set_input(self, state: bool): self.input_states.append(state)
        def safe_off(self): self.input_states.append(False)
        def read_snapshot(self, **_: Any):
            return InstrumentSnapshot("load", {"current": self.currents_commanded[-1] if self.currents_commanded else 0.0, "enabled": True})

    class FakePA:
        def __init__(self): self.identity = "Keysight,PA2201A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def read_snapshot(self):
            return InstrumentSnapshot("pa", {"vin": 48.0, "vout": 12.0, "iin": 2.5, "pin": 120.0})

    from PyQt6 import QtWidgets
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    hub = InstrumentHub(True, {})
    fake_load = FakeLoad()
    hub.instruments["load"] = fake_load
    hub.instruments["pa"] = FakePA()
    hub.instruments["scope"] = FakePA()

    store = WorkbookStore(tmp_path / "test_norm.xlsx")

    # Sweep from 0 to 20 A in 10 A steps
    settings = {
        "run_id": "TEST-NORM-RUN",
        "run_record": {
            "RunID": "TEST-NORM-RUN", "CampaignName": "NormTest", "Created": "2026-08-17T00:00:00Z",
            "Status": "Aborted", "DataSource": "Simulation", "Mode": "Continuous", "VinTarget_V": 48.0,
            "ModulationLabel": "NormTest", "Frequency_Hz": 100000.0, "ModulationMetadata": "",
            "SupplyConfiguration": [], "Length_mm": 24.0, "Width_mm": 16.0, "Height_mm": 3.4,
            "WorkingCap_A": 60.0, "Notes": "", "InstrumentIdentities": {},
            "FPGASnapshotStatus": "Skipped", "FPGASnapshot": {}, "Warnings": "",
        },
        "points": [0.0, 10.0, 20.0],
        "capture_points": set(),
        "mode": "Continuous",
        "settle": 0.01,
        "dwell": 0.01,
        "sample_window": 0.01,
        "sample_count": 1,
        "cooldown": 0.01,
        "working_cap": 60.0,
        "vin_target": 48.0,
        "modulation": "NormTest",
        "frequency": 100000.0,
        "dimensions": (24.0, 16.0, 3.4),
        "supply_channels": [],
        "psu_required": False,
        "data_source": "Simulation",
        "duplicate_action": "keep",
        "return_to_zero_step": 5.0,
    }

    worker = SweepWorker(hub, store, settings)
    ramp_amps = []
    worker.ramp_progress.connect(lambda a: ramp_amps.append(a))
    completed_status = []
    worker.completed.connect(lambda s, w: completed_status.append(s))

    worker.run()

    # Verify normal completion auto-ramped from 20 A -> 15 -> 10 -> 5 -> 0 A -> load OFF
    assert completed_status == ["Valid"]
    assert 15.0 in ramp_amps
    assert 0.0 in ramp_amps
    assert fake_load.input_states[-1] is False


def test_emergency_abort_immediate_safe_off(tmp_path: Path):
    from PyQt6 import QtWidgets
    from sid_bench_gui import SweepWorker, InstrumentHub, WorkbookStore
    from sid_instruments import InstrumentSnapshot
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    class FakeLoad:
        def __init__(self):
            self.currents_commanded = []
            self.input_states = []
            self.safe_off_called = False
            self.identity = "Chroma,63206A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def set_current(self, amps: float): self.currents_commanded.append(amps)
        def set_input(self, state: bool): self.input_states.append(state)
        def safe_off(self):
            self.safe_off_called = True
            self.input_states.append(False)
        def read_snapshot(self, **_: Any):
            return InstrumentSnapshot("load", {"current": self.currents_commanded[-1] if self.currents_commanded else 0.0, "enabled": True})

    hub = InstrumentHub(True, {})
    fake_load = FakeLoad()
    hub.instruments["load"] = fake_load
    hub.instruments["pa"] = FakeLoad()
    hub.instruments["scope"] = FakeLoad()

    store = WorkbookStore(tmp_path / "test_abort.xlsx")

    settings = {
        "run_id": "TEST-ABORT-RUN",
        "run_record": {
            "RunID": "TEST-ABORT-RUN", "CampaignName": "AbortTest", "Created": "2026-08-17T00:00:00Z",
            "Status": "Aborted", "DataSource": "Simulation", "Mode": "Continuous", "VinTarget_V": 48.0,
            "ModulationLabel": "AbortTest", "Frequency_Hz": 100000.0, "ModulationMetadata": "",
            "SupplyConfiguration": [], "Length_mm": 24.0, "Width_mm": 16.0, "Height_mm": 3.4,
            "WorkingCap_A": 60.0, "Notes": "", "InstrumentIdentities": {},
            "FPGASnapshotStatus": "Skipped", "FPGASnapshot": {}, "Warnings": "",
        },
        "points": [30.0, 40.0],
        "capture_points": set(),
        "mode": "Continuous",
        "settle": 0.5,
        "dwell": 0.5,
        "sample_window": 0.1,
        "sample_count": 1,
        "cooldown": 0.1,
        "working_cap": 60.0,
        "vin_target": 48.0,
        "modulation": "AbortTest",
        "frequency": 100000.0,
        "dimensions": (24.0, 16.0, 3.4),
        "supply_channels": [],
        "psu_required": False,
        "data_source": "Simulation",
        "duplicate_action": "keep",
        "return_to_zero_step": 5.0,
    }

    worker = SweepWorker(hub, store, settings)
    completed_status = []
    worker.completed.connect(lambda s, w: completed_status.append(s))

    # Immediate abort without ramping
    worker.abort()
    worker.run()

    assert fake_load.safe_off_called is True
    assert completed_status == ["Aborted"]


def test_four_peer_modes_and_true_disclosure_collapse():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow, DisclosureButton
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.show()

    # 1. Four peer operating modes in segmented selector
    assert window.mode_group.buttons() == [
        window.btn_mode_direct,
        window.btn_mode_step,
        window.btn_mode_cont,
        window.btn_mode_pulse,
    ]
    assert window.btn_mode_direct.text() == "SET CURRENT"
    assert window.btn_mode_step.text() == "STEP CURRENT"
    assert window.btn_mode_cont.text() == "CONTINUOUS"
    assert window.btn_mode_pulse.text() == "PULSE"

    # 2. True disclosure collapse behavior for SET CURRENT (mode 0)
    window.btn_mode_direct.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_stack.currentIndex() == 0
    assert window.btn_adv_direct.text() == "Advanced ▸"
    assert window.direct_adv_box.isHidden()
    # Click disclosure button -> expands
    window.btn_adv_direct.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_direct.text() == "Advanced ▾"
    assert not window.direct_adv_box.isHidden()
    # Click disclosure button -> collapses
    window.btn_adv_direct.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_direct.text() == "Advanced ▸"
    assert window.direct_adv_box.isHidden()

    # 3. True disclosure collapse behavior for STEP CURRENT (mode 1)
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_stack.currentIndex() == 1
    assert window.btn_adv_step.text() == "Advanced ▸"
    assert window.step_adv_box.isHidden()
    window.btn_adv_step.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_step.text() == "Advanced ▾"
    assert not window.step_adv_box.isHidden()
    window.btn_adv_step.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_step.text() == "Advanced ▸"
    assert window.step_adv_box.isHidden()

    # 4. True disclosure collapse behavior for CONTINUOUS (mode 2)
    window.btn_mode_cont.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_stack.currentIndex() == 2
    assert window.btn_adv_cont.text() == "Advanced ▸"
    assert window.cont_adv_box.isHidden()
    window.btn_adv_cont.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_cont.text() == "Advanced ▾"
    assert not window.cont_adv_box.isHidden()
    window.btn_adv_cont.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_cont.text() == "Advanced ▸"
    assert window.cont_adv_box.isHidden()

    # 5. True disclosure collapse behavior for PULSE (mode 3)
    window.btn_mode_pulse.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_stack.currentIndex() == 3
    assert window.btn_adv_pulse.text() == "Advanced ▸"
    assert window.pulse_adv_box.isHidden()
    window.btn_adv_pulse.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_pulse.text() == "Advanced ▾"
    assert not window.pulse_adv_box.isHidden()
    window.btn_adv_pulse.click()
    QtWidgets.QApplication.processEvents()
    assert window.btn_adv_pulse.text() == "Advanced ▸"
    assert window.pulse_adv_box.isHidden()

    # 6. Compact status strip during run and return-to-zero
    window.btn_mode_cont.click()
    QtWidgets.QApplication.processEvents()
    assert window.cont_summary_lbl.height() <= 45
    assert window.pulse_summary_lbl.height() <= 45

    window._run_progress(12, 31, 22.0, 24.0, "Sweep")
    assert "RUNNING" in window.cont_summary_lbl.text()
    assert "Point 12/31" in window.cont_summary_lbl.text()
    assert "22 A" in window.cont_summary_lbl.text()
    assert "remaining" in window.cont_summary_lbl.text()

    window._ramp_progress_received(15.0)
    assert "RETURNING TO ZERO" in window.cont_summary_lbl.text()
    assert "15 A" in window.cont_summary_lbl.text()

    window._run_completed("Valid", "")
    assert "0 → 60 A" in window.cont_summary_lbl.text()

    window.close()


def test_step_current_delayed_auto_recording(tmp_path):
    import os, time
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, WorkbookStore
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.store = WorkbookStore(tmp_path / "step_test.xlsx")
    window.cap_val = 50.0
    window.manual_target_spin.setMaximum(50.0)
    window.show()
    window.simulation.setChecked(True)

    def wait_for_ui(condition, timeout=10.0):
        start = time.time()
        while not condition() and (time.time() - start) < timeout:
            QtWidgets.QApplication.processEvents()
            QtCore.QThread.msleep(20)
        QtWidgets.QApplication.processEvents()

    # 1. Switch to STEP CURRENT mode
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_stack.currentIndex() == 1

    # Check defaults under Advanced
    assert window.step_auto_delay.value() == 4.0
    assert window.step_auto_save.isChecked() is True
    assert window.step_auto_capture.isChecked() is True
    assert window.step_status_lbl.text() == "READY"
    assert window.btn_step_zero.text() == "ZERO / OFF"

    # 2. Step Up (+2 A) from 0 A: 0 A → 2 A
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())

    # Verify load state and command
    assert window._manual_target_current == 2.0
    assert window.hub.environment.current_set == 2.0
    assert window.hub.environment.load_enabled is True
    assert "2.00 A · ON" in window.step_present_lbl.text()

    # Upward actions lock; unload and ZERO/OFF retain priority.
    assert window.btn_plus_step.isEnabled() is False
    assert window.btn_minus_step.isEnabled() is True
    assert window.btn_step_zero.isEnabled() is True
    assert window._step_countdown_timer.isActive() is True
    assert "SETTLING" in window.step_status_lbl.text()

    # Countdown tick down
    window._step_countdown_tick()
    QtWidgets.QApplication.processEvents()
    assert "SETTLING" in window.step_status_lbl.text()

    # Trigger expiration: execute auto-actions (chained measurement + scope capture)
    window._step_remaining_ms = 0
    window._step_countdown_tick()
    wait_for_ui(lambda: window._step_save_done and window._step_capture_done)

    # Verify point recorded and buttons re-enabled
    assert window._step_save_done is True
    assert window._step_capture_done is True
    assert "✓ RECORDED · 2.00 A" in window.step_status_lbl.text()
    assert window.btn_plus_step.isEnabled() is True
    assert window.btn_minus_step.isEnabled() is True
    assert window.btn_step_zero.isEnabled() is True

    # Verify workbook has the recorded point
    runs = window.store.list_runs()
    assert len(runs) >= 1
    assert runs[-1]["Mode"] == "Step Current"

    # 3. Step Up (+2 A): 2 A → 4 A and test Early Manual Override
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())
    assert window._manual_target_current == 4.0
    assert window._step_countdown_timer.isActive() is True
    assert "SETTLING" in window.step_status_lbl.text()

    # Click Save Reading early before timer expires
    window.step_save_btn.click()
    assert not window._step_countdown_timer.isActive()
    wait_for_ui(lambda: "4.00 A" in window.step_status_lbl.text() and not window._step_is_busy())
    assert window._step_save_done is True
    assert window._step_capture_done is True
    assert "✓ RECORDED · 4.00 A" in window.step_status_lbl.text()
    assert window.btn_plus_step.isEnabled() is True

    # 4. Test Post-Record Superseding: Click Save Reading again on the same 4 A point
    prior_measurements = window.store.get_run_measurements(window._step_run_id)
    assert len(prior_measurements) >= 1
    window.step_save_btn.click()
    wait_for_ui(lambda: len(window.store.get_run_measurements(window._step_run_id)) >= 2 and not window._step_is_busy())

    updated_measurements = window.store.get_run_measurements(window._step_run_id)
    statuses = [m["Status"] for m in updated_measurements]
    assert "Superseded" in statuses
    assert "Valid" in statuses
    assert window._manual_target_current == 4.0
    assert window.manual_target_spin.value() == 4.0

    # 5. A downward unload interrupts an unfinished upward point without saving it.
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())
    assert window._manual_target_current == 6.0
    assert window._step_countdown_timer.isActive() is True
    assert window.btn_plus_step.isEnabled() is False
    assert window.btn_minus_step.isEnabled() is True
    window.btn_minus_step.click()
    wait_for_ui(lambda: not window._step_is_busy())
    assert window._manual_target_current == 1.0
    assert 6.0 not in [m["RequestedIout_A"] for m in window.store.get_run_measurements(window._step_run_id)]

    # 6. ZERO/OFF has the same priority during another unfinished upward point.
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())
    assert window._manual_target_current == 3.0
    window.btn_step_zero.click()
    QtWidgets.QApplication.processEvents()

    assert not window._step_countdown_timer.isActive()
    assert window.hub.environment.load_enabled is False
    assert window.hub.environment.current_set == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert window.btn_plus_step.isEnabled() is True
    assert window.btn_minus_step.isEnabled() is True
    closed_runs = window.store.list_runs()
    closed_step = next(run for run in closed_runs if run["RunID"] == runs[-1]["RunID"])
    assert closed_step["Status"] == "Valid"
    assert "_StepCurrent_2to4A" in closed_step["CampaignName"]
    assert len(window.store.get_run_measurements(closed_step["RunID"])) >= 2

    # 7. Distinct Points: Step to 2 A again after zeroing produces a distinct run/point ID
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())
    assert window._manual_target_current == 2.0
    first_2a_run_id = runs[-1]["RunID"]
    new_2a_run_id = window._step_run_id
    assert first_2a_run_id != new_2a_run_id

    window.close()


def test_step_current_ascending_only_recording_and_descending_unload(tmp_path):
    import os, time
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, WorkbookStore
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.store = WorkbookStore(tmp_path / "step_ascend_test.xlsx")
    window.show()
    window.simulation.setChecked(True)

    def wait_for_ui(condition, timeout=10.0):
        start = time.time()
        while not condition() and (time.time() - start) < timeout:
            QtWidgets.QApplication.processEvents()
            QtCore.QThread.msleep(20)
        QtWidgets.QApplication.processEvents()

    # 1. Switch to STEP CURRENT mode
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()
    assert window.run_stack.currentIndex() == 1

    window.manual_step_inc.setValue(2.0)
    window.manual_step_dec.setValue(5.0)
    window.step_auto_save.setChecked(True)
    window.step_auto_capture.setChecked(True)

    # 2. Perform 5 ascending steps (+2 A each): 0 → 2 → 4 → 6 → 8 → 10 A
    for expected_amps in (2.0, 4.0, 6.0, 8.0, 10.0):
        window.btn_plus_step.click()
        wait_for_ui(lambda: window._step_countdown_timer.isActive())
        assert window._manual_target_current == expected_amps
        # Expire countdown to trigger measurement + capture
        window._step_remaining_ms = 0
        window._step_countdown_tick()
        wait_for_ui(lambda: f"{expected_amps:.2f} A" in window.step_status_lbl.text() and not window._step_is_busy())
        assert f"✓ RECORDED · {expected_amps:.2f} A" in window.step_status_lbl.text()
        assert window.btn_plus_step.isEnabled() is True
        assert window.btn_minus_step.isEnabled() is True

    run_id = window._step_run_id
    upward_measurements = window.store.get_run_measurements(run_id)
    assert len(upward_measurements) == 5
    upward_currents = [m["RequestedIout_A"] for m in upward_measurements]
    assert upward_currents == [2.0, 4.0, 6.0, 8.0, 10.0]
    scope_captures = [m for m in upward_measurements if m.get("ScopeCaptureStatus") == "Captured"]
    assert len(scope_captures) == 5

    # 3. Descending unload step 1: 10 A → 5 A (-5 A)
    window.btn_minus_step.click()
    wait_for_ui(lambda: not window._step_is_busy())
    assert window._manual_target_current == 5.0
    assert window.hub.environment.current_set == 5.0
    assert window.hub.environment.load_enabled is True
    assert "5.00 A · ON" in window.step_present_lbl.text()
    assert window._step_countdown_timer.isActive() is False
    assert "RECORDED" not in window.step_status_lbl.text()
    assert "READY" in window.step_status_lbl.text() or "5.00 A" in window.step_status_lbl.text()

    # Verify no new measurement row was added for 5 A
    meas_after_first_down = window.store.get_run_measurements(run_id)
    assert len(meas_after_first_down) == 5
    assert 5.0 not in [m["RequestedIout_A"] for m in meas_after_first_down]

    # 4. Descending unload step 2: 5 A → 0 A (-5 A)
    window.btn_minus_step.click()
    wait_for_ui(lambda: not window._step_is_busy())
    assert window._manual_target_current == 0.0
    assert window.hub.environment.current_set == 0.0
    assert window.hub.environment.load_enabled is False
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert window._step_countdown_timer.isActive() is False
    assert "RECORDED" not in window.step_status_lbl.text()
    assert "READY" in window.step_status_lbl.text()

    # Verify no new measurement row was added for 0 A
    meas_after_second_down = window.store.get_run_measurements(run_id)
    assert len(meas_after_second_down) == 5
    assert [m["RequestedIout_A"] for m in meas_after_second_down] == [2.0, 4.0, 6.0, 8.0, 10.0]

    # 5. Direct return to ZERO / OFF
    window.btn_step_zero.click()
    wait_for_ui(lambda: not window._step_is_busy())
    assert window.hub.environment.load_enabled is False
    assert window.hub.environment.current_set == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert "READY" in window.step_status_lbl.text()

    final_meas = window.store.get_run_measurements(run_id)
    assert len(final_meas) == 5
    assert [m["RequestedIout_A"] for m in final_meas] == [2.0, 4.0, 6.0, 8.0, 10.0]

    # 6. Test ZERO / OFF cancellation during active settling
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())
    assert window._step_countdown_timer.isActive() is True
    zero_cancel_run_id = window._step_run_id

    window.btn_step_zero.click()
    QtWidgets.QApplication.processEvents()
    assert window._step_countdown_timer.isActive() is False
    assert window.hub.environment.load_enabled is False
    assert window.hub.environment.current_set == 0.0
    assert "0.00 A · OFF" in window.step_present_lbl.text()
    assert "READY" in window.step_status_lbl.text()

    zero_meas = window.store.get_run_measurements(zero_cancel_run_id)
    assert len(zero_meas) == 0

    # 7. Test -5 A Step Down cancellation during active settling
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())
    window._step_remaining_ms = 0
    window._step_countdown_tick()
    wait_for_ui(lambda: "✓ RECORDED · 2.00 A" in window.step_status_lbl.text() and not window._step_is_busy())

    step_cancel_run_id = window._step_run_id
    assert len(window.store.get_run_measurements(step_cancel_run_id)) == 1

    # Step up to 4 A, then immediately step down (-5 A) while settling
    window.btn_plus_step.click()
    wait_for_ui(lambda: window._step_countdown_timer.isActive())
    assert window._step_countdown_timer.isActive() is True

    window._step_delta(-1)
    wait_for_ui(lambda: not window._step_is_busy())
    assert window._step_countdown_timer.isActive() is False
    assert window.hub.environment.load_enabled is False
    assert window.hub.environment.current_set == 0.0

    step_cancel_meas = window.store.get_run_measurements(step_cancel_run_id)
    assert len(step_cancel_meas) == 1
    assert step_cancel_meas[0]["RequestedIout_A"] == 2.0

    window.close()


def test_destructive_delete_confirmation_dialog(tmp_path: Path):
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, DeleteRunDialog, WorkbookStore, MEAS_HEADERS
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    # 1. Test DeleteRunDialog component layout, values, and focus
    run_info = {
        "RunID": "20260817-123456-EFFICIENCY-a1b2c",
        "CampaignName": "Efficiency Benchmark",
        "Status": "Valid",
        "DataSource": "Hardware",
        "VinTarget_V": 48.0,
        "Frequency_Hz": 100000.0,
    }
    dialog = DeleteRunDialog(run_info)
    dialog.show()
    QtWidgets.QApplication.processEvents()

    # Check modal and title
    assert dialog.isModal() is True
    assert "Permanently Delete Run" in dialog.windowTitle()

    # Verify no text entry fields exist (no QLineEdit for typing DELETE / RunID)
    line_edits = dialog.findChildren(QtWidgets.QLineEdit)
    assert len(line_edits) == 0, f"Expected 0 text entry fields, found {len(line_edits)}"

    # Verify clear warning wording
    all_text = " ".join([lbl.text() for lbl in dialog.findChildren(QtWidgets.QLabel)])
    assert "Permanently delete this run?" in all_text
    assert "This will remove workbook rows, linked captures, and run metadata. This cannot be undone." in all_text

    # Verify displayed metadata values
    assert "Efficiency Benchmark" in all_text
    assert "48 V" in all_text or "48.0 V" in all_text
    assert "100 kHz" in all_text
    assert "EFFICIENCY-a1b2c" in all_text or "a1b2c" in all_text
    assert "Valid" in all_text
    assert "Hardware" in all_text

    # Verify buttons: Cancel & Delete Permanently
    assert dialog.btn_cancel.text() == "Cancel"
    assert dialog.btn_delete.text() == "Delete Permanently"
    assert dialog.btn_cancel.isDefault() is True
    assert dialog.btn_delete.isDefault() is False
    assert dialog.focusWidget() == dialog.btn_cancel  # Default focus on Cancel!

    # Test reject action
    dialog.btn_cancel.click()
    assert dialog.result() == QtWidgets.QDialog.DialogCode.Rejected
    dialog.close()

    # 2. Test MainWindow integration: Cancel preserves run, Delete Permanently removes run
    excel_path = tmp_path / "test_del_campaign.xlsx"
    store = WorkbookStore(excel_path)
    store.create_run(run_record("RUN-TO-DELETE-1a601"))
    base = {name: "" for name in MEAS_HEADERS}
    base.update({"PointID": "P1", "RunID": "RUN-TO-DELETE-1a601", "Status": "Valid", "DataSource": "Simulation", "VinTarget_V": 48.0, "Frequency_Hz": 100000.0, "RequestedIout_A": 10.0})
    store.append_measurement(base)

    window = MainWindow()
    window.store = store
    window._load_history()
    QtWidgets.QApplication.processEvents()

    assert window.history_table.rowCount() >= 1
    # Check Run Short ID is in column 0
    assert window.history_table.item(0, 0).text() == "1a601"
    assert window.history_table.item(0, 1).text() == "RUN-TO-DELETE-1a601"
    window.history_table.selectRow(0)

    # Cancel case: monkeypatch DeleteRunDialog.exec to return Rejected
    DeleteRunDialog.exec = lambda self: QtWidgets.QDialog.DialogCode.Rejected
    window._history_delete()
    QtWidgets.QApplication.processEvents()
    assert len(store.list_runs()) == 1
    assert len(store.get_run_measurements("RUN-TO-DELETE-1a601")) == 1

    # Delete Permanently case: monkeypatch DeleteRunDialog.exec to return Accepted
    DeleteRunDialog.exec = lambda self: QtWidgets.QDialog.DialogCode.Accepted
    window._history_delete()
    QtWidgets.QApplication.processEvents()
    assert len(store.list_runs()) == 0
    assert len(store.get_run_measurements("RUN-TO-DELETE-1a601")) == 0
    assert "Permanently deleted run 1a601" in window.statusBar().currentMessage()

    window.close()


def test_delete_run_single_selection_and_duplicate_safety(tmp_path: Path, monkeypatch):
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow, DeleteRunDialog, WorkbookStore, MEAS_HEADERS
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    excel_path = tmp_path / "safety_test.xlsx"
    store = WorkbookStore(excel_path)
    store.create_run(run_record("RUN-111-aaa"))
    store.create_run(run_record("RUN-222-bbb"))

    window = MainWindow()
    window.store = store
    window._load_history()
    QtWidgets.QApplication.processEvents()

    # 1. Zero rows selected -> prompt to select one or more runs
    window.history_table.clearSelection()
    info_boxes = []
    monkeypatch.setattr(QtWidgets.QMessageBox, "information", lambda parent, title, text: info_boxes.append((title, text)))
    window._history_delete()
    assert len(info_boxes) == 1
    assert "Select one or more runs to delete." in info_boxes[0][1]
    assert len(store.list_runs()) == 2

    # 2. Dynamic button text update when multiple rows are selected
    window.history_table.selectAll()
    QtWidgets.QApplication.processEvents()
    assert window.history_del_btn.text() == "Delete 2 Runs..."

    # 3. Direct WorkbookStore.delete_run / delete_runs treats duplicate historical rows as one logical run
    # If a duplicate row exists in Runs, delete_run removes all matching rows as 1 logical run
    wb = store._load()
    wb["Runs"].append([store._as_cell(val) for val in run_record("RUN-111-aaa").values()])
    store._save_atomic(wb)

    deleted = store.delete_run("RUN-111-aaa")
    assert len(store.list_runs()) == 1
    assert store.list_runs()[0]["RunID"] == "RUN-222-bbb"

    # 4. Nonexistent RunID is an idempotent stale-selection refresh, not an error.
    assert store.delete_run("RUN-NONEXISTENT") == []
    assert [run["RunID"] for run in store.list_runs()] == ["RUN-222-bbb"]

    window.close()


def test_delete_multiple_selected_runs_batch(tmp_path: Path):
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow, DeleteBatchRunsDialog, WorkbookStore, MEAS_HEADERS
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    excel_path = tmp_path / "batch_del_test.xlsx"
    store = WorkbookStore(excel_path)
    store.create_run(run_record("RUN-BATCH-1-aaa"))
    store.create_run(run_record("RUN-BATCH-2-bbb"))
    store.create_run(run_record("RUN-BATCH-3-ccc"))

    base = {name: "" for name in MEAS_HEADERS}
    store.append_measurement({**base, "PointID": "P1", "RunID": "RUN-BATCH-1-aaa", "Status": "Valid", "RequestedIout_A": 10.0})
    store.append_measurement({**base, "PointID": "P2", "RunID": "RUN-BATCH-2-bbb", "Status": "Valid", "RequestedIout_A": 20.0})
    store.append_measurement({**base, "PointID": "P3", "RunID": "RUN-BATCH-3-ccc", "Status": "Valid", "RequestedIout_A": 30.0})

    window = MainWindow()
    window.store = store
    window._load_history()
    QtWidgets.QApplication.processEvents()

    assert window.history_table.rowCount() == 3

    # Select top 2 rows
    window.history_table.clearSelection()
    window.history_table.item(0, 0).setSelected(True)
    window.history_table.item(1, 0).setSelected(True)
    QtWidgets.QApplication.processEvents()

    assert window.history_del_btn.text() == "Delete 2 Runs..."

    # Test DeleteBatchRunsDialog structure
    runs_to_del = [run_record("RUN-BATCH-1-aaa"), run_record("RUN-BATCH-2-bbb")]
    dialog = DeleteBatchRunsDialog(runs_to_del, window)
    assert "Permanently delete 2 selected runs?" in dialog.findChildren(QtWidgets.QLabel)[0].text()
    assert dialog.btn_cancel.isDefault() is True
    assert dialog.btn_delete.text() == "Delete 2 Runs Permanently"
    dialog.close()

    # Cancel case: monkeypatch DeleteBatchRunsDialog.exec to return Rejected
    DeleteBatchRunsDialog.exec = lambda self: QtWidgets.QDialog.DialogCode.Rejected
    window._history_delete()
    QtWidgets.QApplication.processEvents()
    assert len(store.list_runs()) == 3

    # Delete case: monkeypatch DeleteBatchRunsDialog.exec to return Accepted
    DeleteBatchRunsDialog.exec = lambda self: QtWidgets.QDialog.DialogCode.Accepted
    window._history_delete()
    QtWidgets.QApplication.processEvents()

    # Verify exactly 1 run remains in store (RUN-BATCH-1-aaa since row 0 & 1 were reversed/deleted)
    runs_after = store.list_runs()
    assert len(runs_after) == 1
    assert "Permanently deleted 2 runs" in window.statusBar().currentMessage()
    assert len(store.get_run_measurements(runs_after[0]["RunID"])) == 1

    window.close()


def test_delete_run_preserves_workbook_integrity_when_capture_is_already_missing(tmp_path: Path):
    store = WorkbookStore(tmp_path / "results" / "hardware_campaign.xlsx")
    run_id = "RUN-CAPTURE-CLEANUP"
    store.create_run(run_record(run_id, "Hardware"))
    capture_dir = capture_root_for_source(store.path, "Hardware")
    capture_dir.mkdir(parents=True)
    existing_png = capture_dir / "point.png"
    existing_png.write_bytes(b"png")
    missing_csv = capture_dir / "already_missing.csv"
    record = {name: "" for name in MEAS_HEADERS}
    record.update({
        "PointID": "P1", "RunID": run_id, "Status": "Valid", "DataSource": "Hardware",
        "ScopePNG": str(existing_png), "ScopeCSV": str(missing_csv),
    })
    store.append_measurement(record)

    assert store.delete_run(run_id) == [existing_png]
    assert not existing_png.exists()
    assert store.list_runs() == []
    assert store.get_run_measurements(run_id) == []
    assert store.delete_run(run_id) == []



def test_workbook_corruption_recovery_from_backup(tmp_path: Path):
    excel_path = tmp_path / "campaign_data.xlsx"
    store = WorkbookStore(excel_path)
    store.create_run(run_record("RUN-PRE-CORRUPT"))
    base = {name: "" for name in MEAS_HEADERS}
    base.update({"PointID": "P1", "RunID": "RUN-PRE-CORRUPT", "Status": "Valid", "DataSource": "Simulation", "VinTarget_V": 48.0, "Frequency_Hz": 100000.0, "RequestedIout_A": 10.0})
    store.append_measurement(base)
    store.append_measurement({**base, "PointID": "P2", "RequestedIout_A": 20.0})

    # Verify backup exists
    bak_path = excel_path.with_suffix(".xlsx.bak")
    assert bak_path.exists()

    # Intentionally corrupt live workbook (bad magic number / damaged ZIP)
    with open(excel_path, "wb") as f:
        f.write(b"PK\x00\x00DAMAGED_CORRUPTED_ZIP_MAGIC_BYTES_1234567890")

    # Load store with recovery accepted
    prompt_called = []
    def on_prompt(title, msg):
        prompt_called.append((title, msg))
        assert "Workbook is damaged. Recover from last backup?" in msg
        assert "Results workbook could not be read. The XLSX file appears damaged. No hardware fault occurred." in msg
        return True

    recovered_store = WorkbookStore(excel_path, prompt_fn=on_prompt)
    runs = recovered_store.list_runs()
    assert len(prompt_called) == 1
    assert len(runs) == 1
    assert runs[0]["RunID"] == "RUN-PRE-CORRUPT"
    assert len(recovered_store.get_run_measurements("RUN-PRE-CORRUPT")) >= 1

    # Verify damaged file was preserved with timestamp
    corrupt_files = list(tmp_path.glob("campaign_data_CORRUPT_*.xlsx"))
    assert len(corrupt_files) == 1
    assert corrupt_files[0].read_bytes().startswith(b"PK\x00\x00DAMAGED")


def test_workbook_corruption_recovery_rejected_raises_friendly_error(tmp_path: Path):
    excel_path = tmp_path / "campaign_reject.xlsx"
    store = WorkbookStore(excel_path)
    store.create_run(run_record("RUN-1"))
    store.create_run(run_record("RUN-2"))

    # Corrupt live workbook
    with open(excel_path, "wb") as f:
        f.write(b"NOT_A_VALID_ZIP_FILE")

    # Reject recovery
    def on_reject(title, msg):
        return False

    rejected_store = WorkbookStore(excel_path, prompt_fn=on_reject)
    with pytest.raises(RuntimeError) as exc_info:
        rejected_store.list_runs()

    assert "Results workbook could not be read. The XLSX file appears damaged. No hardware fault occurred." in str(exc_info.value)
    # Original damaged file was NOT overwritten
    assert excel_path.read_bytes() == b"NOT_A_VALID_ZIP_FILE"


def test_workbook_corruption_no_backup_creates_fresh_and_preserves_corrupt(tmp_path: Path):
    excel_path = tmp_path / "no_bak_corrupt.xlsx"
    with open(excel_path, "wb") as f:
        f.write(b"TOTALLY_CORRUPT_NO_BACKUP")

    prompt_called = []
    def on_prompt(title, msg):
        prompt_called.append((title, msg))
        assert "No valid backup was found. Create a new workbook?" in msg
        return True

    store = WorkbookStore(excel_path, prompt_fn=on_prompt)
    runs = store.list_runs()
    assert len(prompt_called) == 1
    assert len(runs) == 0

    # Verify corrupt file preserved
    corrupt_files = list(tmp_path.glob("no_bak_corrupt_CORRUPT_*.xlsx"))
    assert len(corrupt_files) == 1
    assert corrupt_files[0].read_bytes() == b"TOTALLY_CORRUPT_NO_BACKUP"


def test_save_atomic_validates_temporary_file_before_replacing_live(tmp_path: Path, monkeypatch):
    excel_path = tmp_path / "atomic_validate.xlsx"
    store = WorkbookStore(excel_path)
    store.create_run(run_record("ORIGINAL_RUN"))

    # Mock openpyxl load_workbook to fail validation on .tmp files
    from openpyxl import load_workbook as real_load_workbook
    def bad_load_validation(path, **kwargs):
        if str(path).endswith(".tmp.xlsx"):
            raise ValueError("Simulated corrupt temporary file validation error")
        return real_load_workbook(path, **kwargs)

    monkeypatch.setattr("openpyxl.load_workbook", bad_load_validation)

    # Attempting to save another run should fail validation and not corrupt/replace live workbook
    with pytest.raises(RuntimeError) as exc_info:
        store.create_run(run_record("SECOND_RUN"))

    assert "Failed to validate temporary workbook before commit" in str(exc_info.value)
    # Live workbook is intact
    monkeypatch.undo()
    assert len(store.list_runs()) == 1
    assert store.list_runs()[0]["RunID"] == "ORIGINAL_RUN"


def test_workbook_duplicate_run_id_idempotency_and_migration_pass(tmp_path: Path):
    from sid_bench_gui import WorkbookStore, RUN_HEADERS, MEAS_HEADERS, EVENT_HEADERS
    excel_path = tmp_path / "dup_migration_test.xlsx"
    store = WorkbookStore(excel_path)

    # 1. Test create_run idempotency
    rec1 = {
        "RunID": "20260817-062141-test1-3a2b1",
        "CampaignName": "test1",
        "Status": "Aborted",
        "DataSource": "Simulation",
        "VinTarget_V": 48.0,
        "Frequency_Hz": 100000.0,
        "ModulationLabel": "test1",
    }
    created1 = store.create_run(rec1)
    assert created1 is True
    assert len(store.list_runs()) == 1

    # Second call with same RunID must return False and NOT create a second row
    rec2 = dict(rec1)
    rec2["Status"] = "Valid"
    created2 = store.create_run(rec2)
    assert created2 is False
    assert len(store.list_runs()) == 1

    # 2. Simulate historical workbook with multiple duplicate rows for the same RunID
    wb = store._load()
    r_sheet = wb["Runs"]
    # Manually append duplicate rows as an old workbook would have had
    r_sheet.append([store._as_cell(rec1.get(h, "")) for h in RUN_HEADERS])
    rec_valid = dict(rec1)
    rec_valid["Status"] = "Valid"
    rec_valid["Notes"] = "Preserved notes"
    r_sheet.append([store._as_cell(rec_valid.get(h, "")) for h in RUN_HEADERS])
    store._save_atomic(wb)

    # Add measurements for 3a2b1
    store.append_measurement({
        "PointID": "P1",
        "RunID": "20260817-062141-test1-3a2b1",
        "Status": "Valid",
        "DataSource": "Simulation",
        "VinTarget_V": 48.0,
        "Frequency_Hz": 100000.0,
        "RequestedIout_A": 10.0,
    })

    # 3. Reloading through WorkbookStore must consolidate duplicate rows into 1 canonical row
    store2 = WorkbookStore(excel_path)
    runs = store2.list_runs()
    assert len(runs) == 1
    assert runs[0]["RunID"] == "20260817-062141-test1-3a2b1"
    assert runs[0]["Status"] == "Valid"
    assert runs[0]["Notes"] == "Preserved notes"

    # Measurements must remain intact
    meas = store2.get_run_measurements("20260817-062141-test1-3a2b1")
    assert len(meas) == 1
    assert meas[0]["RequestedIout_A"] == 10.0

    # Events sheet must log consolidation
    wb2 = store2._load()
    events = [wb2["Events"].cell(r, 4).value for r in range(2, wb2["Events"].max_row + 1)]
    assert "Duplicate Runs rows consolidated" in events

    # 4. Deleting this run successfully removes the canonical row and its measurements
    deleted = store2.delete_run("20260817-062141-test1-3a2b1")
    assert len(store2.list_runs()) == 0
    assert len(store2.get_run_measurements("20260817-062141-test1-3a2b1")) == 0


def test_set_current_mode_delayed_auto_recording(tmp_path: Path):
    import os, time
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, WorkbookStore

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    excel_path = tmp_path / "set_current_auto_test.xlsx"
    store = WorkbookStore(excel_path)

    window = MainWindow()
    window.store = store
    window.cap_val = 50.0
    window.manual_target_spin.setMaximum(50.0)
    window.show()
    window.simulation.setChecked(True)
    window.btn_mode_direct.click()
    QtWidgets.QApplication.processEvents()

    def wait_for_ui(condition, timeout=5.0):
        start = time.time()
        while not condition() and (time.time() - start) < timeout:
            QtWidgets.QApplication.processEvents()
            QtCore.QThread.msleep(20)
        QtWidgets.QApplication.processEvents()

    # 1. Verify SET CURRENT initial UI elements and Advanced Disclosure
    assert window.direct_adv_box.isHidden()
    assert window.direct_auto_delay.value() == 4.0
    assert window.direct_auto_save.isChecked() is True
    assert window.direct_auto_capture.isChecked() is True
    assert window.direct_status_lbl.text() == "READY"
    assert window.btn_direct_set.isEnabled() is True
    assert window.btn_direct_zero.isEnabled() is True

    # Expand Advanced
    window.btn_adv_direct.click()
    QtWidgets.QApplication.processEvents()
    assert not window.direct_adv_box.isHidden()

    # 2. Enter positive target: 17.0 A and click SET CURRENT
    window.manual_target_spin.setValue(17.0)
    window.btn_direct_set.click()
    wait_for_ui(lambda: window._manual_countdown_timer.isActive())

    # Verify load set and ON
    assert window._manual_target_current == 17.0
    assert window.hub.environment.current_set == 17.0
    assert window.hub.environment.load_enabled is True
    assert "17.00 A · ON" in window.step_present_lbl.text()

    # Verify button locking: SET CURRENT disabled during pending settling/recording
    assert window.btn_direct_set.isEnabled() is False
    assert window.btn_direct_zero.isEnabled() is True  # ZERO / OFF is safety action and stays enabled
    assert window.emergency_stop_btn.isEnabled() is True
    assert window.manual_target_spin.isEnabled() is True
    window.manual_target_spin.setValue(4.0)  # Prepare the next target; it must not auto-run.
    assert window._manual_target_current == 17.0
    assert window._manual_countdown_timer.isActive() is True
    assert "SETTLING" in window.direct_status_lbl.text()

    # Countdown tick
    window._manual_countdown_tick()
    QtWidgets.QApplication.processEvents()
    assert "SETTLING" in window.direct_status_lbl.text()

    # Trigger expiration: execute auto-actions (chained auto-save + auto-capture)
    window._manual_remaining_ms = 0
    window._manual_countdown_tick()
    wait_for_ui(lambda: window._manual_save_done and window._manual_capture_done)

    # Verify point recorded and SET CURRENT re-enabled
    assert window._manual_save_done is True
    assert window._manual_capture_done is True
    assert "✓ RECORDED · 17.00 A" in window.direct_status_lbl.text()
    assert window.btn_direct_set.isEnabled() is True
    assert window.btn_direct_zero.isEnabled() is True

    # Verify workbook record
    runs = window.store.list_runs()
    assert len(runs) >= 1
    assert runs[-1]["Mode"] == "Set Current"
    meas = window.store.get_run_measurements(runs[-1]["RunID"])
    assert len(meas) == 1
    assert meas[0]["RequestedIout_A"] == 17.0
    assert meas[0]["Status"] == "Valid"
    assert meas[0]["ScopeCaptureStatus"] == "Captured"

    # 3. Enter target 15.0 A and test Early Manual Override before auto save occurs
    window.manual_target_spin.setValue(15.0)
    window.btn_direct_set.click()
    wait_for_ui(lambda: window._manual_countdown_timer.isActive())
    assert window._manual_target_current == 15.0
    assert window._manual_countdown_timer.isActive() is True
    assert "SETTLING" in window.direct_status_lbl.text()

    # Click manual Save Reading early
    window.manual_save_btn.click()
    assert not window._manual_countdown_timer.isActive()
    wait_for_ui(lambda: window._manual_save_done and window._manual_capture_done)
    assert window._manual_save_done is True
    assert window._manual_capture_done is True
    assert "✓ RECORDED · 15.00 A" in window.direct_status_lbl.text()
    assert window.btn_direct_set.isEnabled() is True

    # 4. Each SET CURRENT action is a complete single-point run with a useful full name.
    assert window._manual_run_id == ""
    set_runs = [run for run in window.store.list_runs() if run["Mode"] == "Set Current"]
    assert any("_SetCurrent_17A" in run["CampaignName"] for run in set_runs)
    assert any("_SetCurrent_15A" in run["CampaignName"] for run in set_runs)

    # 5. Test Safety ZERO / OFF cancellation during settling
    window.manual_target_spin.setValue(20.0)
    window.btn_direct_set.click()
    wait_for_ui(lambda: window._manual_countdown_timer.isActive())
    assert window._manual_target_current == 20.0
    assert window._manual_countdown_timer.isActive() is True
    assert window.btn_direct_set.isEnabled() is False

    # Click ZERO / OFF
    window.btn_direct_zero.click()
    QtWidgets.QApplication.processEvents()

    assert not window._manual_countdown_timer.isActive()
    assert window.hub.environment.load_enabled is False
    assert window.hub.environment.current_set == 0.0
    window.close()


def test_pa2201a_verified_scpi_command_sequence():
    from sid_instruments import PA2201A

    class FakePASession:
        def __init__(self):
            self.writes = []
            self.queries = []
            self.closed = False
            self.timeout = 3000

        def query(self, command: str) -> str:
            self.queries.append(command)
            if "*IDN?" in command:
                return "Keysight Technologies,PA2201A,TESTPA001,1.0.0"
            if "ANALyze:QUALity1:VOLTage:DC?" in command:
                return "48.02"
            if "ANALyze:QUALity1:CURRent:DC?" in command:
                return "2.49"
            if "ANALyze:QUALity2:VOLTage:DC?" in command:
                return "12.01"
            if "SYSTem:ERRor?" in command:
                return "+0,\"No error\""
            return "0"

        def write(self, command: str) -> None:
            self.writes.append(command)

        def close(self):
            self.closed = True

    class FakeManager:
        def __init__(self, session):
            self.s = session

        def open(self):
            class RM:
                def __init__(self, s):
                    self.s = s

                def open_resource(self, _):
                    return self.s

            return RM(self.s)

    sess = FakePASession()
    pa = PA2201A(FakeManager(sess), "USB::PA", ("PA2201",))
    snap = pa.read_snapshot(settle_s=0.0)
    pa.release()

    # 1. Verify exact configuration sequence: W_1S and DC_MODE on CH1 and CH2
    assert "ANALyze:ENABle ON" in sess.writes
    assert "ANALyze:WINDow W_1S" in sess.writes
    assert "ANALyze:SOURce1:SYNC LINE" in sess.writes
    assert "ANALyze:MODE1 DC_MODE" in sess.writes
    assert "ANALyze:SOURce2:SYNC LINE" in sess.writes
    assert "ANALyze:MODE2 DC_MODE" in sess.writes

    # 2. Verify exact measurement trigger sequence
    assert "ANALyze:QUALity1:MEASure" in sess.writes
    assert "ANALyze:QUALity2:MEASure" in sess.writes

    # 3. Verify exact 3 queries executed
    assert "ANALyze:QUALity1:VOLTage:DC?" in sess.queries
    assert "ANALyze:QUALity1:CURRent:DC?" in sess.queries
    assert "ANALyze:QUALity2:VOLTage:DC?" in sess.queries

    # 4. Verify invalid/unwanted queries were NOT sent
    assert not any(":MEASure:VOLTage:DC?" in q for q in sess.queries)
    assert not any(":MEASure:CURRent:DC?" in q for q in sess.queries)
    assert not any("QUALity2:CURRent" in q for q in sess.queries)
    assert not any("POWer" in q for q in sess.queries)

    # 5. Verify snapshot returned
    assert snap.valid is True
    assert snap.status == "Connected"
    assert snap.values["vin"] == 48.02
    assert snap.values["iin"] == 2.49
    assert snap.values["vout"] == 12.01
    assert "pin" not in snap.values  # Pin is not computed inside PA driver


def test_pa2201a_sentinel_rejection_and_invalid_status():
    from sid_instruments import PA2201A

    class SentinelSession:
        def __init__(self, voltage_val: str):
            self.voltage_val = voltage_val
            self.queries = []
            self.writes = []

        def query(self, command: str) -> str:
            self.queries.append(command)
            if "*IDN?" in command:
                return "Keysight Technologies,PA2201A,TESTPA001,1.0.0"
            if "ANALyze:QUALity1:VOLTage:DC?" in command:
                return self.voltage_val
            if "ANALyze:QUALity1:CURRent:DC?" in command:
                return "2.0"
            if "ANALyze:QUALity2:VOLTage:DC?" in command:
                return "12.0"
            if "SYSTem:ERRor?" in command:
                return "-230,\"Data corrupt or stale\""
            return "0"

        def write(self, command: str) -> None:
            self.writes.append(command)

        def close(self):
            pass

    class Manager:
        def __init__(self, s): self.s = s
        def open(self):
            class RM:
                def __init__(self, s): self.s = s
                def open_resource(self, _): return self.s
            return RM(self.s)

    for sentinel in ["9.910000E+37", "+9.91E+37", "NaN", "Inf", "999999"]:
        sess = SentinelSession(sentinel)
        pa = PA2201A(Manager(sess), "USB::PA", ("PA2201",))
        snap = pa.read_snapshot(settle_s=0.0)
        pa.release()

        assert snap.valid is False
        assert snap.status == "Connected · Invalid Data"
        assert snap.values["vin"] is None
        assert "CH1 DC voltage" in snap.warning
        assert "Instrument error" in snap.warning


def test_pa2201a_read_error_separates_from_offline():
    from sid_instruments import PA2201A

    class FailingReadSession:
        def query(self, command: str) -> str:
            if "*IDN?" in command:
                return "Keysight Technologies,PA2201A,TESTPA001,1.0.0"
            if "ANALyze:QUALity1:VOLTage:DC?" in command:
                raise TimeoutError("VISA timeout querying voltage")
            if "SYSTem:ERRor?" in command:
                return "-113,\"Undefined header\""
            return "0"

        def write(self, command: str) -> None:
            pass

        def close(self):
            pass

    class Manager:
        def __init__(self, s): self.s = s
        def open(self):
            class RM:
                def __init__(self, s): self.s = s
                def open_resource(self, _): return self.s
            return RM(self.s)

    sess = FailingReadSession()
    pa = PA2201A(Manager(sess), "USB::PA", ("PA2201",))
    snap = pa.read_snapshot(settle_s=0.0)
    pa.release()

    assert snap.valid is False
    assert snap.status == "Connected · Read Error"
    assert "PA connected, but CH1 DC voltage query failed" in snap.warning
    assert "Undefined header" in snap.warning


def test_pa2201a_ui_card_and_bench_check_amber_read_error():
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow, InstrumentSnapshot, SUCCESS_GREEN, WARNING_AMBER

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.show()

    # 1. PA card fields: Vin, Iin, Vout (Pin is not on card)
    pa_card = window.cards["pa"]
    assert set(pa_card.value_labels.keys()) == {"Vin", "Iin", "Vout"}

    # 2. Simulate valid snapshot -> Green Connected badge
    valid_snap = InstrumentSnapshot("pa", {"vin": 48.0, "iin": 2.5, "vout": 12.0}, valid=True, status="Connected")
    pa_card._received(valid_snap)
    assert pa_card.status_badge.text() == "Connected"
    assert pa_card.value_labels["Vin"][0].text() == "48"
    assert pa_card.value_labels["Iin"][0].text() == "2.5"
    assert pa_card.value_labels["Vout"][0].text() == "12"

    # 3. Simulate read error snapshot (*IDN? passed, but measurement failed) -> Amber Connected · Read Error
    error_snap = InstrumentSnapshot("pa", {"vin": None, "iin": None, "vout": None}, valid=False, warning="CH1 voltage measurement unavailable", status="Connected · Read Error")
    pa_card._received(error_snap)
    assert pa_card.status_badge.text() == "Connected · Read Error"
    assert pa_card.value_labels["Vin"][0].text() == "—"
    assert pa_card.value_labels["Iin"][0].text() == "—"
    assert pa_card.value_labels["Vout"][0].text() == "—"
    assert "CH1 voltage measurement unavailable" in pa_card.status_badge.toolTip()

    # 4. Simulate actual VISA physical disconnect -> Red Offline badge
    pa_card._failed("VISA open failed: device disconnected")
    assert pa_card.status_badge.text() == "Offline"
    window.close()


from datetime import datetime, timedelta
from PyQt6 import QtWidgets
from sid_bench_gui import MainWindow


def test_bench_startup_not_checked_and_no_premature_stale(tmp_path):
    """Verify that all bench cards start in 'Not Checked' neutral state and never trigger Stale before valid reads."""
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    wb_path = tmp_path / "test_startup.xlsx"
    store = WorkbookStore(wb_path)

    window = MainWindow()
    window.store = store

    # Verify initial card status badges are all 'Not Checked'
    assert window.cards["pa"].status_badge.text() == "Not Checked"
    assert window.load_card.status_badge.text() == "Not Checked"
    assert window.supply_card.status_badge.text() == "Not Checked"
    assert window.scope_card.status_badge.text() == "Not Checked"

    # Verify metrics start with '—' rather than fake zeroes
    for name, (lbl, _) in window.cards["pa"].value_labels.items():
        assert lbl.text() == "—"
    assert "—" in window.load_card.metric_labels["Iout"].text()
    assert "—" in window.load_card.metric_labels["Load V"].text()
    assert "—" in window.load_card.metric_labels["Load P"].text()
    assert window.load_card.metric_labels["Load State"].text() == "—"

    # Call update_age() on all cards: should NOT transition to Stale
    window.cards["pa"].update_age()
    window.load_card.update_age()
    window.supply_card.update_age()
    window.scope_card.update_age()

    assert window.cards["pa"].status_badge.text() == "Not Checked"
    assert window.load_card.status_badge.text() == "Not Checked"
    assert window.supply_card.status_badge.text() == "Not Checked"
    assert window.scope_card.status_badge.text() == "Not Checked"

    window.close()


def test_chroma_load_off_is_connected_not_stale(tmp_path):
    """Verify Chroma with load physically OFF shows Connected green badge with Load State: OFF."""
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    wb_path = tmp_path / "test_chroma_off.xlsx"
    store = WorkbookStore(wb_path)

    window = MainWindow()
    window.store = store
    load_card = window.load_card

    # 1. Valid reading where load is physically OFF
    snap_off = InstrumentSnapshot(
        "load",
        {"iout": 0.0, "vout": 12.02, "pout": 0.0, "input_on": False},
        valid=True,
        status="Connected",
        timestamp=datetime.now().isoformat()
    )
    load_card.last_snapshot = snap_off
    load_card._render_values()

    assert load_card.status_badge.text() == "Connected"
    assert load_card.metric_labels["Load State"].text() == "OFF"
    assert load_card.metric_labels["Iout"].text() == "0.00 A"
    assert load_card.metric_labels["Load V"].text() == "12.02 V"
    assert load_card.metric_labels["Load P"].text() == "0.00 W"

    # update_age() immediately: should remain Connected
    load_card.update_age()
    assert load_card.status_badge.text() == "Connected"

    # Simulate snapshot aged 15 seconds: should transition to Stale
    old_time = (datetime.now() - timedelta(seconds=15)).isoformat()
    snap_old = InstrumentSnapshot(
        "load",
        {"iout": 0.0, "vout": 12.02, "pout": 0.0, "input_on": False},
        valid=True,
        status="Connected",
        timestamp=old_time
    )
    load_card.last_snapshot = snap_old
    load_card.update_age()
    assert load_card.status_badge.text() == "Stale"

    # 2. Valid reading where load is physically ON
    snap_on = InstrumentSnapshot(
        "load",
        {"iout": 17.50, "vout": 11.95, "pout": 209.12, "input_on": True},
        valid=True,
        status="Connected",
        timestamp=datetime.now().isoformat()
    )
    load_card.last_snapshot = snap_on
    load_card._render_values()

    assert load_card.status_badge.text() == "Connected"
    assert load_card.metric_labels["Load State"].text() == "ON"
    assert load_card.metric_labels["Iout"].text() == "17.50 A"

    # 3. Read error snapshot: should show Connected · Read Error and Load State: Unknown
    snap_err = InstrumentSnapshot(
        "load",
        {"iout": None, "vout": None, "pout": None, "input_on": None},
        valid=False,
        status="Connected · Read Error",
        warning="Query timeout on Chroma 63206A"
    )
    load_card.last_snapshot = snap_err
    load_card._render_values()

    assert load_card.status_badge.text() == "Connected · Read Error"
    assert load_card.metric_labels["Load State"].text() == "Unknown"

    window.close()


def test_bench_operation_busy_lock(tmp_path):
    """Verify bench operations lock each other, update button texts, and preserve emergency LOAD OFF."""
    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    wb_path = tmp_path / "test_busy_lock.xlsx"
    store = WorkbookStore(wb_path)

    window = MainWindow()
    window.store = store

    # Initial state: bench buttons enabled, emergency stop enabled
    assert window.check_bench_btn.isEnabled() is True
    assert window.disc_btn.isEnabled() is True
    assert window.release_btn.isEnabled() is True
    assert window.emergency_stop_btn.isEnabled() is True
    assert window.check_bench_btn.text() == "Check / Refresh Entire Bench"
    assert window.disc_btn.text() == "Discover VISA Devices"
    assert window.release_btn.text() == "Release All Devices"

    # Lock bench for "check" operation
    window._set_bench_busy(True, "check")
    assert window.check_bench_btn.isEnabled() is False
    assert window.disc_btn.isEnabled() is False
    assert window.release_btn.isEnabled() is False
    assert window.check_bench_btn.text() == "Checking Bench..."
    assert window.emergency_stop_btn.isEnabled() is True  # Emergency action never disabled!

    # Unlock bench
    window._set_bench_busy(False)
    assert window.check_bench_btn.isEnabled() is True
    assert window.disc_btn.isEnabled() is True
    assert window.release_btn.isEnabled() is True
    assert window.check_bench_btn.text() == "Check / Refresh Entire Bench"
    assert window.disc_btn.text() == "Discover VISA Devices"
    assert window.release_btn.text() == "Release All Devices"

    # Lock bench for "discover" operation
    window._set_bench_busy(True, "discover")
    assert window.disc_btn.text() == "Discovering..."
    assert window.check_bench_btn.isEnabled() is False
    assert window.disc_btn.isEnabled() is False
    assert window.release_btn.isEnabled() is False
    assert window.emergency_stop_btn.isEnabled() is True

    # Lock bench for "release" operation
    window._set_bench_busy(True, "release")
    assert window.release_btn.text() == "Releasing..."
    assert window.check_bench_btn.isEnabled() is False
    assert window.disc_btn.isEnabled() is False
    assert window.release_btn.isEnabled() is False
    assert window.emergency_stop_btn.isEnabled() is True

    # Restore
    window._set_bench_busy(False)
    assert window.check_bench_btn.isEnabled() is True
    assert window.disc_btn.isEnabled() is True
    assert window.release_btn.isEnabled() is True

    window.close()


def test_hard_dwell_timing_and_measure_last_acquisition(tmp_path):
    """Verify that SweepWorker respects Wait/Pulse ON time as hard bounded dwell, and begins acquisition at Wait - Measure Last."""
    import time
    from sid_bench_gui import SweepWorker, InstrumentHub, WorkbookStore
    from sid_instruments import InstrumentSnapshot

    event_timeline = []

    class MockLoad:
        def __init__(self):
            self.identity = "Chroma,63206A,123,1.0"
            self.current = 0.0
            self.input_on = False
        def connect(self, persistent=True): pass
        def release(self): pass
        def set_current(self, amps: float):
            self.current = amps
            event_timeline.append(("set_current", amps, time.monotonic()))
        def set_input(self, state: bool):
            self.input_on = state
            event_timeline.append(("set_input", state, time.monotonic()))
        def safe_off(self):
            self.input_on = False
            event_timeline.append(("safe_off", False, time.monotonic()))
        def read_snapshot(self):
            event_timeline.append(("read_load", self.current, time.monotonic()))
            return InstrumentSnapshot("load", {"current": self.current, "voltage": 12.0, "power": self.current * 12.0, "input_on": self.input_on}, valid=True)

    class MockPA:
        def __init__(self):
            self.identity = "Keysight,PA2201A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def read_snapshot(self):
            event_timeline.append(("read_pa", 48.0, time.monotonic()))
            time.sleep(0.04)  # Simulate small PA measurement time
            return InstrumentSnapshot("pa", {"vin": 48.0, "iin": 2.5, "vout": 12.0}, valid=True)

    class MockScope:
        def __init__(self):
            self.identity = "Keysight,MSOX4024A,123,1.0"
        def connect(self, persistent=True): pass
        def release(self): pass
        def capture(self, png, csv): pass

    hub = InstrumentHub(True, {})
    hub.instruments["load"] = MockLoad()
    hub.instruments["pa"] = MockPA()
    hub.instruments["scope"] = MockScope()

    wb_path = tmp_path / "test_hard_dwell.xlsx"
    store = WorkbookStore(wb_path)

    # 1. Continuous mode: Wait = 0.20 s, Measure last = 0.08 s -> Pre-settle = 0.12 s
    settings_cont = {
        "run_id": "TEST-CONT-DWELL",
        "run_record": {
            "RunID": "TEST-CONT-DWELL", "CampaignName": "ContDwellTest", "Created": "2026-08-17T00:00:00Z",
            "Status": "Aborted", "DataSource": "Simulation", "Mode": "Continuous", "VinTarget_V": 48.0,
            "ModulationLabel": "ContDwellTest", "Frequency_Hz": 100000.0, "ModulationMetadata": "",
            "SupplyConfiguration": [], "Length_mm": 24.0, "Width_mm": 16.0, "Height_mm": 3.4,
            "WorkingCap_A": 60.0, "Notes": "", "InstrumentIdentities": {},
            "FPGASnapshotStatus": "Skipped", "FPGASnapshot": {}, "Warnings": "",
        },
        "points": [10.0],
        "capture_points": set(),
        "mode": "Continuous",
        "settle": 0.20,  # Total hard dwell
        "dwell": 0.20,
        "sample_window": 0.08,  # Measure last
        "sample_count": 1,
        "cooldown": 0.0,
        "working_cap": 60.0,
        "vin_target": 48.0,
        "frequency": 100000.0,
        "dimensions": (24.0, 16.0, 3.4),
        "supply_channels": [],
        "psu_required": False,
        "data_source": "Simulation",
        "duplicate_action": "keep",
        "return_to_zero_step": 10.0,
    }

    t0 = time.monotonic()
    event_timeline.clear()
    worker = SweepWorker(hub, store, settings_cont)
    worker.run()
    total_elapsed = time.monotonic() - t0

    # Total point duration should be at least settle time (0.20s)
    assert total_elapsed >= 0.19

    # Check timing of load set vs measurement start
    set_events = [e for e in event_timeline if e[0] == "set_current" and e[1] == 10.0]
    read_pa_events = [e for e in event_timeline if e[0] == "read_pa"]
    assert len(set_events) == 1
    assert len(read_pa_events) == 1

    t_set = set_events[0][2]
    t_meas = read_pa_events[0][2]
    settle_delay = t_meas - t_set

    # Pre-settle delay should be approximately 0.20 - 0.08 = 0.12 s (>= 0.10 s)
    assert settle_delay >= 0.10

    # 2. Pulse mode: ON time = 0.20 s, Measure last = 0.08 s
    settings_pulse = dict(settings_cont)
    settings_pulse["run_id"] = "TEST-PULSE-DWELL"
    settings_pulse["mode"] = "Pulse"
    settings_pulse["run_record"] = dict(settings_cont["run_record"])
    settings_pulse["run_record"]["RunID"] = "TEST-PULSE-DWELL"
    settings_pulse["run_record"]["Mode"] = "Pulse"

    event_timeline.clear()
    t0 = time.monotonic()
    worker_pulse = SweepWorker(hub, store, settings_pulse)
    worker_pulse.run()
    pulse_elapsed = time.monotonic() - t0

    assert pulse_elapsed >= 0.19
    # Pulse load must be turned OFF at the end of the pulse
    safe_offs = [e for e in event_timeline if e[0] in ("safe_off", "set_input") and e[1] is False]
    assert len(safe_offs) >= 1


def test_sweep_time_estimation_uses_hard_dwell():
    """Verify that sweep summaries calculate total duration as N * Wait (Continuous) and N * (ON + Rest) (Pulse)."""
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()

    # 1. Continuous: 0 to 20 A in 2 A steps = 11 points. Wait = 5.0 s.
    window.cont_start.setValue(0.0)
    window.cont_stop.setValue(20.0)
    window.cont_step.setValue(2.0)
    window.cont_settle.setValue(5.0)
    window.cont_sample_window.setValue(3.0)  # Measure last is inside the 5 s wait
    window._update_sweep_summary()

    # Total estimated time is 11 * 5.0 = 55 s, not 11 * (5.0 + 3.0).
    assert "11 points" in window.cont_summary_lbl.text()
    assert "~55 s" in window.cont_summary_lbl.text()

    # 2. Pulse: 0 to 20 A in 2 A steps = 11 pulses. ON = 5 s, Rest = 5 s.
    window.pulse_start.setValue(0.0)
    window.pulse_stop.setValue(20.0)
    window.pulse_step.setValue(2.0)
    window.pulse_dwell.setValue(5.0)
    window.pulse_cooldown.setValue(5.0)
    window.pulse_sample_window.setValue(3.0)
    window._update_sweep_summary()

    # Measure-last stays inside ON time; estimate is 11 * (5.0 + 5.0) = 110 s.
    assert "11 pulses" in window.pulse_summary_lbl.text()
    assert "~110 s" in window.pulse_summary_lbl.text()

    window.close()


def test_card_status_released_and_historical_value_dimming(tmp_path):
    """Verify that after Release All Devices, cards show 'Released' (gray), dim historical values, and update_age() never overrides Released with Stale."""
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, InstrumentSnapshot, TEXT_MUTED, PRIMARY_BLUE
    from datetime import datetime, timedelta

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()

    pa_card = window.cards["pa"]
    load_card = window.load_card
    supply_card = window.supply_card
    scope_card = window.scope_card

    # 1. Populate cards with valid readings
    pa_snap = InstrumentSnapshot("pa", {"vin": 47.98, "iin": 2.45, "vout": 12.01}, valid=True, status="Connected", timestamp=datetime.now().isoformat())
    pa_card._received(pa_snap)
    assert pa_card.status_badge.text() == "Connected"
    assert pa_card.value_labels["Vin"][0].text() == "47.98"

    load_snap = InstrumentSnapshot("load", {"iout": 10.0, "vout": 12.01, "pout": 120.1, "input_on": True}, valid=True, status="Connected", timestamp=datetime.now().isoformat())
    load_card.last_snapshot = load_snap
    load_card.released = False
    load_card._render_values()
    assert load_card.status_badge.text() == "Connected"
    assert load_card.metric_labels["Iout"].text() == "10.00 A"
    assert load_card.metric_labels["Load State"].text() == "ON"

    psu_snap = InstrumentSnapshot("psu", {"ch1_enabled": False, "ch1_voltage": 0.0, "ch1_current": 0.0, "ch3_enabled": True, "ch3_voltage": 6.0, "ch3_current": 0.12}, valid=True, status="Connected", timestamp=datetime.now().isoformat())
    supply_card.last_snapshot = psu_snap
    supply_card.released = False
    supply_card._render_values()
    assert supply_card.status_badge.text() == "Connected"
    assert "Actual: 6.00 V · 0.120 A · ON" in supply_card.channel_controls[2]["live_lbl"].text()

    # 2. Trigger Release All Devices
    window._release_all_devices()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()

    # All primary status badges must be 'Released'
    assert pa_card.status_badge.text() == "Released"
    assert load_card.status_badge.text() == "Released"
    assert supply_card.status_badge.text() == "Released"
    assert scope_card.status_badge.text() == "Released"

    # Numerical values remain visible, but labeled/dimmed as historical
    assert pa_card.value_labels["Vin"][0].text() == "47.98"
    assert "Last reading before release" in pa_card.value_labels["Vin"][0].toolTip()

    assert load_card.metric_labels["Iout"].text() == "10.00 A"
    assert "Last known" in load_card.metric_labels["Load State"].text()
    assert "Last reading before release" in load_card.metric_labels["Iout"].toolTip()

    assert "Last known" in supply_card.channel_controls[2]["live_lbl"].text()
    assert "Last known state before release" in supply_card.channel_controls[2]["live_lbl"].toolTip()

    # 3. Simulate age expiration (15 seconds old) and call update_age() on all cards
    old_stamp = (datetime.now() - timedelta(seconds=15)).isoformat()
    pa_card.last_snapshot.timestamp = old_stamp
    load_card.last_snapshot.timestamp = old_stamp
    supply_card.last_snapshot.timestamp = old_stamp

    pa_card.update_age()
    load_card.update_age()
    supply_card.update_age()

    # Primary badges MUST REMAIN 'Released' and not be overwritten with 'Stale'!
    assert pa_card.status_badge.text() == "Released"
    assert load_card.status_badge.text() == "Released"
    assert supply_card.status_badge.text() == "Released"
    assert scope_card.status_badge.text() == "Released"

    # 4. Receiving new data restores active 'Connected' badge and clears '(Last known)'
    pa_snap_new = InstrumentSnapshot("pa", {"vin": 48.01, "iin": 0.0, "vout": 12.00}, valid=True, status="Connected", timestamp=datetime.now().isoformat())
    pa_card._received(pa_snap_new)
    assert pa_card.status_badge.text() == "Connected"
    assert pa_card.value_labels["Vin"][0].text() == "48.01"
    assert pa_card.value_labels["Vin"][0].toolTip() == ""

    window.close()


def test_manual_mode_range_and_plot_scaling_authority():
    """Verify that SET CURRENT and STEP CURRENT use the Bench Setup safety cap for progress bar, plot scaling, and clamping, while Continuous/Pulse use sweep parameters."""
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets
    from sid_bench_gui import MainWindow

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()
    window.simulation.setChecked(True)
    window.direct_auto_save.setChecked(False)
    window.direct_auto_capture.setChecked(False)
    window.step_auto_save.setChecked(False)
    window.step_auto_capture.setChecked(False)

    # 1. Configure Bench Setup safety cap to 70 A and Continuous Stop to 60 A
    window.load_card.cap_spin.setValue(70.0)
    window.load_card.apply_cap_btn.click()
    assert window.manual_mode_max_current() == 70.0

    window.cont_start.setValue(0.0)
    window.cont_stop.setValue(60.0)
    window._update_sweep_summary()

    # 2. Check SET CURRENT mode (Mode 0)
    window.btn_mode_direct.click()
    QtWidgets.QApplication.processEvents()

    assert window.manual_target_spin.maximum() == 70.0
    assert window.plot_progress_marker.start_val == 0.0
    assert window.plot_progress_marker.stop_val == 70.0

    # Plot x-axis should be based on safety cap (70 A), not Continuous sweep stop (60 A)
    x_range = window.live_plot_widget.getPlotItem().getViewBox().viewRange()[0]
    assert x_range[0] <= 0.01
    assert x_range[1] >= 69.9  # approx 70 A (with padding)

    # Empty Efficiency y-axis uses default efficiency bounds 90..100%
    y_range = window.live_plot_widget.getPlotItem().getViewBox().viewRange()[1]
    assert y_range[0] <= 90.5
    assert y_range[1] >= 99.5

    # 3. Check STEP CURRENT mode (Mode 1)
    window.btn_mode_step.click()
    QtWidgets.QApplication.processEvents()

    assert window.plot_progress_marker.start_val == 0.0
    assert window.plot_progress_marker.stop_val == 70.0

    # Step up clamping against 70 A safety cap (e.g. 69 A + 2 A = 70 A)
    window._manual_target_current = 69.0
    window.manual_step_inc.setValue(2.0)
    window._step_delta(1)
    assert window._manual_target_current == 70.0
    assert window.manual_target_spin.value() == 70.0

    # 4. Switch to CONTINUOUS mode (Mode 2)
    window.btn_mode_cont.setChecked(True)
    QtWidgets.QApplication.processEvents()

    # Progress bar and plot must restore Continuous sweep range (0..60 A)
    assert window.plot_progress_marker.start_val == 0.0
    assert window.plot_progress_marker.stop_val == 60.0
    x_range_cont = window.live_plot_widget.getPlotItem().getViewBox().viewRange()[0]
    assert x_range_cont[1] < 68.0  # around 60..63 A, NOT 70 A

    # 5. Switch to PULSE mode (Mode 3)
    window.pulse_start.setValue(0.0)
    window.pulse_stop.setValue(40.0)
    window.btn_mode_pulse.setChecked(True)
    QtWidgets.QApplication.processEvents()

    assert window.plot_progress_marker.start_val == 0.0
    assert window.plot_progress_marker.stop_val == 40.0
    x_range_pulse = window.live_plot_widget.getPlotItem().getViewBox().viewRange()[0]
    assert x_range_pulse[1] < 45.0  # around 40..42 A

    # 6. Dynamic update without restart: change Bench Setup safety cap from 70 A -> 80 A
    window.load_card.cap_spin.setValue(80.0)
    window.load_card.apply_cap_btn.click()
    assert window.manual_mode_max_current() == 80.0

    # Switch back to STEP CURRENT and verify immediate dynamic update to 80 A
    window.btn_mode_step.setChecked(True)
    QtWidgets.QApplication.processEvents()
    assert window.plot_progress_marker.stop_val == 80.0
    x_range_step = window.live_plot_widget.getPlotItem().getViewBox().viewRange()[0]
    assert x_range_step[1] >= 79.9

    # Restore default 60 A limit
    window.load_card.cap_spin.setValue(60.0)
    window.load_card.apply_cap_btn.click()
    assert window.cap_val == 60.0

    window.close()


def test_bench_discovery_connection_state_transitions_and_psu_semantics():
    """Verify Discovery, Check, Release, and Discovered transitions across all 4 cards, and E36312A desired-vs-actual semantics."""
    import os, time
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])
    window = MainWindow()

    # Verify get_card_for_instrument and all_instrument_cards covers all 4 devices
    assert window.get_card_for_instrument("scope") is window.scope_card
    assert window.get_card_for_instrument("psu") is window.supply_card
    assert window.get_card_for_instrument("load") is window.load_card
    assert window.get_card_for_instrument("pa") is window.cards["pa"]
    cards = window.all_instrument_cards()
    assert len(cards) == 4
    assert set(cards.keys()) == {"pa", "load", "psu", "scope"}

    # 1. App Startup: All 4 cards must show "Not Checked"
    for k, c in cards.items():
        assert c.status_badge.text() == "Not Checked", f"{k} did not start as Not Checked"

    # E36312A desired configuration semantics:
    # Setpoints: 5.50 V, desired ON: unchecked, actual: — V · — A · —
    psu_card = window.supply_card
    ch1_ctrl = psu_card.channel_controls[0]
    ch1_ctrl["voltage"].setValue(5.50)
    ch1_ctrl["desired_out"].setChecked(False)

    # 2. Simulated Discovery where PA, PSU, and Scope are found, but Chroma is missing
    mock_found = {
        "pa": {"address": "TCPIP0::192.168.1.101::inst0::INSTR", "identity": "Keysight Technologies,PA2201A,MY1234,1.0"},
        "psu": {"address": "TCPIP0::192.168.1.103::inst0::INSTR", "identity": "Keysight Technologies,E36312A,MY5678,1.0"},
        "scope": {"address": "TCPIP0::192.168.1.104::inst0::INSTR", "identity": "Keysight Technologies,MSOX4024A,MY9999,1.0"},
    }
    # Call discovery on_done logic
    window._set_bench_busy(True, "discover")
    # Simulate discovery callback
    for key in ("pa", "load", "psu", "scope"):
        card = window.get_card_for_instrument(key)
        if key in mock_found:
            card._mark_discovered(mock_found[key]["identity"])
        else:
            card._mark_not_found(f"Chroma not responding to *IDN? over VISA")
    window._set_bench_busy(False)

    assert cards["pa"].status_badge.text() == "Discovered"
    assert cards["psu"].status_badge.text() == "Discovered"
    assert cards["scope"].status_badge.text() == "Discovered"
    assert cards["load"].status_badge.text() == "Not Found"

    # 3. Check / Refresh Entire Bench with simulated instruments
    window.simulation.setChecked(True)
    window._check_entire_bench()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()

    # Cards with valid live reads show Connected / Active
    assert cards["pa"].status_badge.text() == "Connected"
    assert cards["load"].status_badge.text() == "Connected"
    assert cards["psu"].status_badge.text() == "Connected"
    assert cards["scope"].status_badge.text() == "Active"

    # Verify E36312A: Refresh did NOT overwrite desired setpoint fields!
    assert ch1_ctrl["voltage"].value() == 5.50
    assert ch1_ctrl["desired_out"].isChecked() is False
    # Actual line shows readback (e.g. 0.00 V · 0.000 A · OFF)
    assert "Actual:" in ch1_ctrl["live_lbl"].text()
    assert "OFF" in ch1_ctrl["live_lbl"].text()

    # 4. Supply Apply Settings: badge becomes "Connected" (NOT "Applied")
    ch1_ctrl["desired_out"].setChecked(True)
    ch1_ctrl["voltage"].setValue(5.20)
    psu_card._apply()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()
    assert psu_card.status_badge.text() == "Connected"
    assert "ON" in ch1_ctrl["live_lbl"].text()

    # 5. Release All Devices: all cards become "Released"
    window._release_all_devices()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()

    for k, c in cards.items():
        assert c.status_badge.text() == "Released", f"{k} did not transition to Released"

    # 6. Discover Again: all found cards transition from Released -> Discovered (none stuck as Released)
    for key in ("pa", "load", "psu", "scope"):
        card = window.get_card_for_instrument(key)
        card._mark_discovered("VISA::INSTR")

    for k, c in cards.items():
        assert c.status_badge.text() == "Discovered", f"{k} remained Released instead of Discovered"

    # 7. Check Again: all successfully read cards become Connected / Active
    window._check_entire_bench()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()

    assert cards["pa"].status_badge.text() == "Connected"
    assert cards["load"].status_badge.text() == "Connected"
    assert cards["psu"].status_badge.text() == "Connected"
    assert cards["scope"].status_badge.text() == "Active"

    window.close()


def test_chroma_63206a_scpi_commands_bench_diagnostics_and_run_behavior(tmp_path: Path, monkeypatch):
    """Verify verified SCPI command set, temporary front-panel local restoration, Bench diagnostics vs run measurement separation, and partial read behavior."""
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore
    from sid_bench_gui import MainWindow, LoadCard, WorkbookStore, calculate_measurement, MEAS_HEADERS
    from sid_instruments import Chroma63206A, VisaManager, InstrumentSnapshot, InstrumentError

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    # 1. Test Chroma63206A SCPI commands and front-panel restoration
    commands_written = []
    queries_made = []

    class MockSession:
        timeout = 3000
        def __init__(self):
            self.closed = False

        def write(self, cmd: str):
            commands_written.append(cmd)

        def query(self, cmd: str):
            queries_made.append(cmd)
            if cmd == "*IDN?":
                return "Chroma,63206A-60-1000,12345,1.0"
            elif cmd == "MEASure:CURRent?":
                return "15.250"
            elif cmd == "MEASure:VOLTage?":
                return "11.950"
            elif cmd == "SYST:ERR?":
                return "+0, \"No error\""
            return "0.0"

        def close(self):
            self.closed = True

    mock_sess = MockSession()

    class MockRM:
        def open_resource(self, addr):
            return mock_sess

    mgr = VisaManager()
    mgr._rm = MockRM()
    load = Chroma63206A(mgr, "GPIB0::2::INSTR", ("63206",))

    # A. Verify local_commands override contains ONLY SYSTem:LOCal
    assert load.local_commands() == ("SYSTem:LOCal",)

    # B. Set current and input: Write commands only (no query)
    load.set_current(15.25)
    assert "MODE CCH" in commands_written
    assert "CURRent:STATic:L1 15.25" in commands_written

    load.set_input(True)
    assert "LOAD ON" in commands_written
    load.set_input(False)
    assert "LOAD OFF" in commands_written

    # Safe shutdown de-energizes first, then clears the stored current setpoint.
    commands_written.clear()
    load.safe_off()
    assert "LOAD OFF" in commands_written
    assert "CURRent:STATic:L1 0" in commands_written
    assert commands_written.index("LOAD OFF") < commands_written.index("CURRent:STATic:L1 0")

    # C. read_snapshot(include_voltage=True) for Bench Setup
    commands_written.clear()
    queries_made.clear()
    snap = load.read_snapshot(include_voltage=True)
    assert "MEASure:CURRent?" in queries_made
    assert "MEASure:VOLTage?" in queries_made
    assert snap.values["current"] == 15.25
    assert snap.values["voltage"] == 11.95
    assert abs(snap.values["power"] - (15.25 * 11.95)) < 1e-4
    assert snap.status == "Connected"
    # Temporary session closed and restored local front-panel
    assert "SYSTem:LOCal" in commands_written

    # D. read_snapshot(include_voltage=False) for Run Measurements: skips voltage query
    commands_written.clear()
    queries_made.clear()
    run_snap = load.read_snapshot(include_voltage=False)
    assert "MEASure:CURRent?" in queries_made
    assert "MEASure:VOLTage?" not in queries_made
    assert run_snap.values["current"] == 15.25
    assert run_snap.values["voltage"] is None
    assert run_snap.values["power"] is None
    assert run_snap.status == "Connected"

    # E. Partial read behavior: MEASure:VOLTage? fails while MEASure:CURRent? succeeds
    def fail_voltage_query(cmd: str):
        queries_made.append(cmd)
        if cmd == "*IDN?":
            return "Chroma,63206A-60-1000,12345,1.0"
        elif cmd == "MEASure:CURRent?":
            return "15.250"
        elif cmd == "MEASure:VOLTage?":
            raise RuntimeError("VI_ERROR_TMO: Timeout expired")
        return "0.0"

    mock_sess.query = fail_voltage_query
    partial_snap = load.read_snapshot(include_voltage=True)
    assert partial_snap.valid is True
    assert partial_snap.values["current"] == 15.25
    assert partial_snap.values["voltage"] is None
    assert partial_snap.values["power"] is None
    assert partial_snap.status == "Connected · Partial Read"
    assert "failed" in partial_snap.warning

    # 2. Test LoadCard UI rendering of diagnostic values and Partial Read
    window = MainWindow()
    load_card = window.load_card
    load_card.last_snapshot = partial_snap
    load_card._render_values()

    assert load_card.metric_labels["Iout"].text() == "15.25 A"
    assert load_card.metric_labels["Load V"].text() == "—"
    assert load_card.metric_labels["Load P"].text() == "—"
    assert load_card.status_badge.text() == "Connected · Partial Read"

    # Full snapshot rendering
    load_card.last_snapshot = snap
    load_card._render_values()
    assert load_card.metric_labels["Iout"].text() == "15.25 A"
    assert load_card.metric_labels["Load V"].text() == "11.95 V"
    assert "182.24 W" in load_card.metric_labels["Load P"].text()
    assert load_card.status_badge.text() == "Connected"

    # Release All Devices tooltip check
    window._release_all_devices()
    QtCore.QThreadPool.globalInstance().waitForDone(3000)
    QtWidgets.QApplication.processEvents()
    assert load_card.status_badge.text() == "Released"
    assert load_card.status_badge.toolTip() == "VISA session closed · front panel restored"

    # 3. Experiment Data Topology: Workbook row verification
    excel_path = tmp_path / "chroma_topology_test.xlsx"
    store = WorkbookStore(excel_path)
    window.store = store

    # Calculate measurement and verify authoritative experiment quantities
    pa_snap = InstrumentSnapshot("pa", {"vin": 48.0, "iin": 2.5, "vout": 12.0})
    derived, warnings = calculate_measurement(pa_snap, run_snap, None, [])

    assert derived["Iout_A"] == 15.25
    assert derived["Vin_V"] == 48.0
    assert derived["Iin_A"] == 2.5
    assert derived["Vout_V"] == 12.0
    assert derived["PinConverter_W"] == 120.0  # 48 * 2.5
    assert derived["Pout_W"] == 183.0  # 12.0 * 15.25
    assert abs(derived["EfficiencyConverter_pct"] - (183.0 / 120.0 * 100.0)) < 1e-4

    # Verify diagnostic Load V / Load P are NOT present in derived measurement or MEAS_HEADERS
    assert "Load V" not in derived
    assert "Load P" not in derived
    assert "ChromaLoadVoltage" not in MEAS_HEADERS
    assert "ChromaLoadPower" not in MEAS_HEADERS

    window.close()


def test_plot_visual_smoothness_and_pen_styling():
    """Verify pyqtgraph antialiasing, round cap/join pens, 6px circular markers, matching colors, and discrete linear data handling."""
    import os
    os.environ["QT_QPA_PLATFORM"] = "offscreen"
    from PyQt6 import QtWidgets, QtCore, QtGui
    import pyqtgraph as pg
    from sid_bench_gui import (
        MainWindow, make_smooth_pen, plot_metric_series, apply_metric_curve_style,
        PLOT_CORE_BLUE, PLOT_SYSTEM_ORANGE, PLOT_AUX_TEAL
    )

    app = QtWidgets.QApplication.instance() or QtWidgets.QApplication([""])

    # 1. Antialiasing is enabled globally
    assert pg.getConfigOption("antialias") is True

    # 2. make_smooth_pen returns QPen with round cap/join and ~2.2px width
    pen = make_smooth_pen("#002676", width=2.2)
    assert isinstance(pen, QtGui.QPen)
    assert abs(pen.widthF() - 2.2) < 1e-4
    assert pen.capStyle() == QtCore.Qt.PenCapStyle.RoundCap
    assert pen.joinStyle() == QtCore.Qt.PenJoinStyle.RoundJoin

    # 3. plot_metric_series helper creates styled PlotDataItem
    pw = pg.PlotWidget(background="w")
    test_item = plot_metric_series(pw, [1.0, 2.0], [10.0, 20.0], color=PLOT_CORE_BLUE, name="Test", symbol="o", symbol_size=6)
    assert test_item.opts["symbol"] == "o"
    assert test_item.opts["symbolSize"] == 6
    assert test_item.opts["symbolBrush"].color().name().lower() == PLOT_CORE_BLUE.lower()
    assert test_item.opts["symbolPen"].color().name().lower() == PLOT_CORE_BLUE.lower()
    assert test_item.opts["pen"].color().name().lower() == PLOT_CORE_BLUE.lower()

    # 4. Live plot curves styling: ~2.2px width, 6px circular markers ("o"), matching colors
    window = MainWindow()
    curves = [
        (window.live_curve, PLOT_CORE_BLUE),
        (window.live_system_curve, PLOT_SYSTEM_ORANGE),
        (window.live_aux_curve, PLOT_AUX_TEAL),
    ]
    for curve, expected_color in curves:
        c_pen = curve.opts["pen"]
        assert isinstance(c_pen, QtGui.QPen)
        assert abs(c_pen.widthF() - 2.2) < 1e-4
        assert c_pen.capStyle() == QtCore.Qt.PenCapStyle.RoundCap
        assert c_pen.joinStyle() == QtCore.Qt.PenJoinStyle.RoundJoin
        assert curve.opts["symbolSize"] == 6
        assert curve.opts["symbol"] == "o"
        assert curve.opts["symbolBrush"].color().name().lower() == expected_color.lower()
        assert curve.opts["symbolPen"].color().name().lower() == expected_color.lower()

    # 5. Discrete straight lines without interpolation
    test_points = [
        {"Status": "Valid", "Iout_A": 5.0, "EfficiencyConverter_pct": 94.2, "EfficiencySystem_pct": 92.1},
        {"Status": "Valid", "Iout_A": 15.0, "EfficiencyConverter_pct": 96.8, "EfficiencySystem_pct": 95.0},
        {"Status": "Valid", "Iout_A": 25.0, "EfficiencyConverter_pct": 95.5, "EfficiencySystem_pct": 93.8},
    ]
    window.plot_rows = test_points
    window._switch_live_plot(0)

    # Actual data array matches exact measured points (no spline/curve-fit points injected)
    xs, ys = window.live_curve.getData()
    assert list(xs) == [5.0, 15.0, 25.0]
    assert list(ys) == [94.2, 96.8, 95.5]

    sys_xs, sys_ys = window.live_system_curve.getData()
    assert list(sys_xs) == [5.0, 15.0, 25.0]
    assert list(sys_ys) == [92.1, 95.0, 93.8]

    # Verify switching metric views maintains exact matching colors and 6px circular markers
    window._switch_live_plot(2)  # Power (W)
    assert window.live_curve.opts["symbol"] == "o"
    assert window.live_curve.opts["symbolSize"] == 6
    assert window.live_curve.opts["symbolBrush"].color().name().lower() == PLOT_CORE_BLUE.lower()
    assert window.live_system_curve.opts["symbolBrush"].color().name().lower() == PLOT_SYSTEM_ORANGE.lower()
    assert window.live_aux_curve.opts["symbolBrush"].color().name().lower() == PLOT_AUX_TEAL.lower()

    window.close()
