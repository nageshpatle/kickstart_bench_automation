"""Kickstart Bench Automation GUI.

Passive startup: no VISA sessions are opened and no polling occurs until requested.
Includes high-resolution instrument photographs from instruments/, embedded Supply Settings
directly inside the E36312A Supply card, 4-column equal-dimension Bench page with no-expand
Scope column, hover tooltips for full VISA controller messages, neutral lab design system,
structured sweep ranges (0->60 A in 2 A steps), buttonless Max Current Cap, and live plotting.
"""

from __future__ import annotations

import csv
import hashlib
import io
import json
import math
import os
import re
import shutil
import sys
import threading
import time
import uuid
import zipfile
import ctypes
from ctypes import wintypes
from dataclasses import asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Callable

from PyQt6 import QtCore, QtGui, QtWidgets
import pyqtgraph as pg

pg.setConfigOptions(antialias=True)


def make_smooth_pen(color: Any, width: float = 2.2) -> QtGui.QPen:
    """Create a polished antialiased pen with rounded caps and joins for discrete curve rendering."""
    pen = pg.mkPen(color, width=width)
    pen.setCapStyle(QtCore.Qt.PenCapStyle.RoundCap)
    pen.setJoinStyle(QtCore.Qt.PenJoinStyle.RoundJoin)
    return pen


from sid_instruments import (
    InstrumentError,
    InstrumentHub,
    InstrumentSnapshot,
    RequiredInstrumentError,
    SupplyChannel,
    utc_now,
)


ROOT = Path(__file__).resolve().parent
INSTRUMENTS_DIR = ROOT / "instruments"
CONFIG_PATH = ROOT / "bench_config.json"
DEFAULT_HARDWARE_WORKBOOK = ROOT / "results" / "hardware_measurements.xlsx"
DEFAULT_SIMULATION_WORKBOOK = ROOT / "results" / "simulation_runs.xlsx"
DEFAULT_FPGA_ROOT = ROOT.parent / "KICKSTART_PILAWA"

# Berkeley Blue & California Gold Design Tokens (matching Kickstart PILAWA GUI)
BERKELEY_BLUE = "#002676"
BLUE_MEDIUM = "#004AAE"
BLUE_LIGHT = "#9FD1FF"
CALIFORNIA_GOLD = "#FDB515"
GOLD_DARK = "#FC9313"
GOLD_LIGHT = "#FFE88D"

PAGE_BG = "#F7F8FA"
CARD_BG = "#FFFFFF"
TEXT_MAIN = "#1F2937"
TEXT_MUTED = "#4B5563"
BORDER = "#D1D5DB"
PRIMARY_BLUE = "#002676"
MODE_CONT = "#004AAE"
MODE_PULSE = "#002676"
SUCCESS_GREEN = "#166534"
WARNING_AMBER = "#B45309"
PLOT_CORE_BLUE = "#002676"
PLOT_SYSTEM_ORANGE = "#D97706"
PLOT_AUX_TEAL = "#0F766E"
DANGER_RED = "#B91C1C"

PALETTE = [
    "#002676", "#004AAE", "#FDB515", "#FC9313", "#166534", "#B45309", "#0891B2", "#B91C1C"
]


def plot_metric_series(
    plot_widget: pg.PlotWidget,
    xs: list[float] | None = None,
    ys: list[float] | None = None,
    color: Any = PLOT_CORE_BLUE,
    name: str = "",
    symbol: str = "o",
    symbol_size: int = 6,
) -> pg.PlotDataItem:
    """Plot a metric series with polished rounded pens, solid line styling, and matching marker colors."""
    xs = xs if xs is not None else []
    ys = ys if ys is not None else []
    pen = make_smooth_pen(color, width=2.2)
    return plot_widget.plot(
        xs,
        ys,
        pen=pen,
        symbol=symbol,
        symbolSize=symbol_size,
        symbolBrush=pg.mkBrush(color),
        symbolPen=pg.mkPen(color, width=1),
        name=name,
    )


def apply_metric_curve_style(
    curve: pg.PlotDataItem,
    color: Any,
    symbol: str = "o",
    symbol_size: int = 6,
    name: str = "",
) -> pg.PlotDataItem:
    """Apply consistent smooth pen and matching marker styling to an existing PlotDataItem."""
    pen = make_smooth_pen(color, width=2.2)
    curve.setPen(pen)
    curve.setSymbol(symbol)
    curve.setSymbolSize(symbol_size)
    curve.setSymbolBrush(pg.mkBrush(color))
    curve.setSymbolPen(pg.mkPen(color, width=1))
    if name:
        curve.opts["name"] = name
    return curve


def apply_forced_light_theme(app: QtWidgets.QApplication):
    """Apply a complete light palette, independent of the OS color scheme."""
    app.setStyle("Fusion")
    try:
        app.styleHints().setColorScheme(QtCore.Qt.ColorScheme.Light)
    except (AttributeError, RuntimeError):
        pass

    palette = QtGui.QPalette()
    colors = {
        QtGui.QPalette.ColorRole.Window: "#F7F8FA",
        QtGui.QPalette.ColorRole.WindowText: "#1F2937",
        QtGui.QPalette.ColorRole.Base: "#FFFFFF",
        QtGui.QPalette.ColorRole.AlternateBase: "#F3F4F6",
        QtGui.QPalette.ColorRole.ToolTipBase: "#FFF8C5",
        QtGui.QPalette.ColorRole.ToolTipText: "#111827",
        QtGui.QPalette.ColorRole.Text: "#1F2937",
        QtGui.QPalette.ColorRole.Button: "#F3F4F6",
        QtGui.QPalette.ColorRole.ButtonText: "#111827",
        QtGui.QPalette.ColorRole.BrightText: "#B91C1C",
        QtGui.QPalette.ColorRole.Highlight: "#002676",
        QtGui.QPalette.ColorRole.HighlightedText: "#FFFFFF",
        QtGui.QPalette.ColorRole.PlaceholderText: "#6B7280",
        QtGui.QPalette.ColorRole.Light: "#FFFFFF",
        QtGui.QPalette.ColorRole.Midlight: "#E5E7EB",
        QtGui.QPalette.ColorRole.Mid: "#9CA3AF",
        QtGui.QPalette.ColorRole.Dark: "#4B5563",
        QtGui.QPalette.ColorRole.Shadow: "#111827",
    }
    for role, color in colors.items():
        palette.setColor(role, QtGui.QColor(color))
    for role in [
        QtGui.QPalette.ColorRole.WindowText,
        QtGui.QPalette.ColorRole.Text,
        QtGui.QPalette.ColorRole.ButtonText,
    ]:
        palette.setColor(QtGui.QPalette.ColorGroup.Disabled, role, QtGui.QColor("#6B7280"))

    app.setPalette(palette)


DEFAULT_CONFIG: dict[str, Any] = {

    "schema": 5,
    "campaign_name": "Efficiency Test",
    "working_current_cap_a": 70.0,
    "vin_safety_enabled": True,
    "last_modulation_label": "",
    "recent_modulations": [],
    "workbooks": {
        "hardware": str(DEFAULT_HARDWARE_WORKBOOK),
        "simulation": str(DEFAULT_SIMULATION_WORKBOOK),
    },
    "fpga_root": str(DEFAULT_FPGA_ROOT),
    "visa_library": r"C:\Windows\System32\visa64.dll",
    "addresses": {
        "pa": "",
        "load": "",
        "psu": "",
        "scope": "",
    },
    "supply_channels": [
        {"channel": 1, "role": "Vdrv_A", "displayed": True, "enabled": False, "contributes_loss": True, "voltage_set": 0.0, "current_limit": 1.0},
        {"channel": 2, "role": "Vdrv_B", "displayed": True, "enabled": False, "contributes_loss": True, "voltage_set": 0.0, "current_limit": 1.0},
        {"channel": 3, "role": "Vdrv_C", "displayed": True, "enabled": True, "contributes_loss": True, "voltage_set": 6.0, "current_limit": 1.0},
    ],
}

RUN_HEADERS = [
    "RunID", "CampaignName", "Created", "Completed", "Status", "DataSource", "Mode", "VinTarget_V",
    "ModulationLabel", "Frequency_Hz", "ModulationMetadata",
    "AuxA_Included", "AuxB_Included", "AuxC_Included",
    "SupplyConfiguration", "WorkingCap_A", "Notes", "InstrumentIdentities",
    "FPGASnapshotStatus", "FPGASnapshot", "Warnings", "SupersedesRunID",
]

MEAS_HEADERS = [
    "PointID", "RunID", "Timestamp", "Status", "DataSource", "Mode", "VinTarget_V",
    "ModulationLabel", "Frequency_Hz", "RequestedIout_A", "Iout_A", "Vin_V", "Iin_A",
    "Vout_V", "PinConverter_W", "Pout_W",
    "Vdrv_A_V", "Idrv_A_A", "Pdrv_A_W",
    "Vdrv_B_V", "Idrv_B_A", "Pdrv_B_W",
    "Vdrv_C_V", "Idrv_C_A", "Pdrv_C_W",
    "Paux_W", "LossConverter_W", "LossSystem_W",
    "EfficiencyConverter_pct", "EfficiencySystem_pct",
    "SupplyMeasurements", "Quality", "Warning", "ScopeCaptureStatus", "ScopeCaptureError",
    "ScopePNG", "ScopeCSV", "SupersedesPointID",
]

EVENT_HEADERS = ["Timestamp", "RunID", "PointID", "Event", "Detail"]


def deep_merge(base: dict[str, Any], update: dict[str, Any]) -> dict[str, Any]:
    result = json.loads(json.dumps(base))
    for key, value in update.items():
        if isinstance(value, dict) and isinstance(result.get(key), dict):
            result[key] = deep_merge(result[key], value)
        else:
            result[key] = value
    return result


def load_config() -> dict[str, Any]:
    if not CONFIG_PATH.exists():
        config = json.loads(json.dumps(DEFAULT_CONFIG))
    else:
        try:
            config = deep_merge(DEFAULT_CONFIG, json.loads(CONFIG_PATH.read_text(encoding="utf-8")))
        except Exception:
            config = json.loads(json.dumps(DEFAULT_CONFIG))
    workbook_override = os.environ.get("KICKSTART_WORKBOOK_PATH", "").strip()
    if workbook_override:
        base = Path(workbook_override)
        config["workbooks"] = {
            "hardware": str(base),
            "simulation": str(base.with_name(f"{base.stem}_simulation{base.suffix}")),
        }
    return config


def save_config(config: dict[str, Any]) -> None:
    if os.environ.get("KICKSTART_WORKBOOK_PATH", "").strip():
        return
    CONFIG_PATH.write_text(json.dumps(config, indent=2), encoding="utf-8")


def format_frequency_khz(frequency_hz: Any) -> str:
    """Format a stored Hz value for operator-facing kHz displays."""
    try:
        return f"{float(frequency_hz) / 1000.0:g} kHz"
    except (TypeError, ValueError):
        return "—"


def capture_root_for_source(workbook_path: Path, source: str) -> Path:
    """Return the one authoritative scope-capture folder for a data source."""
    source_folder = "hardware" if str(source).strip().lower() == "hardware" else "simulation"
    return Path(workbook_path).parent / "captures" / source_folder


def efficiency_axis_bounds(values: list[float], minimum_span: float = 10.0) -> tuple[float, float]:
    """Round live efficiency bounds outward to 5% steps with breathing room."""
    finite = [float(value) for value in values if isinstance(value, (int, float)) and math.isfinite(float(value))]
    if not finite:
        return 90.0, 100.0
    low, high = min(finite), max(finite)
    y_min = math.floor(low / 5.0) * 5.0
    y_max = math.ceil(high / 5.0) * 5.0
    if math.isclose(low, y_min, abs_tol=1e-9):
        y_min -= 5.0
    if math.isclose(high, y_max, abs_tol=1e-9):
        y_max += 5.0
    y_min = max(0.0, y_min)
    y_max = min(105.0, y_max)
    while y_max - y_min < minimum_span - 1e-9:
        if y_min >= 5.0:
            y_min -= 5.0
        elif y_max <= 100.0:
            y_max += 5.0
        else:
            break
    return y_min, y_max


def parse_points(text: str) -> list[float]:
    points: list[float] = []
    for token in (item.strip() for item in text.split(",")):
        if not token:
            continue
        if ":" not in token:
            points.append(float(token))
            continue
        parts = [float(value.strip()) for value in token.split(":")]
        if len(parts) != 3 or parts[2] == 0:
            raise ValueError(f"Invalid range {token!r}; use start:stop:step")
        start, stop, step = parts
        if (stop - start) * step < 0:
            raise ValueError(f"Range step points away from stop in {token!r}")
        current = start
        for _ in range(10000):
            if (step > 0 and current > stop + abs(step) * 1e-9) or (step < 0 and current < stop - abs(step) * 1e-9):
                break
            points.append(round(current, 9))
            current += step
        else:
            raise ValueError("Range creates too many points")
    if not points:
        raise ValueError("Enter at least one current point")
    return points


def generate_points(start: float, stop: float, step: float, cap: float) -> tuple[list[float], str]:
    """Generates an inclusive current point list and returns a human-readable summary."""
    if start < 0:
        raise ValueError("Start current must be non-negative")
    if stop < start:
        raise ValueError("Stop current must be greater than or equal to start current")
    if step <= 0:
        raise ValueError("Step size must be strictly positive")
    if stop > cap:
        raise ValueError(f"Stop current ({stop:g} A) exceeds maximum allowed load current ({cap:g} A)")

    points: list[float] = []
    curr = start
    while curr <= stop + 1e-9:
        points.append(round(curr, 4))
        curr += step
        if len(points) > 5000:
            raise ValueError("Range generates too many points (> 5000)")

    if not points or abs(points[-1] - stop) > 1e-9:
        if stop <= cap:
            points.append(round(stop, 4))

    # Format human-readable summary accurately
    if len(points) == 1:
        summary = f"{start:g} A · 1 point"
    elif step >= (stop - start):
        summary = f"{start:g} A → {stop:g} A · {len(points)} points"
    else:
        summary = f"{start:g} A → {stop:g} A in {step:g} A steps · {len(points)} points"

    return points, summary



def parse_capture_points(text: str) -> set[float]:
    if not text.strip():
        return set()
    pts: set[float] = set()
    for token in text.split(","):
        token = token.strip()
        if token:
            try:
                pts.add(float(token))
            except ValueError as exc:
                raise ValueError(f"Invalid scope capture current: '{token}'") from exc
    return pts


def check_vin_safety(
    measured_vin: float,
    target_vin: float,
    enabled: bool = True,
) -> tuple[bool, str]:
    """Evaluate whether measured Vin is within ±10% of target_vin.

    Returns:
        tuple[bool, str]: (is_safe, fault_description)
    """
    if not enabled:
        return True, ""
    if not isinstance(target_vin, (int, float)) or target_vin <= 0:
        return True, ""
    if not isinstance(measured_vin, (int, float)) or not math.isfinite(float(measured_vin)):
        return True, ""

    low = 0.90 * float(target_vin)
    high = 1.10 * float(target_vin)

    # Floating-point safe comparison: allowing exact boundary with 1e-6 epsilon tolerance
    if float(measured_vin) < low - 1e-6 or float(measured_vin) > high + 1e-6:
        desc = (
            f"Target Vin: {target_vin:.1f} V\n"
            f"Measured Vin: {measured_vin:.1f} V\n"
            f"Allowed range: {low:.1f}–{high:.1f} V"
        )
        return False, desc
    return True, ""




def new_run_id(campaign: str = "", label: str = "") -> str:
    combined = f"{campaign}-{label}".strip("-")
    slug = re.sub(r"[^A-Za-z0-9]+", "-", combined).strip("-")[:32]
    stamp = datetime.now().strftime("%Y%m%d-%H%M%S")
    return f"{stamp}-{slug or 'RUN'}-{uuid.uuid4().hex[:5]}"


def point_id(run_id: str, index: int) -> str:
    return f"{run_id}-P{index + 1:03d}"


def fpga_snapshot(fpga_root: Path, expected_frequency: float | None = None) -> tuple[str, dict[str, Any], str]:
    if not fpga_root.exists():
        return "Unavailable", {}, f"FPGA project not found: {fpga_root}"
    try:
        payload: dict[str, Any] = {"root": str(fpga_root)}
        config_path = fpga_root / "kickstart_pilawa_gui_config.json"
        if config_path.exists():
            raw = config_path.read_bytes()
            config = json.loads(raw.decode("utf-8"))
            payload["config_sha256"] = hashlib.sha256(raw).hexdigest()
            payload["selected_project"] = config.get("selected_profile_project") or config.get("selected_project")
            payload["config_mtime"] = datetime.fromtimestamp(config_path.stat().st_mtime).isoformat(timespec="seconds")
        top_path = fpga_root / "top.v"
        if top_path.exists():
            raw = top_path.read_bytes()
            text = raw.decode("utf-8", errors="replace")
            payload["top_sha256"] = hashlib.sha256(raw).hexdigest()
            payload["top_mtime"] = datetime.fromtimestamp(top_path.stat().st_mtime).isoformat(timespec="seconds")
            freq = re.search(r"PWM_FREQ_HZ\s*=\s*([0-9_]+)", text)
            payload["top_frequency_hz"] = int(freq.group(1).replace("_", "")) if freq else None
            for name in ("PWM_START", "PWM_END", "PWM_INV", "PWM_PIN"):
                matches = re.findall(rf"{name}[^=]*=\s*([^;]+);", text)
                if matches:
                    payload[name.lower()] = matches[:32]
        warning = ""
        status = "Captured"
        actual = payload.get("top_frequency_hz")
        if expected_frequency and actual and abs(float(actual) - expected_frequency) > 0.5:
            status = "Mismatch"
            warning = f"FPGA top.v frequency {format_frequency_khz(actual)} differs from requested {format_frequency_khz(expected_frequency)}"
        return status, payload, warning
    except Exception as exc:
        return "Failed", {}, str(exc)


class WorkbookStore:
    """Single campaign workbook with durable staging and validated commits.

    A cloud-synced or Excel-open workbook can temporarily reject replacement.  In that
    case the newest complete workbook is retained in one local pending file and used as
    the authoritative source until it can be committed.  We never create per-save
    fallback workbooks and never stream-copy over the live XLSX.
    """

    def __init__(self, path: Path, prompt_fn: Callable[[str, str], bool] | None = None):
        self.path = Path(path)
        self._prompt_fn = prompt_fn
        self._lock = threading.RLock()
        self.last_warning = ""

    @property
    def pending_path(self) -> Path:
        root = Path(os.environ.get("LOCALAPPDATA", str(self.path.parent))) / "KickstartBench" / "pending"
        key = hashlib.sha256(str(self.path.resolve()).casefold().encode("utf-8")).hexdigest()[:20]
        return root / f"{key}.pending.xlsx"

    @staticmethod
    def _validate_xlsx(path: Path) -> None:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True)
        _ = wb.sheetnames
        wb.close()

    def _ask_user(self, title: str, message: str, default_yes: bool = True) -> bool:
        if self._prompt_fn is not None:
            return self._prompt_fn(title, message)
        try:
            from PyQt6 import QtWidgets
            app = QtWidgets.QApplication.instance()
            if app is not None:
                default_btn = QtWidgets.QMessageBox.StandardButton.Yes if default_yes else QtWidgets.QMessageBox.StandardButton.No
                res = QtWidgets.QMessageBox.question(
                    None,
                    title,
                    message,
                    QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                    default_btn
                )
                return res == QtWidgets.QMessageBox.StandardButton.Yes
        except Exception:
            pass
        return default_yes

    def _load(self):
        try:
            from openpyxl import Workbook, load_workbook
            from openpyxl.styles import Font, PatternFill
        except ImportError as exc:
            raise RuntimeError("openpyxl is required; run bench_test.py --install") from exc

        backup_path = self.path.with_suffix(self.path.suffix + ".bak")
        pending_path = self.pending_path
        is_new = not self.path.exists()

        if pending_path.exists() and pending_path.stat().st_size > 0:
            try:
                workbook = load_workbook(pending_path)
                _ = workbook.sheetnames
                self.last_warning = (
                    f"Workbook changes are safely queued and visible in History, but {self.path.name} "
                    "is still busy. Close it in Excel so the next save can commit them."
                )
            except (zipfile.BadZipFile, KeyError, ValueError, OSError, EOFError) as pending_exc:
                raise RuntimeError(f"Pending workbook state is unreadable: {pending_path}") from pending_exc
        elif is_new or self.path.stat().st_size == 0:
            workbook = Workbook()
            workbook.remove(workbook.active)
        else:
            try:
                workbook = load_workbook(self.path)
                _ = workbook.sheetnames
            except (zipfile.BadZipFile, KeyError, ValueError, OSError, EOFError) as load_exc:
                self.last_warning = "Results workbook could not be read. The XLSX file appears damaged. No hardware fault occurred."

                # Check whether the existing .bak backup is readable with openpyxl.load_workbook()
                backup_valid = False
                if backup_path.exists() and backup_path.stat().st_size > 0:
                    try:
                        with open(backup_path, "rb") as f:
                            bak_wb = load_workbook(f, read_only=True)
                            _ = bak_wb.sheetnames
                            bak_wb.close()
                        backup_valid = True
                    except Exception:
                        backup_valid = False

                stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                corrupt_target = self.path.with_name(f"{self.path.stem}_CORRUPT_{stamp}{self.path.suffix}")

                if backup_valid:
                    msg = (
                        f"Results workbook could not be read. The XLSX file appears damaged. No hardware fault occurred.\n\n"
                        f"File: {self.path.name}\n\n"
                        f"A valid backup ({backup_path.name}) is available.\n\n"
                        f"Workbook is damaged. Recover from last backup?"
                    )
                    recover = self._ask_user("Workbook Damaged", msg, default_yes=True)
                    if recover:
                        try:
                            self.path.rename(corrupt_target)
                        except Exception:
                            shutil.copy2(self.path, corrupt_target)

                        shutil.copy2(backup_path, self.path)
                        workbook = load_workbook(self.path)
                        self.last_warning = f"Recovered from backup {backup_path.name}. Damaged file preserved as {corrupt_target.name}."
                    else:
                        raise RuntimeError("Results workbook could not be read. The XLSX file appears damaged. No hardware fault occurred.") from load_exc
                else:
                    msg = (
                        f"Results workbook could not be read. The XLSX file appears damaged. No hardware fault occurred.\n\n"
                        f"File: {self.path.name}\n\n"
                        f"No valid backup was found. Create a new workbook?"
                    )
                    create_new = self._ask_user("Workbook Unrecoverable", msg, default_yes=True)
                    if create_new:
                        try:
                            self.path.rename(corrupt_target)
                        except Exception:
                            pass
                        workbook = Workbook()
                        workbook.remove(workbook.active)
                        self.last_warning = f"Created new workbook. Damaged file preserved as {corrupt_target.name}."
                    else:
                        raise RuntimeError("Results workbook could not be read. The XLSX file appears damaged. No hardware fault occurred.") from load_exc

        headers = {"Runs": RUN_HEADERS, "Measurements": MEAS_HEADERS, "Events": EVENT_HEADERS}
        for name, columns in headers.items():
            if name not in workbook.sheetnames:
                sheet = workbook.create_sheet(name)
                sheet.append(columns)
                for cell in sheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF")
                    cell.fill = PatternFill("solid", fgColor="2367D1")
                sheet.freeze_panes = "B2" if name in {"Runs", "Measurements"} else "A2"
                sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(columns)).coordinate}"
                if name == "Runs":
                    sheet.column_dimensions["A"].width = 24
                elif name == "Measurements":
                    sheet.column_dimensions["A"].width = 28
            else:
                sheet = workbook[name]
                # Enforce column A frozen horizontally (B2 freeze)
                target_freeze = "B2" if name in {"Runs", "Measurements"} else "A2"
                sheet.freeze_panes = target_freeze

                # Check for schema updates / missing columns in existing sheets
                existing_headers = [sheet.cell(1, col).value for col in range(1, sheet.max_column + 1)]
                if existing_headers != columns:
                    old_rows: list[dict[str, Any]] = []
                    for row_idx in range(2, sheet.max_row + 1):
                        row_dict: dict[str, Any] = {}
                        for col_idx, col_name in enumerate(existing_headers, start=1):
                            if col_name:
                                row_dict[str(col_name)] = sheet.cell(row_idx, col_idx).value
                        old_rows.append(row_dict)

                    sheet.delete_rows(1, sheet.max_row)
                    sheet.append(columns)
                    for cell in sheet[1]:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill("solid", fgColor="2367D1")
                    sheet.freeze_panes = target_freeze
                    sheet.auto_filter.ref = f"A1:{sheet.cell(1, len(columns)).coordinate}"
                    for row_dict in old_rows:
                        sheet.append([self._as_cell(row_dict.get(col_name, "")) for col_name in columns])
                    if name == "Runs":
                        sheet.column_dimensions["A"].width = 24
                    elif name == "Measurements":
                        sheet.column_dimensions["A"].width = 28

        if "Plots" not in workbook.sheetnames:
            workbook.create_sheet("Plots")

        # Consolidate any duplicate RunIDs in Runs sheet (normalization/migration pass)
        if "Runs" in workbook.sheetnames:
            r_sheet = workbook["Runs"]
            id_col = RUN_HEADERS.index("RunID") + 1
            status_col = RUN_HEADERS.index("Status") + 1

            rows_by_id: dict[str, list[int]] = {}
            for r in range(2, r_sheet.max_row + 1):
                c_val = r_sheet.cell(r, id_col).value
                if c_val is not None and str(c_val).strip():
                    rows_by_id.setdefault(str(c_val).strip(), []).append(r)

            status_priority = {
                "valid": 10,
                "completed": 9,
                "stopped": 8,
                "invalid": 5,
                "aborted": 3,
                "in progress": 2,
            }

            rows_to_delete: list[int] = []
            events_to_log: list[dict[str, Any]] = []

            for rid, row_list in rows_by_id.items():
                if len(row_list) > 1:
                    best_row = row_list[0]
                    best_score = -1
                    merged_values: dict[int, Any] = {}

                    for r in row_list:
                        st = str(r_sheet.cell(r, status_col).value or "").strip().lower()
                        score = status_priority.get(st, 4)
                        if score > best_score:
                            best_score = score
                            best_row = r
                        for col in range(1, len(RUN_HEADERS) + 1):
                            v = r_sheet.cell(r, col).value
                            if v is not None and str(v).strip() != "" and col not in merged_values:
                                merged_values[col] = v

                    for col, v in merged_values.items():
                        cur_v = r_sheet.cell(best_row, col).value
                        if cur_v is None or str(cur_v).strip() == "":
                            r_sheet.cell(best_row, col).value = v

                    for r in row_list:
                        if r != best_row:
                            rows_to_delete.append(r)

                    events_to_log.append({
                        "Timestamp": utc_now(),
                        "RunID": rid,
                        "Event": "Duplicate Runs rows consolidated",
                        "Detail": f"Consolidated {len(row_list)} rows to 1 canonical row ({len(row_list) - 1} duplicates removed)",
                    })

            if rows_to_delete:
                for r in sorted(rows_to_delete, reverse=True):
                    r_sheet.delete_rows(r)
                if "Events" in workbook.sheetnames:
                    e_sheet = workbook["Events"]
                    for evt in events_to_log:
                        e_sheet.append([WorkbookStore._as_cell(evt.get(name, "")) for name in EVENT_HEADERS])
        return workbook

    @staticmethod
    def _as_cell(value: Any) -> Any:
        return json.dumps(value, sort_keys=True) if isinstance(value, (dict, list)) else value

    def _save_atomic(self, workbook) -> bool:
        self.path.parent.mkdir(parents=True, exist_ok=True)
        pending = self.pending_path
        pending.parent.mkdir(parents=True, exist_ok=True)
        temp = pending.with_name(f"{pending.stem}.{uuid.uuid4().hex}.tmp.xlsx")
        backup = self.path.with_suffix(self.path.suffix + ".bak")
        self.last_warning = ""

        # Integrity verification on Runs worksheet
        if "Runs" in workbook.sheetnames:
            r_sheet = workbook["Runs"]
            id_col = RUN_HEADERS.index("RunID") + 1
            run_ids = [
                str(r_sheet.cell(row, id_col).value or "").strip()
                for row in range(2, r_sheet.max_row + 1)
                if r_sheet.cell(row, id_col).value is not None and str(r_sheet.cell(row, id_col).value).strip()
            ]
            if len(run_ids) != len(set(run_ids)):
                self.last_warning = f"Integrity warning: {len(run_ids) - len(set(run_ids))} duplicate RunID(s) detected in Runs."

        try:
            workbook.save(temp)
            self._validate_xlsx(temp)
            os.replace(temp, pending)

            # Build a complete candidate beside the live workbook.  Replacing a file in
            # its own directory is atomic when the filesystem permits it.
            candidate = self.path.with_name(f".{self.path.stem}.{uuid.uuid4().hex}.commit.xlsx")
            shutil.copy2(pending, candidate)
            self._validate_xlsx(candidate)

            if self.path.exists():
                try:
                    self._validate_xlsx(self.path)
                    shutil.copy2(self.path, backup)
                except Exception:
                    pass

            commit_error: OSError | None = None
            for delay in (0.0, 0.05, 0.15, 0.30):
                if delay:
                    time.sleep(delay)
                try:
                    os.replace(candidate, self.path)
                    commit_error = None
                    break
                except OSError as exc:
                    commit_error = exc

            if commit_error is not None:
                candidate.unlink(missing_ok=True)
                self.last_warning = (
                    f"Could not update {self.path.name} because it is open or temporarily locked. "
                    "All changes are safely queued in one pending state and remain visible in History. "
                    "Close the workbook in Excel; the next save will retry."
                )
                return False

            try:
                self._validate_xlsx(self.path)
            except Exception as live_exc:
                self.last_warning = (
                    f"Committed workbook failed final validation; the complete pending state was preserved at {pending}."
                )
                raise RuntimeError(self.last_warning) from live_exc

            pending.unlink(missing_ok=True)
            return True
        except Exception as val_exc:
            try:
                temp.unlink(missing_ok=True)
            except Exception:
                pass
            try:
                if 'candidate' in locals():
                    candidate.unlink(missing_ok=True)
            except Exception:
                pass
            raise RuntimeError(f"Failed to validate temporary workbook before commit: {val_exc}") from val_exc

    def preflight(self, require_live_commit: bool = True) -> None:
        """Validate/recover the store before a sweep is allowed to start."""
        with self._lock:
            wb = self._load()
            committed = self._save_atomic(wb)
            if require_live_commit and not committed:
                raise RuntimeError(self.last_warning)

    @staticmethod
    def _append(sheet, headers: list[str], record: dict[str, Any]) -> None:
        sheet.append([WorkbookStore._as_cell(record.get(name, "")) for name in headers])

    def create_run(self, record: dict[str, Any]) -> bool:
        with self._lock:
            wb = self._load()
            sheet = wb["Runs"]
            id_col = RUN_HEADERS.index("RunID") + 1
            run_id = str(record.get("RunID", "")).strip()
            if not run_id:
                return False

            # Search if RunID already exists in Runs sheet
            for row in range(2, sheet.max_row + 1):
                cell_val = sheet.cell(row, id_col).value
                if cell_val is not None and str(cell_val).strip() == run_id:
                    # Run already exists, do not append duplicate!
                    return False

            self._append(sheet, RUN_HEADERS, record)
            self._append(wb["Events"], EVENT_HEADERS, {"Timestamp": utc_now(), "RunID": run_id, "Event": "Run created", "Detail": str(record.get("Mode", ""))})
            self._save_atomic(wb)
            return True

    def ensure_run_exists(self, record: dict[str, Any]) -> bool:
        return self.create_run(record)

    def append_measurement(self, record: dict[str, Any], duplicate_action: str = "keep") -> None:
        with self._lock:
            wb = self._load()
            sheet = wb["Measurements"]
            if duplicate_action == "supersede":
                for row in self._matching_rows(sheet, record):
                    old_id = sheet.cell(row, MEAS_HEADERS.index("PointID") + 1).value
                    sheet.cell(row, MEAS_HEADERS.index("Status") + 1).value = "Superseded"
                    record["SupersedesPointID"] = old_id
            self._append(sheet, MEAS_HEADERS, record)
            self._save_atomic(wb)

    def update_measurement_scope(self, pid: str, status: str, error: str = "", png: str = "", csv_f: str = "") -> None:
        with self._lock:
            wb = self._load()
            sheet = wb["Measurements"]
            pid_col = MEAS_HEADERS.index("PointID") + 1
            for row in range(2, sheet.max_row + 1):
                if str(sheet.cell(row, pid_col).value or "").strip() == pid:
                    sheet.cell(row, MEAS_HEADERS.index("ScopeCaptureStatus") + 1).value = status
                    sheet.cell(row, MEAS_HEADERS.index("ScopeCaptureError") + 1).value = error
                    sheet.cell(row, MEAS_HEADERS.index("ScopePNG") + 1).value = png
                    sheet.cell(row, MEAS_HEADERS.index("ScopeCSV") + 1).value = csv_f
                    break
            self._save_atomic(wb)

    def discard_interrupted_point(self, pid: str, run_id: str, reason: str) -> bool:
        """Remove an interrupted point and its run shell when no completed point remains."""
        with self._lock:
            wb = self._load()
            sheet = wb["Measurements"]
            pid_col = MEAS_HEADERS.index("PointID") + 1
            removed = False
            for row in range(sheet.max_row, 1, -1):
                if str(sheet.cell(row, pid_col).value or "").strip() == pid:
                    sheet.delete_rows(row, 1)
                    removed = True
            run_col = MEAS_HEADERS.index("RunID") + 1
            run_has_points = any(str(sheet.cell(row, run_col).value or "").strip() == run_id for row in range(2, sheet.max_row + 1))
            if not run_has_points:
                runs = wb["Runs"]
                runs_id_col = RUN_HEADERS.index("RunID") + 1
                for row in range(runs.max_row, 1, -1):
                    if str(runs.cell(row, runs_id_col).value or "").strip() == run_id:
                        runs.delete_rows(row, 1)
            if removed:
                self._append(wb["Events"], EVENT_HEADERS, {"Timestamp": utc_now(), "RunID": run_id, "PointID": pid, "Event": "Point discarded", "Detail": reason})
                self._refresh_plot(wb)
                self._save_atomic(wb)
            return removed

    def _matching_rows(self, sheet, record: dict[str, Any]) -> list[int]:
        # A point is only a duplicate inside the same run type.  A 2 A pulse is
        # not interchangeable with a 2 A continuous or manual measurement.
        indices = {name: MEAS_HEADERS.index(name) + 1 for name in ("DataSource", "Mode", "VinTarget_V", "ModulationLabel", "Frequency_Hz", "RequestedIout_A")}
        matches: list[int] = []
        for row in range(2, sheet.max_row + 1):
            equal = True
            for name, column in indices.items():
                old, new = sheet.cell(row, column).value, record.get(name)
                if isinstance(new, float):
                    try:
                        equal = abs(float(old) - new) <= 1e-9
                    except (TypeError, ValueError):
                        equal = False
                else:
                    equal = (old == new) or (str(old or "") == str(new or ""))
                if not equal:
                    break
            if equal and sheet.cell(row, MEAS_HEADERS.index("Status") + 1).value in {"Valid", "Invalid"}:
                matches.append(row)
        return matches

    def find_duplicates(self, condition: dict[str, Any], points: list[float]) -> list[dict[str, Any]]:
        if not self.path.exists():
            return []
        with self._lock:
            wb = self._load()
            sheet = wb["Measurements"]
            result: list[dict[str, Any]] = []
            for amps in points:
                probe = {**condition, "RequestedIout_A": float(amps)}
                rows = self._matching_rows(sheet, probe)
                if rows:
                    result.append({"current": amps, "count": len(rows)})
            return result

    def finish_run(self, run_id: str, status: str, warning: str = "") -> None:
        with self._lock:
            wb = self._load()
            sheet = wb["Runs"]
            id_col = RUN_HEADERS.index("RunID") + 1
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, id_col).value == run_id:
                    sheet.cell(row, RUN_HEADERS.index("Status") + 1).value = status
                    sheet.cell(row, RUN_HEADERS.index("Completed") + 1).value = utc_now()
                    if warning:
                        sheet.cell(row, RUN_HEADERS.index("Warnings") + 1).value = warning
                    break
            self._append(wb["Events"], EVENT_HEADERS, {"Timestamp": utc_now(), "RunID": run_id, "Event": f"Run {status}", "Detail": warning})
            self._refresh_plot(wb)
            self._save_atomic(wb)

    def update_run_fields(self, run_id: str, fields: dict[str, Any]) -> bool:
        """Update display metadata for an existing run without changing its identity."""
        allowed = {name for name in fields if name in RUN_HEADERS and name != "RunID"}
        if not allowed:
            return False
        with self._lock:
            wb = self._load()
            sheet = wb["Runs"]
            id_col = RUN_HEADERS.index("RunID") + 1
            for row in range(2, sheet.max_row + 1):
                if str(sheet.cell(row, id_col).value or "").strip() == run_id:
                    for name in allowed:
                        sheet.cell(row, RUN_HEADERS.index(name) + 1).value = self._as_cell(fields[name])
                    self._save_atomic(wb)
                    return True
        return False

    def _refresh_plot(self, wb) -> None:
        from openpyxl.chart import ScatterChart, Reference, Series
        plot = wb["Plots"]
        plot.delete_rows(1, max(plot.max_row, 1))
        plot._charts = []
        plot.append(["Iout_A", "EfficiencyConverter_pct", "RunID"])
        plot.freeze_panes = "A2"
        plot.sheet_view.showGridLines = False
        plot.column_dimensions["A"].width = 14
        plot.column_dimensions["B"].width = 28
        plot.column_dimensions["C"].width = 30
        data = wb["Measurements"]
        columns = {name: MEAS_HEADERS.index(name) + 1 for name in ("Status", "DataSource", "Iout_A", "EfficiencyConverter_pct", "RunID")}
        for row in range(2, data.max_row + 1):
            if data.cell(row, columns["Status"]).value == "Valid" and data.cell(row, columns["DataSource"]).value == "Hardware":
                plot.append([data.cell(row, columns["Iout_A"]).value, data.cell(row, columns["EfficiencyConverter_pct"]).value, data.cell(row, columns["RunID"]).value])
        if plot.max_row > 1:
            chart = ScatterChart()
            chart.title = "Valid hardware efficiency"
            chart.x_axis.title = "Iout (A)"
            chart.y_axis.title = "Efficiency (%)"
            chart.y_axis.scaling.min = 0
            chart.y_axis.scaling.max = 100
            chart.width = 16
            chart.height = 8
            chart.legend = None
            chart.series.append(Series(Reference(plot, min_col=2, min_row=2, max_row=plot.max_row), Reference(plot, min_col=1, min_row=2, max_row=plot.max_row)))
            plot.add_chart(chart, "E2")

    def list_runs(self) -> list[dict[str, Any]]:
        if not self.path.exists():
            return []
        with self._lock:
            wb = self._load()
            sheet = wb["Runs"]
            raw_runs = [
                dict(zip(RUN_HEADERS, (sheet.cell(row, col).value for col in range(1, len(RUN_HEADERS) + 1))))
                for row in range(2, sheet.max_row + 1)
            ]
            status_priority = {
                "valid": 10,
                "completed": 9,
                "stopped": 8,
                "invalid": 5,
                "aborted": 3,
                "in progress": 2,
            }
            runs_by_id: dict[str, dict[str, Any]] = {}
            for r in raw_runs:
                rid = str(r.get("RunID", "")).strip()
                if not rid:
                    continue
                if rid not in runs_by_id:
                    runs_by_id[rid] = r
                else:
                    cur_stat = str(runs_by_id[rid].get("Status", "")).strip().lower()
                    new_stat = str(r.get("Status", "")).strip().lower()
                    if status_priority.get(new_stat, 4) > status_priority.get(cur_stat, 4):
                        merged = dict(runs_by_id[rid])
                        for k, v in r.items():
                            if v is not None and str(v).strip() != "":
                                merged[k] = v
                        runs_by_id[rid] = merged
            return list(runs_by_id.values())

    def get_run_measurements(self, run_id: str) -> list[dict[str, Any]]:
        if not self.path.exists():
            return []
        with self._lock:
            wb = self._load()
            sheet = wb["Measurements"]
            results: list[dict[str, Any]] = []
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, MEAS_HEADERS.index("RunID") + 1).value == run_id:
                    results.append(dict(zip(MEAS_HEADERS, (sheet.cell(row, col).value for col in range(1, len(MEAS_HEADERS) + 1)))))
            return results

    def set_run_status(self, run_id: str, status: str) -> None:
        if status not in {"Valid", "Invalid", "Superseded", "Aborted"}:
            raise ValueError(status)
        with self._lock:
            wb = self._load()
            for sheet_name, headers in (("Runs", RUN_HEADERS), ("Measurements", MEAS_HEADERS)):
                sheet = wb[sheet_name]
                id_name = "RunID"
                for row in range(2, sheet.max_row + 1):
                    if sheet.cell(row, headers.index(id_name) + 1).value == run_id:
                        sheet.cell(row, headers.index("Status") + 1).value = status
            self._append(wb["Events"], EVENT_HEADERS, {"Timestamp": utc_now(), "RunID": run_id, "Event": f"Marked {status}", "Detail": "GUI history action"})
            self._refresh_plot(wb)
            self._save_atomic(wb)

    def delete_runs(self, run_ids: list[str]) -> list[Path]:
        deleted_files: list[Path] = []
        unique_run_ids = list(dict.fromkeys([str(r).strip() for r in run_ids if str(r).strip()]))
        if not unique_run_ids:
            return deleted_files

        target_set = set(unique_run_ids)
        with self._lock:
            wb = self._load()
            r_sheet = wb["Runs"]
            run_id_col = RUN_HEADERS.index("RunID") + 1

            # Map matching row indices by RunID in Runs sheet
            matching_run_rows: dict[str, list[int]] = {rid: [] for rid in unique_run_ids}
            for row in range(2, r_sheet.max_row + 1):
                cell_val = r_sheet.cell(row, run_id_col).value
                if cell_val is not None and str(cell_val).strip():
                    val_str = str(cell_val).strip()
                    if val_str in matching_run_rows:
                        matching_run_rows[val_str].append(row)

            # Delete measurements belonging to target runs and collect linked captures
            captures: list[str] = []
            m_sheet = wb["Measurements"]
            m_run_col = MEAS_HEADERS.index("RunID") + 1
            for row in range(m_sheet.max_row, 1, -1):
                m_rid = str(m_sheet.cell(row, m_run_col).value or "").strip()
                if m_rid in target_set:
                    for name in ("ScopePNG", "ScopeCSV"):
                        value = m_sheet.cell(row, MEAS_HEADERS.index(name) + 1).value
                        if value:
                            captures.append(str(value))
                    m_sheet.delete_rows(row)

            # Delete all matching rows from Runs in reverse row order (including any duplicates)
            all_rows_to_delete = sorted([r for rid in unique_run_ids for r in matching_run_rows[rid]], reverse=True)
            for row in all_rows_to_delete:
                r_sheet.delete_rows(row)

            # Remove stale run events/index references before writing one deletion audit event.
            e_sheet = wb["Events"]
            e_run_col = EVENT_HEADERS.index("RunID") + 1
            for row in range(e_sheet.max_row, 1, -1):
                if str(e_sheet.cell(row, e_run_col).value or "").strip() in target_set:
                    e_sheet.delete_rows(row)

            # Post-delete verification: ensure target RunIDs no longer exist in Runs
            final_run_ids: list[str] = []
            for row in range(2, r_sheet.max_row + 1):
                cell_val = r_sheet.cell(row, run_id_col).value
                if cell_val is not None and str(cell_val).strip():
                    final_run_ids.append(str(cell_val).strip())

            if any(rid in final_run_ids for rid in target_set):
                raise RuntimeError("Integrity check failed: some deleted run IDs still present in Runs sheet.")

            event_text = f"Batch permanently deleted ({len(unique_run_ids)} runs)" if len(unique_run_ids) > 1 else "Run permanently deleted"
            self._append(wb["Events"], EVENT_HEADERS, {
                "Timestamp": utc_now(),
                "RunID": ", ".join(unique_run_ids),
                "Event": event_text,
                "Detail": f"Removed runs: {', '.join(unique_run_ids)}"
            })
            self._refresh_plot(wb)
            self._save_atomic(wb)

        capture_root = (self.path.parent / "captures").resolve()
        for value in captures:
            candidate = Path(value)
            if not candidate.is_absolute():
                candidate = (self.path.parent / candidate).resolve()
            try:
                candidate.relative_to(capture_root)
            except ValueError:
                continue
            if candidate.exists() and candidate.is_file():
                candidate.unlink()
                deleted_files.append(candidate)
        return deleted_files

    def delete_run(self, run_id: str) -> list[Path]:
        return self.delete_runs([run_id])


def calculate_measurement(
    pa: InstrumentSnapshot,
    load: InstrumentSnapshot,
    psu: InstrumentSnapshot | None,
    supply_channels: list[SupplyChannel],
    dimensions_mm: tuple[float, float, float] | None = None,
) -> tuple[dict[str, Any], list[str]]:
    warnings: list[str] = [warning for warning in (pa.warning, load.warning, psu.warning if psu else "") if warning]
    if not pa.valid or not load.valid or (psu is not None and not psu.valid):
        return {}, warnings or ["One or more instrument snapshots are invalid"]
    vin = pa.values.get("vin")
    iin = pa.values.get("iin")
    vout = pa.values.get("vout")
    iout = load.values.get("current")
    numeric = (vin, iin, vout, iout)
    if not all(isinstance(value, (int, float)) and math.isfinite(float(value)) for value in numeric):
        return {}, ["Required raw measurement is missing or non-finite"]
    vin, iin, vout, iout = map(float, numeric)
    if vin <= 0 or iin < 0 or vout < 0 or iout < 0:
        return {}, ["Unexpected measurement sign; derived values left blank"]
    pin = vin * iin
    pout = vout * iout
    paux = 0.0
    supply_result: dict[str, Any] = {}
    vdrv_a = None
    idrv_a = None
    pdrv_a = None
    vdrv_b = None
    idrv_b = None
    pdrv_b = None
    vdrv_c = None
    idrv_c = None
    pdrv_c = None

    if psu:
        # Extract individual measured channels 1, 2, 3
        # CH 1 (A)
        v1 = psu.values.get("ch1_voltage")
        i1 = psu.values.get("ch1_current")
        if isinstance(v1, (int, float)) and math.isfinite(float(v1)):
            vdrv_a = float(v1)
        if isinstance(i1, (int, float)) and math.isfinite(float(i1)):
            idrv_a = float(i1)
        if vdrv_a is not None and idrv_a is not None:
            pdrv_a = vdrv_a * idrv_a

        # CH 2 (B)
        v2 = psu.values.get("ch2_voltage")
        i2 = psu.values.get("ch2_current")
        if isinstance(v2, (int, float)) and math.isfinite(float(v2)):
            vdrv_b = float(v2)
        if isinstance(i2, (int, float)) and math.isfinite(float(i2)):
            idrv_b = float(i2)
        if vdrv_b is not None and idrv_b is not None:
            pdrv_b = vdrv_b * idrv_b

        # CH 3 (C)
        v3 = psu.values.get("ch3_voltage")
        i3 = psu.values.get("ch3_current")
        if isinstance(v3, (int, float)) and math.isfinite(float(v3)):
            vdrv_c = float(v3)
        if isinstance(i3, (int, float)) and math.isfinite(float(i3)):
            idrv_c = float(i3)
        if vdrv_c is not None and idrv_c is not None:
            pdrv_c = vdrv_c * idrv_c

        ch_powers = {1: pdrv_a, 2: pdrv_b, 3: pdrv_c}
        for channel in supply_channels:
            voltage = psu.values.get(f"ch{channel.channel}_voltage")
            current = psu.values.get(f"ch{channel.channel}_current")
            supply_result[f"CH{channel.channel}"] = {
                "role": channel.role,
                "voltage": voltage,
                "current": current,
                "contributes_loss": channel.contributes_loss,
            }
            ch_p = ch_powers.get(channel.channel)
            if channel.contributes_loss and ch_p is not None:
                paux += ch_p
    else:
        warnings.append("PSU measurement unavailable; auxiliary loss is incomplete")

    if pin <= 0:
        return {}, warnings + ["Nonpositive input power; derived values left blank"]

    loss_converter = pin - pout
    system_pin = pin + paux
    loss_system = system_pin - pout

    derived = {
        "Iout_A": iout, "Vin_V": vin, "Iin_A": iin, "Vout_V": vout,
        "PinConverter_W": pin, "Pout_W": pout,
        "Vdrv_A_V": vdrv_a, "Idrv_A_A": idrv_a, "Pdrv_A_W": pdrv_a,
        "Vdrv_B_V": vdrv_b, "Idrv_B_A": idrv_b, "Pdrv_B_W": pdrv_b,
        "Vdrv_C_V": vdrv_c, "Idrv_C_A": idrv_c, "Pdrv_C_W": pdrv_c,
        "Paux_W": paux, "LossConverter_W": loss_converter, "LossSystem_W": loss_system,
        "EfficiencyConverter_pct": 100.0 * pout / pin,
        "EfficiencySystem_pct": 100.0 * pout / system_pin if psu is not None and system_pin > 0 else None,
        "SupplyMeasurements": supply_result,
    }
    if derived["EfficiencyConverter_pct"] > 100.0:
        warnings.append("Converter efficiency exceeds 100%; verify signs, mapping, and settling")
    return derived, warnings


class TaskSignals(QtCore.QObject):
    success = QtCore.pyqtSignal(object)
    failure = QtCore.pyqtSignal(str)
    finished = QtCore.pyqtSignal()


class FunctionTask(QtCore.QRunnable):
    def __init__(self, function: Callable[[], Any]):
        super().__init__()
        self.function = function
        self.signals = TaskSignals()

    @QtCore.pyqtSlot()
    def run(self):
        try:
            self.signals.success.emit(self.function())
        except Exception as exc:
            self.signals.failure.emit(str(exc))
        finally:
            self.signals.finished.emit()


def device_thumbnail(device_key: str, w: int = 130, h: int = 75) -> QtGui.QPixmap:
    """Loads actual downloaded hardware photograph or falls back to vector illustration."""
    filename_map = {
        "pa": "Keysight_PA2201A.jpg",
        "load": "Chroma_63206A.jpg",
        "psu": "Keysight_E36312A.png",
        "scope": "Keysight_MSOX4024A.png",
    }
    img_path = INSTRUMENTS_DIR / filename_map.get(device_key, "")

    canvas = QtGui.QPixmap(w, h)
    canvas.fill(QtGui.QColor("#FFFFFF"))
    painter = QtGui.QPainter(canvas)
    painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)
    painter.setRenderHint(QtGui.QPainter.RenderHint.SmoothPixmapTransform)

    outer_rect = QtCore.QRectF(1, 1, w - 2, h - 2)
    painter.setPen(QtGui.QPen(QtGui.QColor(BORDER), 1.0))
    painter.setBrush(QtGui.QBrush(QtGui.QColor("#FAFAFA")))
    painter.drawRoundedRect(outer_rect, 4, 4)

    if img_path.exists():
        src_pix = QtGui.QPixmap(str(img_path))
        if not src_pix.isNull():
            padded_rect = outer_rect.adjusted(3, 3, -3, -3)
            scaled = src_pix.scaled(
                int(padded_rect.width()), int(padded_rect.height()),
                QtCore.Qt.AspectRatioMode.KeepAspectRatio,
                QtCore.Qt.TransformationMode.SmoothTransformation
            )
            x = int(padded_rect.left() + (padded_rect.width() - scaled.width()) / 2)
            y = int(padded_rect.top() + (padded_rect.height() - scaled.height()) / 2)
            painter.drawPixmap(x, y, scaled)
            painter.end()
            return canvas

    # Fallback vector illustration
    painter.setPen(QtGui.QPen(QtGui.QColor(TEXT_MUTED)))
    painter.setFont(QtGui.QFont("Segoe UI", 8, QtGui.QFont.Weight.Bold))
    label_map = {"pa": "PA2201A", "load": "CHROMA 63206A", "psu": "E36312A", "scope": "MSOX4024A"}
    painter.drawText(outer_rect, QtCore.Qt.AlignmentFlag.AlignCenter, label_map.get(device_key, "INSTRUMENT"))
    painter.end()
    return canvas


class NoWheelFilter(QtCore.QObject):
    """Prevents mouse wheel scrolling over numeric spinboxes, comboboxes, and sliders from accidentally changing numbers."""

    def eventFilter(self, obj: QtCore.QObject, event: QtCore.QEvent) -> bool:
        if event.type() == QtCore.QEvent.Type.Wheel:
            if isinstance(obj, (QtWidgets.QAbstractSpinBox, QtWidgets.QComboBox, QtWidgets.QSlider)):
                event.ignore()
                return True
        return super().eventFilter(obj, event)


class CompactDoubleSpinBox(QtWidgets.QDoubleSpinBox):
    """A numeric control that retains precision without showing trailing zeroes."""

    def textFromValue(self, value: float) -> str:
        return f"{value:g}"


class ModeIndicator(QtWidgets.QWidget):
    """Visual waveform indicator based on the Pilawa converter operating slide."""

    def __init__(self, parent: QtWidgets.QWidget | None = None):

        super().__init__(parent)
        self.mode = "Continuous"
        self.setFixedHeight(50)
        self.setMinimumWidth(260)
        self.setSizePolicy(QtWidgets.QSizePolicy.Policy.Preferred, QtWidgets.QSizePolicy.Policy.Fixed)

    def set_mode(self, mode: str):
        self.mode = mode
        self.update()

    def paintEvent(self, event: QtGui.QPaintEvent):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)

        rect = self.rect().adjusted(1, 1, -1, -1)

        if self.mode == "Continuous":
            bg_color = QtGui.QColor("#F0FDF4")
            border_color = QtGui.QColor(MODE_CONT)
            text_color = QtGui.QColor(MODE_CONT)
            title = "CONTINUOUS"
            subtitle = "Load stays ON between points"
            wave_color = QtGui.QColor(MODE_CONT)
        elif self.mode == "Pulse":
            bg_color = QtGui.QColor("#F5F3FF")
            border_color = QtGui.QColor(MODE_PULSE)
            text_color = QtGui.QColor(MODE_PULSE)
            title = "PULSE"
            subtitle = "Returns to 0 A between pulses"
            wave_color = QtGui.QColor(MODE_PULSE)
        else:
            bg_color = QtGui.QColor("#F8FAFC")
            border_color = QtGui.QColor(BORDER)
            text_color = QtGui.QColor(TEXT_MAIN)
            title = "MANUAL"
            subtitle = "Direct load current control"
            wave_color = QtGui.QColor(PRIMARY_BLUE)

        painter.setBrush(QtGui.QBrush(bg_color))
        painter.setPen(QtGui.QPen(border_color, 1.2))
        painter.drawRoundedRect(rect, 5, 5)

        painter.setPen(QtGui.QPen(text_color))
        painter.setFont(QtGui.QFont("Segoe UI", 10, QtGui.QFont.Weight.Bold))
        painter.drawText(rect.left() + 10, rect.top() + 18, title)

        painter.setFont(QtGui.QFont("Segoe UI", 8, QtGui.QFont.Weight.Normal))
        painter.setPen(QtGui.QPen(QtGui.QColor(TEXT_MUTED)))
        painter.drawText(rect.left() + 10, rect.top() + 35, subtitle)

        wave_rect = QtCore.QRectF(rect.right() - 85, rect.top() + 6, 75, 36)
        painter.setPen(QtGui.QPen(QtGui.QColor("#CBD5E1"), 1, QtCore.Qt.PenStyle.DashLine))
        painter.drawLine(QtCore.QPointF(wave_rect.left(), wave_rect.bottom() - 4), QtCore.QPointF(wave_rect.right(), wave_rect.bottom() - 4))

        painter.setPen(QtGui.QPen(wave_color, 2.0))
        path = QtGui.QPainterPath()

        if self.mode == "Continuous":
            base_y = wave_rect.bottom() - 5
            path.moveTo(wave_rect.left(), base_y)
            path.lineTo(wave_rect.left() + 18, base_y)
            path.lineTo(wave_rect.left() + 18, base_y - 12)
            path.lineTo(wave_rect.left() + 44, base_y - 12)
            path.lineTo(wave_rect.left() + 44, base_y - 24)
            path.lineTo(wave_rect.right(), base_y - 24)
        elif self.mode == "Pulse":
            base_y = wave_rect.bottom() - 5
            path.moveTo(wave_rect.left(), base_y)
            path.lineTo(wave_rect.left() + 8, base_y)
            path.lineTo(wave_rect.left() + 8, base_y - 20)
            path.lineTo(wave_rect.left() + 22, base_y - 20)
            path.lineTo(wave_rect.left() + 22, base_y)
            path.lineTo(wave_rect.left() + 40, base_y)
            path.lineTo(wave_rect.left() + 40, base_y - 24)
            path.lineTo(wave_rect.left() + 54, base_y - 24)
            path.lineTo(wave_rect.left() + 54, base_y)
            path.lineTo(wave_rect.right(), base_y)
        else:
            base_y = wave_rect.bottom() - 14
            path.moveTo(wave_rect.left() + 4, base_y)
            path.lineTo(wave_rect.right() - 4, base_y)

        painter.drawPath(path)


class SweepProgressTracker(QtWidgets.QWidget):
    """Horizontal current-progress track with min/max labels, progress dot, and center current badge."""
    def __init__(self, parent: QtWidgets.QWidget | None = None):
        super().__init__(parent)
        self.setFixedHeight(28)
        self.start_val = 0.0
        self.stop_val = 60.0
        self.current_val: float | None = None
        self.active = False

    def set_range(self, start: float, stop: float):
        self.start_val = start
        self.stop_val = max(stop, start + 0.01)
        self.update()

    def update_position(self, current: float | None, start: float | None = None, stop: float | None = None, active: bool = True):
        self.current_val = current
        if start is not None:
            self.start_val = start
        if stop is not None:
            self.stop_val = max(stop, self.start_val + 0.01)
        self.active = active
        self.update()

    def set_idle(self):
        self.active = False
        self.update()


    def paintEvent(self, event: QtGui.QPaintEvent):
        painter = QtGui.QPainter(self)
        painter.setRenderHint(QtGui.QPainter.RenderHint.Antialiasing)

        w = self.width()
        h = self.height()

        left_pad = 42
        right_pad = 46
        track_y = 10

        # Draw Start text
        painter.setFont(QtGui.QFont("Segoe UI", 9, QtGui.QFont.Weight.Bold))
        painter.setPen(QtGui.QPen(QtGui.QColor(TEXT_MUTED)))
        painter.drawText(0, 0, left_pad - 6, 20, QtCore.Qt.AlignmentFlag.AlignRight | QtCore.Qt.AlignmentFlag.AlignVCenter, f"{self.start_val:g} A")

        # Draw Stop text
        painter.drawText(w - right_pad + 6, 0, right_pad - 6, 20, QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter, f"{self.stop_val:g} A")

        track_w = w - left_pad - right_pad
        if track_w <= 10:
            return

        # Base track line
        painter.setPen(QtGui.QPen(QtGui.QColor("#E2E8F0"), 4, QtCore.Qt.PenStyle.SolidLine, QtCore.Qt.PenCapStyle.RoundCap))
        painter.drawLine(left_pad, track_y, left_pad + track_w, track_y)

        if self.current_val is not None and self.active:
            fraction = min(1.0, max(0.0, (self.current_val - self.start_val) / (self.stop_val - self.start_val)))
            fill_x = left_pad + int(fraction * track_w)

            # Active track line
            painter.setPen(QtGui.QPen(QtGui.QColor(CALIFORNIA_GOLD), 4, QtCore.Qt.PenStyle.SolidLine, QtCore.Qt.PenCapStyle.RoundCap))
            painter.drawLine(left_pad, track_y, fill_x, track_y)

            # Dot marker
            painter.setBrush(QtGui.QBrush(QtGui.QColor(BERKELEY_BLUE)))
            painter.setPen(QtGui.QPen(QtGui.QColor("#FFFFFF"), 2))
            painter.drawEllipse(QtCore.QPointF(fill_x, track_y), 6, 6)

            # Current value label below dot
            val_str = f"{self.current_val:g} A"
            painter.setFont(QtGui.QFont("Segoe UI", 8, QtGui.QFont.Weight.Bold))
            painter.setPen(QtGui.QPen(QtGui.QColor(BERKELEY_BLUE)))
            val_w = painter.fontMetrics().horizontalAdvance(val_str)
            tx = max(left_pad, min(w - right_pad - val_w, fill_x - val_w // 2))
            painter.drawText(tx, h - 2, val_str)
        else:
            # Subtle idle dot at start
            painter.setBrush(QtGui.QBrush(QtGui.QColor("#CBD5E1")))
            painter.setPen(QtGui.QPen(QtGui.QColor("#FFFFFF"), 1.5))
            painter.drawEllipse(QtCore.QPointF(left_pad, track_y), 4, 4)


class DisclosureButton(QtWidgets.QToolButton):
    """Clean disclosure button that toggles between 'Advanced ▸' and 'Advanced ▾' and drives visibility of an attached panel."""

    def __init__(self, target_widget: QtWidgets.QWidget | None = None, parent: QtWidgets.QWidget | None = None):
        super().__init__(parent)
        self._target = target_widget
        self.setText("Advanced ▸")
        self.setCheckable(True)
        self.setChecked(False)
        self.setCursor(QtCore.Qt.CursorShape.PointingHandCursor)
        self.setStyleSheet("""
            QToolButton {
                border: 1px solid #D1D5DB;
                background: #F3F4F6;
                color: #374151;
                font-weight: 700;
                font-size: 11px;
                border-radius: 4px;
                padding: 3px 10px;
            }
            QToolButton:hover {
                background: #E5E7EB;
                color: #002676;
            }
            QToolButton:checked {
                background: #EFF6FF;
                color: #1E40AF;
                border-color: #93C5FD;
            }
        """)
        self.toggled.connect(self._on_toggled)

    def set_target(self, target_widget: QtWidgets.QWidget):
        self._target = target_widget
        if self._target:
            self._target.setVisible(self.isChecked())

    def _on_toggled(self, checked: bool):
        self.setText("Advanced ▾" if checked else "Advanced ▸")
        if self._target:
            self._target.setVisible(checked)


class DeleteRunDialog(QtWidgets.QDialog):
    """Destructive confirmation dialog for permanently deleting a run from the workbook and disk."""

    def __init__(self, run_info: dict[str, Any], parent: QtWidgets.QWidget | None = None):
        super().__init__(parent)
        self.setWindowTitle("Permanently Delete Run")
        self.setModal(True)
        self.setFixedWidth(460)
        self.setStyleSheet(f"background: {CARD_BG};")

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(14)

        # Warning Callout Banner
        warn_box = QtWidgets.QFrame()
        warn_box.setStyleSheet("""
            QFrame {
                background: #FEF2F2;
                border: 1.5px solid #FECACA;
                border-radius: 6px;
                padding: 10px 12px;
            }
        """)
        wb_lay = QtWidgets.QVBoxLayout(warn_box)
        wb_lay.setContentsMargins(4, 2, 4, 2)
        wb_lay.setSpacing(4)

        warn_title = QtWidgets.QLabel("Permanently delete this run?")
        warn_title.setStyleSheet("color: #991B1B; font-weight: 800; font-size: 14px; border: none; background: transparent;")
        wb_lay.addWidget(warn_title)

        warn_desc = QtWidgets.QLabel("This will remove workbook rows, linked captures, and run metadata. This cannot be undone.")
        warn_desc.setStyleSheet("color: #B91C1C; font-size: 12px; font-weight: 600; border: none; background: transparent;")
        warn_desc.setWordWrap(True)
        wb_lay.addWidget(warn_desc)
        layout.addWidget(warn_box)

        # Run Metadata Card
        meta_box = QtWidgets.QFrame()
        meta_box.setStyleSheet(f"background: #F8FAFC; border: 1px solid {BORDER}; border-radius: 6px; padding: 8px 12px;")
        grid = QtWidgets.QGridLayout(meta_box)
        grid.setContentsMargins(4, 4, 4, 4)
        grid.setHorizontalSpacing(14)
        grid.setVerticalSpacing(6)

        run_id = str(run_info.get("RunID", ""))
        short_id = run_id.split("-")[-1] if "-" in run_id else run_id
        if len(short_id) < 8 and len(run_id) > 16:
            parts = run_id.split("-")
            short_id = "-".join(parts[-2:]) if len(parts) >= 2 else run_id[-12:]

        test_name = str(run_info.get("CampaignName") or run_info.get("Test Name") or "—")
        vin_val = run_info.get("VinTarget_V")
        vin_str = f"{float(vin_val):g} V" if isinstance(vin_val, (int, float)) or (isinstance(vin_val, str) and vin_val.replace(".", "", 1).isdigit()) else str(vin_val or "—")

        freq_val = run_info.get("Frequency_Hz")
        freq_str = format_frequency_khz(freq_val)

        status_str = str(run_info.get("Status") or "—")
        source_str = str(run_info.get("DataSource") or "—")

        fields = [
            ("Test Name:", test_name),
            ("Vin:", vin_str),
            ("Switching Frequency:", freq_str),
            ("short RunID:", short_id),
            ("Status:", status_str),
            ("Data Source:", source_str),
        ]

        for row_idx, (label_text, value_text) in enumerate(fields):
            lbl = QtWidgets.QLabel(label_text)
            lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px; font-weight: 700; border: none; background: transparent;")
            val = QtWidgets.QLabel(value_text)
            val.setStyleSheet(f"color: {TEXT_MAIN}; font-size: 12px; font-weight: 700; font-family: {'Consolas, monospace' if 'RunID' in label_text else 'Segoe UI'}; border: none; background: transparent;")
            if "RunID" in label_text:
                val.setToolTip(f"Full RunID: {run_id}")
            grid.addWidget(lbl, row_idx, 0, QtCore.Qt.AlignmentFlag.AlignRight)
            grid.addWidget(val, row_idx, 1, QtCore.Qt.AlignmentFlag.AlignLeft)

        layout.addWidget(meta_box)

        # Action Buttons Row: [ Cancel ] [ Delete Permanently ]
        btn_row = QtWidgets.QHBoxLayout()
        btn_row.setSpacing(10)
        btn_row.addStretch()

        self.btn_cancel = QtWidgets.QPushButton("Cancel")
        self.btn_cancel.setStyleSheet(f"""
            QPushButton {{
                background: #FFFFFF;
                color: {TEXT_MAIN};
                border: 1.5px solid {BORDER};
                font-weight: 700;
                font-size: 12px;
                padding: 7px 18px;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background: #F3F4F6;
                border-color: #9CA3AF;
            }}
        """)
        self.btn_cancel.setDefault(True)
        self.btn_cancel.setAutoDefault(True)
        self.btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_cancel)

        self.btn_delete = QtWidgets.QPushButton("Delete Permanently")
        self.btn_delete.setStyleSheet("""
            QPushButton {{
                background: #DC2626;
                color: #FFFFFF;
                border: 1.5px solid #B91C1C;
                font-weight: 800;
                font-size: 12px;
                padding: 7px 18px;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background: #B91C1C;
            }}
            QPushButton:pressed {{
                background: #991B1B;
            }}
        """)
        self.btn_delete.setDefault(False)
        self.btn_delete.setAutoDefault(False)
        self.btn_delete.clicked.connect(self.accept)
        btn_row.addWidget(self.btn_delete)

        layout.addLayout(btn_row)

        # Explicitly set focus to Cancel button for safety
        self.btn_cancel.setFocus()


class DeleteBatchRunsDialog(QtWidgets.QDialog):
    """Destructive confirmation dialog for permanently deleting multiple runs from the workbook and disk."""

    def __init__(self, runs_list: list[dict[str, Any]], parent: QtWidgets.QWidget | None = None):
        super().__init__(parent)
        count = len(runs_list)
        self.setWindowTitle(f"Permanently Delete {count} Runs")
        self.setModal(True)
        self.setFixedWidth(520)
        self.setStyleSheet(f"background: {CARD_BG};")

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(20, 18, 20, 18)
        layout.setSpacing(14)

        # Warning Callout Banner
        warn_box = QtWidgets.QFrame()
        warn_box.setStyleSheet("""
            QFrame {
                background: #FEF2F2;
                border: 1.5px solid #FECACA;
                border-radius: 6px;
                padding: 10px 12px;
            }
        """)
        wb_lay = QtWidgets.QVBoxLayout(warn_box)
        wb_lay.setContentsMargins(4, 2, 4, 2)
        wb_lay.setSpacing(6)

        warn_title = QtWidgets.QLabel(f"Permanently delete {count} selected runs?")
        warn_title.setStyleSheet("color: #991B1B; font-weight: 800; font-size: 14px; border: none; background: transparent;")
        wb_lay.addWidget(warn_title)

        bullets_text = (
            f"This will remove:\n"
            f"• {count} run records\n"
            f"• all measurement rows belonging to those runs\n"
            f"• linked scope PNG/CSV captures\n"
            f"• associated run metadata\n\n"
            f"This cannot be undone."
        )
        warn_desc = QtWidgets.QLabel(bullets_text)
        warn_desc.setStyleSheet("color: #B91C1C; font-size: 12px; font-weight: 600; line-height: 1.4; border: none; background: transparent;")
        warn_desc.setWordWrap(True)
        wb_lay.addWidget(warn_desc)
        layout.addWidget(warn_box)

        # Selected runs summary list / table
        runs_lbl = QtWidgets.QLabel("<b>Selected runs:</b>")
        runs_lbl.setStyleSheet(f"color: {TEXT_MAIN}; font-size: 12px;")
        layout.addWidget(runs_lbl)

        table = QtWidgets.QTableWidget(count, 4)
        table.setHorizontalHeaderLabels(["Run Short ID", "Test Name", "Status", "Source"])
        table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.NoSelection)
        table.setFixedHeight(min(180, 28 * count + 36))
        table.setStyleSheet(f"background: #F8FAFC; border: 1px solid {BORDER}; border-radius: 4px; font-size: 11px;")

        for row, r in enumerate(runs_list):
            rid = str(r.get("RunID", ""))
            short_id = rid.split("-")[-1] if "-" in rid else (rid[-5:] if len(rid) >= 5 else rid)
            item_short = QtWidgets.QTableWidgetItem(short_id)
            item_short.setFont(QtGui.QFont("Consolas", 9, QtGui.QFont.Weight.Bold))
            item_short.setToolTip(f"Full RunID: {rid}")
            table.setItem(row, 0, item_short)
            table.setItem(row, 1, QtWidgets.QTableWidgetItem(str(r.get("CampaignName") or r.get("Test Name") or "—")))
            table.setItem(row, 2, QtWidgets.QTableWidgetItem(str(r.get("Status", "—"))))
            table.setItem(row, 3, QtWidgets.QTableWidgetItem(str(r.get("DataSource", "—"))))

        layout.addWidget(table)

        # Action Buttons Row: [ Cancel ] [ Delete N Runs Permanently ]
        btn_row = QtWidgets.QHBoxLayout()
        btn_row.setSpacing(10)
        btn_row.addStretch()

        self.btn_cancel = QtWidgets.QPushButton("Cancel")
        self.btn_cancel.setStyleSheet(f"""
            QPushButton {{
                background: #FFFFFF;
                color: {TEXT_MAIN};
                border: 1.5px solid {BORDER};
                font-weight: 700;
                font-size: 12px;
                padding: 7px 18px;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background: #F3F4F6;
                border-color: #9CA3AF;
            }}
        """)
        self.btn_cancel.setDefault(True)
        self.btn_cancel.setAutoDefault(True)
        self.btn_cancel.clicked.connect(self.reject)
        btn_row.addWidget(self.btn_cancel)

        self.btn_delete = QtWidgets.QPushButton(f"Delete {count} Runs Permanently")
        self.btn_delete.setStyleSheet("""
            QPushButton {{
                background: #DC2626;
                color: #FFFFFF;
                border: 1.5px solid #B91C1C;
                font-weight: 800;
                font-size: 12px;
                padding: 7px 18px;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background: #B91C1C;
            }}
            QPushButton:pressed {{
                background: #991B1B;
            }}
        """)
        self.btn_delete.setDefault(False)
        self.btn_delete.setAutoDefault(False)
        self.btn_delete.clicked.connect(self.accept)
        btn_row.addWidget(self.btn_delete)
        layout.addLayout(btn_row)

        self.btn_cancel.setFocus()


class LoadCard(QtWidgets.QGroupBox):
    """Vertical card for Chroma 63206A Electronic Load with integrated Safety Limit controls."""
    snapshot = QtCore.pyqtSignal(str, object)
    message = QtCore.pyqtSignal(str)
    cap_applied = QtCore.pyqtSignal(float)

    def __init__(self, hub: InstrumentHub, config: dict[str, Any], save_callback: Callable[[], None]):
        super().__init__("Chroma 63206A Load")
        self.hub = hub
        self.config = config
        self.save_callback = save_callback
        self.last_snapshot: InstrumentSnapshot | None = None
        self.busy = False

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        # Top Header: Photo + Status Badge
        top_row = QtWidgets.QHBoxLayout()
        photo = QtWidgets.QLabel()
        photo.setFixedSize(130, 75)
        photo.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        photo.setPixmap(device_thumbnail("load", 128, 73))
        photo.setStyleSheet(f"background:white; border:1px solid {BORDER}; border-radius:4px;")
        top_row.addWidget(photo)

        self.status_badge = QtWidgets.QLabel("Not Checked")
        self.status_badge.setObjectName("badge_gray")
        self.status_badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.status_badge.setToolTip("Not checked yet")
        self._set_badge("gray", "Not Checked", "Not checked yet")
        top_row.addWidget(self.status_badge, 1)
        layout.addLayout(top_row)

        # Action: Quick Query
        self.read_btn = QtWidgets.QPushButton("Read Load Now")
        self.read_btn.setToolTip("Query present load voltage, current, and power.")
        self.read_btn.clicked.connect(self.read_now)
        layout.addWidget(self.read_btn)

        # Live Metrics display
        self.metric_labels: dict[str, QtWidgets.QLabel] = {}
        grid = QtWidgets.QGridLayout()
        grid.setHorizontalSpacing(8)
        grid.setVerticalSpacing(4)
        for idx, (name, unit) in enumerate([("Iout", "A"), ("Load V", "V"), ("Load P", "W"), ("Load State", "")]):
            lbl_name = QtWidgets.QLabel(f"{name}:")
            lbl_name.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px; font-weight: 600;")
            lbl_val = QtWidgets.QLabel("—" if not unit else f"— {unit}")
            lbl_val.setStyleSheet(f"font-family: Consolas, monospace; font-size: 13px; font-weight: 700; color: {BERKELEY_BLUE};")
            grid.addWidget(lbl_name, idx, 0)
            grid.addWidget(lbl_val, idx, 1)
            self.metric_labels[name] = lbl_val
        layout.addLayout(grid)

        # Prominent Amber ELECTRONIC LOAD SAFETY LIMIT Section
        safety_box = QtWidgets.QFrame()
        safety_box.setObjectName("safety_box")
        safety_box.setStyleSheet(f"""
            QFrame#safety_box {{
                background: #FFFBEB;
                border: 2px solid {WARNING_AMBER};
                border-radius: 6px;
                padding: 6px;
            }}
        """)
        sb_lay = QtWidgets.QVBoxLayout(safety_box)
        sb_lay.setContentsMargins(6, 6, 6, 6)
        sb_lay.setSpacing(4)

        title_lbl = QtWidgets.QLabel("⚡ ELECTRONIC LOAD SAFETY LIMIT")
        title_lbl.setStyleSheet(f"color: #B45309; font-weight: 900; font-size: 11px; letter-spacing: 0.5px;")
        title_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        sb_lay.addWidget(title_lbl)

        cap_lbl = QtWidgets.QLabel("Maximum allowed current:")
        cap_lbl.setStyleSheet("font-size: 11px; font-weight: 700; color: #78350F;")
        sb_lay.addWidget(cap_lbl)

        self.cap_spin = QtWidgets.QDoubleSpinBox()
        self.cap_spin.setRange(0.0, 2000.0)
        self.cap_spin.setValue(float(self.config.get("working_current_cap_a", 60.0)))
        self.cap_spin.setSuffix(" A")
        self.cap_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cap_spin.setStyleSheet(f"""
            QDoubleSpinBox {{
                font-family: Consolas, monospace;
                font-size: 16px;
                font-weight: 900;
                color: #92400E;
                background: #FEF3F2;
                background: #FEF3C7;
                border: 1px solid #FCD34D;
                border-radius: 4px;
                padding: 4px;
            }}
        """)
        sb_lay.addWidget(self.cap_spin)

        self.apply_cap_btn = QtWidgets.QPushButton("Apply Safety Limit")
        self.apply_cap_btn.setStyleSheet(f"""
            QPushButton {{
                background: {WARNING_AMBER};
                color: white;
                font-weight: 800;
                font-size: 12px;
                border: 1px solid #B45309;
                border-radius: 4px;
                padding: 5px;
            }}
            QPushButton:hover {{
                background: #B45309;
            }}
        """)
        self.apply_cap_btn.clicked.connect(self._apply_safety_limit)
        sb_lay.addWidget(self.apply_cap_btn)

        self.chk_load = QtWidgets.QCheckBox("I verified low-current load control")
        self.chk_load.setToolTip("Verify that low-current load control has been checked on this bench before starting tests.")
        self.chk_load.setStyleSheet("font-size: 11px; font-weight: 600; color: #78350F; margin-top: 4px;")
        sb_lay.addWidget(self.chk_load)

        self.chk_vin_safety = QtWidgets.QCheckBox("Auto LOAD OFF if Vin is outside ±10% of Target")
        self.chk_vin_safety.setToolTip("During hardware measurements, automatically turns the electronic load OFF and aborts the run if measured Vin falls below 90% or rises above 110% of Target Vin.")
        self.chk_vin_safety.setStyleSheet("font-size: 11px; font-weight: 600; color: #78350F; margin-top: 2px;")
        self.chk_vin_safety.setChecked(bool(self.config.get("vin_safety_enabled", True)))
        def _on_vin_safety_toggled(checked: bool):
            self.config["vin_safety_enabled"] = checked
            self.save_callback()
        self.chk_vin_safety.toggled.connect(_on_vin_safety_toggled)
        sb_lay.addWidget(self.chk_vin_safety)

        layout.addWidget(safety_box)
        layout.addStretch()

    def _set_badge(self, state: str, text: str, detail: str = ""):
        self.status_badge.setText(text)
        self.status_badge.setToolTip(detail or text)
        style_map = {
            "green": f"color: {SUCCESS_GREEN}; background: #EDFDF5; border: 1px solid {SUCCESS_GREEN}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "gray": f"color: {TEXT_MUTED}; background: #EAECF0; border: 1px solid {BORDER}; border-radius: 4px; padding: 3px 8px; font-weight: 500; font-size: 12px;",
            "blue": f"color: {PRIMARY_BLUE}; background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "amber": f"color: {WARNING_AMBER}; background: #FFF8E6; border: 1px solid {WARNING_AMBER}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "red": f"color: {DANGER_RED}; background: #FEF3F2; border: 1px solid {DANGER_RED}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
        }
        self.status_badge.setStyleSheet(style_map.get(state, style_map["gray"]))

    def _mark_discovered(self, identity: str = ""):
        self.released = False
        self._set_badge("blue", "Discovered", f"Found over VISA · {identity}" if identity else "Found over VISA")

    def _mark_not_found(self, detail: str = ""):
        self.released = False
        self._set_badge("amber", "Not Found", detail or "Chroma electronic load not detected during VISA scan")

    def _apply_safety_limit(self):
        val = self.cap_spin.value()
        self.config["working_current_cap_a"] = val
        self.save_callback()
        self.cap_applied.emit(val)
        self.message.emit(f"Electronic load safety limit updated to {val:g} A")

    def read_now(self):
        if self.busy:
            return
        self.busy = True
        self.read_btn.setEnabled(False)
        self.read_btn.setText("Reading...")

        def query():
            return self.hub.instruments["load"].read_snapshot(include_voltage=True)

        def on_done(snap: InstrumentSnapshot):
            self.released = False
            self.last_snapshot = snap
            self._render_values()
            self.snapshot.emit("load", snap)

        def on_fail(err: str):
            self.released = False
            self._set_badge("red", "Offline", err)
            self.message.emit(f"Load Read Error: {err}")

        def finish():
            self.busy = False
            self.read_btn.setEnabled(True)
            self.read_btn.setText("Read Load Now")

        task = FunctionTask(query)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        task.signals.finished.connect(finish)
        QtCore.QThreadPool.globalInstance().start(task)

    def _render_values(self):
        if not self.last_snapshot:
            return
        vals = self.last_snapshot.values
        is_rel = getattr(self, "released", False)
        if self.last_snapshot.valid:
            style = f"font-family: Consolas, monospace; font-size: 13px; font-weight: 600; color: {TEXT_MUTED};" if is_rel else f"font-family: Consolas, monospace; font-size: 13px; font-weight: 700; color: {BERKELEY_BLUE};"
            tip = "Last reading before release" if is_rel else ""
            cur = vals.get("current") if "current" in vals else vals.get("iout")
            volt = vals.get("voltage") if "voltage" in vals else vals.get("vout")
            pwr = vals.get("power") if "power" in vals else vals.get("pout")

            if "Iout" in self.metric_labels:
                self.metric_labels["Iout"].setText(f"{float(cur):.2f} A" if cur is not None else "—")
                self.metric_labels["Iout"].setStyleSheet(style)
                self.metric_labels["Iout"].setToolTip(tip)
            if "Load V" in self.metric_labels:
                self.metric_labels["Load V"].setText(f"{float(volt):.2f} V" if volt is not None else "—")
                self.metric_labels["Load V"].setStyleSheet(style)
                self.metric_labels["Load V"].setToolTip(tip)
            if "Load P" in self.metric_labels:
                self.metric_labels["Load P"].setText(f"{float(pwr):.2f} W" if pwr is not None else "—")
                self.metric_labels["Load P"].setStyleSheet(style)
                self.metric_labels["Load P"].setToolTip(tip)

            in_on = vals.get("input_on")
            st_text = "ON" if in_on is True else ("OFF" if in_on is False else "Unknown")
            if is_rel:
                st_text = f"{st_text} (Last known)"
            if "Load State" in self.metric_labels:
                self.metric_labels["Load State"].setText(st_text)
                self.metric_labels["Load State"].setToolTip(tip)
                if is_rel:
                    self.metric_labels["Load State"].setStyleSheet(style)
                else:
                    self.metric_labels["Load State"].setStyleSheet(
                        f"font-family: Consolas, monospace; font-size: 13px; font-weight: 800; color: {SUCCESS_GREEN if in_on else TEXT_MUTED};"
                    )
            if not is_rel:
                status_text = getattr(self.last_snapshot, "status", "") or "Connected"
                badge_color = "amber" if "Partial" in status_text else "green"
                self._set_badge(badge_color, status_text, f"{status_text} · {datetime.now().strftime('%H:%M:%S')}")
        else:
            if not is_rel:
                status_text = getattr(self.last_snapshot, "status", "") or "Connected · Read Error"
                self._set_badge("amber", status_text, self.last_snapshot.warning or "Measurement query failed")
            if "Load State" in self.metric_labels:
                self.metric_labels["Load State"].setText("Unknown")
                self.metric_labels["Load State"].setStyleSheet(
                    f"font-family: Consolas, monospace; font-size: 13px; font-weight: 700; color: {WARNING_AMBER};"
                )

    def update_age(self):
        if getattr(self, "released", False):
            return
        if not self.last_snapshot or not self.last_snapshot.valid:
            return
        try:
            stamp = datetime.fromisoformat(self.last_snapshot.timestamp).timestamp()
            age = max(0.0, time.time() - stamp)
            if age > 10.0:
                self._set_badge("amber", "Stale", f"Last valid read {age:.0f}s ago")
        except Exception:
            pass

    def _mark_released(self):
        self.released = True
        self._set_badge("gray", "Released", "VISA session closed · front panel restored")
        self._render_values()


class SupplyCard(QtWidgets.QGroupBox):



    """Vertical card for Keysight E36312A with embedded channel configuration."""
    snapshot = QtCore.pyqtSignal(str, object)
    message = QtCore.pyqtSignal(str)
    settings_applied = QtCore.pyqtSignal()

    def __init__(self, hub: InstrumentHub, config: dict[str, Any], save_callback: Callable[[], None]):
        super().__init__("E36312A Supply")
        self.hub = hub
        self.config = config
        self.save_callback = save_callback
        self.pool = QtCore.QThreadPool.globalInstance()
        self.last_snapshot: InstrumentSnapshot | None = None
        self.busy = False
        self.channel_controls: list[dict[str, Any]] = []

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        # Top Header: Photo + Status Badge
        top_row = QtWidgets.QHBoxLayout()
        photo = QtWidgets.QLabel()
        photo.setFixedSize(130, 75)
        photo.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        photo.setPixmap(device_thumbnail("psu", 128, 73))
        photo.setStyleSheet(f"background:white; border:1px solid {BORDER}; border-radius:4px;")
        top_row.addWidget(photo)

        self.status_badge = QtWidgets.QLabel("Not Checked")
        self.status_badge.setObjectName("badge_gray")
        self.status_badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.status_badge.setToolTip("Supply not checked yet")
        self._set_badge("gray", "Not Checked", "Supply not checked yet")
        top_row.addWidget(self.status_badge, 1)
        layout.addLayout(top_row)

        # Embedded Channel Controls for CH1, CH2, CH3 (Vertical stack with clean typography & subtle dividers)
        raw_channels = self.config.get("supply_channels", [])
        for ch_idx, ch_data in enumerate(raw_channels):
            ch_num = ch_data.get("channel", 1)
            is_ch1 = (ch_num == 1)
            v_max = 6.0 if is_ch1 else 25.0
            i_max = 5.0 if is_ch1 else 1.0
            rating_str = "6 V / 5 A" if is_ch1 else "25 V / 1 A"

            if ch_idx > 0:
                sep = QtWidgets.QFrame()
                sep.setFrameShape(QtWidgets.QFrame.Shape.HLine)
                sep.setStyleSheet(f"border: none; background: {BORDER}; max-height: 1px; margin: 4px 0;")
                layout.addWidget(sep)

            ch_widget = QtWidgets.QWidget()
            b_lay = QtWidgets.QVBoxLayout(ch_widget)
            b_lay.setContentsMargins(2, 2, 2, 2)
            b_lay.setSpacing(3)

            # Row 1: Channel Tag + Rating & Desired ON Checkbox
            r1 = QtWidgets.QHBoxLayout()
            r1.setSpacing(6)
            ch_lbl = QtWidgets.QLabel(f"<b>CH{ch_num}</b>  <span style='color:{TEXT_MUTED}; font-size:11px;'>{rating_str}</span>")
            ch_lbl.setStyleSheet(f"color: {TEXT_MAIN}; font-size: 12px;")

            desired_chk = QtWidgets.QCheckBox("Desired ON")
            desired_chk.setChecked(ch_data.get("enabled", False))
            desired_chk.setToolTip("Intended output state to apply to E36312A on next Apply Settings.")
            desired_chk.setStyleSheet("font-size: 11px; font-weight: 700; color: #1E293B;")

            r1.addWidget(ch_lbl)
            r1.addStretch()
            r1.addWidget(desired_chk)
            b_lay.addLayout(r1)

            # Row 2: Role Editor
            r2 = QtWidgets.QHBoxLayout()
            r2.setSpacing(6)
            role_lbl = QtWidgets.QLabel("Role:")
            role_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px; font-weight: 600;")
            role_edit = QtWidgets.QLineEdit(ch_data.get("role", f"Vdrv_{chr(64+ch_num)}"))
            role_edit.setPlaceholderText("Role name")
            role_edit.setToolTip(f"Identifier for CH{ch_num} saved in workbook and measurement records.")
            role_edit.setStyleSheet(f"padding: 2px 5px; font-size: 11px; border: 1px solid {BORDER}; border-radius: 3px;")
            r2.addWidget(role_lbl)
            r2.addWidget(role_edit, 1)
            b_lay.addLayout(r2)

            # Row 3: Setpoints (Voltage & Current Lim) in a clean grid
            sp_grid = QtWidgets.QGridLayout()
            sp_grid.setHorizontalSpacing(8)
            sp_grid.setVerticalSpacing(2)
            sp_grid.setContentsMargins(0, 2, 0, 2)

            v_lbl = QtWidgets.QLabel("Set Voltage")
            v_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px; font-weight: 600;")
            v_spin = QtWidgets.QDoubleSpinBox()
            v_spin.setRange(0.0, v_max)
            v_spin.setValue(float(ch_data.get("voltage_set", 0.0)))
            v_spin.setSuffix(" V")
            v_spin.setDecimals(2)
            v_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
            v_spin.setToolTip(f"CH{ch_num} Set Voltage (0.0 to {v_max:g} V)")
            v_spin.setStyleSheet(f"padding: 2px 4px; font-size: 11px; font-family: Consolas, monospace;")

            i_lbl = QtWidgets.QLabel("Current Limit")
            i_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px; font-weight: 600;")
            i_spin = QtWidgets.QDoubleSpinBox()
            i_spin.setRange(0.001, i_max)
            i_spin.setValue(float(ch_data.get("current_limit", 1.0)))
            i_spin.setSuffix(" A")
            i_spin.setDecimals(2)
            i_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
            i_spin.setToolTip(f"CH{ch_num} Current Limit (0.001 to {i_max:g} A)")
            i_spin.setStyleSheet(f"padding: 2px 4px; font-size: 11px; font-family: Consolas, monospace;")

            sp_grid.addWidget(v_lbl, 0, 0)
            sp_grid.addWidget(v_spin, 0, 1)
            sp_grid.addWidget(i_lbl, 1, 0)
            sp_grid.addWidget(i_spin, 1, 1)
            b_lay.addLayout(sp_grid)

            # Row 4: Actual Readback (Clean typography without nested container)
            live_lbl = QtWidgets.QLabel("Actual: — V · — A · —")
            live_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 11px; font-weight: 600; padding: 2px 0;")
            b_lay.addWidget(live_lbl)

            # Row 5: Include in Paux Loss Checkbox
            loss_chk = QtWidgets.QCheckBox("Include in Paux loss")
            loss_chk.setChecked(ch_data.get("contributes_loss", True))
            loss_chk.setStyleSheet(f"font-size: 11px; color: {TEXT_MAIN};")
            loss_chk.setToolTip("When checked, this channel's power (V * I) is added to Paux auxiliary loss in System Loss and System Efficiency calculations.")
            b_lay.addWidget(loss_chk)

            layout.addWidget(ch_widget)
            self.channel_controls.append({
                "channel": ch_num, "role_edit": role_edit, "desired_out": desired_chk,
                "voltage": v_spin, "limit": i_spin, "live_lbl": live_lbl, "loss_chk": loss_chk
            })

        # Apply to Supply Button
        self.apply_btn = QtWidgets.QPushButton("Apply Settings to Supply")
        self.apply_btn.setObjectName("primary_action")
        self.apply_btn.clicked.connect(self._apply)
        layout.addWidget(self.apply_btn)
        layout.addStretch()

    def _set_badge(self, state: str, text: str, detail: str = ""):
        self.status_badge.setText(text)
        self.status_badge.setToolTip(detail or text)
        style_map = {
            "green": f"color: {SUCCESS_GREEN}; background: #EDFDF5; border: 1px solid {SUCCESS_GREEN}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "gray": f"color: {TEXT_MUTED}; background: #EAECF0; border: 1px solid {BORDER}; border-radius: 4px; padding: 3px 8px; font-weight: 500; font-size: 12px;",
            "blue": f"color: {PRIMARY_BLUE}; background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "amber": f"color: {WARNING_AMBER}; background: #FFF8E6; border: 1px solid {WARNING_AMBER}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "red": f"color: {DANGER_RED}; background: #FEF3F2; border: 1px solid {DANGER_RED}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
        }
        self.status_badge.setStyleSheet(style_map.get(state, style_map["gray"]))

    def _mark_discovered(self, identity: str = ""):
        self.released = False
        self._set_badge("blue", "Discovered", f"Found over VISA · {identity}" if identity else "Found over VISA")

    def _mark_not_found(self, detail: str = ""):
        self.released = False
        self._set_badge("amber", "Not Found", detail or "E36312A power supply not detected during VISA scan")

    def _apply(self):
        if self.busy:
            return
        self.busy = True
        self.apply_btn.setEnabled(False)
        self.apply_btn.setText("Applying...")

        new_channels: list[SupplyChannel] = []
        for c in self.channel_controls:
            role_text = c["role_edit"].text().strip() or f"CH{c['channel']}"
            new_channels.append(SupplyChannel(
                channel=c["channel"],
                role=role_text,
                displayed=True,
                enabled=c["desired_out"].isChecked(),
                contributes_loss=c["loss_chk"].isChecked(),
                voltage_set=c["voltage"].value(),
                current_limit=c["limit"].value(),
            ))

        self.config["supply_channels"] = [asdict(ch) for ch in new_channels]
        self.save_callback()

        def apply_task():
            for ch in new_channels:
                self.hub.instruments["psu"].configure_channel(ch.channel, ch.voltage_set or 0.0, ch.current_limit or 0.0, ch.enabled)
            return self.hub.instruments["psu"].read_snapshot(channels=[1, 2, 3])

        def on_done(snap: InstrumentSnapshot):
            self.released = False
            self.last_snapshot = snap
            self._render_values()
            self._set_badge("green", "Connected", f"Settings applied · Connected · {datetime.now().strftime('%H:%M:%S')}")
            self.message.emit("Supply settings applied successfully")
            self.snapshot.emit("psu", snap)
            self.settings_applied.emit()

        def on_fail(err: str):
            self._set_badge("red", "Apply Fault", err)
            self.message.emit(f"Supply Apply Error: {err}")

        def finish():
            self.busy = False
            self.apply_btn.setEnabled(True)
            self.apply_btn.setText("Apply Settings to Supply")

        task = FunctionTask(apply_task)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        task.signals.finished.connect(finish)
        QtCore.QThreadPool.globalInstance().start(task)


    def read_once(self):
        if self.busy:
            return
        self.busy = True

        def query():
            return self.hub.instruments["psu"].read_snapshot(channels=[1, 2, 3])
        def on_done(snap: InstrumentSnapshot):
            self.released = False
            self.last_snapshot = snap
            self._render_values()
            self.snapshot.emit("psu", snap)
        def on_fail(err: str):
            self.released = False
            self._set_badge("red", "Offline", err)
            self.message.emit(f"Supply: {err}")
        def finish():
            self.busy = False

        task = FunctionTask(query)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        task.signals.finished.connect(finish)
        QtCore.QThreadPool.globalInstance().start(task)

    def _render_values(self):
        if not self.last_snapshot or not self.last_snapshot.valid:
            return
        vals = self.last_snapshot.values
        is_rel = getattr(self, "released", False)
        for c in self.channel_controls:
            ch = c["channel"]
            outp = vals.get(f"ch{ch}_enabled")
            v = vals.get(f"ch{ch}_voltage", 0.0)
            i = vals.get(f"ch{ch}_current", 0.0)
            st_text = "ON" if outp else "OFF"
            if is_rel:
                c["live_lbl"].setText(f"Actual: {v:.2f} V · {i:.3f} A · {st_text} (Last known)")
                c["live_lbl"].setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-weight: 600; font-size: 11px; padding: 2px 0;")
                c["live_lbl"].setToolTip("Last known state before release")
            else:
                c["live_lbl"].setText(f"Actual: {v:.2f} V · {i:.3f} A · {st_text}")
                c["live_lbl"].setToolTip("")
                if outp:
                    c["live_lbl"].setStyleSheet(f"color: {SUCCESS_GREEN}; font-family: Consolas, monospace; font-weight: 700; font-size: 11px; padding: 2px 0;")
                else:
                    c["live_lbl"].setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-weight: 600; font-size: 11px; padding: 2px 0;")

        if not is_rel:
            self._set_badge("green", "Connected", f"Connected · {datetime.now().strftime('%H:%M:%S')}")

    def _mark_released(self):
        self.released = True
        self._set_badge("gray", "Released", "Supply released; session idle")
        self._render_values()

    def update_age(self):
        if getattr(self, "released", False):
            return
        if not self.last_snapshot or not self.last_snapshot.valid:
            return
        try:
            stamp = datetime.fromisoformat(self.last_snapshot.timestamp).timestamp()
            age = max(0.0, time.time() - stamp)
            if age > 10.0:
                self._set_badge("amber", "Stale", f"Last read {age:.0f}s ago")
        except Exception:
            pass
        except Exception:
            pass


class ScopeCard(QtWidgets.QGroupBox):
    """Vertical card for Keysight MSOX4024A oscilloscope with fixed non-expanding thumbnail."""
    snapshot = QtCore.pyqtSignal(str, object)
    message = QtCore.pyqtSignal(str)

    def __init__(self, hub: InstrumentHub, store_getter: Callable[[], WorkbookStore],
                 source_getter: Callable[[], str] | None = None):
        super().__init__("MSOX4024A Scope")
        self.hub = hub
        self.store_getter = store_getter
        self.source_getter = source_getter or (lambda: "Hardware")
        self.last_snapshot: InstrumentSnapshot | None = None
        self.last_png_path: Path | None = None
        self.busy = False

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(6)

        # Top Header: Photo + Status Badge
        top_row = QtWidgets.QHBoxLayout()
        photo = QtWidgets.QLabel()
        photo.setFixedSize(130, 75)
        photo.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        photo.setPixmap(device_thumbnail("scope", 128, 73))
        photo.setStyleSheet(f"background:white; border:1px solid {BORDER}; border-radius:4px;")
        top_row.addWidget(photo)

        self.status_badge = QtWidgets.QLabel("Not Checked")
        self.status_badge.setObjectName("badge_gray")
        self.status_badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.status_badge.setToolTip("Scope not checked yet")
        self._set_badge("gray", "Not Checked", "Scope not checked yet")
        top_row.addWidget(self.status_badge, 1)
        layout.addLayout(top_row)

        # Action button
        self.capture_btn = QtWidgets.QPushButton("Capture screen + data")
        self.capture_btn.setObjectName("primary_action")
        self.capture_btn.setToolTip("Freezes scope with :STOP, saves PNG and CSV from current displayed state, and restores live :RUN mode.")
        self.capture_btn.clicked.connect(self.capture_now)
        layout.addWidget(self.capture_btn)

        # Fixed 16:9 Waveform Preview Thumbnail (cannot expand column width)
        self.thumb_label = QtWidgets.QLabel("No Capture")
        self.thumb_label.setFixedSize(130, 75)
        self.thumb_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.thumb_label.setStyleSheet("background: #EAECF0; border: 1px solid #D7DEE8; border-radius: 4px; color: #5F6B7A; font-size: 11px;")
        self.thumb_label.setPixmap(device_thumbnail("scope", 128, 73))
        layout.addWidget(self.thumb_label, 0, QtCore.Qt.AlignmentFlag.AlignCenter)

        # Details
        self.chan_label = QtWidgets.QLabel("Displayed: —")
        self.chan_label.setStyleSheet(f"font-size: 12px; font-weight: 600; color: {TEXT_MAIN};")
        self.file_label = QtWidgets.QLabel("Saved: None")
        self.file_label.setStyleSheet(f"font-size: 11px; color: {TEXT_MUTED};")
        self.file_label.setWordWrap(True)
        self.time_label = QtWidgets.QLabel("Captured: —")
        self.time_label.setStyleSheet(f"font-size: 11px; color: {TEXT_MUTED};")

        layout.addWidget(self.chan_label)
        layout.addWidget(self.file_label)
        layout.addWidget(self.time_label)

        self.open_folder_btn = QtWidgets.QPushButton("Open Captures Folder")
        self.open_folder_btn.clicked.connect(self.open_captures_folder)
        layout.addWidget(self.open_folder_btn)
        layout.addStretch()

    def _set_badge(self, state: str, text: str, detail: str = ""):
        self.status_badge.setText(text)
        self.status_badge.setToolTip(detail or text)
        style_map = {
            "green": f"color: {SUCCESS_GREEN}; background: #EDFDF5; border: 1px solid {SUCCESS_GREEN}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "gray": f"color: {TEXT_MUTED}; background: #EAECF0; border: 1px solid {BORDER}; border-radius: 4px; padding: 3px 8px; font-weight: 500; font-size: 12px;",
            "blue": f"color: {PRIMARY_BLUE}; background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "amber": f"color: {WARNING_AMBER}; background: #FFF8E6; border: 1px solid {WARNING_AMBER}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "red": f"color: {DANGER_RED}; background: #FEF3F2; border: 1px solid {DANGER_RED}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
        }
        self.status_badge.setStyleSheet(style_map.get(state, style_map["gray"]))

    def _mark_discovered(self, identity: str = ""):
        self.released = False
        self._set_badge("blue", "Discovered", f"Found over VISA · {identity}" if identity else "Found over VISA")

    def _mark_not_found(self, detail: str = ""):
        self.released = False
        self._set_badge("amber", "Not Found", detail or "MSOX4024A oscilloscope not detected during VISA scan")

    def _mark_released(self):
        self.released = True
        self._set_badge("gray", "Released", "Scope released; session idle")

    def check_scope(self):
        def query():
            return self.hub.instruments["scope"].read_snapshot()
        def on_done(snap: InstrumentSnapshot):
            self.released = False
            self.last_snapshot = snap
            chans = snap.values.get("displayed_channels", "None")
            self.chan_label.setText(f"Displayed: CH {chans}")
            self._set_badge("green", "Active", f"Scope Active · CH {chans}")
            self.snapshot.emit("scope", snap)
        def on_fail(err: str):
            self.released = False
            self._set_badge("red", "Offline", err)
            self.message.emit(f"Scope Check: {err}")

        task = FunctionTask(query)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        QtCore.QThreadPool.globalInstance().start(task)

    def capture_now(self):
        if self.busy:
            return
        self.busy = True
        self.capture_btn.setEnabled(False)
        self.capture_btn.setText("Capturing...")

        store = self.store_getter()
        cap_dir = capture_root_for_source(store.path, self.source_getter())
        cap_dir.mkdir(parents=True, exist_ok=True)
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        png_path = cap_dir / f"scope_manual_{stamp}.png"
        csv_path = cap_dir / f"scope_manual_{stamp}.csv"

        def do_capture():
            return self.hub.instruments["scope"].capture(png_path, csv_path)

        def on_done(result):
            self.released = False
            channels, samples = result
            self.last_png_path = png_path
            short_name = png_path.name if len(png_path.name) <= 24 else (png_path.name[:21] + "...")
            self.file_label.setText(f"Saved: {short_name}")
            self.file_label.setToolTip(str(png_path))
            self.time_label.setText(f"Captured: {datetime.now().strftime('%H:%M:%S')} ({samples} pts)")
            self._set_badge("green", "Captured", f"Captured {png_path.name}")

            pix = QtGui.QPixmap(str(png_path))
            if not pix.isNull():
                self.thumb_label.setPixmap(pix.scaled(self.thumb_label.size(), QtCore.Qt.AspectRatioMode.KeepAspectRatio, QtCore.Qt.TransformationMode.SmoothTransformation))

        def on_fail(err: str):
            self.released = False
            self._set_badge("amber", "Capture Fault", err)
            self.message.emit(f"Manual Scope Capture: {err}")

        def finish():
            self.busy = False
            self.capture_btn.setEnabled(True)
            self.capture_btn.setText("Capture screen + data")

        task = FunctionTask(do_capture)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        task.signals.finished.connect(finish)
        QtCore.QThreadPool.globalInstance().start(task)

    def open_captures_folder(self):
        store = self.store_getter()
        cap_dir = capture_root_for_source(store.path, self.source_getter())
        cap_dir.mkdir(parents=True, exist_ok=True)
        QtGui.QDesktopServices.openUrl(QtCore.QUrl.fromLocalFile(str(cap_dir.resolve())))

    def update_age(self):
        pass


class InstrumentCard(QtWidgets.QGroupBox):
    """Vertical instrument card with spacious number displays and real hardware photos."""
    snapshot = QtCore.pyqtSignal(str, object)
    message = QtCore.pyqtSignal(str)

    def __init__(self, key: str, title: str, fields: list[tuple[str, str]], get_instrument: Callable[[], Any], read_kwargs: Callable[[], dict[str, Any]] | None = None):
        super().__init__(title)
        self.key = key
        self.get_instrument = get_instrument
        self.read_kwargs = read_kwargs or (lambda: {})
        self.pool = QtCore.QThreadPool.globalInstance()
        self.last_snapshot: InstrumentSnapshot | None = None
        self.busy = False
        self.value_labels: dict[str, tuple[QtWidgets.QLabel, str]] = {}

        layout = QtWidgets.QVBoxLayout(self)
        layout.setContentsMargins(10, 10, 10, 10)
        layout.setSpacing(8)

        # Top Header: Photo + Status Badge
        top_row = QtWidgets.QHBoxLayout()
        photo = QtWidgets.QLabel()
        photo.setFixedSize(130, 75)
        photo.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        photo.setPixmap(device_thumbnail(key, 128, 73))
        photo.setStyleSheet(f"background:white; border:1px solid {BORDER}; border-radius:4px;")
        top_row.addWidget(photo)

        self.status_badge = QtWidgets.QLabel("Not Checked")
        self.status_badge.setObjectName("badge_gray")
        self.status_badge.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.status_badge.setToolTip("Instrument not checked yet")
        self._set_badge("gray", "Not Checked", "Instrument not checked yet")
        top_row.addWidget(self.status_badge, 1)
        layout.addLayout(top_row)

        # Primary live measurements readout in spacious vertical rows
        grid = QtWidgets.QVBoxLayout()
        grid.setContentsMargins(2, 6, 2, 2)
        grid.setSpacing(8)
        for name, unit in fields:
            box = QtWidgets.QFrame()
            box.setStyleSheet(f"background: #F8FAFC; border: 1px solid {BORDER}; border-radius: 4px; padding: 4px 8px;")
            b_lay = QtWidgets.QVBoxLayout(box)
            b_lay.setContentsMargins(4, 2, 4, 2)
            b_lay.setSpacing(2)

            name_lbl = QtWidgets.QLabel(f"{name} ({unit})")
            name_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-weight: 700; font-size: 11px;")
            b_lay.addWidget(name_lbl)

            val_lbl = QtWidgets.QLabel("—")
            val_lbl.setObjectName("measurement")
            val_lbl.setStyleSheet(f"color: {PRIMARY_BLUE}; font-family: Consolas; font-weight: 800; font-size: 22px;")
            b_lay.addWidget(val_lbl)

            grid.addWidget(box)
            self.value_labels[name] = (val_lbl, unit)

        layout.addLayout(grid)
        layout.addStretch()

    def _set_badge(self, state: str, text: str, detail: str = ""):
        self.status_badge.setText(text)
        self.status_badge.setToolTip(detail or text)
        style_map = {
            "green": f"color: {SUCCESS_GREEN}; background: #EDFDF5; border: 1px solid {SUCCESS_GREEN}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "gray": f"color: {TEXT_MUTED}; background: #EAECF0; border: 1px solid {BORDER}; border-radius: 4px; padding: 3px 8px; font-weight: 500; font-size: 12px;",
            "blue": f"color: {PRIMARY_BLUE}; background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "amber": f"color: {WARNING_AMBER}; background: #FFF8E6; border: 1px solid {WARNING_AMBER}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
            "red": f"color: {DANGER_RED}; background: #FEF3F2; border: 1px solid {DANGER_RED}; border-radius: 4px; padding: 3px 8px; font-weight: 700; font-size: 12px;",
        }
        self.status_badge.setStyleSheet(style_map.get(state, style_map["gray"]))

    def _mark_discovered(self, identity: str = ""):
        self.released = False
        self._set_badge("blue", "Discovered", f"Found over VISA · {identity}" if identity else "Found over VISA")

    def _mark_not_found(self, detail: str = ""):
        self.released = False
        self._set_badge("amber", "Not Found", detail or f"{self.title()} not detected during VISA scan")

    def _run(self, function: Callable[[], Any], success: Callable[[Any], None] | None = None):
        if self.busy:
            return
        self.busy = True
        task = FunctionTask(function)
        if success:
            task.signals.success.connect(success)
        task.signals.failure.connect(self._failed)
        task.signals.finished.connect(lambda: setattr(self, "busy", False))
        self.pool.start(task)

    def read_once(self):
        self._run(lambda: self.get_instrument().read_snapshot(**self.read_kwargs()), self._received)

    def _received(self, snap: InstrumentSnapshot):
        self.released = False
        self.last_snapshot = snap
        self._render_values()
        if snap.valid:
            self._set_badge("green", "Connected", f"Connected · {datetime.now().strftime('%H:%M:%S')}")
        else:
            status_text = getattr(snap, "status", "") or "Connected · Read Error"
            self._set_badge("amber", status_text, snap.warning or "Measurement query failed")
            self.message.emit(f"{self.title()}: {snap.warning}")
        self.snapshot.emit(self.key, snap)

    def _failed(self, error: str):
        self.released = False
        self._set_badge("red", "Offline", error)
        self.message.emit(f"{self.title()}: {error}")

    def _mark_released(self):
        self.released = True
        self._set_badge("gray", "Released", "Instrument released; session idle")
        self._render_values()

    def _render_values(self):
        if not self.last_snapshot:
            return
        values = self.last_snapshot.values
        aliases = {
            "Vin": "vin", "Iin": "iin", "Vout": "vout", "Pin": "pin", "Iout": "current", "Load V": "voltage",
            "Load P": "power",
        }
        is_rel = getattr(self, "released", False)
        for label_name, (label, _) in self.value_labels.items():
            value = values.get(aliases.get(label_name, label_name))
            if value is None:
                label.setText("—")
                label.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas; font-weight: 700; font-size: 22px;")
                label.setToolTip("")
            elif isinstance(value, (float, int)):
                if abs(float(value)) < 0.0005 and "I" in label_name:
                    disp_str = "0.000"
                else:
                    disp_str = f"{value:.4g}"
                label.setText(disp_str)
                if is_rel:
                    label.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas; font-weight: 700; font-size: 22px;")
                    label.setToolTip("Last reading before release")
                else:
                    label.setStyleSheet(f"color: {PRIMARY_BLUE}; font-family: Consolas; font-weight: 800; font-size: 22px;")
                    label.setToolTip("")
            else:
                label.setText(str(value))
                if is_rel:
                    label.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas; font-weight: 700; font-size: 22px;")
                    label.setToolTip("Last reading before release")
                else:
                    label.setStyleSheet(f"color: {PRIMARY_BLUE}; font-family: Consolas; font-weight: 800; font-size: 22px;")
                    label.setToolTip("")

    def update_age(self):
        if getattr(self, "released", False):
            return
        if not self.last_snapshot or not self.last_snapshot.valid:
            return
        try:
            stamp = datetime.fromisoformat(self.last_snapshot.timestamp).timestamp()
            age = max(0.0, time.time() - stamp)
            if age > 10.0:
                self._set_badge("amber", "Stale", f"Last valid read {age:.0f}s ago")
        except Exception:
            pass


class SweepWorker(QtCore.QThread):
    progress = QtCore.pyqtSignal(int, int, float, float, str)
    ramp_progress = QtCore.pyqtSignal(float)
    state_changed = QtCore.pyqtSignal(str, str)
    measurement = QtCore.pyqtSignal(dict)
    completed = QtCore.pyqtSignal(str, str)
    warning = QtCore.pyqtSignal(str)
    vin_safety_tripped = QtCore.pyqtSignal(float, float)

    def __init__(self, hub: InstrumentHub, store: WorkbookStore, settings: dict[str, Any]):
        super().__init__()
        self.hub = hub
        self.store = store
        self.settings = settings
        self.abort_event = threading.Event()
        self.stop_event = threading.Event()

    def abort(self):
        """Immediate emergency abort without gradual ramp."""
        self.abort_event.set()
        try:
            self.hub.instruments["load"].safe_off()
        except Exception:
            pass

    def cancel(self):
        """Backward-compatible alias for abort."""
        self.abort()

    def stop_and_return_to_zero(self):
        """Controlled stop: stop taking new test points, then gracefully ramp to 0 A."""
        self.stop_event.set()

    def _wait(self, seconds: float, load_active: bool = False):
        deadline = time.monotonic() + max(0.0, seconds)
        while time.monotonic() < deadline:
            if self.abort_event.is_set():
                raise InterruptedError("Operator emergency abort")
            if self.stop_event.is_set():
                raise InterruptedError("Operator stop and return to zero")
            time.sleep(min(0.05, max(0.0, deadline - time.monotonic())))

    def _ramp_down_to_zero(self, starting_amps: float):
        if starting_amps <= 0:
            try:
                self.hub.instruments["load"].safe_off()
            except Exception:
                pass
            return
        self.state_changed.emit("RETURNING TO ZERO", f"Ramping {starting_amps:g} A → 0 A")
        step = max(0.1, float(self.settings.get("return_to_zero_step", 5.0)))
        current = starting_amps
        load = self.hub.instruments["load"]
        while current > 0:
            if self.abort_event.is_set():
                try:
                    load.safe_off()
                except Exception:
                    pass
                return
            current = max(0.0, current - step)
            try:
                load.set_current(current)
            except Exception:
                pass
            self.ramp_progress.emit(current)
            time.sleep(0.15)
        load.safe_off()

    def _measure_average(self, count: int, window: float) -> tuple[InstrumentSnapshot, InstrumentSnapshot, InstrumentSnapshot | None]:
        channels = [channel.channel for channel in self.settings["supply_channels"] if channel.displayed]
        if count <= 1:
            if self.abort_event.is_set():
                raise InterruptedError("Operator emergency abort")
            if self.stop_event.is_set():
                raise InterruptedError("Operator stop and return to zero")
            pa_snap = self.hub.instruments["pa"].read_snapshot()
            load_snap = self.hub.instruments["load"].read_snapshot(include_voltage=False)
            psu_snap = None
            try:
                psu_snap = self.hub.instruments["psu"].read_snapshot(channels=channels)
            except Exception as exc:
                if self.settings["psu_required"]:
                    raise RequiredInstrumentError(f"Run-critical PSU measurement failed: {exc}") from exc
                self.warning.emit(f"PSU snapshot unavailable: {exc}")
            return pa_snap, load_snap, psu_snap

        pa_values: list[dict[str, Any]] = []
        load_values: list[dict[str, Any]] = []
        psu_values: list[dict[str, Any]] = []
        interval = window / max(count - 1, 1)
        for index in range(count):
            if self.abort_event.is_set():
                raise InterruptedError("Operator emergency abort")
            if self.stop_event.is_set():
                raise InterruptedError("Operator stop and return to zero")
            pa_values.append(self.hub.instruments["pa"].read_snapshot().values)
            load_values.append(self.hub.instruments["load"].read_snapshot(include_voltage=False).values)
            try:
                psu_values.append(self.hub.instruments["psu"].read_snapshot(channels=channels).values)
            except Exception as exc:
                if self.settings["psu_required"]:
                    raise RequiredInstrumentError(f"Run-critical PSU measurement failed: {exc}") from exc
                self.warning.emit(f"PSU snapshot unavailable: {exc}")
            if index < count - 1:
                self._wait(interval, load_active=True)

        def average(rows: list[dict[str, Any]]) -> dict[str, Any]:
            result: dict[str, Any] = {}
            for key in {key for row in rows for key in row}:
                numeric = [float(row[key]) for row in rows if isinstance(row.get(key), (int, float)) and math.isfinite(float(row[key]))]
                result[key] = sum(numeric) / len(numeric) if numeric else rows[-1].get(key)
            return result

        return InstrumentSnapshot("pa", average(pa_values)), InstrumentSnapshot("load", average(load_values)), (InstrumentSnapshot("psu", average(psu_values)) if psu_values else None)

    def run(self):
        run_id = self.settings["run_id"]
        warnings: list[str] = []
        status = "Valid"
        any_invalid = False
        load = self.hub.instruments["load"]
        last_commanded_amps = 0.0
        try:
            # Persist the run envelope before any instrument connection or load command.
            # The finalizer below updates this same row on success, stop, or abort.
            if not self.store.create_run(self.settings["run_record"]):
                raise RuntimeError(f"RunID already exists: {run_id}")
            for required in ("pa", "load"):
                self.hub.instruments[required].connect(persistent=True)
            if self.settings["psu_required"]:
                self.hub.instruments["psu"].connect(persistent=True)
            self.settings["run_record"]["InstrumentIdentities"] = {
                key: getattr(inst, "identity", "") for key, inst in self.hub.instruments.items()
            }
            points = self.settings["points"]
            self.state_changed.emit("RUNNING", "Test sequence in progress")

            for index, amps in enumerate(points):
                if self.abort_event.is_set():
                    raise InterruptedError("Operator emergency abort")
                if self.stop_event.is_set():
                    break

                next_amps = points[index + 1] if index + 1 < len(points) else 0.0
                self.progress.emit(index + 1, len(points), amps, next_amps, f"Applying {amps:g} A")
                if amps < 0 or amps > self.settings["working_cap"]:
                    raise RequiredInstrumentError(f"{amps:g} A exceeds active working cap")

                last_commanded_amps = amps
                dwell_total = float(self.settings["dwell"] if self.settings["mode"] == "Pulse" else self.settings["settle"])
                measure_last = min(dwell_total, max(0.05, float(self.settings.get("sample_window", 1.5))))
                pre_settle = max(0.0, dwell_total - measure_last)

                if self.settings["mode"] == "Pulse":
                    load.set_input(False)
                    load.set_current(amps)
                    load.set_input(True)
                else:
                    load.set_current(amps)
                    load.set_input(True)

                t_start = time.monotonic()
                t_deadline = t_start + dwell_total

                # Settle converter before acquisition begins
                if pre_settle > 0:
                    self._wait(pre_settle, load_active=True)

                # Execute measurement near the end of the dwell
                pa_snap, load_snap, psu_snap = self._measure_average(self.settings["sample_count"], measure_last)

                # Evaluate Vin safety shutdown before processing/saving the point
                pa_vin = pa_snap.values.get("vin") if pa_snap and pa_snap.valid else None
                if pa_vin is not None and isinstance(pa_vin, (int, float)) and math.isfinite(float(pa_vin)):
                    is_safe, vin_fault_desc = check_vin_safety(
                        float(pa_vin),
                        float(self.settings.get("vin_target", 0.0)),
                        self.settings.get("vin_safety_enabled", True),
                    )
                    if not is_safe:
                        load.safe_off()
                        last_commanded_amps = 0.0
                        self.abort_event.set()
                        self.vin_safety_tripped.emit(float(self.settings.get("vin_target", 0.0)), float(pa_vin))
                        status = "Aborted"
                        warnings.append(f"Vin safety shutdown; {vin_fault_desc.replace(chr(10), '; ')}")
                        break

                derived, point_warnings = calculate_measurement(pa_snap, load_snap, psu_snap, self.settings["supply_channels"])

                pid = point_id(run_id, index)
                capture_status, capture_error, png_value, csv_value = "Skipped", "", "", ""
                if any(abs(amps - point) <= 0.05 for point in self.settings["capture_points"]):
                    capture_dir = capture_root_for_source(self.store.path, self.settings["data_source"])
                    capture_dir.mkdir(parents=True, exist_ok=True)
                    png_path = capture_dir / f"{pid}.png"
                    csv_path = capture_dir / f"{pid}.csv"
                    try:
                        self.hub.instruments["scope"].capture(png_path, csv_path)
                        capture_status = "Captured"
                        png_value, csv_value = str(png_path), str(csv_path)
                    except Exception as exc:
                        capture_status, capture_error = "Failed", str(exc)
                        png_value = str(png_path) if png_path.exists() else ""
                        csv_value = str(csv_path) if csv_path.exists() else ""
                        point_warnings.append(f"Scope capture failed: {exc}")

                # Bounded dwell: if acquisition finished before hard deadline, wait remainder
                time_left = t_deadline - time.monotonic()
                if time_left > 0:
                    self._wait(time_left, load_active=True)

                if self.settings["mode"] == "Pulse":
                    load.safe_off()
                    last_commanded_amps = 0.0

                record = {
                    "PointID": pid, "RunID": run_id, "Timestamp": utc_now(), "Status": "Valid" if derived else "Invalid",
                    "DataSource": self.settings["data_source"], "Mode": self.settings["mode"], "VinTarget_V": self.settings["vin_target"],
                    "ModulationLabel": self.settings.get("modulation", ""), "Frequency_Hz": self.settings["frequency"],
                    "RequestedIout_A": amps, **derived, "Quality": "Valid" if derived and not point_warnings else "Warning",
                    "Warning": "; ".join(point_warnings), "ScopeCaptureStatus": capture_status,
                    "ScopeCaptureError": capture_error, "ScopePNG": png_value, "ScopeCSV": csv_value,
                }
                any_invalid = any_invalid or record["Status"] == "Invalid"
                self.store.append_measurement(record, self.settings["duplicate_action"])
                self.measurement.emit(record)

                if self.settings["mode"] == "Pulse":
                    if index < len(points) - 1 and not self.stop_event.is_set() and not self.abort_event.is_set():
                        self._wait(self.settings["cooldown"])

            # Graceful ramp-down to zero for Continuous mode normal completion or early stop
            if not self.abort_event.is_set() and last_commanded_amps > 0:
                self._ramp_down_to_zero(last_commanded_amps)
            else:
                try:
                    load.safe_off()
                except Exception:
                    pass

            if self.stop_event.is_set():
                status = "Stopped"
                warnings.append("Stopped by operator; returned to 0 A")
            elif status == "Aborted":
                pass
            elif any_invalid:
                status = "Invalid"
            else:
                status = "Valid"
        except InterruptedError as exc:
            if self.abort_event.is_set():
                status = "Aborted"
                warnings.append("Emergency abort by operator")
                try:
                    load.safe_off()
                except Exception:
                    pass
            elif self.stop_event.is_set():
                if last_commanded_amps > 0:
                    try:
                        self._ramp_down_to_zero(last_commanded_amps)
                    except Exception:
                        try:
                            load.safe_off()
                        except Exception:
                            pass
                else:
                    try:
                        load.safe_off()
                    except Exception:
                        pass
                status = "Stopped"
                warnings.append("Stopped by operator; returned to 0 A")
            else:
                status = "Aborted"
                warnings.append(str(exc))
                try:
                    load.safe_off()
                except Exception:
                    pass
        except Exception as exc:
            status = "Aborted"
            warnings.append(str(exc))
            try:
                load.safe_off()
            except Exception:
                pass
            self.warning.emit(str(exc))
        finally:
            try:
                load.safe_off()
            except Exception:
                pass
            try:
                self.store.finish_run(run_id, status, "; ".join(warnings))
            except Exception as exc:
                warnings.append(f"Workbook save warning: {exc}")
            self.completed.emit(status, "; ".join(warnings))



class MainWindow(QtWidgets.QMainWindow):
    def __init__(self):
        super().__init__()
        self._native_chrome_applied = False
        self.config = load_config()
        self.hub = InstrumentHub(False, self.config)
        workbook_paths = self.config["workbooks"]
        self.hardware_store = WorkbookStore(Path(workbook_paths["hardware"]))
        self.simulation_store = WorkbookStore(Path(workbook_paths["simulation"]))
        self.store = self.hardware_store
        self.worker: SweepWorker | None = None
        self.demo_timer: QtCore.QTimer | None = None
        self.demo_index = 0
        self.demo_points: list[float] = []
        self.last_values: dict[str, InstrumentSnapshot] = {}
        self.plot_rows: list[dict[str, Any]] = []
        self._manual_target_current = 0.0
        self.cap_val = float(self.config.get("working_current_cap_a", 60.0))
        self.bench_operation_busy = False

        # Unified Manual / Step Current auto-recording state
        self._manual_step_action = "idle"
        self._manual_point_token = 0
        self._manual_countdown_timer = QtCore.QTimer(self)
        self._manual_countdown_timer.setInterval(100)
        self._manual_countdown_timer.timeout.connect(self._manual_countdown_tick)
        self._manual_remaining_ms = 0
        self._manual_save_done = False
        self._manual_capture_done = False
        self._manual_target_amps = 0.0
        self._manual_run_id = ""
        self._manual_point_id = ""
        self._manual_last_record: dict[str, Any] | None = None
        self._manual_active_task = False
        self.point_action_busy = False
        self._manual_mode_name = "Set Current"
        self._manual_point_index = -1
        self._manual_run_created = False
        self._manual_run_label = ""
        self._manual_base_campaign = ""
        self._manual_recorded_currents: list[float] = []
        self._manual_any_invalid = False
        self._manual_store: WorkbookStore | None = None
        self._selected_mode_id: int | None = None

        self.setWindowTitle("KICKSTART BENCH")

        self.resize(1280, 768)
        self.setMinimumSize(1150, 720)


        self.wheel_filter = NoWheelFilter(self)
        app = QtWidgets.QApplication.instance()
        if app:
            app.installEventFilter(self.wheel_filter)

        self._build_ui()
        self._apply_style()
        self._load_history()
        self.update_enabled_states()

        self.age_timer = QtCore.QTimer(self)
        self.age_timer.timeout.connect(lambda: [card.update_age() for card in self.cards.values()])
        self.age_timer.start(1000)

    @staticmethod
    def _windows_colorref(hex_color: str) -> int:
        """Convert #RRGGBB to the Windows COLORREF byte order 0x00BBGGRR."""
        value = str(hex_color).lstrip("#")
        if len(value) != 6:
            raise ValueError("Expected an RGB hex color")
        red = int(value[0:2], 16)
        green = int(value[2:4], 16)
        blue = int(value[4:6], 16)
        return red | (green << 8) | (blue << 16)

    def _apply_native_windows_chrome(self):
        if getattr(self, "_native_chrome_applied", False) or sys.platform != "win32":
            return
        if os.environ.get("QT_QPA_PLATFORM") == "offscreen":
            return
        try:
            handle = self.windowHandle()
            if handle is None:
                return
            hwnd = int(handle.winId())
            if not hwnd:
                return
            dwmapi = ctypes.WinDLL("dwmapi")
            set_attribute = dwmapi.DwmSetWindowAttribute
            set_attribute.argtypes = [wintypes.HWND, wintypes.DWORD, ctypes.c_void_p, wintypes.DWORD]
            set_attribute.restype = wintypes.LONG

            DWMWA_BORDER_COLOR = 34
            DWMWA_CAPTION_COLOR = 35
            DWMWA_TEXT_COLOR = 36

            def apply(attribute, color):
                value = ctypes.c_uint32(self._windows_colorref(color))
                set_attribute(hwnd, attribute, ctypes.byref(value), ctypes.sizeof(value))

            apply(DWMWA_BORDER_COLOR, BERKELEY_BLUE)
            apply(DWMWA_CAPTION_COLOR, BERKELEY_BLUE)
            apply(DWMWA_TEXT_COLOR, "#FFFFFF")
            self._native_chrome_applied = True
        except Exception:
            pass

    def showEvent(self, event: QtGui.QShowEvent):
        super().showEvent(event)
        self._apply_native_windows_chrome()

    @property
    def _step_point_token(self):
        return self._manual_point_token

    @_step_point_token.setter
    def _step_point_token(self, v):
        self._manual_point_token = v

    @property
    def _step_countdown_timer(self):
        return self._manual_countdown_timer

    @property
    def _step_remaining_ms(self):
        return self._manual_remaining_ms

    @_step_remaining_ms.setter
    def _step_remaining_ms(self, v):
        self._manual_remaining_ms = v

    @property
    def _step_save_done(self):
        return self._manual_save_done

    @_step_save_done.setter
    def _step_save_done(self, v):
        self._manual_save_done = v

    @property
    def _step_capture_done(self):
        return self._manual_capture_done

    @_step_capture_done.setter
    def _step_capture_done(self, v):
        self._manual_capture_done = v

    @property
    def _step_target_amps(self):
        return self._manual_target_amps

    @_step_target_amps.setter
    def _step_target_amps(self, v):
        self._manual_target_amps = v

    @property
    def _step_run_id(self):
        return self._manual_run_id

    @_step_run_id.setter
    def _step_run_id(self, v):
        self._manual_run_id = v

    @property
    def _step_point_id(self):
        return self._manual_point_id

    @_step_point_id.setter
    def _step_point_id(self, v):
        self._manual_point_id = v

    @property
    def _step_last_record(self):
        return self._manual_last_record

    @_step_last_record.setter
    def _step_last_record(self, v):
        self._manual_last_record = v

    @property
    def _step_active_task(self):
        return self._manual_active_task

    @_step_active_task.setter
    def _step_active_task(self, v):
        self._manual_active_task = v

    def _apply_style(self):

        self.setStyleSheet(f"""
            QMainWindow, QDialog, QMessageBox {{
                background-color: {PAGE_BG};
                color: {TEXT_MAIN};
            }}
            QWidget {{ font-family: 'Segoe UI', Arial, sans-serif; font-size: 13px; color: {TEXT_MAIN}; }}
            
            /* Tab Navigation */
            QTabWidget::pane {{ border: 1px solid {BORDER}; background: {CARD_BG}; border-radius: 6px; top: -1px; }}
            QTabBar::tab {{
                background: #E5E7EB; color: {TEXT_MUTED}; font-weight: 700; font-size: 13px;
                padding: 8px 20px; border-top-left-radius: 6px; border-top-right-radius: 6px;
                margin-right: 4px; border: 1px solid {BORDER}; border-bottom: 0;
            }}
            QTabBar::tab:selected {{ background: {CARD_BG}; color: {BERKELEY_BLUE}; border-bottom: 3px solid {BERKELEY_BLUE}; }}
            QTabBar::tab:hover:!selected {{ background: #D0D5DD; }}
            
            /* Card Containers */
            QGroupBox {{
                background: {CARD_BG}; border: 1px solid {BORDER}; border-radius: 6px;
                margin-top: 14px; padding-top: 10px; font-weight: 700; font-size: 13px;
            }}
            QGroupBox::title {{
                color: {BERKELEY_BLUE}; subcontrol-origin: margin; left: 12px; padding: 0 4px; background: {PAGE_BG};
            }}
            
            /* Buttons */
            QPushButton {{
                padding: 6px 14px; border: 1px solid {BORDER}; border-radius: 5px;
                background: {CARD_BG}; color: {TEXT_MAIN}; font-weight: 600; font-size: 13px;
            }}
            QPushButton:hover {{ border-color: {BERKELEY_BLUE}; background: #EEF4FF; color: {BERKELEY_BLUE}; }}
            QPushButton:pressed {{ background: #DBEAFE; }}
            QPushButton:disabled {{ background: #F3F4F6; color: #9CA3AF; border-color: #E5E7EB; }}
            QPushButton:checked {{ background: {BERKELEY_BLUE}; color: white; border-color: {BERKELEY_BLUE}; }}
            
            QPushButton#primary_action, QPushButton#check_bench_btn {{
                background-color: {BERKELEY_BLUE}; color: #FFFFFF; border: 1px solid #001B57;
                font-weight: 700; font-size: 13px; padding: 8px 18px; border-radius: 6px;
            }}
            QPushButton#primary_action:hover, QPushButton#check_bench_btn:hover {{
                background-color: #0B3C91;
            }}
            QPushButton#primary_action:pressed, QPushButton#check_bench_btn:pressed {{
                background-color: #001B57;
            }}
            QPushButton#primary_action:disabled, QPushButton#check_bench_btn:disabled {{
                background-color: #93C5FD; color: #EFF6FF; border-color: #BFDBFE;
            }}
            
            QPushButton#emergency_stop_btn {{
                background: {CARD_BG}; color: {DANGER_RED}; border: 1px solid #DC2626;
                font-weight: 900; font-size: 14px; padding: 6px 18px; border-radius: 6px; min-width: 95px;
            }}
            QPushButton#emergency_stop_btn:hover {{ background: #FEF2F2; }}
            
            /* LineEdit & ComboBox */
            QLineEdit, QComboBox {{
                background: {CARD_BG}; border: 1px solid {BORDER}; border-radius: 5px;
                padding: 5px 8px; color: {TEXT_MAIN}; font-size: 13px;
            }}
            QLineEdit:focus, QComboBox:focus {{
                border: 2px solid {BERKELEY_BLUE}; background: #EEF4FF;
            }}
            
            /* SpinBoxes: Clean manual numerical entry (no mini buttons) */
            QSpinBox, QDoubleSpinBox {{
                background: {CARD_BG}; border: 1px solid {BORDER}; border-radius: 5px;
                padding: 5px 8px; color: {TEXT_MAIN}; font-size: 13px;
            }}
            QSpinBox:focus, QDoubleSpinBox:focus {{
                border: 2px solid {BERKELEY_BLUE}; background: #EEF4FF;
            }}
            QSpinBox::up-button, QDoubleSpinBox::up-button,
            QSpinBox::down-button, QDoubleSpinBox::down-button {{
                width: 0px;
                height: 0px;
                border: none;
            }}
            
            /* Live Measurements */
            QLabel#measurement {{
                font-family: Consolas, 'Courier New', monospace; font-size: 20px;
                font-weight: 700; color: {BERKELEY_BLUE};
            }}
            
            /* Badges & Banners */
            QLabel#banner_sim {{
                background: {CALIFORNIA_GOLD}; color: #111827; font-weight: 800;
                font-size: 12px; padding: 4px 10px; border-radius: 4px;
            }}
            QStatusBar {{ background: #F1F5F9; color: {TEXT_MUTED}; font-size: 12px; font-weight: 600; border-top: 1px solid {BORDER}; }}
            QProgressBar {{
                background: #EAECF0; border: 1px solid {BORDER}; border-radius: 4px;
                text-align: center; font-weight: 700; font-size: 12px; color: {TEXT_MAIN};
            }}
            QProgressBar::chunk {{ background: {BERKELEY_BLUE}; border-radius: 3px; }}
            
            /* Tables */
            QTableWidget {{
                background: {CARD_BG}; alternate-background-color: #F8FAFC;
                border: 1px solid {BORDER}; gridline-color: #EAECF0;
                selection-background-color: #EEF4FF; selection-color: {BERKELEY_BLUE}; font-size: 13px;
            }}
            QHeaderView::section {{
                background: #F1F5F9; color: {TEXT_MAIN}; font-weight: 700; padding: 6px;
                border: 0; border-right: 1px solid {BORDER}; border-bottom: 1px solid {BORDER}; font-size: 12px;
            }}
            QToolTip {{
                color: #111827;
                background-color: #FFF8C5;
                border: 1px solid #6B7280;
                border-radius: 3px;
                padding: 4px 6px;
            }}
        """)

    def _build_ui(self):
        # Pre-initialize developer/simulation & setup controls
        self.simulation = QtWidgets.QCheckBox("Demo Mode")
        self.simulation.setObjectName("demo_mode_toggle")
        self.simulation.setToolTip("Enable Demo Mode to run simulated tests with synthetic data (no hardware commands).")
        self.simulation.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {TEXT_MUTED}; padding: 2px 6px;")
        self.simulation.toggled.connect(self._switch_mode)

        self.fpga_check = QtWidgets.QCheckBox("Attach FPGA snapshot")
        self.fpga_check.setChecked(True)

        self.notes = QtWidgets.QLineEdit()
        self.notes.setPlaceholderText("Operator notes")

        central = QtWidgets.QWidget()
        outer = QtWidgets.QVBoxLayout(central)
        outer.setContentsMargins(12, 8, 12, 8)
        outer.setSpacing(6)

        # 1. Top Header Row
        outer.addLayout(self._header())

        # Gold Accent Divider Line (matching Kickstart PILAWA GUI)
        self.toolbar_accent = QtWidgets.QFrame()
        self.toolbar_accent.setFixedHeight(2)
        self.toolbar_accent.setStyleSheet("background: #FDB515; border: none;")
        outer.addWidget(self.toolbar_accent)

        # High-Contrast Orange Simulation Warning Banner (Hidden in Hardware Mode)
        self.sim_warning_banner = QtWidgets.QFrame()
        self.sim_warning_banner.setObjectName("sim_warning_banner")
        self.sim_warning_banner.setStyleSheet("""
            QFrame#sim_warning_banner {
                background: #D97706;
                border: 1px solid #B45309;
                border-radius: 6px;
                padding: 4px 12px;
            }
        """)
        swb_lay = QtWidgets.QHBoxLayout(self.sim_warning_banner)
        swb_lay.setContentsMargins(8, 4, 8, 4)
        swb_lay.setSpacing(6)

        t_banner = QtWidgets.QLabel("DEMO MODE · SYNTHETIC DATA · NO HARDWARE COMMANDS")
        t_banner.setStyleSheet("color: #FFFFFF; font-size: 12px; font-weight: 800; letter-spacing: 0.8px;")
        t_banner.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        swb_lay.addWidget(t_banner)

        outer.addWidget(self.sim_warning_banner)
        self.sim_warning_banner.setVisible(False)


        # 2. Persistent Global Run-Status Strip (visible during active sweeps)
        self.run_strip = QtWidgets.QFrame()
        self.run_strip.setStyleSheet(f"background: #EEF2F7; border: 1px solid {BORDER}; border-radius: 5px;")
        strip_layout = QtWidgets.QHBoxLayout(self.run_strip)
        strip_layout.setContentsMargins(6, 2, 6, 2)
        strip_layout.setSpacing(10)

        self.strip_label = QtWidgets.QLabel("Status: IDLE")
        self.strip_label.setStyleSheet(f"color: {TEXT_MUTED}; font-weight: 700; font-size: 13px;")
        strip_layout.addWidget(self.strip_label)


        self.strip_progress = QtWidgets.QProgressBar()
        self.strip_progress.setFixedHeight(16)
        self.strip_progress.setFixedWidth(220)
        self.strip_progress.setFormat("Idle")
        strip_layout.addWidget(self.strip_progress)

        self.strip_next = QtWidgets.QLabel("")
        self.strip_next.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 700; font-size: 13px;")
        strip_layout.addWidget(self.strip_next)

        strip_layout.addStretch()

        self.strip_stop_btn = QtWidgets.QPushButton("■ STOP & RETURN TO ZERO")
        self.strip_stop_btn.setStyleSheet(f"background: #FFFBEB; color: {WARNING_AMBER}; border: 1px solid {WARNING_AMBER}; font-weight: 800; font-size: 12px; padding: 4px 10px; border-radius: 4px;")
        self.strip_stop_btn.clicked.connect(self.stop_and_return_to_zero)
        strip_layout.addWidget(self.strip_stop_btn)

        outer.addWidget(self.run_strip)

        self.run_strip.setVisible(False)

        # 3. 4 Clean Tabs: Bench Setup, Run (with Test Setup on top), History, Help
        self.tabs = QtWidgets.QTabWidget()
        self.tabs.addTab(self._bench_tab(), "Bench Setup")
        self.tabs.addTab(self._run_tab(), "Run")
        self.tabs.addTab(self._history_tab(), "History")
        self.tabs.addTab(self._help_tab(), "Help")
        self.tabs.currentChanged.connect(lambda _: self.update_enabled_states())

        outer.addWidget(self.tabs)
        self.setCentralWidget(central)
        self._remove_all_spinbox_buttons()
        self.statusBar().showMessage("Kickstart Bench · Ready")


    def _remove_all_spinbox_buttons(self):
        targets = set(self.findChildren(QtWidgets.QAbstractSpinBox))
        if self.centralWidget():
            targets.update(self.centralWidget().findChildren(QtWidgets.QAbstractSpinBox))
        for sp in targets:
            sp.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)



    def _header(self):
        row = QtWidgets.QHBoxLayout()
        row.setSpacing(10)

        title = QtWidgets.QLabel("KICKSTART BENCH")
        title.setStyleSheet(f"font-size: 16px; font-weight: 800; color: {PRIMARY_BLUE}; letter-spacing: 0.5px;")
        row.addWidget(title)

        self.banner = QtWidgets.QLabel("")  # Internal reference maintained
        self.kpi_sim_badge = QtWidgets.QLabel("")  # Internal reference maintained

        row.addStretch()


        # KPIs in Header (direct measurements in navy, derived power metrics in amber)
        self.kpi_labels: dict[str, QtWidgets.QLabel] = {}
        direct_kpis = QtWidgets.QHBoxLayout()
        direct_kpis.setSpacing(14)
        derived_kpis = QtWidgets.QHBoxLayout()
        derived_kpis.setSpacing(12)

        def add_kpi(target: QtWidgets.QHBoxLayout, name: str, value_color: str):
            kpi_col = QtWidgets.QVBoxLayout()
            kpi_col.setContentsMargins(2, 0, 2, 0)
            kpi_col.setSpacing(1)

            tag = QtWidgets.QLabel(name)
            tag.setStyleSheet("color: #64748B; font-weight: 700; font-size: 11px; letter-spacing: 0.5px;")
            tag.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)

            val = QtWidgets.QLabel("—")
            val.setStyleSheet(f"color: {value_color}; font-family: Consolas, 'Segoe UI', monospace; font-weight: 800; font-size: 20px; letter-spacing: -0.5px;")
            val.setAlignment(QtCore.Qt.AlignmentFlag.AlignLeft | QtCore.Qt.AlignmentFlag.AlignVCenter)

            kpi_col.addWidget(tag)
            kpi_col.addWidget(val)
            self.kpi_labels[name] = val
            target.addLayout(kpi_col)

        for name in ("Vin", "Iin", "Vout", "Iout"):
            add_kpi(direct_kpis, name, BERKELEY_BLUE)
        for name in ("Pin", "Pout", "Eff"):
            add_kpi(derived_kpis, name, WARNING_AMBER)

        row.addLayout(direct_kpis)
        separator = QtWidgets.QFrame()
        separator.setFrameShape(QtWidgets.QFrame.Shape.VLine)
        separator.setStyleSheet(f"color: {BORDER};")
        separator.setFixedHeight(34)
        row.addWidget(separator)
        row.addLayout(derived_kpis)
        row.addSpacing(8)

        # Single Emergency Stop Button (Top-Right)
        self.emergency_stop_btn = QtWidgets.QPushButton("LOAD OFF")
        self.emergency_stop_btn.setObjectName("emergency_stop_btn")
        self.emergency_stop_btn.setToolTip("Cancels the active test and commands the electronic load OFF.")
        self.emergency_stop_btn.setShortcut(QtGui.QKeySequence(QtCore.Qt.Key.Key_Escape))
        self.emergency_stop_btn.clicked.connect(self.emergency_stop_action)
        row.addWidget(self.emergency_stop_btn)

        return row


    def keyPressEvent(self, event: QtGui.QKeyEvent):
        if event.key() == QtCore.Qt.Key.Key_Escape:
            self.emergency_stop_action()
            event.accept()
        else:
            super().keyPressEvent(event)

    def emergency_stop_action(self):
        """Immediate emergency abort without gradual ramp."""
        self._cancel_manual_automation("0.00 A · OFF")
        self._finalize_manual_session("Emergency LOAD OFF", status_override="Stopped")
        self._clear_live_run_view()
        if self.demo_timer and self.demo_timer.isActive():
            self.demo_timer.stop()
            self._demo_completed("Aborted")
        if self.worker and self.worker.isRunning():
            self.worker.abort()
            self._state_changed("ABORTED", "Emergency abort")
            self.statusBar().showMessage("Emergency abort requested; electronic load immediately turned OFF")
        self.manual_target_spin.setValue(0.0)
        self._manual_target_current = 0.0
        self.step_present_lbl.setText("0.00 A · OFF")
        self.step_present_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
        if hasattr(self, "manual_actual_lbl"):
            self.manual_actual_lbl.setText("0.00 A · OFF")
            self.manual_actual_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 22px; font-weight: 800;")
        if hasattr(self, "step_actual_lbl"):
            self.step_actual_lbl.setText("0.00 A · OFF")
            self.step_actual_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 22px; font-weight: 800;")
        def stop_cmd():
            load = self.hub.instruments["load"]
            load.safe_off()
        self._run_function(stop_cmd, lambda _: self.statusBar().showMessage("Electronic load OFF (0.00 A)"))



    # ----------------- TAB 1: BENCH SETUP (4-COLUMN EQUAL LAYOUT) -----------------
    def _bench_tab(self):
        scroll = QtWidgets.QScrollArea()
        scroll.setWidgetResizable(True)
        scroll.setFrameShape(QtWidgets.QFrame.Shape.NoFrame)

        widget = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(widget)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(10)

        # Pre-instantiate instrument cards
        self.load_card = LoadCard(self.hub, self.config, lambda: save_config(self.config))
        self.load_card.cap_applied.connect(self._on_safety_limit_applied)
        self.chk_load = self.load_card.chk_load
        self.chk_load.toggled.connect(lambda _: self.update_enabled_states())
        self.chk_vin_safety = self.load_card.chk_vin_safety

        self.supply_card = SupplyCard(self.hub, self.config, lambda: save_config(self.config))
        self.scope_card = ScopeCard(
            self.hub,
            lambda: self.store,
            lambda: "Simulation" if self.simulation.isChecked() else "Hardware",
        )
        self.cards = {
            "pa": InstrumentCard("pa", "PA2201A Analyzer", [("Vin", "V"), ("Iin", "A"), ("Vout", "V")], lambda: self.hub.instruments["pa"]),
            "load": self.load_card,
            "psu": self.supply_card,
        }

        # Top Bar: Check Entire Bench & Compact Readiness Status Indicator
        top_banner = QtWidgets.QFrame()
        top_banner.setObjectName("top_banner")
        top_banner.setStyleSheet(f"QFrame#top_banner {{ background: #FFFFFF; border: 1px solid {BORDER}; border-radius: 6px; }}")
        b_layout = QtWidgets.QHBoxLayout(top_banner)
        b_layout.setContentsMargins(8, 6, 8, 6)
        b_layout.setSpacing(12)

        self.check_bench_btn = QtWidgets.QPushButton("Check / Refresh Entire Bench")
        self.check_bench_btn.setObjectName("check_bench_btn")
        self.check_bench_btn.clicked.connect(self._check_entire_bench)
        b_layout.addWidget(self.check_bench_btn)

        self.readiness_status = QtWidgets.QLabel("● Not checked")
        self.readiness_status.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 800; font-size: 13px; margin-left: 4px;")
        b_layout.addWidget(self.readiness_status)

        b_layout.addStretch()
        layout.addWidget(top_banner)

        # Compact Readiness Checklist Strip (No duplicate checkbox)
        chk_frame = QtWidgets.QFrame()
        chk_frame.setStyleSheet(f"background: #F1F5F9; border-radius: 5px; padding: 5px 10px;")
        c_layout = QtWidgets.QHBoxLayout(chk_frame)
        c_layout.setContentsMargins(4, 2, 4, 2)
        c_layout.setSpacing(16)

        self.chk_inst = QtWidgets.QLabel("Instruments: Unchecked")
        self.chk_vin = QtWidgets.QLabel("Target Vin: Unchecked")
        self.chk_cap = QtWidgets.QLabel(f"Current Cap: {float(self.config['working_current_cap_a']):g} A")

        for lbl in (self.chk_inst, self.chk_vin, self.chk_cap):
            lbl.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {TEXT_MUTED};")
            c_layout.addWidget(lbl)
        c_layout.addStretch()
        layout.addWidget(chk_frame)

        # 4 Equal-Dimension Vertical Columns Side-by-Side
        cols_container = QtWidgets.QWidget()
        cols_layout = QtWidgets.QHBoxLayout(cols_container)
        cols_layout.setContentsMargins(0, 0, 0, 0)
        cols_layout.setSpacing(10)

        for card in (self.cards["pa"], self.load_card, self.supply_card, self.scope_card):
            card.setSizePolicy(QtWidgets.QSizePolicy.Policy.Expanding, QtWidgets.QSizePolicy.Policy.Preferred)
            cols_layout.addWidget(card, 1)
            card.snapshot.connect(self._snapshot_received)
            card.message.connect(self.statusBar().showMessage)

        layout.addWidget(cols_container)

        # Bottom Actions Row: [ Discover VISA Devices ] [ Release All Devices ]
        btn_row = QtWidgets.QHBoxLayout()
        btn_row.setSpacing(10)
        self.disc_btn = QtWidgets.QPushButton("Discover VISA Devices")
        self.disc_btn.setToolTip("Scan VISA resource managers for connected instruments.")
        self.disc_btn.clicked.connect(self._discover_devices)
        btn_row.addWidget(self.disc_btn)

        self.release_btn = QtWidgets.QPushButton("Release All Devices")
        self.release_btn.setToolTip("Close bench VISA sessions and return instruments to front-panel/local control.")
        self.release_btn.clicked.connect(self._release_all_devices)
        btn_row.addWidget(self.release_btn)

        btn_row.addStretch()
        layout.addLayout(btn_row)
        layout.addStretch()

        scroll.setWidget(widget)
        return scroll

    def get_card_for_instrument(self, key: str):
        if key == "scope":
            return getattr(self, "scope_card", None)
        if key in ("supply", "psu"):
            return getattr(self, "supply_card", None) or self.cards.get("supply") or self.cards.get("psu")
        if key == "load":
            return getattr(self, "load_card", None) or self.cards.get("load")
        if key == "pa":
            return self.cards.get("pa")
        return self.cards.get(key)

    def all_instrument_cards(self) -> dict[str, Any]:
        cards = {}
        for k in ("pa", "load", "psu", "scope"):
            c = self.get_card_for_instrument(k)
            if c is not None:
                cards[k] = c
        return cards

    def _set_bench_busy(self, busy: bool, op: str = ""):
        self.bench_operation_busy = busy
        if busy:
            if op == "check":
                self.check_bench_btn.setText("Checking Bench...")
            elif op == "discover":
                self.disc_btn.setText("Discovering...")
            elif op == "release":
                self.release_btn.setText("Releasing...")
        else:
            self.check_bench_btn.setText("Check / Refresh Entire Bench")
            self.disc_btn.setText("Discover VISA Devices")
            self.release_btn.setText("Release All Devices")
        self.update_enabled_states()

    def _release_all_devices(self):
        if self.bench_operation_busy:
            return
        self._set_bench_busy(True, "release")
        self.statusBar().showMessage("Closing VISA sessions...")

        def do_release():
            self.hub.release_all()

        def on_done(_):
            try:
                for key, card in self.all_instrument_cards().items():
                    if hasattr(card, "_mark_released"):
                        card._mark_released()
                    elif hasattr(card, "_set_badge"):
                        card._set_badge("gray", "Released", f"{key.upper()} released; session idle")

                self.readiness_status.setText("● Not checked")
                self.readiness_status.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                self.chk_inst.setText("Instruments: Unchecked")
                self.chk_inst.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {TEXT_MUTED};")
                self.chk_vin.setText("Target Vin: Unchecked")
                self.chk_vin.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {TEXT_MUTED};")
                self.statusBar().showMessage("All KICKSTART VISA sessions released")
            finally:
                self._set_bench_busy(False)

        def on_fail(err: str):
            try:
                self.statusBar().showMessage(f"Release error · {err}")
            finally:
                self._set_bench_busy(False)

        task = FunctionTask(do_release)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        QtCore.QThreadPool.globalInstance().start(task)

    def _check_entire_bench(self):
        if self.bench_operation_busy:
            return
        self._set_bench_busy(True, "check")
        self.statusBar().showMessage("Checking instruments...")

        for key, card in self.all_instrument_cards().items():
            if hasattr(card, "_set_badge"):
                card._set_badge("gray", "Checking...", f"Querying {key.upper()}...")

        def query_all():
            results = {}
            for key in ("pa", "load", "psu", "scope"):
                try:
                    snap = self.hub.instruments[key].read_snapshot(**(self._psu_read_kwargs() if key == "psu" else {}))
                    results[key] = snap
                except Exception as exc:
                    results[key] = InstrumentSnapshot(key, {}, valid=False, status="Offline", warning=str(exc))
            return results

        def on_done(results: dict[str, InstrumentSnapshot]):
            try:
                for key in ("pa", "load", "psu", "scope"):
                    card = self.get_card_for_instrument(key)
                    if not card or key not in results:
                        continue
                    snap = results[key]
                    card.released = False
                    card.last_snapshot = snap
                    if hasattr(card, "_render_values"):
                        card._render_values()

                    if key == "scope":
                        if snap.valid:
                            chans = snap.values.get("displayed_channels", "")
                            if hasattr(card, "chan_label"):
                                card.chan_label.setText(f"Displayed: CH {chans}")
                            card._set_badge("green", "Active", f"Active · CH {chans}")
                        else:
                            card._set_badge("red", "Offline", snap.warning or "Scope communication failure")
                    else:
                        if snap.valid:
                            card._set_badge("green", "Connected", f"Connected · {datetime.now().strftime('%H:%M:%S')}")
                            self._snapshot_received(key, snap)
                        else:
                            st = getattr(snap, "status", "")
                            if "Connected" in st or "Read Error" in st or "Invalid Data" in st:
                                card._set_badge("amber", st or "Connected · Read Error", snap.warning or "Measurement query failed")
                            else:
                                card._set_badge("red", "Offline", snap.warning or "Communication failure")

                pa_snap = results.get("pa", InstrumentSnapshot("pa", {}, valid=False, status="Offline"))
                load_snap = results.get("load", InstrumentSnapshot("load", {}, valid=False, status="Offline"))
                pa_ok = pa_snap.valid
                load_ok = load_snap.valid
                pa_connected = pa_ok or ("Connected" in getattr(pa_snap, "status", ""))
                load_connected = load_ok or ("Connected" in getattr(load_snap, "status", ""))
                pa_vin = float(pa_snap.values.get("vin", 0.0)) if (pa_ok and isinstance(pa_snap.values.get("vin"), (int, float))) else 0.0
                vin_ok = abs(pa_vin - self.vin_target.value()) <= 1.0 if pa_ok else False
                load_verified = self.chk_load.isChecked()

                if pa_ok and load_ok:
                    self.chk_inst.setText("Instruments: PA+Load OK")
                    self.chk_inst.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {SUCCESS_GREEN};")
                elif pa_connected and load_connected:
                    err_part = "PA Read Error" if not pa_ok else "Load Read Error"
                    self.chk_inst.setText(f"Instruments: {err_part}")
                    self.chk_inst.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {WARNING_AMBER};")
                else:
                    self.chk_inst.setText("Instruments: PA/Load Offline")
                    self.chk_inst.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {DANGER_RED};")

                if pa_ok:
                    self.chk_vin.setText(f"Vin: {pa_vin:.2f}V ({'Matches Target' if vin_ok else 'Mismatch'})")
                    self.chk_vin.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {SUCCESS_GREEN if vin_ok else WARNING_AMBER};")
                elif pa_connected:
                    self.chk_vin.setText("Vin: — (Read Error)")
                    self.chk_vin.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {WARNING_AMBER};")
                else:
                    self.chk_vin.setText("Target Vin: Unchecked")
                    self.chk_vin.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {TEXT_MUTED};")

                now_str = datetime.now().strftime('%H:%M:%S')
                if pa_ok and load_ok and vin_ok and load_verified:
                    self.readiness_status.setText("● READY")
                    self.readiness_status.setStyleSheet(f"color: {SUCCESS_GREEN}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                    self.statusBar().showMessage(f"Bench check complete · {now_str} · READY")
                elif not pa_connected or not load_connected:
                    self.readiness_status.setText("● NOT READY (Instruments offline)")
                    self.readiness_status.setStyleSheet(f"color: {DANGER_RED}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                    self.statusBar().showMessage(f"Bench check complete · {now_str} · NOT READY (Instruments offline)")
                elif not pa_ok:
                    self.readiness_status.setText("● NOT READY (PA read error)")
                    self.readiness_status.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                    self.statusBar().showMessage(f"Bench check complete · {now_str} · NOT READY (PA read error)")
                elif not load_ok:
                    self.readiness_status.setText("● NOT READY (Load read error)")
                    self.readiness_status.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                    self.statusBar().showMessage(f"Bench check complete · {now_str} · NOT READY (Load read error)")
                elif not vin_ok:
                    self.readiness_status.setText("● NOT READY (Vin mismatch)")
                    self.readiness_status.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                    self.statusBar().showMessage(f"Bench check complete · {now_str} · NOT READY (Vin mismatch)")
                elif not load_verified:
                    self.readiness_status.setText("● NOT READY (Load not verified)")
                    self.readiness_status.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                    self.statusBar().showMessage(f"Bench check complete · {now_str} · NOT READY (Load not verified)")
                else:
                    self.readiness_status.setText("● NOT READY")
                    self.readiness_status.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 800; font-size: 13px; margin-left: 4px;")
                    self.statusBar().showMessage(f"Bench check complete · {now_str} · NOT READY")
            finally:
                self._set_bench_busy(False)

        def on_fail(err: str):
            try:
                self.statusBar().showMessage(f"Bench check failed · {err}")
            finally:
                self._set_bench_busy(False)

        task = FunctionTask(query_all)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        QtCore.QThreadPool.globalInstance().start(task)

    # ----------------- TAB 2: RUN (INTEGRATED SETUP + SPLIT-VIEW) -----------------
    def _run_tab(self):
        widget = QtWidgets.QWidget()
        outer = QtWidgets.QVBoxLayout(widget)
        outer.setContentsMargins(10, 8, 10, 8)
        outer.setSpacing(8)

        # 1. Compact Single Horizontal Strip for Test Setup
        setup_strip = QtWidgets.QFrame()
        setup_strip.setObjectName("run_setup_strip")
        setup_strip.setStyleSheet(f"QFrame#run_setup_strip {{ background: #FFFFFF; border: 1px solid {BORDER}; border-radius: 6px; }}")
        ss_lay = QtWidgets.QHBoxLayout(setup_strip)
        ss_lay.setContentsMargins(12, 6, 12, 6)
        ss_lay.setSpacing(14)

        # Test Name
        ss_lay.addWidget(QtWidgets.QLabel("<b>Test Name:</b>"))
        self.test_name = QtWidgets.QLineEdit(self.config.get("campaign_name", "Efficiency Test"))
        self.test_name.setPlaceholderText("e.g. SID 48V SplitPhase fsw100k")
        self.test_name.setToolTip("Descriptive name or leave blank for auto-generated timestamp (e.g. Test_20260817_0420).")
        self.test_name.setMinimumWidth(220)
        ss_lay.addWidget(self.test_name, 1)

        # Target Vin
        ss_lay.addWidget(QtWidgets.QLabel("<b>Target Vin:</b>"))
        self.vin_target = QtWidgets.QDoubleSpinBox()
        self.vin_target.setRange(0, 1000)
        self.vin_target.setValue(float(self.config.get("vin_target_v", 48.0)))
        self.vin_target.setSuffix(" V")
        self.vin_target.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.vin_target.setFixedWidth(80)
        self.vin_target.setToolTip("Target input voltage to the converter under test.")
        ss_lay.addWidget(self.vin_target)

        # Switching Frequency
        ss_lay.addWidget(QtWidgets.QLabel("<b>Switching Frequency (kHz):</b>"))
        self.frequency = CompactDoubleSpinBox()
        self.frequency.setRange(0.001, 10_000)
        self.frequency.setDecimals(3)
        self.frequency.setValue(float(self.config.get("frequency_hz", 100_000)) / 1000.0)
        self.frequency.setSuffix(" kHz")
        self.frequency.setSingleStep(10.0)
        self.frequency.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.frequency.setFixedWidth(105)
        self.frequency.setToolTip("Switching frequency in kHz. Stored run data remains in Hz for compatibility.")
        ss_lay.addWidget(self.frequency)

        outer.addWidget(setup_strip)

        # 2. Prominent Compact Segmented Mode Selector: SET CURRENT | STEP CURRENT | CONTINUOUS | PULSE ... Demo Mode ○
        mode_bar = QtWidgets.QHBoxLayout()
        mode_bar.setContentsMargins(0, 0, 0, 2)
        mode_bar.setSpacing(0)

        mode_frame = QtWidgets.QFrame()
        mode_frame.setFixedHeight(34)
        mode_frame.setStyleSheet(f"background: #E5E7EB; border: 1px solid {BORDER}; border-radius: 6px; padding: 1px;")
        mb_layout = QtWidgets.QHBoxLayout(mode_frame)
        mb_layout.setContentsMargins(2, 2, 2, 2)
        mb_layout.setSpacing(3)

        self.mode_group = QtWidgets.QButtonGroup(self)
        self.btn_mode_direct = QtWidgets.QPushButton("SET CURRENT")
        self.btn_mode_direct.setToolTip("Command arbitrary target current setpoint in Amps.")
        self.btn_mode_step = QtWidgets.QPushButton("STEP CURRENT")
        self.btn_mode_step.setToolTip("Increase/decrease current manually using configurable steps.")
        self.btn_mode_cont = QtWidgets.QPushButton("CONTINUOUS")
        self.btn_mode_cont.setToolTip("Automated staircase sweep across current steps. Load remains ON between points.")
        self.btn_mode_pulse = QtWidgets.QPushButton("PULSE")
        self.btn_mode_pulse.setToolTip("Pulsed sweep across current steps. Load returns to 0 A / OFF between pulses.")

        mode_buttons = [
            (self.btn_mode_direct, 0),
            (self.btn_mode_step, 1),
            (self.btn_mode_cont, 2),
            (self.btn_mode_pulse, 3),
        ]

        for btn, idx in mode_buttons:
            btn.setCheckable(True)
            btn.setStyleSheet("""
                QPushButton {
                    background: transparent;
                    color: #4B5563;
                    font-weight: 800;
                    border: none;
                    font-size: 11px;
                    border-radius: 4px;
                    padding: 4px 14px;
                    letter-spacing: 0.5px;
                }
                QPushButton:checked {
                    background: #002676;
                    color: #FFFFFF;
                }
                QPushButton:hover:!checked {
                    background: #D1D5DB;
                }
            """)
            self.mode_group.addButton(btn, idx)
            mb_layout.addWidget(btn)

        # Backward compatibility aliases
        self.btn_mode_manual = self.btn_mode_direct
        self.radio_manual = self.btn_mode_direct
        self.radio_cont = self.btn_mode_cont
        self.radio_pulse = self.btn_mode_pulse
        self.btn_sub_direct = self.btn_mode_direct
        self.btn_sub_step = self.btn_mode_step
        self.btn_submode_direct = self.btn_mode_direct
        self.btn_submode_step = self.btn_mode_step

        self.btn_mode_cont.setChecked(True)
        mode_bar.addWidget(mode_frame)
        mode_bar.addStretch()

        # Right-aligned Demo Mode Toggle
        mode_bar.addWidget(self.simulation)

        outer.addLayout(mode_bar)
        self.mode_group.idToggled.connect(lambda id_, chk: self._mode_selected(id_) if chk else None)

        # 3. Main Splitter: Left Controls / Right Live Plot (~48% Controls, ~52% Plot)
        self.run_splitter = QtWidgets.QSplitter(QtCore.Qt.Orientation.Horizontal)

        # LEFT PANEL: Parameters & Controls in a Scroll Area (no scrollbar in normal 1080p operation)
        left_scroll = QtWidgets.QScrollArea()
        left_scroll.setWidgetResizable(True)
        left_scroll.setFrameShape(QtWidgets.QFrame.Shape.NoFrame)
        left_scroll.setHorizontalScrollBarPolicy(QtCore.Qt.ScrollBarPolicy.ScrollBarAlwaysOff)
        left_widget = QtWidgets.QWidget()
        l_layout = QtWidgets.QVBoxLayout(left_widget)
        l_layout.setContentsMargins(2, 2, 4, 2)
        l_layout.setSpacing(6)

        self.run_stack = QtWidgets.QStackedWidget()

        # 1. SET CURRENT View
        direct_w = QtWidgets.QWidget()
        d_lay = QtWidgets.QVBoxLayout(direct_w)
        d_lay.setContentsMargins(2, 2, 2, 2)
        d_lay.setSpacing(6)

        d_hdr = QtWidgets.QHBoxLayout()
        d_title = QtWidgets.QLabel("SET CURRENT")
        d_title.setStyleSheet(f"font-size: 13px; font-weight: 800; color: {BERKELEY_BLUE}; letter-spacing: 0.5px;")
        d_hdr.addWidget(d_title)
        d_hdr.addStretch()
        self.btn_adv_direct = DisclosureButton()
        self.direct_adv = self.btn_adv_direct
        d_hdr.addWidget(self.btn_adv_direct)
        d_lay.addLayout(d_hdr)

        d_box = QtWidgets.QGroupBox("Target Current")
        db_lay = QtWidgets.QHBoxLayout(d_box)
        db_lay.setContentsMargins(8, 6, 8, 6)
        db_lay.setSpacing(8)

        self.manual_target_spin = QtWidgets.QDoubleSpinBox()
        self.manual_target_spin.setRange(0.0, self.manual_mode_max_current())
        self.manual_target_spin.setValue(0.0)
        self.manual_target_spin.setSuffix(" A")
        self.manual_target_spin.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.manual_target_spin.setStyleSheet("font-size: 15px; font-weight: 700; padding: 4px;")
        db_lay.addWidget(self.manual_target_spin, 1)

        self.btn_direct_set = QtWidgets.QPushButton("SET CURRENT")
        self.btn_direct_set.setStyleSheet(f"background: {PRIMARY_BLUE}; color: white; font-weight: 800; font-size: 12px; padding: 8px 14px; border-radius: 4px;")
        self.btn_direct_set.clicked.connect(self._manual_apply_direct)
        db_lay.addWidget(self.btn_direct_set)

        self.btn_direct_zero = QtWidgets.QPushButton("ZERO / OFF")
        self.btn_direct_zero.setStyleSheet(f"background: #FFFBEB; color: {WARNING_AMBER}; border: 1.5px solid {WARNING_AMBER}; font-weight: 800; font-size: 12px; padding: 8px 12px; border-radius: 4px;")
        self.btn_direct_zero.clicked.connect(self._manual_zero)
        db_lay.addWidget(self.btn_direct_zero)
        d_lay.addWidget(d_box)

        # Status Pill for SET CURRENT
        self.direct_status_lbl = QtWidgets.QLabel("READY")
        self.direct_status_lbl.setObjectName("direct_status_lbl")
        self.direct_status_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.direct_status_lbl.setStyleSheet("color: #475569; font-weight: 800; font-size: 11px; padding: 3px 8px; border-radius: 4px; background: #F1F5F9; border: 1px solid #CBD5E1;")
        d_lay.addWidget(self.direct_status_lbl)

        # Actual Readout & State Box
        readout_box = QtWidgets.QFrame()
        readout_box.setStyleSheet(f"background: #F8FAFC; border: 1px solid {BORDER}; border-radius: 6px; padding: 4px 8px;")
        rb_lay = QtWidgets.QVBoxLayout(readout_box)
        rb_lay.setContentsMargins(4, 2, 4, 2)
        rb_lay.setSpacing(1)
        act_title = QtWidgets.QLabel("Actual Load State:")
        act_title.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 10px; font-weight: 700;")
        self.manual_actual_lbl = QtWidgets.QLabel("0.00 A · OFF")
        self.manual_actual_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 18px; font-weight: 800;")
        self.manual_actual_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        rb_lay.addWidget(act_title)
        rb_lay.addWidget(self.manual_actual_lbl)
        d_lay.addWidget(readout_box)

        # Advanced Box (Direct Set) - Collapsed by default
        self.direct_adv_box = QtWidgets.QGroupBox("Advanced Settings")
        self.direct_adv_box.setVisible(False)
        self.btn_adv_direct.set_target(self.direct_adv_box)
        da_lay = QtWidgets.QVBoxLayout(self.direct_adv_box)
        da_lay.setContentsMargins(8, 6, 8, 6)
        da_lay.setSpacing(6)

        d_delay_row = QtWidgets.QHBoxLayout()
        d_delay_row.addWidget(QtWidgets.QLabel("Auto-record after:"))
        self.direct_auto_delay = QtWidgets.QDoubleSpinBox()
        self.direct_auto_delay.setRange(0.5, 60.0)
        self.direct_auto_delay.setSingleStep(0.5)
        self.direct_auto_delay.setValue(4.0)
        self.direct_auto_delay.setSuffix(" s")
        self.direct_auto_delay.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.direct_auto_delay.setToolTip("Wait time after setting current before recording measurement and/or scope capture.")
        d_delay_row.addWidget(self.direct_auto_delay)
        da_lay.addLayout(d_delay_row)

        d_chk_row = QtWidgets.QHBoxLayout()
        self.direct_auto_save = QtWidgets.QCheckBox("Auto Save Reading")
        self.direct_auto_save.setChecked(True)
        self.direct_auto_save.setToolTip("Automatically record electrical measurement once settle delay expires.")
        self.direct_auto_capture = QtWidgets.QCheckBox("Auto Scope Capture")
        self.direct_auto_capture.setChecked(True)
        self.direct_auto_capture.setToolTip("Automatically execute oscilloscope capture once settle delay expires.")
        d_chk_row.addWidget(self.direct_auto_save)
        d_chk_row.addWidget(self.direct_auto_capture)
        da_lay.addLayout(d_chk_row)

        da_act_row = QtWidgets.QHBoxLayout()
        self.manual_save_btn = QtWidgets.QPushButton("Save Reading")
        self.manual_save_btn.setToolTip("Explicit manual override to store current operating point in Excel workbook.")
        self.manual_save_btn.clicked.connect(self._manual_save_direct_override)
        self.manual_capture_btn = QtWidgets.QPushButton("Scope Capture")
        self.manual_capture_btn.setToolTip("Explicit manual override to capture frozen oscilloscope screenshot and CSV.")
        self.manual_capture_btn.clicked.connect(self._manual_capture_direct_override)
        self.direct_save_btn = self.manual_save_btn
        self.direct_capture_btn = self.manual_capture_btn
        da_act_row.addWidget(self.manual_save_btn)
        da_act_row.addWidget(self.manual_capture_btn)
        da_lay.addLayout(da_act_row)

        d_lay.addWidget(self.direct_adv_box)
        d_lay.addStretch()
        self.run_stack.addWidget(direct_w)

        # 2. STEP CURRENT View
        step_w = QtWidgets.QWidget()
        s_lay = QtWidgets.QVBoxLayout(step_w)
        s_lay.setContentsMargins(2, 2, 2, 2)
        s_lay.setSpacing(6)

        s_hdr = QtWidgets.QHBoxLayout()
        s_title = QtWidgets.QLabel("STEP CURRENT")
        s_title.setStyleSheet(f"font-size: 13px; font-weight: 800; color: {BERKELEY_BLUE}; letter-spacing: 0.5px;")
        s_hdr.addWidget(s_title)
        s_hdr.addStretch()
        self.btn_adv_step = DisclosureButton()
        self.step_adv = self.btn_adv_step
        s_hdr.addWidget(self.btn_adv_step)
        s_lay.addLayout(s_hdr)

        # Commanded / Present current display
        pres_box = QtWidgets.QFrame()
        pres_box.setStyleSheet(f"background: #FFFFFF; border: 1.5px solid {BORDER}; border-radius: 6px; padding: 4px 8px;")
        pb_lay = QtWidgets.QVBoxLayout(pres_box)
        pb_lay.setContentsMargins(4, 2, 4, 2)
        pb_lay.setSpacing(1)
        pres_title = QtWidgets.QLabel("Present Commanded Setpoint:")
        pres_title.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 10px; font-weight: 700;")
        self.step_present_lbl = QtWidgets.QLabel("0.00 A · OFF")
        self.step_present_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 20px; font-weight: 900;")
        self.step_present_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        pb_lay.addWidget(pres_title)
        pb_lay.addWidget(self.step_present_lbl)
        s_lay.addWidget(pres_box)

        # Compact Step Point Status Pill
        self.step_status_lbl = QtWidgets.QLabel("READY")
        self.step_status_lbl.setObjectName("step_status_lbl")
        self.step_status_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.step_status_lbl.setStyleSheet("color: #475569; font-weight: 800; font-size: 11px; padding: 3px 8px; border-radius: 4px; background: #F1F5F9; border: 1px solid #CBD5E1;")
        s_lay.addWidget(self.step_status_lbl)

        # Step Buttons Row: [ −5 A ] [ ZERO / OFF ] [ +2 A ]
        step_btn_box = QtWidgets.QGroupBox("Step Controls")
        sb_lay = QtWidgets.QHBoxLayout(step_btn_box)
        sb_lay.setContentsMargins(8, 6, 8, 6)
        sb_lay.setSpacing(8)

        self.btn_minus_step = QtWidgets.QPushButton("−5 A")
        self.btn_minus_step.setStyleSheet(f"background: #F1F5F9; color: {BERKELEY_BLUE}; border: 1.5px solid {BORDER}; font-weight: 800; font-size: 13px; padding: 8px 14px; border-radius: 4px;")
        self.btn_minus_step.clicked.connect(self._step_down)
        sb_lay.addWidget(self.btn_minus_step, 1)

        self.btn_step_zero = QtWidgets.QPushButton("ZERO / OFF")
        self.btn_step_zero.setStyleSheet(f"background: #FFFBEB; color: {WARNING_AMBER}; border: 1.5px solid {WARNING_AMBER}; font-weight: 800; font-size: 12px; padding: 8px 12px; border-radius: 4px;")
        self.btn_step_zero.clicked.connect(self._manual_zero)
        sb_lay.addWidget(self.btn_step_zero, 1)

        self.btn_plus_step = QtWidgets.QPushButton("+2 A")
        self.btn_plus_step.setStyleSheet(f"background: #F1F5F9; color: {BERKELEY_BLUE}; border: 1.5px solid {BORDER}; font-weight: 800; font-size: 13px; padding: 8px 14px; border-radius: 4px;")
        self.btn_plus_step.clicked.connect(self._step_up)
        sb_lay.addWidget(self.btn_plus_step, 1)
        s_lay.addWidget(step_btn_box)

        # Actual Readout & State Box
        step_readout_box = QtWidgets.QFrame()
        step_readout_box.setStyleSheet(f"background: #F8FAFC; border: 1px solid {BORDER}; border-radius: 6px; padding: 4px 8px;")
        srb_lay = QtWidgets.QVBoxLayout(step_readout_box)
        srb_lay.setContentsMargins(4, 2, 4, 2)
        srb_lay.setSpacing(1)
        step_act_title = QtWidgets.QLabel("Actual Load State:")
        step_act_title.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 10px; font-weight: 700;")
        self.step_actual_lbl = QtWidgets.QLabel("0.00 A · OFF")
        self.step_actual_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 18px; font-weight: 800;")
        self.step_actual_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        srb_lay.addWidget(step_act_title)
        srb_lay.addWidget(self.step_actual_lbl)
        s_lay.addWidget(step_readout_box)

        # Advanced Box (Step Settings) - Collapsed by default
        self.step_adv_box = QtWidgets.QGroupBox("Advanced Settings")
        self.step_adv_box.setVisible(False)
        self.btn_adv_step.set_target(self.step_adv_box)
        sa_lay = QtWidgets.QVBoxLayout(self.step_adv_box)
        sa_lay.setContentsMargins(8, 6, 8, 6)
        sa_lay.setSpacing(6)

        inc_row = QtWidgets.QHBoxLayout()
        inc_row.addWidget(QtWidgets.QLabel("Increase step (+):"))
        self.manual_step_inc = QtWidgets.QDoubleSpinBox()
        self.manual_step_inc.setRange(0.01, 100)
        self.manual_step_inc.setValue(2.0)
        self.manual_step_inc.setSuffix(" A")
        self.manual_step_inc.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.manual_step_inc.valueChanged.connect(self._update_step_inc_button)
        inc_row.addWidget(self.manual_step_inc)
        sa_lay.addLayout(inc_row)

        dec_row = QtWidgets.QHBoxLayout()
        dec_row.addWidget(QtWidgets.QLabel("Decrease step (−):"))
        self.manual_step_dec = QtWidgets.QDoubleSpinBox()
        self.manual_step_dec.setRange(0.01, 100)
        self.manual_step_dec.setValue(5.0)
        self.manual_step_dec.setSuffix(" A")
        self.manual_step_dec.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.manual_step_dec.valueChanged.connect(self._update_step_dec_button)
        dec_row.addWidget(self.manual_step_dec)
        sa_lay.addLayout(dec_row)

        # Auto-record delay & feature checkboxes
        delay_row = QtWidgets.QHBoxLayout()
        delay_row.addWidget(QtWidgets.QLabel("Auto-record after:"))
        self.step_auto_delay = QtWidgets.QDoubleSpinBox()
        self.step_auto_delay.setRange(0.5, 60.0)
        self.step_auto_delay.setSingleStep(0.5)
        self.step_auto_delay.setValue(4.0)
        self.step_auto_delay.setSuffix(" s")
        self.step_auto_delay.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.step_auto_delay.setToolTip("Wait time after setting current before recording measurement and/or scope capture.")
        delay_row.addWidget(self.step_auto_delay)
        sa_lay.addLayout(delay_row)

        chk_row = QtWidgets.QHBoxLayout()
        self.step_auto_save = QtWidgets.QCheckBox("Auto Save Reading")
        self.step_auto_save.setChecked(True)
        self.step_auto_save.setToolTip("Automatically record electrical measurement once settle delay expires.")
        self.step_auto_capture = QtWidgets.QCheckBox("Auto Scope Capture")
        self.step_auto_capture.setChecked(True)
        self.step_auto_capture.setToolTip("Automatically execute oscilloscope capture once settle delay expires.")
        chk_row.addWidget(self.step_auto_save)
        chk_row.addWidget(self.step_auto_capture)
        sa_lay.addLayout(chk_row)

        self.manual_step = self.manual_step_inc

        sa_act_row = QtWidgets.QHBoxLayout()
        self.step_save_btn = QtWidgets.QPushButton("Save Reading")
        self.step_save_btn.setToolTip("Explicit manual override to store current operating point in Excel workbook.")
        self.step_save_btn.clicked.connect(self._manual_save_step_override)
        self.step_capture_btn = QtWidgets.QPushButton("Scope Capture")
        self.step_capture_btn.setToolTip("Explicit manual override to capture frozen oscilloscope screenshot and CSV.")
        self.step_capture_btn.clicked.connect(self._manual_capture_step_override)
        sa_act_row.addWidget(self.step_save_btn)
        sa_act_row.addWidget(self.step_capture_btn)
        sa_lay.addLayout(sa_act_row)

        s_lay.addWidget(self.step_adv_box)
        s_lay.addStretch()
        self.run_stack.addWidget(step_w)

        self.manual_current = self.manual_target_spin

        # 3. CONTINUOUS View (2-Column Layout: Basic | Advanced)
        cont_widget = QtWidgets.QWidget()
        cont_outer = QtWidgets.QVBoxLayout(cont_widget)
        cont_outer.setContentsMargins(2, 2, 2, 2)
        cont_outer.setSpacing(6)

        cont_hdr = QtWidgets.QHBoxLayout()
        cont_title = QtWidgets.QLabel("CONTINUOUS SWEEP")
        cont_title.setStyleSheet(f"font-size: 13px; font-weight: 800; color: {MODE_CONT}; letter-spacing: 0.5px;")
        cont_hdr.addWidget(cont_title)
        cont_hdr.addStretch()
        self.btn_adv_cont = DisclosureButton()
        self.cont_adv = self.btn_adv_cont
        cont_hdr.addWidget(self.btn_adv_cont)
        cont_outer.addLayout(cont_hdr)

        cont_cols = QtWidgets.QHBoxLayout()
        cont_cols.setSpacing(8)

        # Left Column: Basic
        cont_basic_box = QtWidgets.QGroupBox("Basic")
        cb_form = QtWidgets.QFormLayout(cont_basic_box)
        cb_form.setContentsMargins(8, 8, 8, 8)
        cb_form.setHorizontalSpacing(10)
        cb_form.setVerticalSpacing(6)

        self.cont_start = QtWidgets.QDoubleSpinBox(); self.cont_start.setRange(0, 2000); self.cont_start.setValue(0.0); self.cont_start.setSuffix(" A"); self.cont_start.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cont_stop = QtWidgets.QDoubleSpinBox(); self.cont_stop.setRange(0, 2000); self.cont_stop.setValue(60.0); self.cont_stop.setSuffix(" A"); self.cont_stop.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cont_step = QtWidgets.QDoubleSpinBox(); self.cont_step.setRange(0.01, 100); self.cont_step.setValue(2.0); self.cont_step.setSuffix(" A"); self.cont_step.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cont_settle = QtWidgets.QDoubleSpinBox(); self.cont_settle.setRange(0.1, 120); self.cont_settle.setValue(5.0); self.cont_settle.setSuffix(" s"); self.cont_settle.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cont_settle.setToolTip("Total dwell at each current point. Measure last occurs inside this interval.")

        for sp in (self.cont_start, self.cont_stop, self.cont_step, self.cont_settle):
            sp.valueChanged.connect(self._update_sweep_summary)

        cb_form.addRow("Start", self.cont_start)
        cb_form.addRow("Stop", self.cont_stop)
        cb_form.addRow("Step", self.cont_step)
        cb_form.addRow("Wait", self.cont_settle)
        cont_cols.addWidget(cont_basic_box, 1)

        # Right Column: Advanced (Collapsed by default)
        self.cont_adv_box = QtWidgets.QGroupBox("Advanced Settings")
        self.cont_adv_box.setVisible(False)
        self.btn_adv_cont.set_target(self.cont_adv_box)
        ca_form = QtWidgets.QFormLayout(self.cont_adv_box)
        ca_form.setContentsMargins(8, 8, 8, 8)
        ca_form.setHorizontalSpacing(10)
        ca_form.setVerticalSpacing(4)

        self.cont_sample_count = QtWidgets.QSpinBox(); self.cont_sample_count.setRange(1, 20); self.cont_sample_count.setValue(1); self.cont_sample_count.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cont_sample_count.setToolTip("Number of readings averaged per point (1 = single snapshot).")
        self.cont_sample_count.valueChanged.connect(self._update_sample_controls)

        self.cont_sample_window = QtWidgets.QDoubleSpinBox(); self.cont_sample_window.setRange(0.05, 30); self.cont_sample_window.setValue(3.0); self.cont_sample_window.setSuffix(" s"); self.cont_sample_window.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cont_sample_window.setToolTip("Acquisition region during the final part of the total Wait interval.")
        self.cont_sample_window.valueChanged.connect(self._update_sweep_summary)

        self.cont_capture_points = QtWidgets.QLineEdit("0, 10, 20, 30")
        self.cont_capture_points.setToolTip("Comma-separated currents where oscilloscope triggers a frozen capture.")
        self.cont_capture_points.textChanged.connect(self._update_sweep_summary)

        self.cont_scope_summary_lbl = QtWidgets.QLabel("4 scope captures")
        self.cont_scope_summary_lbl.setStyleSheet(f"color: {SUCCESS_GREEN}; font-weight: 700; font-size: 11px;")

        self.cont_return_step = QtWidgets.QDoubleSpinBox()
        self.cont_return_step.setRange(0.1, 100.0)
        self.cont_return_step.setValue(5.0)
        self.cont_return_step.setSuffix(" A")
        self.cont_return_step.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.cont_return_step.setToolTip("Current decrement per step when returning to 0 A (default: 5 A).")

        self.cont_psu_req = QtWidgets.QCheckBox("Stop test if gate PSU fails")
        self.cont_psu_req.setStyleSheet("font-size: 11px;")

        ca_form.addRow("Readings", self.cont_sample_count)
        ca_form.addRow("Measure last", self.cont_sample_window)
        ca_form.addRow("Scope caps", self.cont_capture_points)
        ca_form.addRow("", self.cont_scope_summary_lbl)
        ca_form.addRow("Return step", self.cont_return_step)
        ca_form.addRow("", self.cont_psu_req)
        cont_cols.addWidget(self.cont_adv_box, 1)

        cont_outer.addLayout(cont_cols)

        # Compact Status strip (35-45px high)
        self.cont_summary_lbl = QtWidgets.QLabel("0 → 60 A   |   31 points   |   ~155 s")
        self.cont_summary_lbl.setFixedHeight(38)
        self.cont_summary_lbl.setStyleSheet(f"color: #166534; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #F0FDF4; border: 1px solid #BBF7D0; border-radius: 4px;")
        self.cont_summary_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        cont_outer.addWidget(self.cont_summary_lbl)

        self.run_stack.addWidget(cont_widget)

        # 4. PULSE View (2-Column Layout: Basic | Advanced)
        pulse_widget = QtWidgets.QWidget()
        pulse_outer = QtWidgets.QVBoxLayout(pulse_widget)
        pulse_outer.setContentsMargins(2, 2, 2, 2)
        pulse_outer.setSpacing(6)

        pulse_hdr = QtWidgets.QHBoxLayout()
        pulse_title = QtWidgets.QLabel("PULSE SWEEP")
        pulse_title.setStyleSheet(f"font-size: 13px; font-weight: 800; color: {MODE_PULSE}; letter-spacing: 0.5px;")
        pulse_hdr.addWidget(pulse_title)
        pulse_hdr.addStretch()
        self.btn_adv_pulse = DisclosureButton()
        self.pulse_adv = self.btn_adv_pulse
        pulse_hdr.addWidget(self.btn_adv_pulse)
        pulse_outer.addLayout(pulse_hdr)

        pulse_cols = QtWidgets.QHBoxLayout()
        pulse_cols.setSpacing(8)

        # Left Column: Basic
        pulse_basic_box = QtWidgets.QGroupBox("Basic")
        pb_form = QtWidgets.QFormLayout(pulse_basic_box)
        pb_form.setContentsMargins(8, 8, 8, 8)
        pb_form.setHorizontalSpacing(10)
        pb_form.setVerticalSpacing(6)

        self.pulse_start = QtWidgets.QDoubleSpinBox(); self.pulse_start.setRange(0, 2000); self.pulse_start.setValue(0.0); self.pulse_start.setSuffix(" A"); self.pulse_start.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_stop = QtWidgets.QDoubleSpinBox(); self.pulse_stop.setRange(0, 2000); self.pulse_stop.setValue(60.0); self.pulse_stop.setSuffix(" A"); self.pulse_stop.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_step = QtWidgets.QDoubleSpinBox(); self.pulse_step.setRange(0.01, 100); self.pulse_step.setValue(2.0); self.pulse_step.setSuffix(" A"); self.pulse_step.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_dwell = QtWidgets.QDoubleSpinBox(); self.pulse_dwell.setRange(0.1, 120); self.pulse_dwell.setValue(5.0); self.pulse_dwell.setSuffix(" s"); self.pulse_dwell.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_dwell.setToolTip("Pulse duration that load holds target current before returning to 0 A.")
        self.pulse_cooldown = QtWidgets.QDoubleSpinBox(); self.pulse_cooldown.setRange(0.1, 600); self.pulse_cooldown.setValue(5.0); self.pulse_cooldown.setSuffix(" s"); self.pulse_cooldown.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_cooldown.setToolTip("Rest cooldown period at 0 A between pulses.")

        for sp in (self.pulse_start, self.pulse_stop, self.pulse_step, self.pulse_dwell, self.pulse_cooldown):
            sp.valueChanged.connect(self._update_sweep_summary)

        pb_form.addRow("Start", self.pulse_start)
        pb_form.addRow("Stop", self.pulse_stop)
        pb_form.addRow("Step", self.pulse_step)
        pb_form.addRow("ON time", self.pulse_dwell)
        pb_form.addRow("Rest", self.pulse_cooldown)
        pulse_cols.addWidget(pulse_basic_box, 1)

        # Right Column: Advanced (Collapsed by default)
        self.pulse_adv_box = QtWidgets.QGroupBox("Advanced Settings")
        self.pulse_adv_box.setVisible(False)
        self.btn_adv_pulse.set_target(self.pulse_adv_box)
        pa_form = QtWidgets.QFormLayout(self.pulse_adv_box)
        pa_form.setContentsMargins(8, 8, 8, 8)
        pa_form.setHorizontalSpacing(10)
        pa_form.setVerticalSpacing(4)

        self.pulse_sample_count = QtWidgets.QSpinBox(); self.pulse_sample_count.setRange(1, 20); self.pulse_sample_count.setValue(1); self.pulse_sample_count.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_sample_count.setToolTip("Number of readings averaged per point (1 = single snapshot).")
        self.pulse_sample_count.valueChanged.connect(self._update_sample_controls)

        self.pulse_sample_window = QtWidgets.QDoubleSpinBox(); self.pulse_sample_window.setRange(0.05, 30); self.pulse_sample_window.setValue(3.0); self.pulse_sample_window.setSuffix(" s"); self.pulse_sample_window.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_sample_window.setToolTip("Acquisition region during the final part of the pulse ON interval.")
        self.pulse_sample_window.valueChanged.connect(self._update_sweep_summary)

        self.pulse_capture_points = QtWidgets.QLineEdit("0, 10, 20, 30")
        self.pulse_capture_points.setToolTip("Comma-separated currents where oscilloscope triggers a frozen capture.")
        self.pulse_capture_points.textChanged.connect(self._update_sweep_summary)

        self.pulse_scope_summary_lbl = QtWidgets.QLabel("4 scope captures")
        self.pulse_scope_summary_lbl.setStyleSheet(f"color: {SUCCESS_GREEN}; font-weight: 700; font-size: 11px;")

        self.pulse_return_step = QtWidgets.QDoubleSpinBox()
        self.pulse_return_step.setRange(0.1, 100.0)
        self.pulse_return_step.setValue(5.0)
        self.pulse_return_step.setSuffix(" A")
        self.pulse_return_step.setButtonSymbols(QtWidgets.QAbstractSpinBox.ButtonSymbols.NoButtons)
        self.pulse_return_step.setToolTip("Current decrement per step when returning to 0 A (default: 5 A).")

        self.pulse_psu_req = QtWidgets.QCheckBox("Stop test if gate PSU fails")
        self.pulse_psu_req.setStyleSheet("font-size: 11px;")

        pa_form.addRow("Readings", self.pulse_sample_count)
        pa_form.addRow("Measure last", self.pulse_sample_window)
        pa_form.addRow("Scope caps", self.pulse_capture_points)
        pa_form.addRow("", self.pulse_scope_summary_lbl)
        pa_form.addRow("Return step", self.pulse_return_step)
        pa_form.addRow("", self.pulse_psu_req)
        pulse_cols.addWidget(self.pulse_adv_box, 1)

        pulse_outer.addLayout(pulse_cols)

        # Compact Status strip (35-45px high)
        self.pulse_summary_lbl = QtWidgets.QLabel("0 → 60 A   |   31 pulses   |   ~310 s")
        self.pulse_summary_lbl.setFixedHeight(38)
        self.pulse_summary_lbl.setStyleSheet(f"color: #166534; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #F0FDF4; border: 1px solid #BBF7D0; border-radius: 4px;")
        self.pulse_summary_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        pulse_outer.addWidget(self.pulse_summary_lbl)

        self.run_stack.addWidget(pulse_widget)
        l_layout.addWidget(self.run_stack)

        # Sweep Action Box (Side-by-side action buttons to eliminate vertical scroll)
        self.sweep_exec_box = QtWidgets.QGroupBox("Execution Controls")
        s_layout_box = QtWidgets.QVBoxLayout(self.sweep_exec_box)
        s_layout_box.setContentsMargins(8, 6, 8, 6)
        s_layout_box.setSpacing(6)

        # 1. Action Buttons Row: [ ▶ START SWEEP / ▶ START DEMO ]  [ ■ STOP & RETURN TO ZERO ]
        btn_row = QtWidgets.QHBoxLayout()
        btn_row.setSpacing(8)

        self.run_sweep_btn = QtWidgets.QPushButton("▶  START SWEEP")
        self.run_sweep_btn.setObjectName("primary_action")
        self.run_sweep_btn.setStyleSheet(f"background: {SUCCESS_GREEN}; color: white; font-weight: 800; font-size: 13px; padding: 8px 12px; border-radius: 6px;")
        self.run_sweep_btn.clicked.connect(self.start_run)
        btn_row.addWidget(self.run_sweep_btn, 1)

        self.stop_sweep_btn = QtWidgets.QPushButton("■  STOP & RETURN TO ZERO")
        self.stop_sweep_btn.setStyleSheet(f"background: #FFFBEB; color: {WARNING_AMBER}; border: 2px solid {WARNING_AMBER}; font-weight: 800; font-size: 12px; padding: 8px 10px; border-radius: 6px;")
        self.stop_sweep_btn.setToolTip("Stop taking new measurement points and gracefully ramp current down to 0 A.")
        self.stop_sweep_btn.clicked.connect(self.stop_and_return_to_zero)
        btn_row.addWidget(self.stop_sweep_btn, 1)

        s_layout_box.addLayout(btn_row)

        status_sub = QtWidgets.QHBoxLayout()
        self.lbl_curr_point = QtWidgets.QLabel("Current: —")
        self.lbl_curr_point.setStyleSheet(f"font-weight: 700; font-size: 12px; color: {PRIMARY_BLUE};")
        self.lbl_next_point = QtWidgets.QLabel("Next: —")
        self.lbl_next_point.setStyleSheet(f"font-weight: 600; font-size: 11px; color: {TEXT_MUTED};")
        status_sub.addWidget(self.lbl_curr_point)
        status_sub.addWidget(self.lbl_next_point)
        s_layout_box.addLayout(status_sub)

        self.run_progress_bar = QtWidgets.QProgressBar()
        self.run_progress_bar.setFixedHeight(16)
        self.run_progress_bar.setFormat("Idle")
        s_layout_box.addWidget(self.run_progress_bar)

        l_layout.addWidget(self.sweep_exec_box)
        l_layout.addStretch()

        left_scroll.setWidget(left_widget)
        self.run_splitter.addWidget(left_scroll)

        # RIGHT PANEL: Live Interactive Plots & Controls
        right_panel = QtWidgets.QWidget()
        r_layout = QtWidgets.QVBoxLayout(right_panel)
        r_layout.setContentsMargins(4, 4, 4, 4)
        r_layout.setSpacing(6)

        # 1. Compact Live Status & Current-Progress Marker
        self.plot_status_box = QtWidgets.QFrame()
        self.plot_status_box.setStyleSheet(f"background: #F8FAFC; border: 1px solid {BORDER}; border-radius: 6px; padding: 4px 8px;")
        psb_lay = QtWidgets.QVBoxLayout(self.plot_status_box)
        psb_lay.setContentsMargins(6, 4, 6, 4)
        psb_lay.setSpacing(2)

        stat_r = QtWidgets.QHBoxLayout()
        stat_r.setSpacing(10)
        self.live_stat_tag = QtWidgets.QLabel("IDLE ○")
        self.live_stat_tag.setStyleSheet(f"font-weight: 800; font-size: 12px; color: {TEXT_MUTED};")

        self.live_point_lbl = QtWidgets.QLabel("Point — / —")
        self.live_point_lbl.setStyleSheet(f"font-weight: 700; font-size: 12px; color: {TEXT_MAIN};")

        self.live_cmd_lbl = QtWidgets.QLabel("Command: —")
        self.live_cmd_lbl.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {BERKELEY_BLUE};")

        self.live_act_lbl = QtWidgets.QLabel("Actual: —")
        self.live_act_lbl.setStyleSheet(f"font-weight: 700; font-size: 12px; color: {SUCCESS_GREEN};")

        stat_r.addWidget(self.live_stat_tag)
        stat_r.addWidget(self.live_point_lbl)
        stat_r.addWidget(self.live_cmd_lbl)
        stat_r.addWidget(self.live_act_lbl)
        stat_r.addStretch()

        self.plot_progress_marker = SweepProgressTracker()

        psb_lay.addLayout(stat_r)
        psb_lay.addWidget(self.plot_progress_marker)
        r_layout.addWidget(self.plot_status_box)

        # 2. Plot Controls Row: Metric Dropdown | [Reset View]
        p_top = QtWidgets.QHBoxLayout()
        p_top.setSpacing(8)
        self.live_metric_combo = QtWidgets.QComboBox()
        self.live_metric_combo.setMinimumWidth(130)
        self.live_metric_combo.addItems([
            "Efficiency (%)",
            "Loss (W)",
            "Power (W)",
        ])
        self.live_metric_combo.currentIndexChanged.connect(self._switch_live_plot)
        p_top.addWidget(self.live_metric_combo)

        self.reset_view_btn = QtWidgets.QPushButton("Reset View")
        self.reset_view_btn.setToolTip("Restores the controlled current and metric axis ranges.")
        self.reset_view_btn.clicked.connect(self._reset_live_plot_view)
        p_top.addWidget(self.reset_view_btn)

        self.sim_watermark = QtWidgets.QLabel("")  # Internal reference maintained
        p_top.addStretch()

        r_layout.addLayout(p_top)

        # 3. PyQtGraph Plot with Semi-Transparent "DEMO DATA" Watermark
        self.live_plot_widget = pg.PlotWidget(background="w")
        self.live_plot_widget.showGrid(x=True, y=True, alpha=0.3)
        self.live_plot_widget.setLabel("bottom", "Output Current (Iout)", "A")
        self.live_plot_widget.setLabel("left", "Efficiency (%)")
        self.live_plot_widget.setMouseEnabled(x=False, y=False)
        self.live_legend = self.live_plot_widget.addLegend(offset=(-12, -12))
        self.live_legend.anchor((1, 1), (1, 1), offset=(-12, -12))
        self.live_legend.setBrush(pg.mkBrush(255, 255, 255, 220))
        self.live_legend.setPen(pg.mkPen(BORDER))
        self.live_curve = plot_metric_series(
            self.live_plot_widget, [], [], PLOT_CORE_BLUE, name="Converter", symbol="o", symbol_size=6
        )
        self.live_system_curve = plot_metric_series(
            self.live_plot_widget, [], [], PLOT_SYSTEM_ORANGE, name="System", symbol="o", symbol_size=6
        )
        self.live_aux_curve = plot_metric_series(
            self.live_plot_widget, [], [], PLOT_AUX_TEAL, name="Paux", symbol="o", symbol_size=6
        )

        # Centered Watermark Overlay on Plot Canvas (Translucent Orange)
        self.plot_watermark_lbl = QtWidgets.QLabel("DEMO DATA", self.live_plot_widget)
        self.plot_watermark_lbl.setAttribute(QtCore.Qt.WidgetAttribute.WA_TransparentForMouseEvents)
        self.plot_watermark_lbl.setStyleSheet("color: rgba(217, 119, 6, 0.18); font-size: 52px; font-weight: 900; letter-spacing: 8px;")
        self.plot_watermark_lbl.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.plot_watermark_lbl.setVisible(False)

        pw_lay = QtWidgets.QVBoxLayout(self.live_plot_widget)
        pw_lay.setContentsMargins(0, 0, 0, 0)
        pw_lay.addWidget(self.plot_watermark_lbl, 0, QtCore.Qt.AlignmentFlag.AlignCenter)

        r_layout.addWidget(self.live_plot_widget)

        self.run_splitter.addWidget(right_panel)
        self.run_splitter.setStretchFactor(0, 1)
        self.run_splitter.setStretchFactor(1, 1)
        self.run_splitter.setSizes([560, 600])

        outer.addWidget(self.run_splitter)
        self._mode_selected(self.mode_group.checkedId())
        self._update_sweep_summary()

        return widget

    def manual_mode_max_current(self) -> float:
        """Single source of truth for maximum allowed current in manual modes (SET CURRENT and STEP CURRENT)."""
        return float(self.cap_val)

    def frequency_hz(self) -> float:
        """Convert the operator-facing kHz control at the storage/hardware boundary."""
        return float(self.frequency.value()) * 1000.0

    def _update_progress_range(self):
        if hasattr(self, "plot_progress_marker"):
            mode_id = self._selected_mode_id if self._selected_mode_id is not None else (self.mode_group.checkedId() if hasattr(self, "mode_group") else 2)
            if mode_id in (0, 1):  # SET CURRENT (0) and STEP CURRENT (1)
                max_safe = self.manual_mode_max_current()
                self.plot_progress_marker.set_range(0.0, max_safe)
                cur = getattr(self, "_manual_target_current", 0.0)
                self.plot_progress_marker.update_position(cur, 0.0, max_safe, active=(cur > 0.0))
            elif mode_id == 3:  # PULSE
                self.plot_progress_marker.set_range(self.pulse_start.value(), self.pulse_stop.value())
            else:  # CONTINUOUS
                self.plot_progress_marker.set_range(self.cont_start.value(), self.cont_stop.value())

    def _mode_selected(self, button_id: int):
        previous = self._selected_mode_id
        self._selected_mode_id = button_id
        if previous is not None and previous != button_id:
            self._finalize_manual_session("Mode changed")
            self._clear_live_run_view()
        self.run_stack.setCurrentIndex(button_id)
        # button_id 0 = SET CURRENT (Manual)
        # button_id 1 = STEP CURRENT (Manual)
        # button_id 2 = CONTINUOUS
        # button_id 3 = PULSE
        modes = ["Set Current", "Step Current", "Continuous", "Pulse"]
        mode_str = modes[button_id] if button_id < len(modes) else "Continuous"
        if hasattr(self, "mode_indicator"):
            self.mode_indicator.set_mode(mode_str)
        self.sweep_exec_box.setVisible(button_id in (2, 3))
        self._update_run_button_text()
        self._update_sweep_summary()
        self._update_progress_range()
        if hasattr(self, "live_plot_widget"):
            self._reset_live_plot_view()
        self.update_enabled_states()


    def _update_run_button_text(self):
        if not hasattr(self, "run_sweep_btn"):
            return
        mode_idx = self.mode_group.checkedId() if hasattr(self, "mode_group") else 2
        if hasattr(self, "simulation") and self.simulation.isChecked():
            self.run_sweep_btn.setText("▶  START DEMO")
            self.run_sweep_btn.setStyleSheet(f"background: #D97706; color: white; font-weight: 800; font-size: 13px; padding: 8px 12px; border-radius: 6px;")
            self.run_sweep_btn.setToolTip("Start simulated demo test run using synthetic data (no hardware commands).")
        else:
            self.run_sweep_btn.setStyleSheet(f"background: {SUCCESS_GREEN}; color: white; font-weight: 800; font-size: 13px; padding: 8px 12px; border-radius: 6px;")
            if mode_idx == 3:
                self.run_sweep_btn.setText("▶  START PULSE TEST")
                self.run_sweep_btn.setToolTip("Start automated pulse sweep sequence.")
            else:
                self.run_sweep_btn.setText("▶  START SWEEP")
                self.run_sweep_btn.setToolTip("Start automated continuous sweep sequence.")


    def _update_step_inc_button(self, val: float):
        self.btn_plus_step.setText(f"+{val:g} A")

    def _update_step_dec_button(self, val: float):
        self.btn_minus_step.setText(f"−{val:g} A")

    def _update_step_buttons(self, val: float):
        self._update_step_inc_button(val)


    def _update_sample_controls(self):
        self._update_sweep_summary()

    def _update_sweep_summary(self):
        # Continuous
        try:
            start_c = self.cont_start.value()
            stop_c = self.cont_stop.value()
            step_c = self.cont_step.value()
            pts_c, _ = generate_points(start_c, stop_c, step_c, self.cap.value())
            est_time_c = len(pts_c) * self.cont_settle.value()
            self.cont_summary_lbl.setText(f"{start_c:g} → {stop_c:g} A   |   {len(pts_c)} points   |   ~{est_time_c:.0f} s")
            self.cont_summary_lbl.setStyleSheet("color: #166534; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #F0FDF4; border: 1px solid #BBF7D0; border-radius: 4px;")

            caps_c_txt = self.cont_capture_points.text().strip()
            if not caps_c_txt:
                self.cont_scope_summary_lbl.setText("No scope captures")
                self.cont_scope_summary_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px;")
            else:
                try:
                    caps_c = sorted(list(parse_capture_points(caps_c_txt)))
                    matching_c = [c for c in caps_c if any(math.isclose(c, p, abs_tol=1e-3) for p in pts_c)]
                    unaligned_c = [c for c in caps_c if not any(math.isclose(c, p, abs_tol=1e-3) for p in pts_c)]
                    if not unaligned_c:
                        self.cont_scope_summary_lbl.setText(f"{len(caps_c)} scope captures")
                        self.cont_scope_summary_lbl.setStyleSheet(f"color: {SUCCESS_GREEN}; font-weight: 700; font-size: 11px;")
                    else:
                        un_str = ", ".join(f"{x:g}A" for x in unaligned_c)
                        self.cont_scope_summary_lbl.setText(f"{len(matching_c)} of {len(caps_c)} valid · ⚠️ Unaligned: {un_str}")
                        self.cont_scope_summary_lbl.setStyleSheet(f"color: {DANGER_RED}; font-weight: 700; font-size: 11px;")
                except Exception as exc:
                    self.cont_scope_summary_lbl.setText(f"Invalid scope capture: {exc}")
                    self.cont_scope_summary_lbl.setStyleSheet(f"color: {DANGER_RED}; font-weight: 700; font-size: 11px;")
        except Exception as exc:
            self.cont_summary_lbl.setText(f"Error: {exc}")
            self.cont_summary_lbl.setStyleSheet("color: #B91C1C; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #FEF2F2; border: 1px solid #FECACA; border-radius: 4px;")

        # Pulse
        try:
            start_p = self.pulse_start.value()
            stop_p = self.pulse_stop.value()
            step_p = self.pulse_step.value()
            pts_p, _ = generate_points(start_p, stop_p, step_p, self.cap.value())
            est_time_p = len(pts_p) * (self.pulse_dwell.value() + self.pulse_cooldown.value())
            self.pulse_summary_lbl.setText(f"{start_p:g} → {stop_p:g} A   |   {len(pts_p)} pulses   |   ~{est_time_p:.0f} s")
            self.pulse_summary_lbl.setStyleSheet("color: #166534; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #F0FDF4; border: 1px solid #BBF7D0; border-radius: 4px;")

            caps_p_txt = self.pulse_capture_points.text().strip()
            if not caps_p_txt:
                self.pulse_scope_summary_lbl.setText("No scope captures")
                self.pulse_scope_summary_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-size: 11px;")
            else:
                try:
                    caps_p = sorted(list(parse_capture_points(caps_p_txt)))
                    matching_p = [c for c in caps_p if any(math.isclose(c, p, abs_tol=1e-3) for p in pts_p)]
                    unaligned_p = [c for c in caps_p if not any(math.isclose(c, p, abs_tol=1e-3) for p in pts_p)]
                    if not unaligned_p:
                        self.pulse_scope_summary_lbl.setText(f"{len(caps_p)} scope captures")
                        self.pulse_scope_summary_lbl.setStyleSheet(f"color: {SUCCESS_GREEN}; font-weight: 700; font-size: 11px;")
                    else:
                        un_str = ", ".join(f"{x:g}A" for x in unaligned_p)
                        self.pulse_scope_summary_lbl.setText(f"{len(matching_p)} of {len(caps_p)} valid · ⚠️ Unaligned: {un_str}")
                        self.pulse_scope_summary_lbl.setStyleSheet(f"color: {DANGER_RED}; font-weight: 700; font-size: 11px;")
                except Exception as exc:
                    self.pulse_scope_summary_lbl.setText(f"Invalid scope capture: {exc}")
                    self.pulse_scope_summary_lbl.setStyleSheet(f"color: {DANGER_RED}; font-weight: 700; font-size: 11px;")
        except Exception as exc:
            self.pulse_summary_lbl.setText(f"Error: {exc}")
            self.pulse_summary_lbl.setStyleSheet("color: #B91C1C; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #FEF2F2; border: 1px solid #FECACA; border-radius: 4px;")

        self._update_progress_range()


    def _reset_live_plot_view(self):
        self._switch_live_plot(self.live_metric_combo.currentIndex())

    def _clear_live_run_view(self):
        """Clear data/status that belongs to the previous logical run."""
        self.plot_rows.clear()
        if hasattr(self, "live_curve"):
            self.live_curve.setData([], [])
        if hasattr(self, "live_system_curve"):
            self.live_system_curve.setData([], [])
        if hasattr(self, "live_aux_curve"):
            self.live_aux_curve.setData([], [])
        if hasattr(self, "live_stat_tag"):
            self.live_stat_tag.setText("READY ◌")
            self.live_stat_tag.setStyleSheet(f"font-weight: 800; font-size: 12px; color: {TEXT_MUTED};")
        if hasattr(self, "live_point_lbl"):
            self.live_point_lbl.setText("Fresh run")
        if hasattr(self, "live_cmd_lbl"):
            self.live_cmd_lbl.setText("Command: 0 A")
        if hasattr(self, "run_progress_bar"):
            self.run_progress_bar.setRange(0, 1)
            self.run_progress_bar.setValue(0)
            self.run_progress_bar.setFormat("READY")
        if hasattr(self, "strip_progress"):
            self.strip_progress.setRange(0, 1)
            self.strip_progress.setValue(0)
            self.strip_progress.setFormat("READY")
        if hasattr(self, "plot_progress_marker"):
            self.plot_progress_marker.set_idle()
        if hasattr(self, "live_plot_widget"):
            self._reset_live_plot_view()

    def _switch_live_plot(self, index: int):
        metric_specs = [
            ("Efficiency (%)", [
                ("EfficiencyConverter_pct", "Converter", PLOT_CORE_BLUE),
                ("EfficiencySystem_pct", "System", PLOT_SYSTEM_ORANGE),
            ]),
            ("Loss (W)", [
                ("LossConverter_W", "Converter Loss", PLOT_CORE_BLUE),
                ("LossSystem_W", "System Loss", PLOT_SYSTEM_ORANGE),
            ]),
            ("Power (W)", [
                ("PinConverter_W", "Pin", PLOT_CORE_BLUE),
                ("Pout_W", "Pout", PLOT_SYSTEM_ORANGE),
                ("Paux_W", "Paux", PLOT_AUX_TEAL),
            ]),
        ]
        y_label, series_specs = metric_specs[index]
        self.live_plot_widget.setLabel("left", y_label)
        curves = [self.live_curve, self.live_system_curve, self.live_aux_curve]
        self.live_legend.clear()
        xs: list[float] = []
        ys: list[float] = []
        positive_efficiencies: list[float] = []

        for curve_index, curve in enumerate(curves):
            if curve_index >= len(series_specs):
                curve.setData([], [])
                continue
            field, legend_name, base_color = series_specs[curve_index]
            pairs = [
                (row.get("Iout_A"), row.get(field)) for row in self.plot_rows
                if row.get("Status") == "Valid" and isinstance(row.get("Iout_A"), (int, float))
                and isinstance(row.get(field), (int, float))
                and math.isfinite(float(row.get("Iout_A"))) and math.isfinite(float(row.get(field)))
            ]
            apply_metric_curve_style(curve, base_color, symbol="o", symbol_size=6, name=legend_name)
            curve.setData([p[0] for p in pairs], [p[1] for p in pairs])
            self.live_legend.addItem(curve, legend_name)
            xs.extend(float(p[0]) for p in pairs)
            ys.extend(float(p[1]) for p in pairs)
            if index == 0:
                positive_efficiencies.extend(float(y) for x, y in pairs if float(x) > 1e-6)

        self.live_legend.setVisible(True)
        if index == 0:
            y_min, y_max = efficiency_axis_bounds(positive_efficiencies)
            self.live_plot_widget.setYRange(y_min, y_max, padding=0.02)

        # Controlled ranges for X-axis
        mode_id = self._selected_mode_id if self._selected_mode_id is not None else (self.mode_group.checkedId() if hasattr(self, "mode_group") else 2)
        if mode_id in (0, 1):  # SET CURRENT (0) and STEP CURRENT (1)
            start_x = 0.0
            stop_x = self.manual_mode_max_current()
        elif mode_id == 3:  # PULSE
            start_x = min(0.0, self.pulse_start.value())
            stop_x = self.pulse_stop.value()
        else:  # CONTINUOUS
            start_x = min(0.0, self.cont_start.value())
            stop_x = self.cont_stop.value()

        max_x = max(stop_x, max(xs) if xs else 10.0, 1.0)
        self.live_plot_widget.setXRange(start_x, max_x, padding=0.04)

        if index != 0:  # Loss / Power use ordinary numeric bounds.
            if ys:
                y_min = min([0.0] + ys)
                y_max = max([1.0] + ys) * 1.1
                self.live_plot_widget.setYRange(y_min, y_max, padding=0.02)
            else:
                self.live_plot_widget.setYRange(0.0, 10.0, padding=0.02)

    # ----------------- TAB 3: HISTORY & MULTI-RUN COMPARE -----------------
    def _history_tab(self):
        widget = QtWidgets.QWidget()
        layout = QtWidgets.QHBoxLayout(widget)
        layout.setContentsMargins(12, 10, 12, 10)
        layout.setSpacing(12)

        # Left Table
        left_box = QtWidgets.QGroupBox("Campaign Runs (Select Multiple to Overlay)")
        l_layout = QtWidgets.QVBoxLayout(left_box)
        l_layout.setContentsMargins(10, 10, 10, 10)
        l_layout.setSpacing(8)

        self.history_table = QtWidgets.QTableWidget(0, 9)
        self.history_table.setHorizontalHeaderLabels(["Run Short ID", "RunID", "Full Test Name", "Status", "Source", "Vin", "Freq (kHz)", "Base Test", "Mode"])
        self.history_table.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.history_table.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.history_table.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.history_table.setSelectionBehavior(QtWidgets.QAbstractItemView.SelectionBehavior.SelectRows)
        self.history_table.setSelectionMode(QtWidgets.QAbstractItemView.SelectionMode.ExtendedSelection)
        self.history_table.setEditTriggers(QtWidgets.QAbstractItemView.EditTrigger.NoEditTriggers)
        self.history_table.itemSelectionChanged.connect(self._history_selection_changed)
        l_layout.addWidget(self.history_table)

        act_row = QtWidgets.QHBoxLayout()
        copy_btn = QtWidgets.QPushButton("Copy Run to Clipboard (TSV)")
        copy_btn.clicked.connect(self._copy_run_clipboard)
        self.history_del_btn = QtWidgets.QPushButton("Delete Permanently...")
        self.history_del_btn.setStyleSheet(f"""
            QPushButton {{
                background: #FEF2F2;
                color: #B91C1C;
                border: 1px solid #FECACA;
                font-weight: 700;
                padding: 4px 10px;
                border-radius: 4px;
            }}
            QPushButton:hover {{
                background: #FEE2E2;
                border-color: #FCA5A5;
            }}
        """)
        self.history_del_btn.clicked.connect(self._history_delete)
        act_row.addWidget(self.history_del_btn)
        act_row.addStretch()
        l_layout.addLayout(act_row)

        layout.addWidget(left_box, 1)

        # Right Multi-Curve Comparison Plot
        right_box = QtWidgets.QGroupBox("Multi-Run Comparison Curves")
        r_layout = QtWidgets.QVBoxLayout(right_box)
        r_layout.setContentsMargins(10, 10, 10, 10)
        r_layout.setSpacing(8)

        top_sel = QtWidgets.QHBoxLayout()
        top_sel.addWidget(QtWidgets.QLabel("<b>Comparison Metric:</b>"))
        self.comp_metric_combo = QtWidgets.QComboBox()
        self.comp_metric_combo.setMinimumWidth(130)
        self.comp_metric_combo.addItems([
            "Efficiency (%)",
            "Loss (W)",
            "Power (W)",
        ])
        self.comp_metric_combo.currentIndexChanged.connect(lambda: self._history_selection_changed())
        top_sel.addWidget(self.comp_metric_combo)
        top_sel.addStretch()
        r_layout.addLayout(top_sel)

        self.comp_plot_widget = pg.PlotWidget(background="w")
        self.comp_plot_widget.showGrid(x=True, y=True, alpha=0.3)
        self.comp_plot_widget.setLabel("bottom", "Output Current (Iout)", "A")
        self.comp_plot_widget.setLabel("left", "Efficiency (%)")
        self.comp_plot_widget.setMouseEnabled(x=False, y=False)
        self.comp_legend = self.comp_plot_widget.addLegend(offset=(-12, -12))
        self.comp_legend.anchor((1, 1), (1, 1), offset=(-12, -12))
        self.comp_legend.setBrush(pg.mkBrush(255, 255, 255, 220))
        self.comp_legend.setPen(pg.mkPen(BORDER))
        r_layout.addWidget(self.comp_plot_widget)

        layout.addWidget(right_box, 1)
        return widget

    def _history_selection_changed(self):
        self.comp_plot_widget.clear()
        self.comp_legend.clear()

        selected_rows = sorted(list({idx.row() for idx in self.history_table.selectedIndexes()}))
        if hasattr(self, "history_del_btn"):
            if len(selected_rows) > 1:
                self.history_del_btn.setText(f"Delete {len(selected_rows)} Runs...")
            else:
                self.history_del_btn.setText("Delete Permanently...")

        if not selected_rows:
            return

        metric_index = self.comp_metric_combo.currentIndex()
        metric_specs = [
            ("Efficiency (%)", [
                ("EfficiencyConverter_pct", "Converter", PLOT_CORE_BLUE),
                ("EfficiencySystem_pct", "System", PLOT_SYSTEM_ORANGE),
            ]),
            ("Loss (W)", [
                ("LossConverter_W", "Converter Loss", PLOT_CORE_BLUE),
                ("LossSystem_W", "System Loss", PLOT_SYSTEM_ORANGE),
            ]),
            ("Power (W)", [
                ("PinConverter_W", "Pin", PLOT_CORE_BLUE),
                ("Pout_W", "Pout", PLOT_SYSTEM_ORANGE),
                ("Paux_W", "Paux", PLOT_AUX_TEAL),
            ]),
        ]
        y_label, series_specs = metric_specs[metric_index]
        self.comp_plot_widget.setLabel("left", y_label)

        all_xs: list[float] = []
        all_ys: list[float] = []

        for i, row in enumerate(selected_rows):
            short_id = self.history_table.item(row, 0).text() if self.history_table.item(row, 0) else ""
            run_id = self.history_table.item(row, 0).data(QtCore.Qt.ItemDataRole.UserRole)
            if not run_id and self.history_table.item(row, 1):
                run_id = self.history_table.item(row, 1).text()
            if not run_id:
                continue
            meas = self._store_for_history_row(row).get_run_measurements(run_id)
            if not meas:
                continue
            vin = self.history_table.item(row, 5).text() if self.history_table.item(row, 5) else ""
            freq = self.history_table.item(row, 6).text() if self.history_table.item(row, 6) else ""
            mod = self.history_table.item(row, 7).text() if self.history_table.item(row, 7) else ""
            label_name = f"{vin} V, {freq}, {mod} ({short_id or run_id.split('-')[-1]})"

            valid_meas = [r for r in meas if r.get("Status") == "Valid"]
            run_markers = ["o", "s", "t", "d", "+", "x", "star"]
            marker = "o" if len(selected_rows) == 1 else run_markers[i % len(run_markers)]
            for field, series_name, base_color in series_specs:
                pairs = [
                    (r.get("Iout_A"), r.get(field)) for r in valid_meas
                    if isinstance(r.get("Iout_A"), (int, float)) and isinstance(r.get(field), (int, float))
                    and math.isfinite(float(r.get("Iout_A"))) and math.isfinite(float(r.get(field)))
                ]
                if not pairs:
                    continue
                color = QtGui.QColor(base_color)
                if i > 0:
                    color = color.lighter(100 + min(45, i * 12)) if i % 2 else color.darker(100 + min(35, i * 10))
                xs = [float(p[0]) for p in pairs]
                ys = [float(p[1]) for p in pairs]
                all_xs.extend(xs)
                if metric_index == 0:
                    all_ys.extend(y for x, y in zip(xs, ys) if x > 1e-6)
                else:
                    all_ys.extend(ys)
                legend_name = series_name if len(selected_rows) == 1 else f"{series_name} · {short_id or run_id.split('-')[-1]}"
                plot_metric_series(
                    self.comp_plot_widget,
                    xs,
                    ys,
                    color=color,
                    name=legend_name,
                    symbol=marker,
                    symbol_size=6,
                )

        if all_xs:
            self.comp_plot_widget.setXRange(0.0, max(max(all_xs), 1.0), padding=0.04)
            if metric_index == 0:
                y_min, y_max = efficiency_axis_bounds(all_ys)
                self.comp_plot_widget.setYRange(y_min, y_max, padding=0.02)
            else:
                self.comp_plot_widget.setYRange(min([0.0] + all_ys), max([1.0] + all_ys) * 1.1, padding=0.02)

    def _copy_run_clipboard(self):
        selected_rows = sorted(list({idx.row() for idx in self.history_table.selectedIndexes()}))
        if not selected_rows:
            QtWidgets.QMessageBox.information(self, "Copy Run", "Select at least one run from the history table.")
            return
        row = selected_rows[0]
        run_id = self.history_table.item(row, 0).data(QtCore.Qt.ItemDataRole.UserRole)
        if not run_id and self.history_table.item(row, 1):
            run_id = self.history_table.item(row, 1).text()
        meas = self._store_for_history_row(row).get_run_measurements(run_id)
        if not meas:
            QtWidgets.QMessageBox.information(self, "Copy Run", f"No measurement records found for {run_id}.")
            return
        output = io.StringIO()
        writer = csv.writer(output, delimiter="\t")
        writer.writerow(MEAS_HEADERS)
        for r in meas:
            writer.writerow([r.get(h, "") for h in MEAS_HEADERS])
        QtWidgets.QApplication.clipboard().setText(output.getvalue())
        self.statusBar().showMessage(f"Copied {len(meas)} points of {run_id} to clipboard")

    # ----------------- TAB 4: HELP / DEVELOPER -----------------
    def _help_tab(self):
        widget = QtWidgets.QWidget()
        layout = QtWidgets.QVBoxLayout(widget)
        layout.setContentsMargins(14, 10, 14, 10)
        layout.setSpacing(10)

        # Bench Configuration & Operator Notes Group
        dev_box = QtWidgets.QGroupBox("Bench Configuration & Operator Notes")
        dev_box.setStyleSheet(f"QGroupBox {{ font-weight: 800; color: {BERKELEY_BLUE}; }}")
        d_layout = QtWidgets.QHBoxLayout(dev_box)
        d_layout.setContentsMargins(12, 8, 12, 8)
        d_layout.setSpacing(12)
        d_layout.addWidget(self.fpga_check)
        d_layout.addWidget(self.notes, 1)
        layout.addWidget(dev_box)


        text = QtWidgets.QTextBrowser()
        text.setOpenExternalLinks(True)
        text.setHtml(f"""
        <h2 style='color:{BERKELEY_BLUE}'>Kickstart Bench · Operating Reference</h2>
        
        <h3>Five-Minute Quick Start</h3>
        <ol>
            <li><b>Bench Setup:</b> Click <i>Check / Refresh Entire Bench</i>. Verify all instruments respond and Vin matches target.</li>
            <li><b>Supply:</b> Set voltages, current limits, role names, and desired output states directly in the E36312A Supply card and click <i>Apply Settings to Supply</i>.</li>
            <li><b>Run:</b> Confirm Test Name, Vin, Modulation, and Frequency in the header bar. Select Manual, Continuous, or Pulse mode. Click <i>Review and Start Sweep</i>.</li>
        </ol>

        <h3>Gate Driver & Auxiliary Supply Calculation</h3>
        <ul>
            <li><b>Channel Role:</b> Editable name for each channel (e.g. <code>Vdrv_A</code>, <code>Vdrv_B</code>, <code>Vdrv_C</code>, <code>Vaux_5V</code>). Stored in the <code>SupplyMeasurements</code> column with each point.</li>
            <li><b>Include in Paux loss:</b> When checked for a channel, its measured power consumption (<i>P = V &times; I</i>) is added to the auxiliary loss:
                <br><b>Paux_W</b> = &Sigma; (V<sub>ch</sub> &times; I<sub>ch</sub>) for all checked channels.</li>
            <li><b>Converter Efficiency (%):</b> Pure power stage efficiency: <b>100 &times; Pout / Pin</b> (calculated from Power Analyzer PA2201A).</li>
            <li><b>System Efficiency (%):</b> Total system efficiency including gate-drive & auxiliary bias power: <b>100 &times; Pout / (Pin + Paux)</b>.</li>
            <li><b>E36312A Ratings:</b> CH1 is <b>6V / 5A</b>; CH2 and CH3 are <b>25V / 1A</b>.</li>
        </ul>

        <h3>Modes & Waveforms</h3>
        <ul>
            <li><b>Manual Mode:</b> Direct load control with two dedicated sub-modes and a unified state model (<b>0 A always means LOAD OFF</b>):
                <ul>
                    <li><b>Direct Set:</b> Enter any current setpoint up to max cap. <i>ZERO 0 A</i> or setting 0 A turns load OFF; setting &gt;0 A sets target current and automatically turns load ON.</li>
                    <li><b>Step Control:</b> 1-click step buttons (<i>&minus;5 A</i>, <i>ZERO 0 A</i>, <i>+2 A</i>) that immediately command the load. Stepping to/below 0 A clamps to 0 A and turns load OFF; stepping to positive current turns load ON. Independent asymmetric increase (+2 A default) and decrease (−5 A default) step sizes are configurable in Advanced.</li>
                </ul>
            </li>
            <li><b>Continuous Mode:</b> Automated staircase sweep from start to stop current with dwell delay. Load remains ON between points. Default: 0 &rarr; 60 A in 2 A steps.</li>
            <li><b>Pulse Mode:</b> Pulsed sweep where load pulses to target current for Pulse ON time and rests at 0 A between pulses. Supports fast 2-point pulse tests (0 A &rarr; Max) or multi-point sweeps. Default: 0 &rarr; 60 A in 2 A steps.</li>
        </ul>


        <h3>Timing & Measurement Fields</h3>
        <ul>
            <li><b>Wait at each current:</b> Hard safety-bounded dwell time (s) load holds current. Total time per point in Continuous mode.</li>
            <li><b>Pulse ON time:</b> Hard safety-bounded pulse duration (s) load holds target current before turning OFF.</li>
            <li><b>Measure last:</b> Acquisition region inside the final 3 s of the total dwell/pulse ON interval. Pre-measurement settling time is <i>Wait or ON time − Measure last</i>.</li>
            <li><b>Readings to average:</b> Number of sequential snapshots averaged during the <i>Measure last</i> window (1 = single snapshot).</li>
            <li><b>Rest between pulses:</b> Cooldown duration (s) at 0 A between pulses.</li>
        </ul>

        <h3>Instrument Sessions & Safety</h3>
        <ul>
            <li><b>LOAD OFF / Emergency Stop:</b> Immediately commands the electronic load OFF and aborts active sweeps. Shortcut: <b>Esc</b>.</li>
            <li><b>Vin Safety Shutdown:</b> When enabled in Bench Setup, automatically turns the electronic load OFF and aborts active runs if measured Vin falls outside ±10% of Target Vin. <i>Note: Vin safety shutdown is evaluated from PA measurements and therefore is not instantaneous hardware protection. It is intended to prevent the load from remaining active after a major input-voltage abnormality.</i></li>
            <li><b>Thermal Checks:</b> Infrared / thermal camera verification remains the operator's responsibility.</li>
        </ul>
        """)
        layout.addWidget(text, 1)
        return widget


    # ----------------- SMART ENABLE/DISABLE LOGIC -----------------
    @property
    def cap(self):
        class CapProxy:
            def __init__(self, owner): self.owner = owner
            def value(self): return self.owner.cap_val
            def setValue(self, v): self.owner._on_safety_limit_applied(float(v))
        return CapProxy(self)

    def _on_safety_limit_applied(self, val: float):
        self.cap_val = val
        self.config["working_current_cap_a"] = val
        save_config(self.config)
        if hasattr(self, "run_cap_badge"):
            self.run_cap_badge.setText(f"Load limit: {val:g} A")
        if hasattr(self, "chk_cap"):
            self.chk_cap.setText(f"Current Cap: {val:g} A")
        if hasattr(self, "manual_target_spin"):
            self.manual_target_spin.setRange(0.0, val)
        if hasattr(self, "cont_stop"):
            self.cont_stop.setMaximum(val)
        if hasattr(self, "pulse_stop"):
            self.pulse_stop.setMaximum(val)
        self._update_sweep_summary()
        self._update_progress_range()
        if hasattr(self, "live_plot_widget"):
            self._reset_live_plot_view()
        self.statusBar().showMessage(f"Electronic load safety limit set to {val:g} A", 4000)

    def update_enabled_states(self):
        is_running = (self.worker is not None and self.worker.isRunning()) or (self.demo_timer is not None and self.demo_timer.isActive())
        is_sim = self.simulation.isChecked()
        mode_id = self.mode_group.checkedId()
        manual_busy = self._manual_is_busy()

        # Emergency Stop Button Appearance
        if is_running:
            self.emergency_stop_btn.setText("LOAD OFF / ABORT")
            self.emergency_stop_btn.setStyleSheet(f"background: {DANGER_RED}; color: white; font-weight: 900;")
        else:
            self.emergency_stop_btn.setText("LOAD OFF")
            self.emergency_stop_btn.setStyleSheet(f"background: {CARD_BG}; color: {DANGER_RED}; border: 1px solid {DANGER_RED}; font-weight: 700;")

        # Stop & Return to Zero buttons
        self.stop_sweep_btn.setEnabled(is_running)
        self.strip_stop_btn.setEnabled(is_running)

        # Mode segmented buttons
        for rb in (self.btn_mode_direct, self.btn_mode_step, self.btn_mode_cont, self.btn_mode_pulse):
            rb.setEnabled(not is_running and not manual_busy)

        # Top Setup parameters
        for w in (self.test_name, self.vin_target, self.frequency, self.simulation, self.fpga_check, self.notes):
            w.setEnabled(not is_running)

        # Bench tab
        bench_busy = getattr(self, "bench_operation_busy", False)
        self.check_bench_btn.setEnabled(not is_running and not bench_busy)
        self.disc_btn.setEnabled(not is_running and not is_sim and not bench_busy)
        self.release_btn.setEnabled(not is_running and not bench_busy)

        # Supply card controls
        for c in self.supply_card.channel_controls:
            c["role_edit"].setEnabled(not is_running)
            c["desired_out"].setEnabled(not is_running)
            c["loss_chk"].setEnabled(not is_running)
            c["voltage"].setEnabled(not is_running)
            c["limit"].setEnabled(not is_running)
        self.supply_card.apply_btn.setEnabled(not is_running)

        # Load card controls
        if hasattr(self, "load_card"):
            self.load_card.read_btn.setEnabled(not is_running)
            self.load_card.cap_spin.setEnabled(not is_running)
            self.load_card.apply_cap_btn.setEnabled(not is_running)

        # SET/STEP CURRENT actions use one centralized priority/busy policy.
        direct_active = (not is_running and mode_id == 0)
        step_active = (not is_running and mode_id == 1)
        self._update_current_action_enabled_state(is_running)
        for w in (self.manual_target_spin, self.direct_auto_delay, self.direct_auto_save,
                   self.direct_auto_capture, self.manual_save_btn, self.manual_capture_btn, self.btn_adv_direct):
            w.setEnabled(direct_active)
        self.manual_save_btn.setEnabled(direct_active and not self._manual_active_task)
        self.manual_capture_btn.setEnabled(direct_active and not self._manual_active_task)

        for w in (self.manual_step_inc, self.manual_step_dec, self.step_auto_delay,
                   self.step_auto_save, self.step_auto_capture,
                   self.step_save_btn, self.step_capture_btn, self.btn_adv_step):
            w.setEnabled(step_active)
        self.step_save_btn.setEnabled(step_active and not self._manual_active_task)
        self.step_capture_btn.setEnabled(step_active and not self._manual_active_task)

        # CONTINUOUS controls (mode 2)
        cont_active = (not is_running and mode_id == 2)
        for w in (self.cont_start, self.cont_stop, self.cont_step, self.cont_settle,
                  self.cont_sample_count, self.cont_sample_window, self.cont_capture_points, self.cont_return_step,
                  self.cont_psu_req, self.btn_adv_cont):
            w.setEnabled(cont_active)

        # PULSE controls (mode 3)
        pulse_active = (not is_running and mode_id == 3)
        for w in (self.pulse_start, self.pulse_stop, self.pulse_step, self.pulse_dwell,
                  self.pulse_cooldown, self.pulse_sample_count, self.pulse_sample_window, self.pulse_capture_points,
                  self.pulse_return_step, self.pulse_psu_req, self.btn_adv_pulse):
            w.setEnabled(pulse_active)

        # Start Sweep button
        can_start = not is_running
        self.run_sweep_btn.setEnabled(can_start)
        if is_running:
            self.run_sweep_btn.setText("SWEEP RUNNING...")
        else:
            self._update_run_button_text()
            if not is_sim and not self.chk_load.isChecked():
                self.run_sweep_btn.setToolTip("Hardware write locked: Check 'I verified low-current load control on this bench' in Bench Setup first.")
            else:
                self.run_sweep_btn.setToolTip("Review parameters and begin automated sweep.")

    def _update_current_action_enabled_state(self, is_running: bool | None = None):
        """Apply the safety priority for all manual current-changing actions."""
        if is_running is None:
            is_running = (self.worker is not None and self.worker.isRunning()) or (self.demo_timer is not None and self.demo_timer.isActive())
        mode_id = self.mode_group.checkedId()
        direct_active = not is_running and mode_id == 0
        step_active = not is_running and mode_id == 1
        acquisition_busy = bool(self.point_action_busy)
        command_busy = bool(self._manual_active_task) and not acquisition_busy
        waiting = self._manual_countdown_timer.isActive()

        self.btn_direct_zero.setEnabled(direct_active)
        self.btn_step_zero.setEnabled(step_active)
        self.btn_direct_set.setEnabled(direct_active and not acquisition_busy and not command_busy and not waiting)
        self.btn_plus_step.setEnabled(step_active and not acquisition_busy and not command_busy and not waiting)
        # A downward step may interrupt an upward acquisition, but not another
        # zero/down hardware command that is already being dispatched.
        self.btn_minus_step.setEnabled(step_active and not command_busy)



    # ----------------- SLOTS & CORE OPERATIONS -----------------
    def _store_for_source(self, source: str) -> WorkbookStore:
        # Tests/tools may explicitly inject one standalone store. In normal GUI
        # operation, however, source identity—not the currently selected mode—
        # determines which workbook owns a run.
        current_store = getattr(self, "store", None)
        known_stores = {getattr(self, "hardware_store", None), getattr(self, "simulation_store", None)}
        if current_store is not None and current_store not in known_stores:
            return current_store
        return self.simulation_store if str(source).strip().lower() == "simulation" else self.hardware_store

    def _store_for_history_row(self, row: int) -> WorkbookStore:
        source_item = self.history_table.item(row, 4)
        return self._store_for_source(source_item.text() if source_item else "Hardware")

    def _supply_channels(self) -> list[SupplyChannel]:
        return [SupplyChannel(raw["channel"], raw.get("role", f"CH{raw['channel']}"), True, raw.get("enabled", False), raw.get("contributes_loss", True), raw.get("voltage_set", 0.0), None, raw.get("current_limit", 1.0), None) for raw in self.config["supply_channels"]]

    def _psu_read_kwargs(self):
        return {"channels": [1, 2, 3]}

    def _switch_mode(self, enabled: bool):
        if self.worker and self.worker.isRunning():
            self.simulation.blockSignals(True)
            self.simulation.setChecked(not enabled)
            self.simulation.blockSignals(False)
            return
        self.hub.release_all()
        config = {**self.config, "simulation_scenario": "Nominal"}
        self.hub = InstrumentHub(enabled, config)
        self.store = self.simulation_store if enabled else self.hardware_store
        self.supply_card.hub = self.hub
        self.scope_card.hub = self.hub

        self.simulation.setText("Demo Mode")
        if enabled:
            self.simulation.setStyleSheet("font-weight: 700; font-size: 12px; color: #D97706; padding: 2px 6px;")
        else:
            self.simulation.setStyleSheet(f"font-weight: 600; font-size: 12px; color: {TEXT_MUTED}; padding: 2px 6px;")

        if hasattr(self, "sim_warning_banner"):
            self.sim_warning_banner.setVisible(enabled)
        if hasattr(self, "plot_watermark_lbl"):
            self.plot_watermark_lbl.setVisible(enabled)

        for card in self.cards.values():
            card.last_snapshot = None
            card._mark_released()
        self._update_run_button_text()
        self.statusBar().showMessage("Demo Mode active: synthetic data, no hardware commands" if enabled else "Hardware idle")
        self.update_enabled_states()

    def _discover_devices(self):
        if self.bench_operation_busy:
            return
        if self.simulation.isChecked():
            self.statusBar().showMessage("Demo Mode active: simulation identities active")
            return
        self._set_bench_busy(True, "discover")
        self.statusBar().showMessage("Scanning VISA resources...")

        for key, card in self.all_instrument_cards().items():
            if hasattr(card, "_set_badge"):
                card._set_badge("gray", "Discovering...", f"Scanning for {key.upper()}...")

        def do_discover():
            return self.hub.discover()

        def on_done(found: dict[str, dict[str, str]]):
            try:
                found = found or {}
                found_names = []
                missing_names = []
                device_display_names = {
                    "pa": "PA2201A",
                    "load": "Chroma 63206A",
                    "psu": "E36312A",
                    "scope": "MSOX4024A",
                }

                for key in ("pa", "load", "psu", "scope"):
                    card = self.get_card_for_instrument(key)
                    display_name = device_display_names.get(key, key.upper())
                    if key in found:
                        item = found[key]
                        if "addresses" in self.config and item.get("address"):
                            self.config["addresses"][key] = item["address"]
                        found_names.append(display_name)
                        if hasattr(card, "_mark_discovered"):
                            card._mark_discovered(item.get("identity", item.get("address", "")))
                        elif hasattr(card, "_set_badge"):
                            card._set_badge("blue", "Discovered", f"Found over VISA · {item.get('identity', '')}")
                    else:
                        missing_names.append(display_name)
                        if hasattr(card, "_mark_not_found"):
                            card._mark_not_found(f"No {display_name} detected during VISA scan")
                        elif hasattr(card, "_set_badge"):
                            card._set_badge("amber", "Not Found", f"No {display_name} detected during VISA scan")

                save_config(self.config)

                count = len(found_names)
                if count == 4:
                    summary = f"VISA discovery complete · All 4 instruments found ({', '.join(found_names)})"
                elif count > 0:
                    summary = f"VISA discovery complete · Found {', '.join(found_names)} · {', '.join(missing_names)} not found"
                else:
                    summary = "VISA discovery complete · 0/4 supported instruments found"
                self.statusBar().showMessage(summary)
            finally:
                self._set_bench_busy(False)

        def on_fail(err: str):
            try:
                for key, card in self.all_instrument_cards().items():
                    if hasattr(card, "_mark_not_found"):
                        card._mark_not_found(f"Discovery error: {err}")
                self.statusBar().showMessage(f"VISA discovery failed · {err}")
            finally:
                self._set_bench_busy(False)

        task = FunctionTask(do_discover)
        task.signals.success.connect(on_done)
        task.signals.failure.connect(on_fail)
        QtCore.QThreadPool.globalInstance().start(task)

    def _snapshot_received(self, key: str, snap: InstrumentSnapshot):
        self.last_values[key] = snap
        if key == "pa":
            if snap.valid:
                if isinstance(snap.values.get("vin"), (int, float)):
                    self.kpi_labels["Vin"].setText(f"{snap.values['vin']:.2f} V")
                if isinstance(snap.values.get("iin"), (int, float)):
                    iin_val = snap.values["iin"]
                    self.kpi_labels["Iin"].setText("0.000 A" if abs(iin_val) < 0.0005 else f"{iin_val:.3f} A")
                if isinstance(snap.values.get("vout"), (int, float)):
                    self.kpi_labels["Vout"].setText(f"{snap.values['vout']:.2f} V")

        if key == "load":
            current = snap.values.get("current") if "current" in snap.values else snap.values.get("iout")
            enabled = snap.values.get("input_on") if "input_on" in snap.values else snap.values.get("enabled", None)
            if snap.valid and isinstance(current, (int, float)):
                if current > 0.001 and enabled is not False:
                    state_str = f"{current:.2f} A · ON"
                    state_color = SUCCESS_GREEN
                else:
                    state_str = "0.00 A · OFF"
                    state_color = TEXT_MUTED
                self.kpi_labels["Iout"].setText(state_str)
                if hasattr(self, "manual_actual_lbl"):
                    self.manual_actual_lbl.setText(state_str)
                    self.manual_actual_lbl.setStyleSheet(f"color: {state_color}; font-family: Consolas, monospace; font-size: 22px; font-weight: 800;")
                if hasattr(self, "step_actual_lbl"):
                    self.step_actual_lbl.setText(state_str)
                    self.step_actual_lbl.setStyleSheet(f"color: {state_color}; font-family: Consolas, monospace; font-size: 22px; font-weight: 800;")
                if hasattr(self, "live_act_lbl"):
                    self.live_act_lbl.setText(f"Actual: {current:.2f} A")

    def _update_derived_kpis(self, record: dict[str, Any]) -> None:
        """Update Pin/Pout/converter efficiency from one valid calculated point."""
        if record.get("Status") != "Valid":
            return
        values = (
            record.get("PinConverter_W"),
            record.get("Pout_W"),
            record.get("EfficiencyConverter_pct"),
        )
        if not all(isinstance(value, (int, float)) and math.isfinite(float(value)) for value in values):
            return
        pin, pout, efficiency = (float(value) for value in values)
        self.kpi_labels["Pin"].setText(f"{pin:.2f} W")
        self.kpi_labels["Pout"].setText(f"{pout:.2f} W")
        self.kpi_labels["Eff"].setText(f"{efficiency:.2f}%")

    def _validate_current(self, amps: float, current_value: float | None = None) -> bool:
        cap = self.manual_mode_max_current() if hasattr(self, "manual_mode_max_current") else self.cap_val
        if amps < 0 or amps > cap:
            QtWidgets.QMessageBox.critical(self, "Current rejected", f"{amps:g} A exceeds maximum allowed current ({cap:g} A).")
            return False
        if current_value is not None:
            ref = current_value
        elif hasattr(self, "_manual_target_current"):
            ref = self._manual_target_current
        elif hasattr(self, "manual_target_spin"):
            ref = self.manual_target_spin.value()
        elif hasattr(self, "manual_current"):
            ref = self.manual_current.value()
        else:
            ref = 0.0
        if amps >= 100 or abs(amps - ref) >= 50:
            ans = QtWidgets.QMessageBox.warning(self, "Large current command", f"Confirm command of {amps:g} A.", QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No, QtWidgets.QMessageBox.StandardButton.No)
            return ans == QtWidgets.QMessageBox.StandardButton.Yes
        return True

    def require_load_control_verified(self, action_name: str = "") -> bool:
        if self.simulation.isChecked():
            return True
        if self.chk_load.isChecked():
            return True
        QtWidgets.QMessageBox.warning(
            self,
            "Hardware write locked",
            "Check 'I verified low-current load control on this bench' in Bench Setup first."
        )
        return False

    def _hardware_write_allowed(self) -> bool:
        return self.require_load_control_verified()

    def _run_function(self, function: Callable[[], Any], success: Callable[[Any], None] | None = None,
                      failure: Callable[[str], None] | None = None):
        task = FunctionTask(function)
        if success: task.signals.success.connect(success)
        def on_failure(error: str):
            if failure:
                failure(error)
            QtWidgets.QMessageBox.warning(self, "Bench operation", error)
        task.signals.failure.connect(on_failure)
        QtCore.QThreadPool.globalInstance().start(task)

    def _manual_is_busy(self) -> bool:
        if getattr(self, "point_action_busy", False):
            return True
        if hasattr(self, "_manual_countdown_timer") and self._manual_countdown_timer.isActive():
            return True
        if getattr(self, "_manual_active_task", False):
            return True
        return False

    def _step_is_busy(self) -> bool:
        return self._manual_is_busy()

    def _update_manual_status(self, text: str, style_mode: str = "normal"):
        styles = {
            "ready": "color: #475569; font-weight: 800; font-size: 11px; padding: 3px 8px; border-radius: 4px; background: #F1F5F9; border: 1px solid #CBD5E1;",
            "settling": "color: #2563EB; font-weight: 800; font-size: 11px; padding: 3px 8px; border-radius: 4px; background: #EFF6FF; border: 1px solid #BFDBFE;",
            "saving": "color: #D97706; font-weight: 800; font-size: 11px; padding: 3px 8px; border-radius: 4px; background: #FFFBEB; border: 1px solid #FDE68A;",
            "recorded": "color: #166534; font-weight: 800; font-size: 11px; padding: 3px 8px; border-radius: 4px; background: #F0FDF4; border: 1px solid #BBF7D0;",
            "error": "color: #B91C1C; font-weight: 800; font-size: 11px; padding: 3px 8px; border-radius: 4px; background: #FEF2F2; border: 1px solid #FECACA;",
        }
        st = styles.get(style_mode, styles["ready"])
        for lbl_name in ("direct_status_lbl", "step_status_lbl"):
            if hasattr(self, lbl_name):
                lbl = getattr(self, lbl_name)
                lbl.setText(text)
                lbl.setStyleSheet(st)

    def _cancel_manual_automation(self, status_text: str = "READY", quiet: bool = False):
        interrupted_store = self._manual_store
        interrupted_pid = self._manual_point_id
        interrupted_run_id = self._manual_run_id
        invalidate_saved_point = bool(self.point_action_busy and self._manual_save_done and interrupted_store and interrupted_pid)
        if hasattr(self, "_manual_countdown_timer") and self._manual_countdown_timer.isActive():
            self._manual_countdown_timer.stop()
        self._manual_point_token += 1
        self._manual_remaining_ms = 0
        self.point_action_busy = False
        self._manual_active_task = False
        if status_text.startswith("✓"):
            self._update_manual_status(status_text, "recorded")
        elif "OFF" in status_text or status_text == "READY":
            self._update_manual_status(status_text, "ready")
        else:
            self._update_manual_status(status_text, "ready")
        self.update_enabled_states()
        if invalidate_saved_point:
            if self._manual_target_amps in self._manual_recorded_currents:
                self._manual_recorded_currents.remove(self._manual_target_amps)
            self._run_function(
                lambda: interrupted_store.discard_interrupted_point(interrupted_pid, interrupted_run_id, "Interrupted by operator current reduction / LOAD OFF")
            )

    def _manual_action_failed(self, error: str, token: int):
        if token != self._manual_point_token:
            return
        self._manual_active_task = False
        self.point_action_busy = False
        self._update_manual_status(f"Error: {error}", "error")
        self.update_enabled_states()

    def _cancel_step_automation(self, status_text: str = "READY", quiet: bool = False):
        self._cancel_manual_automation(status_text, quiet)

    @staticmethod
    def _run_number(value: float) -> str:
        return f"{float(value):g}".replace("-", "neg").replace(".", "p")

    def _base_test_name(self) -> str:
        campaign = self.test_name.text().strip()
        if not campaign:
            campaign = f"Test_{datetime.now().strftime('%Y%m%d_%H%M')}"
            self.test_name.setText(campaign)
        return campaign

    def _manual_label(self, mode_name: str, amps: float | None = None) -> str:
        base = self._manual_base_campaign or self._base_test_name()
        if mode_name == "Set Current":
            return f"{base}_SetCurrent_{self._run_number(amps or 0.0)}A"
        currents = self._manual_recorded_currents
        if currents:
            return f"{base}_StepCurrent_{self._run_number(min(currents))}to{self._run_number(max(currents))}A"
        return f"{base}_StepCurrent_session"

    def _begin_manual_session(self, mode_name: str, amps: float):
        self._manual_mode_name = mode_name
        self._manual_base_campaign = self._base_test_name()
        self._manual_recorded_currents = []
        self._manual_any_invalid = False
        self._manual_run_created = False
        self._manual_point_index = -1
        self._manual_run_label = self._manual_label(mode_name, amps)
        self._manual_run_id = new_run_id(self._manual_run_label)
        self._manual_store = self._store_for_source("Simulation" if self.simulation.isChecked() else "Hardware")
        self._clear_live_run_view()

    def _finalize_manual_session(self, reason: str = "Manual run closed", status_override: str | None = None):
        run_id = self._manual_run_id
        if run_id and self._manual_run_created and self._manual_store is not None:
            final_label = self._manual_label(self._manual_mode_name, self._manual_target_amps)
            status = status_override or ("Invalid" if self._manual_any_invalid else "Valid")
            try:
                self._manual_store.update_run_fields(run_id, {"CampaignName": final_label})
                self._manual_store.finish_run(run_id, status, reason)
                self._load_history()
            except Exception as exc:
                self.statusBar().showMessage(f"Run close could not be saved: {exc}", 8000)
        self._manual_run_id = ""
        self._manual_point_id = ""
        self._manual_point_index = -1
        self._manual_run_created = False
        self._manual_run_label = ""
        self._manual_base_campaign = ""
        self._manual_recorded_currents = []
        self._manual_any_invalid = False
        self._manual_store = None

    def _start_manual_point_sequence(self, new_val: float, mode_name: str, step_action: str = "ascending_measurement"):
        self._cancel_manual_automation("PREPARING", quiet=True)
        self._manual_point_token += 1
        current_token = self._manual_point_token
        self.point_action_busy = True
        self._manual_step_action = step_action
        self._manual_target_amps = new_val
        self._manual_save_done = False
        self._manual_capture_done = False
        self._manual_last_record = None
        if not self._manual_run_id or self._manual_mode_name != mode_name or mode_name == "Set Current":
            self._finalize_manual_session("New manual run started")
            self._begin_manual_session(mode_name, new_val)
        self._manual_mode_name = mode_name
        self._manual_point_index += 1
        self._manual_point_id = point_id(self._manual_run_id, self._manual_point_index)
        self.update_enabled_states()

        load = self.hub.instruments["load"]
        def cmd():
            load.set_current(new_val)
            load.set_input(True)

        def on_set_done(_):
            if current_token != self._manual_point_token:
                return
            self.statusBar().showMessage(f"Load set to {new_val:g} A · ON")
            if self._manual_save_done or self._manual_capture_done or getattr(self, "_manual_active_task", False):
                return
            if self._manual_step_action != "ascending_measurement" and self._manual_mode_name != "Set Current":
                self._manual_point_completed()
                return
            is_direct = (self.mode_group.checkedId() == 0)
            auto_save = self.direct_auto_save.isChecked() if is_direct else self.step_auto_save.isChecked()
            auto_cap = self.direct_auto_capture.isChecked() if is_direct else self.step_auto_capture.isChecked()
            delay_sec = self.direct_auto_delay.value() if is_direct else self.step_auto_delay.value()

            if not auto_save and not auto_cap:
                self._manual_point_completed()
                return

            self._manual_remaining_ms = int(round(delay_sec * 1000))
            self._update_manual_status_countdown()
            self._manual_countdown_timer.start()
            self.update_enabled_states()

        self._run_function(cmd, on_set_done, lambda error: self._manual_action_failed(error, current_token))

    def _update_manual_status_countdown(self):
        sec = max(0.0, self._manual_remaining_ms / 1000.0)
        self._update_manual_status(f"SETTLING · {sec:.1f} s", "settling")

    def _update_step_status_countdown(self):
        self._update_manual_status_countdown()

    def _manual_countdown_tick(self):
        self._manual_remaining_ms -= 100
        if self._manual_remaining_ms > 0:
            self._update_manual_status_countdown()
        else:
            self._manual_countdown_timer.stop()
            self._manual_execute_next_auto_action()

    def _step_countdown_tick(self):
        self._manual_countdown_tick()

    def _manual_execute_next_auto_action(self):
        current_token = self._manual_point_token
        if self._manual_step_action != "ascending_measurement" and self._manual_mode_name != "Set Current":
            self._manual_point_completed()
            return
        is_direct = (self.mode_group.checkedId() == 0)
        auto_save = self.direct_auto_save.isChecked() if is_direct else self.step_auto_save.isChecked()
        auto_cap = self.direct_auto_capture.isChecked() if is_direct else self.step_auto_capture.isChecked()

        if auto_save and not self._manual_save_done:
            self._execute_manual_measurement(is_auto=True, token=current_token)
        elif auto_cap and not self._manual_capture_done:
            self._execute_manual_capture(is_auto=True, token=current_token)
        else:
            self._manual_point_completed()

    def _step_execute_next_auto_action(self):
        self._manual_execute_next_auto_action()

    def _execute_manual_measurement(self, is_auto: bool = True, token: int | None = None):
        if token is None:
            token = self._manual_point_token
        if hasattr(self, "_manual_countdown_timer") and self._manual_countdown_timer.isActive():
            self._manual_countdown_timer.stop()

        self.point_action_busy = True
        self._manual_active_task = True
        self._update_manual_status("SAVING READING", "saving")
        self.update_enabled_states()

        try:
            settings = self._collect_settings(manual=True)
        except Exception as exc:
            self._manual_action_failed(str(exc), token)
            return

        run_id = self._manual_run_id or settings["run_id"]
        run_rec = dict(settings["run_record"])
        run_rec["RunID"] = run_id
        run_rec["Mode"] = self._manual_mode_name
        run_rec["CampaignName"] = self._manual_run_label or self._manual_label(self._manual_mode_name, self._manual_target_amps)
        run_rec["ModulationLabel"] = self._manual_base_campaign or settings["modulation"]
        store = self._store_for_source(settings["data_source"])
        amps = self._manual_target_amps if self._manual_target_amps > 0 else self.manual_current.value()
        pid = self._manual_point_id or point_id(run_id, 0)
        supersede = self._manual_save_done  # If already saved for this point instance, supersede!

        def task():
            pa = self.hub.instruments["pa"].read_snapshot()
            load = self.hub.instruments["load"].read_snapshot(include_voltage=False)
            try:
                psu = self.hub.instruments["psu"].read_snapshot(channels=[1, 2, 3])
            except Exception:
                psu = None
            if token != self._manual_point_token:
                return {}

            # Evaluate Vin safety shutdown if PA snapshot is valid and numeric Vin is present
            pa_vin = pa.values.get("vin") if pa and pa.valid else None
            if pa_vin is not None and isinstance(pa_vin, (int, float)) and math.isfinite(float(pa_vin)):
                is_safe, vin_fault_desc = check_vin_safety(
                    float(pa_vin),
                    float(settings["vin_target"]),
                    settings.get("vin_safety_enabled", True),
                )
                if not is_safe:
                    try:
                        self.hub.instruments["load"].safe_off()
                    except Exception:
                        pass
                    return {
                        "vin_safety_tripped": True,
                        "target_vin": float(settings["vin_target"]),
                        "measured_vin": float(pa_vin),
                        "fault_desc": vin_fault_desc,
                    }

            store.create_run(run_rec)
            derived, warnings = calculate_measurement(pa, load, psu, settings["supply_channels"])
            record = {
                "PointID": pid,
                "RunID": run_id,
                "Timestamp": utc_now(),
                "Status": "Valid" if derived else "Invalid",
                "DataSource": settings["data_source"],
                "Mode": self._manual_mode_name,
                "VinTarget_V": settings["vin_target"],
                "ModulationLabel": self._manual_base_campaign or settings["modulation"],
                "Frequency_Hz": settings["frequency"],
                "RequestedIout_A": amps,
                **derived,
                "Quality": "Valid" if derived and not warnings else "Warning",
                "Warning": "; ".join(warnings),
                "ScopeCaptureStatus": "Captured" if self._manual_capture_done else "Skipped",
            }
            dup_action = "supersede" if supersede else "keep"
            store.append_measurement(record, duplicate_action=dup_action)
            return record

        def on_done(record: dict[str, Any]):
            if record and record.get("vin_safety_tripped"):
                self._manual_point_token += 1
                self._manual_active_task = False
                self.point_action_busy = False
                self._cancel_manual_automation("ABORTED", quiet=True)
                self.step_present_lbl.setText("0.00 A · OFF")
                self.step_present_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
                self._manual_target_current = 0.0
                self.manual_target_spin.setValue(0.0)
                if hasattr(self, "plot_progress_marker"):
                    self.plot_progress_marker.update_position(0.0, 0.0, self.manual_mode_max_current(), active=False)
                self._update_manual_status("ABORTED · Vin Safety", "error")
                self.statusBar().showMessage("Vin safety shutdown · LOAD OFF", 10000)
                self._finalize_manual_session(f"Vin safety shutdown; {record['fault_desc'].replace(chr(10), '; ')}")
                self.show_vin_safety_dialog(record["target_vin"], record["measured_vin"])
                self.update_enabled_states()
                return

            if token != self._manual_point_token:
                if record:
                    self._run_function(lambda: store.discard_interrupted_point(pid, run_id, "Interrupted by operator current reduction / LOAD OFF"))
                return
            if not record:
                self._manual_action_failed("Measurement was cancelled before it could be saved", token)
                return
            self._manual_active_task = False
            self._manual_save_done = True
            self._manual_run_created = True
            self._manual_last_record = record
            if amps not in self._manual_recorded_currents:
                self._manual_recorded_currents.append(amps)
            self._manual_any_invalid = self._manual_any_invalid or record.get("Status") != "Valid"
            self._measurement_received(record)

            is_direct = (self.mode_group.checkedId() == 0)
            auto_cap = self.direct_auto_capture.isChecked() if is_direct else self.step_auto_capture.isChecked()
            if auto_cap and not self._manual_capture_done:
                self._execute_manual_capture(is_auto=True, token=token)
            else:
                self._manual_point_completed()

        self._run_function(task, on_done, lambda error: self._manual_action_failed(error, token))

    def _execute_step_measurement(self, is_auto: bool = True, token: int | None = None):
        self._execute_manual_measurement(is_auto, token)

    def _execute_manual_capture(self, is_auto: bool = True, token: int | None = None):
        if token is None:
            token = self._manual_point_token
        if hasattr(self, "_manual_countdown_timer") and self._manual_countdown_timer.isActive():
            self._manual_countdown_timer.stop()

        self.point_action_busy = True
        self._manual_active_task = True
        self._update_manual_status("CAPTURING SCOPE", "saving")
        self.update_enabled_states()

        try:
            settings = self._collect_settings(manual=True)
        except Exception as exc:
            self._manual_action_failed(str(exc), token)
            return

        run_id = self._manual_run_id or settings["run_id"]
        run_rec = dict(settings["run_record"])
        run_rec["RunID"] = run_id
        run_rec["Mode"] = self._manual_mode_name
        run_rec["CampaignName"] = self._manual_run_label or self._manual_label(self._manual_mode_name, self._manual_target_amps)
        run_rec["ModulationLabel"] = self._manual_base_campaign or settings["modulation"]
        store = self._store_for_source(settings["data_source"])
        pid = self._manual_point_id or point_id(run_id, 0)
        capture_dir = capture_root_for_source(store.path, settings["data_source"])
        capture_dir.mkdir(parents=True, exist_ok=True)
        png = capture_dir / f"{pid}.png"
        csv_f = capture_dir / f"{pid}.csv"
        amps = self._manual_target_amps if self._manual_target_amps > 0 else self.manual_current.value()

        def task():
            if token != self._manual_point_token:
                return {}
            status, error = "Captured", ""
            try:
                self.hub.instruments["scope"].capture(png, csv_f)
            except Exception as exc:
                status, error = "Failed", str(exc)
                for artifact in (png, csv_f):
                    if artifact.exists():
                        artifact.unlink()
            if token != self._manual_point_token:
                for artifact in (png, csv_f):
                    if artifact.exists():
                        artifact.unlink()
                return {}
            store.create_run(run_rec)
            record = {
                "PointID": pid,
                "RunID": run_id,
                "Timestamp": utc_now(),
                "Status": "Valid" if status == "Captured" else "Invalid",
                "DataSource": settings["data_source"],
                "Mode": self._manual_mode_name,
                "VinTarget_V": settings["vin_target"],
                "ModulationLabel": self._manual_base_campaign or settings["modulation"],
                "Frequency_Hz": settings["frequency"],
                "RequestedIout_A": amps,
                "Quality": "Capture only" if not self._manual_save_done else "Valid",
                "Warning": error,
                "ScopeCaptureStatus": status,
                "ScopeCaptureError": error,
                "ScopePNG": str(png) if status == "Captured" else "",
                "ScopeCSV": str(csv_f) if status == "Captured" else "",
            }
            if not self._manual_save_done:
                store.append_measurement(record)
            else:
                store.update_measurement_scope(
                    pid, status, error, str(png) if status == "Captured" else "", str(csv_f) if status == "Captured" else ""
                )
            return record

        def on_done(record: dict[str, Any]):
            if token != self._manual_point_token:
                if record:
                    self._run_function(lambda: store.discard_interrupted_point(pid, run_id, "Interrupted by operator current reduction / LOAD OFF"))
                return
            if not record:
                self._manual_action_failed("Scope capture was cancelled", token)
                return
            self._manual_active_task = False
            self._manual_capture_done = True
            self._manual_run_created = True
            if amps not in self._manual_recorded_currents:
                self._manual_recorded_currents.append(amps)
            self._manual_any_invalid = self._manual_any_invalid or record.get("Status") != "Valid"
            self.statusBar().showMessage(f"Scope capture {record.get('ScopeCaptureStatus', 'Done')}")

            is_direct = (self.mode_group.checkedId() == 0)
            auto_save = self.direct_auto_save.isChecked() if is_direct else self.step_auto_save.isChecked()
            if auto_save and not self._manual_save_done:
                self._execute_manual_measurement(is_auto=True, token=token)
            else:
                self._manual_point_completed()

        self._run_function(task, on_done, lambda error: self._manual_action_failed(error, token))

    def _execute_step_capture(self, is_auto: bool = True, token: int | None = None):
        self._execute_manual_capture(is_auto, token)

    def _manual_point_completed(self):
        self._manual_active_task = False
        self.point_action_busy = False
        amps = self._manual_target_amps if self._manual_target_amps > 0 else self.manual_current.value()
        if amps > 0 and (self._manual_step_action == "ascending_measurement" or self._manual_mode_name == "Set Current"):
            self._update_manual_status(f"✓ RECORDED · {amps:.2f} A", "recorded")
        elif amps > 0:
            self._update_manual_status(f"READY · {amps:.2f} A", "ready")
        else:
            self._update_manual_status("READY", "ready")
        if self._manual_mode_name == "Set Current":
            self._finalize_manual_session("Set current point completed")
        self.update_enabled_states()

    def _step_point_completed(self):
        self._manual_point_completed()

    def _manual_save_direct_override(self):
        self._manual_save_override("Set Current")

    def _manual_capture_direct_override(self):
        self._manual_capture_override("Set Current")

    def _manual_save_step_override(self):
        self._manual_save_override("Step Current")

    def _manual_capture_step_override(self):
        self._manual_capture_override("Step Current")

    def _manual_save_override(self, mode_name: str = "Set Current"):
        if hasattr(self, "_manual_countdown_timer") and self._manual_countdown_timer.isActive():
            self._manual_countdown_timer.stop()
        self._manual_step_action = "ascending_measurement"
        self._manual_mode_name = mode_name
        if not self._manual_run_id:
            self._manual_target_amps = self.manual_current.value()
            self._begin_manual_session(mode_name, self._manual_target_amps)
            self._manual_point_index = 0
            self._manual_point_id = point_id(self._manual_run_id, self._manual_point_index)
        self._execute_manual_measurement(is_auto=False)

    def _manual_capture_override(self, mode_name: str = "Set Current"):
        if hasattr(self, "_manual_countdown_timer") and self._manual_countdown_timer.isActive():
            self._manual_countdown_timer.stop()
        self._manual_step_action = "ascending_measurement"
        self._manual_mode_name = mode_name
        if not self._manual_run_id:
            self._manual_target_amps = self.manual_current.value()
            self._begin_manual_session(mode_name, self._manual_target_amps)
            self._manual_point_index = 0
            self._manual_point_id = point_id(self._manual_run_id, self._manual_point_index)
        self._execute_manual_capture(is_auto=False)

    def _manual_measure(self):
        self._manual_save_direct_override()

    def _capture_now(self):
        self._manual_capture_direct_override()

    def _manual_direct_set(self):
        max_safe = self.manual_mode_max_current()
        amps = min(max_safe, max(0.0, self.manual_target_spin.value()))
        self.manual_target_spin.setValue(amps)
        if amps <= 0.0:
            self._manual_step_action = "zero_off"
            self._cancel_manual_automation("READY")
            self._finalize_manual_session("ZERO / OFF")
            self._clear_live_run_view()
            self._manual_target_current = 0.0
            self.manual_target_spin.setValue(0.0)
            if hasattr(self, "plot_progress_marker"):
                self.plot_progress_marker.update_position(0.0, 0.0, max_safe, active=False)
            self.step_present_lbl.setText("0.00 A · OFF")
            self.step_present_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
            self._manual_active_task = True
            safety_token = self._manual_point_token
            def cmd():
                load = self.hub.instruments["load"]
                load.safe_off()
            def on_done(_):
                self._manual_active_task = False
                self.statusBar().showMessage("Load set to 0.00 A · OFF")
                self._update_manual_status("READY", "ready")
                self.update_enabled_states()
            self._run_function(cmd, on_done, lambda error: self._manual_action_failed(error, safety_token))
        else:
            if not self.require_load_control_verified("Set Current") or not self._validate_current(amps, current_value=self._manual_target_current):
                return
            self._manual_target_current = amps
            if hasattr(self, "plot_progress_marker"):
                self.plot_progress_marker.update_position(amps, 0.0, max_safe, active=(amps > 0.0))
            self._manual_step_action = "ascending_measurement"
            self.step_present_lbl.setText(f"{amps:.2f} A · ON")
            self.step_present_lbl.setStyleSheet(f"color: {BERKELEY_BLUE}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
            self._start_manual_point_sequence(amps, "Set Current", step_action="ascending_measurement")

    def _manual_direct_zero(self):
        self._manual_step_action = "zero_off"
        self._cancel_manual_automation("READY")
        self._finalize_manual_session("ZERO / OFF")
        self._clear_live_run_view()
        self.manual_target_spin.setValue(0.0)
        self._manual_target_current = 0.0
        if hasattr(self, "plot_progress_marker"):
            self.plot_progress_marker.update_position(0.0, 0.0, self.manual_mode_max_current(), active=False)
        self.step_present_lbl.setText("0.00 A · OFF")
        self.step_present_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
        self._update_manual_status("READY", "ready")
        self._manual_active_task = True
        safety_token = self._manual_point_token
        load = self.hub.instruments["load"]
        def cmd():
            load.safe_off()
        def on_done(_):
            self._manual_active_task = False
            self.statusBar().showMessage("Load zeroed (0.00 A · OFF)")
            self._update_manual_status("READY", "ready")
            self.update_enabled_states()
        self._run_function(cmd, on_done, lambda error: self._manual_action_failed(error, safety_token))

    def _step_delta(self, direction: int):
        max_safe = self.manual_mode_max_current()
        if direction > 0:
            delta = self.manual_step_inc.value()
            new_val = min(max_safe, self._manual_target_current + delta)
            if new_val == self._manual_target_current and new_val > 0:
                self.statusBar().showMessage(f"Load already at maximum safe limit ({max_safe:g} A)", 4000)
                return
            if not self.require_load_control_verified("Step Current") or not self._validate_current(new_val, current_value=self._manual_target_current):
                return
            self._manual_step_action = "ascending_measurement"
            self._manual_target_current = new_val
            self.manual_target_spin.setValue(new_val)
            if hasattr(self, "plot_progress_marker"):
                self.plot_progress_marker.update_position(new_val, 0.0, max_safe, active=(new_val > 0.0))
            self.step_present_lbl.setText(f"{new_val:.2f} A · ON")
            self.step_present_lbl.setStyleSheet(f"color: {BERKELEY_BLUE}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
            self._start_manual_point_sequence(new_val, "Step Current", step_action="ascending_measurement")
        else:
            delta = self.manual_step_dec.value()
            new_val = max(0.0, self._manual_target_current - delta)
            self._manual_target_current = new_val
            self.manual_target_spin.setValue(new_val)
            if hasattr(self, "plot_progress_marker"):
                self.plot_progress_marker.update_position(new_val, 0.0, max_safe, active=(new_val > 0.0))

            load = self.hub.instruments["load"]
            if new_val <= 0.0:
                self._manual_step_action = "zero_off"
                self._cancel_manual_automation("READY")
                self._finalize_manual_session("ZERO / OFF")
                self._clear_live_run_view()
                self.step_present_lbl.setText("0.00 A · OFF")
                self.step_present_lbl.setStyleSheet(f"color: {TEXT_MUTED}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
                self._update_manual_status("READY", "ready")
                self._manual_active_task = True
                safety_token = self._manual_point_token
                def cmd_zero():
                    load.safe_off()
                def on_zero_done(_):
                    self._manual_active_task = False
                    self.statusBar().showMessage("Load stepped to 0.00 A · OFF")
                    self._update_manual_status("READY", "ready")
                    self.update_enabled_states()
                self._run_function(cmd_zero, on_zero_done, lambda error: self._manual_action_failed(error, safety_token))
                return

            if not self.require_load_control_verified("Step Current") or not self._validate_current(new_val, current_value=self._manual_target_current):
                return
            self._manual_step_action = "descending_unload"
            self._cancel_manual_automation(f"READY · {new_val:.2f} A", quiet=True)
            self.step_present_lbl.setText(f"{new_val:.2f} A · ON")
            self.step_present_lbl.setStyleSheet(f"color: {BERKELEY_BLUE}; font-family: Consolas, monospace; font-size: 24px; font-weight: 900;")
            self._update_manual_status(f"RAMPING DOWN · {new_val:.2f} A", "settling")
            self.update_enabled_states()
            self._manual_active_task = True
            safety_token = self._manual_point_token
            def cmd_down():
                load.set_current(new_val)
                load.set_input(True)
            def on_down_done(_):
                self._manual_active_task = False
                self.statusBar().showMessage(f"Load stepped down to {new_val:g} A · ON")
                self._update_manual_status(f"READY · {new_val:.2f} A", "ready")
                self.update_enabled_states()
            self._run_function(cmd_down, on_down_done, lambda error: self._manual_action_failed(error, safety_token))

    def _manual_set(self):
        self._manual_direct_set()

    def _manual_apply_direct(self):
        self._manual_direct_set()

    def _manual_zero(self):
        self._manual_direct_zero()

    def _manual_delta(self, direction: int):
        self._step_delta(direction)

    def _step_up(self):
        self._step_delta(1)

    def _step_down(self):
        self._step_delta(-1)

    @property
    def _step_run_id(self) -> str:
        return self._manual_run_id

    @_step_run_id.setter
    def _step_run_id(self, val: str):
        self._manual_run_id = val

    @property
    def _step_point_id(self) -> str:
        return self._manual_point_id

    @_step_point_id.setter
    def _step_point_id(self, val: str):
        self._manual_point_id = val

    @property
    def _step_save_done(self) -> bool:
        return self._manual_save_done

    @_step_save_done.setter
    def _step_save_done(self, val: bool):
        self._manual_save_done = val

    @property
    def _step_capture_done(self) -> bool:
        return self._manual_capture_done

    @_step_capture_done.setter
    def _step_capture_done(self, val: bool):
        self._manual_capture_done = val

    @property
    def _step_countdown_timer(self) -> QtCore.QTimer:
        return self._manual_countdown_timer

    @property
    def _step_remaining_ms(self) -> int:
        return self._manual_remaining_ms

    @_step_remaining_ms.setter
    def _step_remaining_ms(self, val: int):
        self._manual_remaining_ms = val

    def _collect_settings(self, manual: bool = False) -> dict[str, Any]:
        channels = self._supply_channels()
        campaign = self.test_name.text().strip()
        if not campaign:
            campaign = f"Test_{datetime.now().strftime('%Y%m%d_%H%M')}"
            self.test_name.setText(campaign)
            self.statusBar().showMessage(f"Notice: Auto-generated Test Name '{campaign}'", 4000)

        base_campaign = campaign
        mode_idx = self.mode_group.checkedId()
        if manual:
            mode = self._manual_mode_name
        else:
            mode = ["Set Current", "Step Current", "Continuous", "Pulse"][mode_idx]
        cap = self.cap_val

        if mode in ("Set Current", "Step Current"):
            points = [self.manual_current.value()]
            capture_points = set()
            settle, dwell, s_win, s_cnt, cooldown, psu_req = 3.0, 3.0, 0.5, 1, 3.0, False
            return_step = 5.0
        elif mode == "Continuous":
            points, _ = generate_points(self.cont_start.value(), self.cont_stop.value(), self.cont_step.value(), cap)
            capture_points = parse_capture_points(self.cont_capture_points.text())
            settle = self.cont_settle.value()
            dwell = 3.0
            s_win = self.cont_sample_window.value()
            s_cnt = self.cont_sample_count.value()
            cooldown = 3.0
            psu_req = self.cont_psu_req.isChecked()
            return_step = self.cont_return_step.value()
        else: # Pulse
            points, _ = generate_points(self.pulse_start.value(), self.pulse_stop.value(), self.pulse_step.value(), cap)
            capture_points = parse_capture_points(self.pulse_capture_points.text())
            settle = 3.0
            dwell = self.pulse_dwell.value()
            s_win = self.pulse_sample_window.value()
            s_cnt = self.pulse_sample_count.value()
            cooldown = self.pulse_cooldown.value()
            psu_req = self.pulse_psu_req.isChecked()
            return_step = self.pulse_return_step.value()
            if s_win > dwell:
                raise ValueError("Pulse measurement window cannot exceed dwell time")

        if mode == "Set Current":
            campaign = f"{base_campaign}_SetCurrent_{self._run_number(self._manual_target_amps or self.manual_current.value())}A"
        elif mode == "Step Current":
            campaign = self._manual_run_label or f"{base_campaign}_StepCurrent_session"
        elif mode == "Continuous":
            campaign = (
                f"{base_campaign}_Continuous_{self._run_number(self.cont_start.value())}to"
                f"{self._run_number(self.cont_stop.value())}A_step{self._run_number(self.cont_step.value())}A"
            )
        else:
            campaign = (
                f"{base_campaign}_Pulse_{self._run_number(self.pulse_start.value())}to"
                f"{self._run_number(self.pulse_stop.value())}A_step{self._run_number(self.pulse_step.value())}A_"
                f"on{self._run_number(self.pulse_dwell.value())}s_rest{self._run_number(self.pulse_cooldown.value())}s"
            )

        run_id = new_run_id(campaign)
        frequency_hz = self.frequency_hz()
        fpga_status, fpga_data, fpga_warning = (fpga_snapshot(Path(self.config["fpga_root"]), frequency_hz) if self.fpga_check.isChecked() else ("Unavailable", {}, "Skipped by operator"))
        data_source = "Simulation" if self.simulation.isChecked() else "Hardware"
        identities = {key: getattr(inst, "identity", "") for key, inst in self.hub.instruments.items()}
        supply_json = [asdict(channel) for channel in channels]

        aux_a_inc = any(ch.channel == 1 and ch.contributes_loss for ch in channels)
        aux_b_inc = any(ch.channel == 2 and ch.contributes_loss for ch in channels)
        aux_c_inc = any(ch.channel == 3 and ch.contributes_loss for ch in channels)

        run_record = {
            "RunID": run_id, "CampaignName": campaign, "Created": utc_now(), "Status": "Aborted", "DataSource": data_source, "Mode": mode,
            "VinTarget_V": self.vin_target.value(), "ModulationLabel": base_campaign, "Frequency_Hz": frequency_hz,
            "ModulationMetadata": "",
            "AuxA_Included": aux_a_inc, "AuxB_Included": aux_b_inc, "AuxC_Included": aux_c_inc,
            "SupplyConfiguration": supply_json,
            "WorkingCap_A": cap, "Notes": self.notes.text(), "InstrumentIdentities": identities,
            "FPGASnapshotStatus": fpga_status, "FPGASnapshot": fpga_data, "Warnings": fpga_warning,
        }

        return {
            "run_id": run_id, "run_record": run_record, "points": points, "capture_points": capture_points,
            "mode": mode, "settle": settle, "dwell": dwell, "sample_window": s_win, "sample_count": s_cnt,
            "cooldown": cooldown, "working_cap": cap, "vin_target": self.vin_target.value(),
            "vin_safety_enabled": self.chk_vin_safety.isChecked() if hasattr(self, "chk_vin_safety") else bool(self.config.get("vin_safety_enabled", True)),
            "modulation": base_campaign, "frequency": frequency_hz,
            "supply_channels": channels, "psu_required": psu_req,
            "data_source": data_source, "duplicate_action": "keep",
            "return_to_zero_step": return_step,
        }

    def start_run(self):
        if self.worker and self.worker.isRunning(): return
        mode_idx = self.mode_group.checkedId() if hasattr(self, "mode_group") else 2
        action_name = "Pulse Sweep" if mode_idx == 3 else "Continuous Sweep"
        if not self.require_load_control_verified(action_name): return
        try: settings = self._collect_settings()
        except Exception as exc: QtWidgets.QMessageBox.warning(self, "Run preflight", str(exc)); return
        run_store = self._store_for_source(settings["data_source"])

        try:
            run_store.preflight(require_live_commit=True)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(
                self,
                "Run not started · Workbook unavailable",
                f"The campaign workbook could not accept a validated save. No sweep or load command was started.\n\n{exc}",
            )
            return

        # Validate scope capture points against requested sweep points
        if settings.get("capture_points"):
            unaligned = [
                c for c in sorted(list(settings["capture_points"]))
                if not any(math.isclose(c, p, abs_tol=0.05) for p in settings["points"])
            ]
            if unaligned:
                un_str = ", ".join(f"{x:g} A" for x in unaligned)
                reply = QtWidgets.QMessageBox.warning(
                    self,
                    "Unaligned Scope Capture Currents",
                    f"The following scope capture current(s) do not match any point in the requested sweep points:\n\n"
                    f"  {un_str}\n\n"
                    f"Scope captures will only execute at points included in the sweep.\n"
                    f"Do you want to proceed?",
                    QtWidgets.QMessageBox.StandardButton.Yes | QtWidgets.QMessageBox.StandardButton.No,
                    QtWidgets.QMessageBox.StandardButton.No
                )
                if reply != QtWidgets.QMessageBox.StandardButton.Yes:
                    return

        if not self.simulation.isChecked():
            try:
                pa = self.hub.instruments["pa"].read_snapshot()
                actual = float(pa.values["vin"])
                if abs(actual - settings["vin_target"]) > 1.0:
                    QtWidgets.QMessageBox.warning(self, "Vin preflight failed", f"PA reports {actual:g} V; target is {settings['vin_target']:g} V (±1.0 V).")
                    return
            except Exception as exc:
                QtWidgets.QMessageBox.warning(self, "Vin preflight failed", str(exc)); return

        condition = {"DataSource": settings["data_source"], "Mode": settings["mode"], "VinTarget_V": settings["vin_target"], "ModulationLabel": settings["modulation"], "Frequency_Hz": settings["frequency"]}
        duplicates = run_store.find_duplicates(condition, settings["points"])
        if duplicates:
            box = QtWidgets.QMessageBox(self); box.setWindowTitle("Duplicate points"); box.setText("Existing measurements match: " + ", ".join(f"{item['current']:g} A" for item in duplicates)); supersede = box.addButton("Replace / Supersede", QtWidgets.QMessageBox.ButtonRole.AcceptRole); keep = box.addButton("Keep both", QtWidgets.QMessageBox.ButtonRole.ActionRole); cancel = box.addButton(QtWidgets.QMessageBox.StandardButton.Cancel); box.exec()
            if box.clickedButton() == cancel: return
            settings["duplicate_action"] = "supersede" if box.clickedButton() == supersede else "keep"

        suspicious = any(value >= 100 for value in settings["points"]) or any(abs(b - a) >= 50 for a, b in zip(settings["points"], settings["points"][1:]))
        if suspicious:
            phrase, ok = QtWidgets.QInputDialog.getText(self, "Large sweep command", f"Points contain a current ≥ 100 A or step jump ≥ 50 A. Type RUN MAX {max(settings['points']):g} A")
            if not ok or phrase.strip() != f"RUN MAX {max(settings['points']):g} A": return

        summary = f"Test: {settings['run_record']['CampaignName']}\nMode: {settings['mode']}\nVin: {settings['vin_target']:g} V\nPoints: {', '.join(f'{x:g}' for x in settings['points'])} A\nCap: {settings['working_cap']:g} A\n\nConfirm to execute sweep."
        if QtWidgets.QMessageBox.question(self, "Confirm sweep execution", summary) != QtWidgets.QMessageBox.StandardButton.Yes: return

        self.config["campaign_name"] = self.test_name.text()
        self.config["vin_target_v"] = self.vin_target.value()
        self.config["frequency_hz"] = self.frequency_hz()
        save_config(self.config)

        self.run_strip.setVisible(True)
        self.worker = SweepWorker(self.hub, run_store, settings)
        self.worker.progress.connect(self._run_progress)
        self.worker.ramp_progress.connect(self._ramp_progress_received)
        self.worker.state_changed.connect(self._state_changed)
        self.worker.measurement.connect(self._measurement_received)
        self.worker.warning.connect(lambda msg: self.statusBar().showMessage(msg))
        self.worker.completed.connect(self._run_completed)
        self.worker.vin_safety_tripped.connect(self._on_vin_safety_tripped)

        self.plot_rows.clear()
        self.live_curve.setData([], [])
        self.live_system_curve.setData([], [])
        self.live_aux_curve.setData([], [])
        self.strip_label.setText(f"Status: RUNNING")
        self.strip_label.setStyleSheet(f"color: {SUCCESS_GREEN}; font-weight: 800; font-size: 13px;")
        self.stop_sweep_btn.setText("■  STOP & RETURN TO ZERO")
        self.stop_sweep_btn.setEnabled(True)
        self.strip_stop_btn.setText("■ STOP & RETURN TO ZERO")
        self.strip_stop_btn.setEnabled(True)
        self.worker.start()
        self.update_enabled_states()

    def stop_and_return_to_zero(self):
        if self.worker and self.worker.isRunning():
            self.worker.stop_and_return_to_zero()
            self._state_changed("RETURNING TO ZERO", "Stopping sweep and returning load to 0 A")
            self.statusBar().showMessage("Stopping sweep; gracefully returning current to 0 A...")

    def _state_changed(self, state: str, detail: str = ""):
        self.strip_label.setText(f"Status: {state}")
        color_map = {
            "RUNNING": SUCCESS_GREEN,
            "RETURNING TO ZERO": WARNING_AMBER,
            "STOPPED": "#475569",
            "COMPLETED": SUCCESS_GREEN,
            "ABORTED": DANGER_RED,
        }
        color = color_map.get(state, TEXT_MUTED)
        self.strip_label.setStyleSheet(f"color: {color}; font-weight: 800; font-size: 13px;")

        if hasattr(self, "live_stat_tag"):
            icon = "●" if state in ("RUNNING", "COMPLETED") else ("◌" if state == "RETURNING TO ZERO" else ("⏹" if state == "STOPPED" else "✖"))
            self.live_stat_tag.setText(f"{state} {icon}")
            self.live_stat_tag.setStyleSheet(f"font-weight: 800; font-size: 12px; color: {color};")
            if detail and hasattr(self, "live_cmd_lbl") and state == "RETURNING TO ZERO":
                self.live_cmd_lbl.setText(detail)

        if state == "RETURNING TO ZERO":
            self.stop_sweep_btn.setText("■ RETURNING TO ZERO...")
            self.stop_sweep_btn.setEnabled(False)
            self.strip_stop_btn.setText("■ RETURNING TO ZERO...")
            self.strip_stop_btn.setEnabled(False)

    def _on_vin_safety_tripped(self, target_vin: float, measured_vin: float):
        self._state_changed("ABORTED", "Vin safety shutdown · LOAD OFF")
        self.statusBar().showMessage("Vin safety shutdown · LOAD OFF", 10000)
        self.show_vin_safety_dialog(target_vin, measured_vin)

    def show_vin_safety_dialog(self, target_vin: float, measured_vin: float):
        low = target_vin * 0.90
        high = target_vin * 1.10
        msg = (
            "The electronic load was turned OFF because Vin moved outside\n"
            "±10% of Target Vin.\n\n"
            f"Target Vin: {target_vin:.1f} V\n"
            f"Measured Vin: {measured_vin:.1f} V\n"
            f"Allowed range: {low:.1f}–{high:.1f} V\n\n"
            "Check the input supply current limit, source protection,\n"
            "wiring, and converter condition before restarting."
        )
        QtWidgets.QMessageBox.critical(
            self,
            "INPUT VOLTAGE SAFETY SHUTDOWN",
            msg,
            QtWidgets.QMessageBox.StandardButton.Ok,
        )

    def _ramp_progress_received(self, amps: float):
        if hasattr(self, "live_cmd_lbl"):
            self.live_cmd_lbl.setText(f"Command: {amps:g} A")
        if hasattr(self, "lbl_curr_point"):
            self.lbl_curr_point.setText(f"Ramping: {amps:g} A")
        mode_idx = self.mode_group.checkedId()
        is_pulse = (mode_idx == 3)
        start_val = self.pulse_start.value() if is_pulse else self.cont_start.value()
        stop_val = self.pulse_stop.value() if is_pulse else self.cont_stop.value()
        if hasattr(self, "plot_progress_marker"):
            self.plot_progress_marker.update_position(amps, start_val, stop_val, active=True)

        lbl = self.pulse_summary_lbl if is_pulse else self.cont_summary_lbl
        lbl.setText(f"RETURNING TO ZERO   |   Current command: {amps:g} A")
        lbl.setStyleSheet("color: #B45309; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #FFFBEB; border: 1px solid #FDE68A; border-radius: 4px;")

    def _run_progress(self, current: int, total: int, amps: float, next_amps: float, text: str):
        self.run_progress_bar.setRange(0, total)
        self.run_progress_bar.setValue(current)
        self.run_progress_bar.setFormat(f"{current}/{total} ({amps:g} A)")
        self.strip_progress.setRange(0, total)
        self.strip_progress.setValue(current)
        self.strip_progress.setFormat(f"{current}/{total} ({amps:g} A)")
        self.lbl_curr_point.setText(f"Current: {amps:g} A ({current}/{total})")
        self.lbl_next_point.setText(f"Next: {next_amps:g} A" if next_amps else "Next: Done")
        self.strip_next.setText(f"Next: {next_amps:g} A" if next_amps else "")

        mode_idx = self.mode_group.checkedId()
        is_pulse = (mode_idx == 3)
        start_val = self.pulse_start.value() if is_pulse else self.cont_start.value()
        stop_val = self.pulse_stop.value() if is_pulse else self.cont_stop.value()

        # Compact live status above plot
        if hasattr(self, "live_stat_tag"):
            self.live_stat_tag.setText("RUNNING ●")
            self.live_stat_tag.setStyleSheet(f"font-weight: 800; font-size: 12px; color: {SUCCESS_GREEN};")
            self.live_point_lbl.setText(f"Point {current} / {total}")
            self.live_cmd_lbl.setText(f"Command: {amps:g} A")
            if hasattr(self, "plot_progress_marker"):
                self.plot_progress_marker.update_position(amps, start_val, stop_val, active=True)

        # Update compact status strip below controls
        lbl = self.pulse_summary_lbl if is_pulse else self.cont_summary_lbl
        if is_pulse:
            dwell = self.pulse_dwell.value()
            cooldown = self.pulse_cooldown.value()
            rem_s = max(0, (total - current)) * (dwell + cooldown)
        else:
            settle = self.cont_settle.value()
            rem_s = max(0, (total - current)) * settle
        lbl.setText(f"RUNNING   |   Point {current}/{total}   |   {amps:g} A   |   ~{rem_s:.0f} s remaining")
        lbl.setStyleSheet("color: #1E40AF; font-weight: 700; font-size: 12px; padding: 4px 8px; background: #EFF6FF; border: 1px solid #BFDBFE; border-radius: 4px;")

    def _measurement_received(self, record: dict[str, Any]):
        self.plot_rows.append(record)
        self._switch_live_plot(self.live_metric_combo.currentIndex())
        if isinstance(record.get("Vin_V"), (int, float)):
            self.kpi_labels["Vin"].setText(f"{record['Vin_V']:.2f} V")
        if isinstance(record.get("Iin_A"), (int, float)):
            iin_val = record["Iin_A"]
            self.kpi_labels["Iin"].setText("0.000 A" if abs(iin_val) < 0.0005 else f"{iin_val:.3f} A")
        if isinstance(record.get("Vout_V"), (int, float)):
            self.kpi_labels["Vout"].setText(f"{record['Vout_V']:.2f} V")
        if isinstance(record.get("Iout_A"), (int, float)):
            iout_val = record["Iout_A"]
            self.kpi_labels["Iout"].setText(f"{iout_val:.2f} A · ON" if iout_val > 0.001 else "0.00 A · OFF")
            if hasattr(self, "live_act_lbl"):
                self.live_act_lbl.setText(f"Actual: {iout_val:.2f} A")
        self._update_derived_kpis(record)
        record_store = self._store_for_source(record.get("DataSource", "Hardware"))
        if record_store.last_warning:
            self.statusBar().showMessage(record_store.last_warning)

    def _run_completed(self, status: str, warning: str):
        disp_status = "COMPLETED" if status == "Valid" else ("STOPPED" if status == "Stopped" else "ABORTED")
        color = SUCCESS_GREEN if status == "Valid" else (TEXT_MUTED if status == "Stopped" else DANGER_RED)
        self.strip_label.setText(f"Status: {disp_status}")
        self.strip_label.setStyleSheet(f"color: {color}; font-weight: 800; font-size: 13px;")
        self.run_progress_bar.setFormat(f"{disp_status}" + (f" · {warning}" if warning else ""))
        if hasattr(self, "live_stat_tag"):
            icon = "●" if status == "Valid" else ("⏹" if status == "Stopped" else "✖")
            self.live_stat_tag.setText(f"{disp_status} {icon}")
            self.live_stat_tag.setStyleSheet(f"font-weight: 800; font-size: 12px; color: {color};")
        if hasattr(self, "plot_progress_marker"):
            self.plot_progress_marker.set_idle()
        self.stop_sweep_btn.setText("■  STOP & RETURN TO ZERO")
        self.strip_stop_btn.setText("■ STOP & RETURN TO ZERO")
        self._update_sweep_summary()
        self._load_history()
        self.update_enabled_states()

    def _start_demo_run(self):
        """Interactive simulated sweep preview running step-by-step with real-time UI animation."""
        if self.demo_timer and self.demo_timer.isActive():
            self.demo_timer.stop()
        self.demo_points = [0.0, 10.0, 20.0, 30.0, 40.0, 50.0, 60.0]
        self.demo_index = 0
        self.plot_rows.clear()
        self.live_curve.setData([], [])
        self.live_system_curve.setData([], [])
        self.live_aux_curve.setData([], [])
        self.tabs.setCurrentIndex(1)  # Switch to Run tab
        self.run_strip.setVisible(True)
        self.strip_label.setText("SIMULATION DEMO: Generic converter profile")
        self.strip_label.setStyleSheet(f"color: {WARNING_AMBER}; font-weight: 700;")

        self.demo_timer = QtCore.QTimer(self)
        self.demo_timer.timeout.connect(self._demo_tick)
        self.demo_timer.start(350)
        self.update_enabled_states()

    def _demo_tick(self):
        if self.demo_index >= len(self.demo_points):
            self.demo_timer.stop()
            self._demo_completed("Valid")
            return

        amps = self.demo_points[self.demo_index]
        next_amps = self.demo_points[self.demo_index + 1] if self.demo_index + 1 < len(self.demo_points) else 0.0
        self.demo_index += 1

        vin = 48.0
        vout = max(11.5, 12.05 - (amps * 0.005))
        eff = 90.0 + (10.0 * (amps / 60.0)) - (6.0 * ((amps / 60.0) ** 2)) if amps > 0 else 0.0
        pin = (vout * amps) / (eff / 100.0) if eff > 0 else 2.0
        loss = pin - (vout * amps)

        rec = {
            "PointID": f"DEMO-P{self.demo_index:03d}", "RunID": "DEMO-RUN", "Timestamp": utc_now(),
            "Status": "Valid", "DataSource": "Simulation", "Mode": "Continuous", "VinTarget_V": 48.0,
            "ModulationLabel": "Generic converter demo", "Frequency_Hz": 100000.0, "RequestedIout_A": amps,
            "Iout_A": amps, "Vin_V": vin, "Vout_V": vout, "Iin_A": pin / vin, "PinConverter_W": pin,
            "Pout_W": vout * amps, "Paux_W": 0.15, "LossConverter_W": loss, "LossSystem_W": loss + 0.15,
            "EfficiencyConverter_pct": eff, "EfficiencySystem_pct": eff * 0.995,
        }

        self._run_progress(self.demo_index, len(self.demo_points), amps, next_amps, "Simulation demo")
        self._measurement_received(rec)
        self.kpi_labels["Vin"].setText(f"{vin:.2f} V")
        self.kpi_labels["Iin"].setText(f"{pin / vin:.3f} A" if amps > 0 else "0.000 A")
        self.kpi_labels["Vout"].setText(f"{vout:.2f} V")
        self.kpi_labels["Iout"].setText(f"{amps:.2f} A · ON" if amps > 0 else "0.00 A · OFF")

    def _demo_completed(self, status: str):
        self.strip_label.setText(f"DEMO COMPLETED ({status})")
        self.strip_label.setStyleSheet(f"color: {SUCCESS_GREEN if status == 'Valid' else DANGER_RED}; font-weight: 700;")
        self.run_progress_bar.setFormat("Demo completed")
        if hasattr(self, "live_stat_tag"):
            self.live_stat_tag.setText("COMPLETE ●")
            self.live_stat_tag.setStyleSheet(f"font-weight: 800; font-size: 12px; color: {SUCCESS_GREEN};")
        if hasattr(self, "plot_progress_marker"):
            self.plot_progress_marker.set_idle()
        self.update_enabled_states()



    def _load_history(self):
        try:
            stores = []
            if getattr(self, "store", None) is not None:
                stores.append(self.store)
            if hasattr(self, "hardware_store") and self.hardware_store not in stores:
                stores.append(self.hardware_store)
            if hasattr(self, "simulation_store") and self.simulation_store not in stores:
                stores.append(self.simulation_store)
            runs = []
            seen_ids = set()
            for s in stores:
                if s is None: continue
                try:
                    for r in s.list_runs():
                        rid = r.get("RunID")
                        if rid and rid not in seen_ids:
                            seen_ids.add(rid)
                            runs.append(r)
                except Exception:
                    pass
            runs.sort(key=lambda item: str(item.get("Created") or ""))
        except Exception as exc:
            self.statusBar().showMessage(str(exc))
            return
        self.history_table.setRowCount(len(runs))
        for row, run in enumerate(reversed(runs)):
            run_id = str(run.get("RunID", ""))
            short_id = run_id.split("-")[-1] if "-" in run_id else (run_id[-5:] if len(run_id) >= 5 else run_id)
            item_short = QtWidgets.QTableWidgetItem(short_id)
            item_short.setData(QtCore.Qt.ItemDataRole.UserRole, run_id)
            item_short.setFont(QtGui.QFont("Consolas", 9, QtGui.QFont.Weight.Bold))
            self.history_table.setItem(row, 0, item_short)

            cols = (
                ("RunID", run_id),
                ("CampaignName", str(run.get("CampaignName", ""))),
                ("Status", str(run.get("Status", ""))),
                ("DataSource", str(run.get("DataSource", ""))),
                ("VinTarget_V", str(run.get("VinTarget_V", ""))),
                ("Frequency_Hz", format_frequency_khz(run.get("Frequency_Hz"))),
                ("ModulationLabel", str(run.get("ModulationLabel", ""))),
                ("Mode", str(run.get("Mode", ""))),
            )
            for c_idx, (_, val) in enumerate(cols, start=1):
                self.history_table.setItem(row, c_idx, QtWidgets.QTableWidgetItem(val))

    def _history_delete(self):
        selected_rows = sorted(list({idx.row() for idx in self.history_table.selectedIndexes()}))
        if len(selected_rows) == 0:
            QtWidgets.QMessageBox.information(self, "Delete Runs", "Select one or more runs to delete.")
            return

        # Collect unique full RunIDs and short IDs
        selected_run_ids: list[str] = []
        short_id_map: dict[str, str] = {}
        store_map: dict[str, WorkbookStore] = {}
        for row in selected_rows:
            short_id = self.history_table.item(row, 0).text() if self.history_table.item(row, 0) else ""
            run_id = self.history_table.item(row, 0).data(QtCore.Qt.ItemDataRole.UserRole)
            if not run_id and self.history_table.item(row, 1):
                run_id = self.history_table.item(row, 1).text()
            run_id = str(run_id or "").strip()
            if run_id and run_id not in selected_run_ids:
                selected_run_ids.append(run_id)
                short_id_map[run_id] = short_id or (run_id.split("-")[-1] if "-" in run_id else run_id)
                store_map[run_id] = self._store_for_history_row(row)

        if not selected_run_ids:
            QtWidgets.QMessageBox.warning(self, "Delete Runs", "Could not identify RunID for selected rows.")
            return

        try:
            stores = []
            if getattr(self, "store", None) is not None:
                stores.append(self.store)
            if hasattr(self, "hardware_store") and self.hardware_store not in stores:
                stores.append(self.hardware_store)
            if hasattr(self, "simulation_store") and self.simulation_store not in stores:
                stores.append(self.simulation_store)
            runs = []
            seen_ids = set()
            for s in stores:
                if s is None: continue
                for r in s.list_runs():
                    rid = r.get("RunID")
                    if rid and rid not in seen_ids:
                        seen_ids.add(rid)
                        runs.append(r)
        except Exception as exc:
            QtWidgets.QMessageBox.critical(self, "Delete Runs", f"Error reading workbook: {exc}")
            return

        # Runs sheet is authoritative. A stale UI row is harmless: refresh it and
        # idempotently remove any orphan measurements/capture references.
        matching_runs_map = {str(r.get("RunID", "")).strip(): r for r in runs if str(r.get("RunID", "")).strip() in selected_run_ids}
        matching_runs = [matching_runs_map[rid] for rid in selected_run_ids if rid in matching_runs_map]

        def delete_selected() -> bool:
            try:
                grouped: dict[WorkbookStore, list[str]] = {}
                for rid in selected_run_ids:
                    grouped.setdefault(store_map[rid], []).append(rid)
                for selected_store, run_ids in grouped.items():
                    selected_store.delete_runs(run_ids)
            except Exception as del_err:
                QtWidgets.QMessageBox.critical(self, "Delete Runs Error", str(del_err))
                self._load_history()
                return False
            self._load_history()
            self._history_selection_changed()
            self.history_table.clearSelection()
            return True

        if not matching_runs:
            if delete_selected():
                self.statusBar().showMessage("Selected run was already removed; History refreshed")
        elif len(selected_run_ids) == 1:
            dialog = DeleteRunDialog(matching_runs[0], self)
            if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
                if delete_selected():
                    short_str = short_id_map.get(selected_run_ids[0], selected_run_ids[0])
                    self.statusBar().showMessage(f"Permanently deleted run {short_str}")
        else:
            # Multi-run batch
            dialog = DeleteBatchRunsDialog(matching_runs, self)
            if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
                if delete_selected():
                    self.statusBar().showMessage(f"Permanently deleted {len(selected_run_ids)} runs")

    def closeEvent(self, event: QtGui.QCloseEvent):
        app = QtWidgets.QApplication.instance()
        if app and hasattr(self, "wheel_filter"):
            try:
                app.removeEventFilter(self.wheel_filter)
            except Exception:
                pass
        if hasattr(self, "age_timer") and self.age_timer and self.age_timer.isActive():
            self.age_timer.stop()
        if hasattr(self, "demo_timer") and self.demo_timer and self.demo_timer.isActive():
            self.demo_timer.stop()
        if self.worker and self.worker.isRunning():
            self.worker.cancel()
            self.worker.wait(5000)
        QtCore.QThreadPool.globalInstance().waitForDone(1000)
        QtWidgets.QApplication.processEvents()
        self._finalize_manual_session("Application closed", status_override="Stopped")
        self.hub.safe_shutdown()
        self.config["frequency_hz"] = self.frequency_hz()
        save_config(self.config)
        event.accept()




def main() -> int:
    app = QtWidgets.QApplication(sys.argv)
    apply_forced_light_theme(app)
    app.setApplicationName("Kickstart Bench")
    window = MainWindow()
    window.show()
    return app.exec()


if __name__ == "__main__":
    raise SystemExit(main())
